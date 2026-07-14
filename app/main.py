import hashlib
import html
import re
import json
import os
import uuid
import unicodedata
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from urllib.parse import urlparse
from typing import Any, Dict, Optional
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import psycopg
from minio import Minio
from psycopg.types.json import Json
from fastapi import FastAPI, Header, HTTPException, Request, UploadFile, File, Form
from pydantic import BaseModel, Field


APP_NAME = "Construction Science Agents API"
APP_VERSION = "0.1.0"

# BASELINE VISUAL APROVADA — OBRA_CAIO_SUM_V1.
# Não alterar múltiplos blocos visuais na mesma rodada.
# Ajustes futuros devem ser feitos primeiro em scripts/dev/preview_placa_template.py e só depois migrados para este template.
TEMPLATE_VISUAL_OBRA_CAIO_SUM_V1: dict[str, Any] = {
    "id": "OBRA_CAIO_SUM_V1",
    "pagina": {
        "largura_cm": 21.0,
        "altura_cm": 29.7,
        "orientacao": "A4_VERTICAL",
    },
    "cores": {
        "azul_petroleo": "#005A64",
        "amarelo_atencao": "#F2C230",
        "branco": "#FFFFFF",
        "preto": "#000000",
        "cinza_texto": "#333333",
    },
    "fontes": {
        "familia_base": "Helvetica",
        "familia_negrito": "Helvetica-Bold",
        "familia_sum": "Helvetica",
        "cabecalho_max": 50,
        "cabecalho_min": 34,
        "sum": 46,
        "sum_gap_su": -7.0,
        "sum_gap_um": -9.0,
        "pictograma_placeholder_grande": 95,
        "pictograma_placeholder_textual": 72,
        "texto_principal_max": 34,
        "texto_principal_min": 24,
        "texto_secundario_max": 18,
        "texto_secundario_min": 14,
        "rodape": 7,
    },
    "margens": {
        "corpo_lateral_cm": 1.0,
        "corpo_inferior_cm": 1.8,
    },
    "cabecalho": {
        "x_cm": 0.0,
        "y_cm": 24.2,
        "largura_cm": 21.0,
        "altura_cm": 5.5,
    },
    "triangulo_atencao": {
        "x_cm": 1.2,
        "y_cm": 25.55,
        "largura_cm": 3.0,
        "altura_cm": 3.0,
        "exclamacao_y_offset_cm": 0.72,
    },
    "titulo_cabecalho": {
        "x_cm": 4.55,
        "y_cm": 26.05,
        "largura_cm": 9.05,
        "altura_cm": 1.8,
    },
    "separador_vertical": {
        "x_cm": 13.95,
        "y_cm": 25.45,
        "largura_cm": 0.05,
        "altura_cm": 3.15,
    },
    "sum": {
        "texto": "SUM",
        "x_cm": 14.65,
        "y_cm": 26.08,
        "largura_cm": 4.6,
        "altura_cm": 1.8,
    },
    "corpo_branco": {
        "x_cm": 1.0,
        "y_cm": 1.8,
        "largura_cm": 19.0,
        "altura_cm": 22.4,
    },
    "circulo_pictograma": {
        "centro_x_cm": 10.5,
        "centro_y_cm": 14.3,
        "raio_cm": 7.3,
        "exclamacao_y_cm": 13.22,
    },
    "texto_principal": {
        "x_cm": 1.0,
        "y_cm": 4.75,
        "largura_cm": 19.0,
        "altura_cm": 1.0,
        "max_linhas": 2,
    },
    "texto_secundario": {
        "x_cm": 1.0,
        "y_cm": 3.45,
        "largura_cm": 19.0,
        "altura_cm": 1.0,
        "max_linhas": 2,
    },
    "rodape": {
        "texto": "",
        "x_centro_cm": 10.5,
        "y_cm": 0.65,
    },
}


DATABASE_URL = os.getenv("DATABASE_URL")
REDIS_URL = os.getenv("REDIS_URL")
TZ = os.getenv("TZ", "America/Sao_Paulo")
TELEGRAM_ALLOWED_USER_IDS = os.getenv("TELEGRAM_ALLOWED_USER_IDS")
TELEGRAM_ALLOWED_CHAT_IDS = os.getenv("TELEGRAM_ALLOWED_CHAT_IDS")
TELEGRAM_EXECUTIVE_USER_ID = os.getenv("TELEGRAM_EXECUTIVE_USER_ID")
TELEGRAM_EXECUTIVE_CHAT_ID = os.getenv("TELEGRAM_EXECUTIVE_CHAT_ID")


app = FastAPI(
    title=APP_NAME,
    version=APP_VERSION,
    description="API core do MVP Obra-Caio Control Tower.",
)


class EntradaPayload(BaseModel):
    tenant_id: Optional[str] = Field(default="construtora-piloto")
    obra_codigo: Optional[str] = Field(default="OBRA-CAIO")
    canal: Optional[str] = Field(default="desconhecido")
    remetente_nome: Optional[str] = None
    remetente_identificador: Optional[str] = None
    tipo_mensagem: Optional[str] = "texto"
    conteudo: Optional[str] = None
    anexos: Optional[list[dict[str, Any]]] = Field(default_factory=list)
    payload_original: Optional[Dict[str, Any]] = Field(default_factory=dict)


class TelegramEntradaPayload(BaseModel):
    tenant_id: Optional[str] = Field(default="construtora-piloto")
    obra_codigo: Optional[str] = Field(default="OBRA-CAIO")
    telegram_update_id: Optional[str] = None
    telegram_message_id: Optional[str] = None
    telegram_user_id: Optional[str] = None
    telegram_username: Optional[str] = None
    chat_id: Optional[str] = None
    chat_type: Optional[str] = None
    remetente_nome: Optional[str] = None
    remetente_identificador: Optional[str] = None
    usuario_autorizado: Optional[bool] = Field(default=False)
    motivo_autorizacao: Optional[str] = None
    tipo_mensagem: Optional[str] = "texto"
    conteudo: Optional[str] = None
    anexos: Optional[list[dict[str, Any]]] = Field(default_factory=list)
    payload_original: Optional[Dict[str, Any]] = Field(default_factory=dict)


class ProcessarComandoRDORequest(BaseModel):
    id_comando: Optional[str] = None


class ProcessarComandoComunicacaoObraRequest(BaseModel):
    id_comando: Optional[str] = None


class ProcessarComandoGestaoOperacionalRequest(BaseModel):
    id_comando: Optional[str] = None


class GestaoOperacionalObraRequest(BaseModel):
    obra_codigo: str


class GestaoOperacionalAreaRequest(BaseModel):
    obra_codigo: str
    codigo_area: str


class GestaoOperacionalDocumentosResumoRequest(BaseModel):
    obra_codigo: str


class GestaoOperacionalDocumentosIndexadosRequest(BaseModel):
    obra_codigo: str
    disciplina: str | None = None
    extensao: str | None = None
    termo: str | None = None
    limite: int = Field(default=50, ge=1, le=200)


class GestaoOperacionalClassificarDocumentosRequest(BaseModel):
    obra_codigo: str
    limite: int = Field(default=200, ge=1, le=1000)
    reprocessar: bool = False


class GestaoOperacionalDocumentosClassificadosRequest(BaseModel):
    obra_codigo: str
    area: str | None = None
    disciplina: str | None = None
    status_revisao: str | None = None
    eh_obsoleto: bool | None = None
    eh_as_built: bool | None = None
    termo: str | None = None
    limite: int = Field(default=50, ge=1, le=200)


class GestaoOperacionalRelatorioDocumentalRequest(BaseModel):
    obra_codigo: str
    incluir_amostras: bool = True
    limite_amostras: int = Field(default=5, ge=1, le=5)


class GestaoOperacionalRiscosDocumentaisRequest(BaseModel):
    obra_codigo: str
    area: str | None = None
    disciplina: str | None = None
    limite_amostras: int = Field(default=10, ge=1, le=50)


class GestaoOperacionalPlanoSaneamentoDocumentalRequest(BaseModel):
    obra_codigo: str
    area: str | None = None
    disciplina: str | None = None
    limite_acoes: int = Field(default=10, ge=1, le=50)


class GestaoOperacionalCriarPendenciaDocumentalRequest(BaseModel):
    obra_codigo: str
    area: str | None = None
    disciplina: str | None = None
    motivo: str
    descricao: str | None = None


class GestaoOperacionalUltimaRevisaoDocumentalRequest(BaseModel):
    obra_codigo: str
    disciplina: str | None = None
    area: str | None = None
    limite_candidatos: int = Field(default=10, ge=1, le=100)


class GestaoOperacionalValidarDocumentoCampoRequest(BaseModel):
    obra_codigo: str
    documento_id: int = Field(gt=0)


class GestaoOperacionalDiagnosticoRequest(BaseModel):
    obra_codigo: str
    area: str | None = None
    incluir_documentos: bool = True
    incluir_pendencias: bool = True
    incluir_restricoes: bool = True


class GestaoOperacionalPlanoOperacionalRequest(BaseModel):
    obra_codigo: str
    area: str | None = None
    limite_acoes: int = Field(default=10, ge=1, le=10)


class GestaoOperacionalResumoExecutivoRequest(BaseModel):
    obra_codigo: str
    area: str | None = None
    limite_itens: int = Field(default=10, ge=1, le=50)


class GestaoOperacionalBriefingDiarioRequest(BaseModel):
    obra_codigo: str
    area: str | None = None
    limite_itens: int = Field(default=10, ge=1, le=50)
    incluir_acoes: bool = True
    incluir_documentos: bool = True
    incluir_historico: bool = True


class GestaoOperacionalRelatorioSemanalRequest(BaseModel):
    obra_codigo: str
    area: str | None = None
    data_inicio: date | None = None
    data_fim: date | None = None
    limite_itens: int = Field(default=10, ge=1, le=50)
    salvar_relatorio: bool = True


class GestaoOperacionalExportarRelatorioSemanalRequest(BaseModel):
    obra_codigo: str
    area: str | None = None
    data_inicio: date | None = None
    data_fim: date | None = None
    relatorio_semanal_id: int | None = Field(default=None, gt=0)
    formato: str = "MARKDOWN"
    limite_itens: int = Field(default=10, ge=1, le=50)
    salvar_exportacao: bool = True


class GestaoOperacionalGerarPdfRelatorioSemanalRequest(BaseModel):
    obra_codigo: str
    area: str | None = None
    data_inicio: date | None = None
    data_fim: date | None = None
    exportacao_relatorio_id: int | None = Field(default=None, gt=0)
    relatorio_semanal_id: int | None = Field(default=None, gt=0)
    limite_itens: int = Field(default=10, ge=1, le=50)
    salvar_pdf: bool = True
    armazenar_minio: bool = True


class GestaoOperacionalAprovarPdfRelatorioSemanalRequest(BaseModel):
    obra_codigo: str
    pdf_relatorio_id: int = Field(gt=0)
    decisao: str = Field(pattern="^(APROVAR|REJEITAR|SOLICITAR_AJUSTES)$")
    motivo: str | None = None
    observacao: str | None = None
    decisor_nome: str | None = None
    decisor_telegram_user_id: str | None = None
    decisor_telegram_username: str | None = None
    decisor_chat_id: str | None = None


class GestaoOperacionalSolicitarEnvioRelatorioSemanalRequest(BaseModel):
    obra_codigo: str
    pdf_relatorio_id: int = Field(gt=0)
    canal_pretendido: str
    destinatario_nome: str | None = None
    destinatario_contato: str | None = None
    assunto: str | None = None
    mensagem: str | None = None
    solicitado_por: str | None = None
    solicitante_telegram_user_id: str | None = None
    solicitante_telegram_username: str | None = None
    solicitante_chat_id: str | None = None


class GestaoOperacionalBriefingDiarioAgendadoRequest(BaseModel):
    forcar: bool = False


class GestaoOperacionalConfirmarBriefingDiarioRequest(BaseModel):
    envio_id: int = Field(gt=0)
    telegram_message_id: str | None = None
    status: str = Field(default="CONCLUIDO", pattern="^(CONCLUIDO|ERRO)$")
    mensagem_erro: str | None = None


class GestaoOperacionalCriarAcaoRequest(BaseModel):
    obra_codigo: str
    area: str | None = None
    disciplina: str | None = None
    titulo: str = Field(min_length=1)
    descricao: str | None = None
    prioridade: str = Field(default="MEDIA", pattern="^(BAIXA|MEDIA|ALTA|CRITICA)$")
    responsavel: str | None = None
    prazo: date | None = None
    referencia_documento_id: int | None = Field(default=None, gt=0)


class GestaoOperacionalListarAcoesRequest(BaseModel):
    obra_codigo: str
    area: str | None = None
    disciplina: str | None = None
    status: str | None = Field(default="ABERTA", pattern="^(ABERTA|EM_ANDAMENTO|CONCLUIDA|CANCELADA)$")
    limite: int = Field(default=20, ge=1, le=100)


class GestaoOperacionalAtualizarAcaoRequest(BaseModel):
    obra_codigo: str
    acao_id: int = Field(gt=0)
    status: str | None = Field(default=None, pattern="^(ABERTA|EM_ANDAMENTO|CONCLUIDA|CANCELADA)$")
    prioridade: str | None = Field(default=None, pattern="^(BAIXA|MEDIA|ALTA|CRITICA)$")
    responsavel: str | None = None
    prazo: date | None = None
    observacao: str | None = None


class GestaoOperacionalDetalheAcaoRequest(BaseModel):
    obra_codigo: str
    acao_id: int = Field(gt=0)
    incluir_historico: bool = True


class GestaoOperacionalRegistrarRevisaoDocumentalRequest(BaseModel):
    obra_codigo: str = Field(min_length=1)
    documento_minio_id: int | None = Field(default=None, gt=0)
    bucket: str | None = None
    object_key: str | None = None
    nome_arquivo_original: str = Field(min_length=1)
    disciplina: str | None = None
    area: str | None = None
    codigo_documento: str | None = None
    titulo_documento: str | None = None
    revisao_detectada: str | None = None
    data_documento: date | None = None
    tipo_documento: str | None = "PROJETO"
    motivo_alteracao: str | None = None
    responsavel_upload: str | None = None
    observacao: str | None = None


class GestaoOperacionalRevisoesDocumentaisRequest(BaseModel):
    obra_codigo: str = Field(min_length=1)
    area: str | None = None
    disciplina: str | None = None
    status_revisao: str | None = Field(default=None, pattern="^(RECEBIDO_PARA_ANALISE|EM_ANALISE_TECNICA|APROVADO_COMO_VIGENTE|REJEITADO|AJUSTES_SOLICITADOS|SUBSTITUIDO|OBSOLETO|AS_BUILT)$")
    status_vigencia: str | None = Field(default=None, pattern="^(NAO_VIGENTE|VIGENTE|SUBSTITUIDO|OBSOLETO|HISTORICO)$")
    limite_itens: int = Field(default=20, ge=1, le=200)


class GestaoOperacionalImportarRevisoesDocumentaisRequest(BaseModel):
    obra_codigo: str = Field(min_length=1)
    bucket: str = Field(min_length=1)
    prefixo: str = Field(min_length=1)
    limite_itens: int = Field(default=50, ge=1, le=500)


class GestaoOperacionalAprovarRevisaoDocumentalRequest(BaseModel):
    obra_codigo: str = Field(min_length=1)
    revisao_documental_id: int = Field(gt=0)
    decisao: str = Field(pattern="^(APROVAR_COMO_VIGENTE|REJEITAR|SOLICITAR_AJUSTES|MARCAR_EM_ANALISE)$")
    motivo: str | None = None
    observacao: str | None = None
    decisor_nome: str | None = None
    decisor_telegram_user_id: str | None = None
    decisor_telegram_username: str | None = None
    decisor_chat_id: str | None = None


class GestaoOperacionalLiberarRevisaoDocumentalCampoRequest(BaseModel):
    obra_codigo: str = Field(min_length=1)
    revisao_documental_id: int = Field(gt=0)
    decisao: str = Field(pattern="^(LIBERAR_PARA_CAMPO|SUSPENDER_LIBERACAO_CAMPO|REVOGAR_LIBERACAO_CAMPO)$")
    motivo: str | None = None
    observacao: str | None = None
    instrucoes_campo: str | None = None
    decisor_nome: str | None = None
    decisor_telegram_user_id: str | None = None
    decisor_telegram_username: str | None = None
    decisor_chat_id: str | None = None


class GestaoOperacionalLiberacoesRevisoesDocumentaisCampoRequest(BaseModel):
    obra_codigo: str = Field(min_length=1)
    area: str | None = None
    disciplina: str | None = None
    status_liberacao: str | None = Field(default=None, pattern="^(LIBERADO_PARA_USO_DOCUMENTAL_EM_CAMPO|SUSPENSA|REVOGADA)$")
    limite_itens: int = Field(default=20, ge=1, le=200)


def get_db_connection():
    if not DATABASE_URL:
        raise RuntimeError("DATABASE_URL não definida.")
    return psycopg.connect(DATABASE_URL)


def stable_hash(data: Any) -> str:
    raw = json.dumps(data, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def parse_correlation_id(value: Optional[str]) -> str:
    if not value:
        return str(uuid.uuid4())

    try:
        return str(uuid.UUID(value))
    except ValueError as exc:
        raise HTTPException(
            status_code=422,
            detail={
                "message": "X-Correlation-Id inválido. Use um UUID válido.",
                "error": str(exc),
            },
        )


def has_any_term(texto: str, termos: list[str]) -> bool:
    return any(termo in texto for termo in termos)


def parse_instrucao_aprovacao_comando(conteudo: Optional[str]) -> Optional[dict[str, Any]]:
    texto = (conteudo or "").strip().lower()
    if not texto:
        return None

    match = re.search(
        r"\b(aprovar|aprova|autorizar|confirmar|cancelar|cancela|rejeitar)\s+comando\s+([0-9a-fA-F-]+)\b",
        texto,
    )
    if not match:
        return None

    verbo = match.group(1)
    identificador = match.group(2)
    acao = "CANCELAR" if verbo in {"cancelar", "cancela", "rejeitar"} else "APROVAR"
    tipo_identificador = "id_comando" if "-" in identificador else "id"

    if tipo_identificador == "id":
        try:
            comando_id = int(identificador)
        except ValueError:
            return None
        return {
            "acao": acao,
            "comando_id": comando_id,
            "id_comando": None,
            "identificador": str(comando_id),
            "tipo_identificador": tipo_identificador,
        }

    try:
        id_comando = str(uuid.UUID(identificador))
    except ValueError:
        return None

    return {
        "acao": acao,
        "comando_id": None,
        "id_comando": id_comando,
        "identificador": id_comando,
        "tipo_identificador": tipo_identificador,
    }


def classificar_instrucao_aprovacao(instrucao: dict[str, Any]) -> dict[str, Any]:
    if instrucao["acao"] == "APROVAR":
        return {
            "intencao": "APROVAR_COMANDO_EXECUTIVO",
            "agente_destino": "AGENTE_007_ORQUESTRADOR_EXECUTIVO",
            "tipo_comando": "APROVAR_COMANDO_EXECUTIVO",
            "requer_aprovacao": False,
            "confianca": 0.98,
            "justificativa": "Mensagem solicita aprovação auditável de comando executivo existente.",
        }

    return {
        "intencao": "CANCELAR_COMANDO_EXECUTIVO",
        "agente_destino": "AGENTE_007_ORQUESTRADOR_EXECUTIVO",
        "tipo_comando": "CANCELAR_COMANDO_EXECUTIVO",
        "requer_aprovacao": False,
        "confianca": 0.98,
        "justificativa": "Mensagem solicita cancelamento auditável de comando executivo existente.",
    }


def parse_env_allowlist(*values: Optional[str]) -> set[str]:
    allowed: set[str] = set()
    for value in values:
        if not value:
            continue
        allowed.update(part.strip() for part in value.split(",") if part.strip())
    return allowed


def avaliar_autorizacao_telegram(payload: TelegramEntradaPayload) -> tuple[bool, str]:
    allowed_user_ids = parse_env_allowlist(
        TELEGRAM_ALLOWED_USER_IDS,
        TELEGRAM_EXECUTIVE_USER_ID,
    )
    allowed_chat_ids = parse_env_allowlist(
        TELEGRAM_ALLOWED_CHAT_IDS,
        TELEGRAM_EXECUTIVE_CHAT_ID,
    )

    if payload.telegram_user_id and payload.telegram_user_id in allowed_user_ids:
        return True, "Autorizado por telegram_user_id configurado em variável de ambiente."

    if payload.chat_id and payload.chat_id in allowed_chat_ids:
        return True, "Autorizado por chat_id configurado em variável de ambiente."

    if not allowed_user_ids and not allowed_chat_ids:
        return (
            False,
            "Bloqueado: nenhuma allowlist Telegram configurada nas variáveis de ambiente.",
        )

    if not payload.telegram_user_id and not payload.chat_id:
        return (
            False,
            "Bloqueado: payload não trouxe telegram_user_id nem chat_id para validação.",
        )

    return (
        False,
        "Bloqueado: telegram_user_id/chat_id não constam nas allowlists de ambiente.",
    )


def classificar_intencao_executiva(conteudo: Optional[str]) -> dict[str, Any]:
    texto = (conteudo or "").lower()
    termos_rdo = ["rdo", "diário de obra", "diario de obra", "diário", "diario", "relatório diário", "relatorio diario"]
    menciona_rdo = has_any_term(texto, termos_rdo)
    termos_placa = [
        "placa",
        "placa de aviso",
        "aviso para o canteiro",
        "comunicado para a obra",
        "sinalização da obra",
        "sinalizacao da obra",
        "placa de segurança",
        "placa de seguranca",
    ]
    menciona_placa = has_any_term(texto, termos_placa)

    comandos_documentos_resumo = [
        "documentos",
        "documentos da obra",
        "resumo documentos",
        "resumo dos documentos",
        "quais documentos temos",
        "listar documentos da obra",
    ]
    texto_comando = re.sub(r"\s+", " ", texto.strip().rstrip("?.!"))
    texto_comando_normalizado = normalizar_texto_comparacao(texto_comando)
    liberacao_campo = extrair_liberacao_revisao_campo_telegram(conteudo)
    if liberacao_campo:
        return {
            "intencao": "LIBERAR_REVISAO_DOCUMENTAL_CAMPO_OBRA",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "LIBERAR_REVISAO_DOCUMENTAL_CAMPO_OBRA",
            **liberacao_campo,
            "requer_aprovacao": False,
            "confianca": 0.99,
            "justificativa": "Liberação documental de campo explícita e auditável, sem execução de serviço.",
        }
    decisao_revisao = extrair_decisao_revisao_documental_telegram(conteudo)
    if decisao_revisao:
        return {
            "intencao": "APROVAR_REVISAO_DOCUMENTAL_OBRA",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "APROVAR_REVISAO_DOCUMENTAL_OBRA",
            "decisao": decisao_revisao["decisao"],
            "revisao_documental_id": decisao_revisao["revisao_documental_id"],
            "requer_aprovacao": False,
            "confianca": 0.99,
            "justificativa": "Decisão técnica documental explícita e auditável, sem execução externa.",
        }
    decisoes_pdf_relatorio = (
        (
            "SOLICITAR_AJUSTES",
            r"(?:solicitar ajustes no pdf|pedir ajustes no pdf|corrigir pdf|ajustar pdf|"
            r"solicitar revisao do pdf)",
        ),
        (
            "REJEITAR",
            r"(?:rejeitar pdf|rejeitar relatorio semanal|nao aprovar pdf|recusar pdf)",
        ),
        (
            "APROVAR",
            r"(?:aprovar pdf(?: do relatorio semanal)?|aprovar relatorio semanal|"
            r"aprovo o pdf|pode aprovar o pdf)",
        ),
    )
    for decisao, padrao in decisoes_pdf_relatorio:
        if re.match(rf"^{padrao}(?:\s|$)", texto_comando_normalizado):
            return {
                "intencao": "APROVAR_PDF_RELATORIO_SEMANAL_EXECUTIVO_OBRA",
                "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
                "tipo_comando": "APROVAR_PDF_RELATORIO_SEMANAL_EXECUTIVO_OBRA",
                "decisao": decisao,
                "requer_aprovacao": False,
                "confianca": 0.99,
                "justificativa": (
                    "Decisão executiva auditável sobre PDF para uso interno, "
                    "sem envio ou alteração de sistemas externos."
                ),
            }
    if re.match(
        r"^(?:solicitar envio do|pedir envio do|preparar envio do|registrar envio do|"
        r"enviar|encaminhar) (?:pdf|relatorio)(?:\s|$)",
        texto_comando_normalizado,
    ):
        return {
            "intencao": "SOLICITAR_ENVIO_RELATORIO_SEMANAL_EXECUTIVO_OBRA",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "SOLICITAR_ENVIO_RELATORIO_SEMANAL_EXECUTIVO_OBRA",
            "requer_aprovacao": False,
            "confianca": 0.99,
            "justificativa": (
                "Registro controlado de solicitação de envio futuro, sem envio, "
                "arquivo, link ou execução externa."
            ),
        }
    comandos_diagnostico_operacional = {
        "diagnostico operacional da obra", "situacao operacional da obra",
        "status operacional da obra", "como esta a obra", "risco operacional da obra",
        "diagnostico operacional do refeitorio", "status do refeitorio",
        "o que impede o avanco do refeitorio",
    }
    comandos_plano_operacional = {
        "gerar plano operacional da obra", "plano operacional da obra",
        "plano de acao da obra", "acoes prioritarias da obra",
        "proximos passos da obra", "plano de acao do refeitorio",
        "o que fazer agora no refeitorio", "proximos passos do refeitorio",
    }
    comandos_listar_acoes = {
        "listar acoes operacionais abertas", "acoes pendentes da obra",
        "acoes pendentes do refeitorio", "status das acoes operacionais",
    }
    comandos_resumo_executivo = {
        "resumo operacional da obra", "resumo executivo da obra",
        "resumo executivo operacional", "resumo das acoes operacionais",
        "o que esta pendente na obra", "status executivo da obra",
        "panorama operacional da obra", "resumo operacional do refeitorio",
        "status executivo do refeitorio",
    }
    comandos_briefing_diario = {
        "briefing diario da obra", "briefing executivo da obra",
        "resumo do dia da obra", "o que preciso olhar hoje",
        "pauta operacional de hoje", "pauta executiva de hoje",
        "briefing operacional", "briefing executivo", "briefing do refeitorio",
        "resumo do dia do refeitorio",
    }
    comandos_relatorio_semanal = {
        "relatorio semanal da obra", "resumo semanal executivo",
        "balanco semanal da obra", "o que aconteceu na obra esta semana",
        "o que aconteceu na obra essa semana", "relatorio semanal executivo",
        "relatorio semanal do refeitorio", "balanco semanal do refeitorio",
    }
    comandos_exportar_relatorio_semanal = {
        "exportar relatorio semanal da obra",
        "gerar markdown do relatorio semanal",
        "preparar relatorio semanal para revisao",
        "exportar relatorio semanal do refeitorio",
        "preparar relatorio semanal do refeitorio",
    }
    comandos_pdf_relatorio_semanal = {
        "gerar pdf do relatorio semanal",
        "pdf do relatorio semanal da obra",
        "preparar pdf do relatorio semanal",
        "gerar pdf do relatorio semanal do refeitorio",
        "preparar pdf para revisao",
    }
    comandos_consultar_revisoes = {
        "revisoes documentais da obra", "documentos recebidos para analise",
        "projetos recebidos para analise", "ultimas revisoes de projeto",
        "status das revisoes documentais",
    }
    comandos_consultar_liberacoes_campo = {
        "revisoes liberadas para campo", "documentos liberados para campo",
        "projetos liberados para campo", "status das liberacoes de campo",
    }
    comandos_importar_revisoes = {
        "importar revisoes documentais pendentes",
        "registrar documentos recebidos para analise",
        "registrar projetos recebidos para analise",
    }
    if texto_comando_normalizado in comandos_importar_revisoes:
        return {
            "intencao": "IMPORTAR_REVISOES_DOCUMENTAIS_PENDENTES_OBRA",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "IMPORTAR_REVISOES_DOCUMENTAIS_PENDENTES_OBRA",
            "requer_aprovacao": False, "confianca": 0.99,
            "justificativa": "Registro interno de arquivos já indexados, sem alteração no MinIO.",
        }
    if texto_comando_normalizado in comandos_consultar_revisoes:
        return {
            "intencao": "CONSULTAR_REVISOES_DOCUMENTAIS_OBRA",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "CONSULTAR_REVISOES_DOCUMENTAIS_OBRA",
            "requer_aprovacao": False, "confianca": 0.99,
            "justificativa": "Consulta ao controle documental sem liberação de execução.",
        }
    if texto_comando_normalizado in comandos_consultar_liberacoes_campo:
        return {
            "intencao": "CONSULTAR_LIBERACOES_REVISOES_DOCUMENTAIS_CAMPO_OBRA",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "CONSULTAR_LIBERACOES_REVISOES_DOCUMENTAIS_CAMPO_OBRA",
            "requer_aprovacao": False, "confianca": 0.99,
            "justificativa": "Consulta auditável das liberações documentais de campo.",
        }
    if texto_comando_normalizado in comandos_pdf_relatorio_semanal:
        return {
            "intencao": "GERAR_PDF_RELATORIO_SEMANAL_EXECUTIVO_OBRA",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "GERAR_PDF_RELATORIO_SEMANAL_EXECUTIVO_OBRA",
            "requer_aprovacao": False,
            "confianca": 0.99,
            "justificativa": (
                "Geração de PDF privado para revisão interna, sem envio ou execução externa."
            ),
        }
    if texto_comando_normalizado in comandos_exportar_relatorio_semanal:
        return {
            "intencao": "EXPORTAR_RELATORIO_SEMANAL_EXECUTIVO_OBRA",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "EXPORTAR_RELATORIO_SEMANAL_EXECUTIVO_OBRA",
            "requer_aprovacao": False,
            "confianca": 0.99,
            "justificativa": (
                "Exportação Markdown consultiva para revisão interna, sem envio ou execução externa."
            ),
        }
    if texto_comando_normalizado in comandos_relatorio_semanal:
        return {
            "intencao": "GERAR_RELATORIO_SEMANAL_EXECUTIVO_OBRA",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "GERAR_RELATORIO_SEMANAL_EXECUTIVO_OBRA",
            "requer_aprovacao": False,
            "confianca": 0.99,
            "justificativa": "Relatório semanal executivo consultivo, sem alteração ou execução externa.",
        }
    if texto_comando_normalizado in comandos_briefing_diario:
        return {
            "intencao": "GERAR_BRIEFING_DIARIO_OBRA",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "GERAR_BRIEFING_DIARIO_OBRA",
            "requer_aprovacao": False,
            "confianca": 0.99,
            "justificativa": "Briefing diário executivo consultivo, sem alteração ou execução externa.",
        }
    if texto_comando_normalizado in comandos_resumo_executivo:
        return {
            "intencao": "GERAR_RESUMO_EXECUTIVO_OPERACIONAL_OBRA",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "GERAR_RESUMO_EXECUTIVO_OPERACIONAL_OBRA",
            "requer_aprovacao": False,
            "confianca": 0.99,
            "justificativa": "Resumo executivo consultivo, sem alteração ou execução externa.",
        }
    if texto_comando_normalizado in comandos_listar_acoes:
        return {
            "intencao": "LISTAR_ACOES_OPERACIONAIS_OBRA",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "LISTAR_ACOES_OPERACIONAIS_OBRA",
            "requer_aprovacao": False,
            "confianca": 0.98,
            "justificativa": "Consulta interna de ações operacionais, sem alteração de sistemas externos.",
        }
    detalhe_acao = re.fullmatch(r"(?:detalhar|historico da) acao\s+\d+", texto_comando_normalizado)
    if detalhe_acao:
        return {
            "intencao": "DETALHAR_ACAO_OPERACIONAL_OBRA",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "DETALHAR_ACAO_OPERACIONAL_OBRA",
            "requer_aprovacao": False, "confianca": 0.99,
            "justificativa": "Consulta ao detalhe e histórico auditável de ação operacional interna.",
        }
    atualizar_acao = re.fullmatch(
        r"(?:definir responsavel da acao \d+ como .+|atribuir acao \d+ para .+|"
        r"definir prazo da acao \d+ para \d{2} \d{2} \d{4}|prazo da acao \d+ \d{2} \d{2} \d{4}|"
        r"alterar prioridade da acao \d+ para (?:baixa|media|alta|critica)|prioridade da acao \d+ (?:baixa|media|alta|critica)|"
        r"colocar acao \d+ em andamento|marcar acao \d+ como concluida|cancelar acao \d+|"
        r"(?:adicionar|registrar) observacao na acao \d+\s+.+)", texto_comando_normalizado,
    )
    if atualizar_acao:
        return {
            "intencao": "ATUALIZAR_ACAO_OPERACIONAL_OBRA",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "ATUALIZAR_ACAO_OPERACIONAL_OBRA",
            "requer_aprovacao": False,
            "confianca": 0.99,
            "justificativa": "Atualização auditável de registro operacional interno.",
        }
    if (
        texto_comando_normalizado.startswith("registrar acao operacional")
        or texto_comando_normalizado.startswith("abrir acao para")
        or texto_comando_normalizado.startswith("criar acao para")
        or texto_comando_normalizado.startswith("registrar acao critica")
    ):
        return {
            "intencao": "CRIAR_ACAO_OPERACIONAL_OBRA",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "CRIAR_ACAO_OPERACIONAL_OBRA",
            "requer_aprovacao": False,
            "confianca": 0.97,
            "justificativa": "Criação auditável de ação operacional interna, sem execução externa.",
        }
    if texto_comando_normalizado in comandos_plano_operacional:
        return {
            "intencao": "GERAR_PLANO_OPERACIONAL_OBRA",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "GERAR_PLANO_OPERACIONAL_OBRA",
            "requer_aprovacao": False,
            "confianca": 0.98,
            "justificativa": "Plano operacional consultivo, sem liberação de frente ou alteração externa.",
        }
    if texto_comando_normalizado in comandos_diagnostico_operacional:
        return {
            "intencao": "GERAR_DIAGNOSTICO_OPERACIONAL_OBRA",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "GERAR_DIAGNOSTICO_OPERACIONAL_OBRA",
            "requer_aprovacao": False,
            "confianca": 0.98,
            "justificativa": "Diagnóstico operacional consultivo, sem liberação de frente ou alteração externa.",
        }
    comandos_plano_saneamento = {
        "gerar plano de saneamento documental",
        "plano de saneamento documental",
        "plano para corrigir documentos obsoletos",
        "listar acoes recomendadas para documentacao",
        "acoes recomendadas documentacao",
        "como corrigir riscos documentais",
    }
    if texto_comando_normalizado in comandos_plano_saneamento:
        return {
            "intencao": "GERAR_PLANO_SANEAMENTO_DOCUMENTAL",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "GERAR_PLANO_SANEAMENTO_DOCUMENTAL",
            "requer_aprovacao": False,
            "confianca": 0.98,
            "justificativa": "Plano consultivo de saneamento documental, sem abertura automática de pendências.",
        }
    if re.fullmatch(
        r"(?:abrir|criar|registrar) pendencia (?:documental|de documento obsoleto|para revisar)(?: .+)?",
        texto_comando_normalizado,
    ):
        return {
            "intencao": "CRIAR_PENDENCIA_DOCUMENTAL",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "CRIAR_PENDENCIA_DOCUMENTAL",
            "requer_aprovacao": False,
            "confianca": 0.98,
            "justificativa": "Pedido explícito para registro de pendência interna de gestão documental.",
        }
    comandos_riscos_documentais = {
        "verificar riscos documentais da obra", "riscos documentais da obra",
        "documentos criticos para campo", "tem conflito documental no refeitorio",
        "riscos documentais do refeitorio", "tem projeto obsoleto de luminotecnico",
        "risco documental luminotecnico", "riscos do projeto hidraulico",
        "riscos do projeto hidralico", "riscos estrutura", "riscos eletrica",
    }
    if texto_comando_normalizado in comandos_riscos_documentais:
        return {
            "intencao": "AVALIAR_RISCOS_DOCUMENTAIS_OBRA",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "AVALIAR_RISCOS_DOCUMENTAIS_OBRA",
            "requer_aprovacao": False,
            "confianca": 0.98,
            "justificativa": "Avaliação consultiva de riscos documentais, sem alteração ou liberação para uso.",
        }
    if re.fullmatch(
        r"(?:posso usar o |validar |verificar uso |esse )?documento\s+\d+(?:\s+(?:em campo|esta obsoleto|pode ir para campo))?",
        texto_comando_normalizado,
    ) or re.fullmatch(r"esse documento\s+\d+\s+pode ir para campo", texto_comando_normalizado):
        return {
            "intencao": "VALIDAR_DOCUMENTO_CAMPO",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "VALIDAR_DOCUMENTO_CAMPO",
            "requer_aprovacao": False,
            "confianca": 0.98,
            "justificativa": "Validação consultiva de documento para uso em campo, sem liberação definitiva.",
        }
    if (
        texto_comando_normalizado.startswith("ultima revisao ")
        or re.fullmatch(r"qual a ultima revisao do .+", texto_comando_normalizado)
        or re.fullmatch(r"qual projeto .+ mais recente", texto_comando_normalizado)
    ):
        return {
            "intencao": "CONSULTAR_ULTIMA_REVISAO_DOCUMENTAL",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "CONSULTAR_ULTIMA_REVISAO_DOCUMENTAL",
            "requer_aprovacao": False,
            "confianca": 0.98,
            "justificativa": "Consulta somente leitura da revisão documental mais recente não obsoleta.",
        }
    comandos_relatorio_documental = {
        "relatorio documental da obra",
        "resumo documental da obra",
        "situacao dos projetos",
        "diagnostico documental",
        "relatorio dos documentos",
    }
    if texto_comando_normalizado in comandos_relatorio_documental:
        return {
            "intencao": "GERAR_RELATORIO_DOCUMENTAL_OBRA",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "GERAR_RELATORIO_DOCUMENTAL_OBRA",
            "requer_aprovacao": False,
            "confianca": 0.98,
            "justificativa": "Relatório executivo somente leitura da classificação documental da obra.",
        }
    if texto_comando_normalizado in {"classificar documentos", "reclassificar documentos"}:
        return {
            "intencao": "CLASSIFICAR_DOCUMENTOS_OBRA",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "CLASSIFICAR_DOCUMENTOS_OBRA",
            "requer_aprovacao": False,
            "confianca": 0.98,
            "justificativa": "Classificação determinística de metadados documentais, sem acesso ou alteração no MinIO.",
        }
    comandos_classificados = {
        "documentos obsoletos", "listar documentos obsoletos", "tem as built",
        "documentos as built", "ultima revisao luminotecnico",
        "documentos do refeitorio", "documentos refeitorio",
        "documentos de hidraulica", "documentos de hidralica", "documentos de hidr",
        "documentos de eletrica", "documentos luminotecnico",
    }
    if texto_comando_normalizado in comandos_classificados:
        return {
            "intencao": "CONSULTAR_DOCUMENTOS_CLASSIFICADOS",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "CONSULTAR_DOCUMENTOS_CLASSIFICADOS",
            "requer_aprovacao": False,
            "confianca": 0.97,
            "justificativa": "Consulta somente leitura da classificação documental técnica da obra.",
        }
    if re.fullmatch(
        r"(?:analisar|resumir|verificar)\s+(?:documento|projeto)\s+.+",
        texto_comando,
    ):
        return {
            "intencao": "ANALISAR_DOCUMENTO_OBRA",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "ANALISAR_DOCUMENTO_OBRA",
            "requer_aprovacao": False,
            "confianca": 0.96,
            "justificativa": "Análise preliminar, somente leitura, de documento indexado da obra.",
        }
    if texto_comando in comandos_documentos_resumo:
        return {
            "intencao": "CONSULTAR_DOCUMENTOS_OBRA_RESUMO",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "CONSULTAR_DOCUMENTOS_OBRA_RESUMO",
            "requer_aprovacao": False,
            "confianca": 0.95,
            "justificativa": "Consulta somente leitura do resumo documental indexado da obra.",
        }

    comandos_documentos_indexados = [
        "documentos arquitetura",
        "documentos eletrica",
        "documentos elétrica",
        "documentos hidraulica",
        "documentos hidráulica",
        "documentos estrutura",
        "documentos luminotecnico",
        "documentos luminotécnico",
        "documentos dwg",
        "documentos pdf",
        "buscar refeitório",
        "buscar alteracoes refeitório",
        "buscar alterações refeitório",
        "procurar refeitório",
        "procurar projeto refeitório",
    ]
    if texto_comando in comandos_documentos_indexados:
        return {
            "intencao": "CONSULTAR_DOCUMENTOS_OBRA_INDEXADOS",
            "agente_destino": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
            "tipo_comando": "CONSULTAR_DOCUMENTOS_OBRA_INDEXADOS",
            "requer_aprovacao": False,
            "confianca": 0.95,
            "justificativa": "Busca somente leitura no índice documental da obra.",
        }

    if has_any_term(texto, ["confirmar", "confirmo", "aprovado", "aprovar"]):
        return {
            "intencao": "CONFIRMAR_COMANDO",
            "agente_destino": "AGENTE_007_ORQUESTRADOR_EXECUTIVO",
            "tipo_comando": "CONFIRMAR_COMANDO_EXECUTIVO",
            "requer_aprovacao": False,
            "confianca": 0.90,
            "justificativa": "Mensagem indica confirmação ou aprovação executiva.",
        }

    if has_any_term(texto, ["cancelar", "cancele", "rejeitar", "rejeito", "não aprovar", "nao aprovar"]):
        return {
            "intencao": "CANCELAR_COMANDO",
            "agente_destino": "AGENTE_007_ORQUESTRADOR_EXECUTIVO",
            "tipo_comando": "CANCELAR_COMANDO_EXECUTIVO",
            "requer_aprovacao": False,
            "confianca": 0.90,
            "justificativa": "Mensagem indica cancelamento ou rejeição executiva.",
        }

    if menciona_placa and has_any_term(
        texto,
        [
            "gerar pdf",
            "gerar o pdf",
            "gere o pdf",
            "pdf da placa",
            "pdf do aviso",
            "emitir pdf",
            "emita o pdf",
        ],
    ):
        return {
            "intencao": "GERAR_PDF_PLACA_AVISO",
            "agente_destino": "AGENTE_006_COMUNICACAO_VISUAL_OBRA",
            "tipo_comando": "GERAR_PDF_PLACA_AVISO",
            "requer_aprovacao": True,
            "confianca": 0.88,
            "justificativa": "Geração de PDF de placa é ação sensível e exige aprovação; nenhum PDF real é gerado nesta etapa.",
        }

    if menciona_placa and has_any_term(
        texto,
        ["imprimir", "impressão", "impressao", "imprima", "mande imprimir", "solicite impressão", "solicite impressao"],
    ):
        return {
            "intencao": "IMPRIMIR_PLACA_AVISO",
            "agente_destino": "AGENTE_006_COMUNICACAO_VISUAL_OBRA",
            "tipo_comando": "IMPRIMIR_PLACA_AVISO",
            "requer_aprovacao": True,
            "confianca": 0.88,
            "justificativa": "Impressão de placa é ação sensível e exige aprovação; nenhuma impressão real é executada nesta etapa.",
        }

    if menciona_placa and has_any_term(
        texto,
        ["instalar", "instale", "instala", "fixar", "fixe", "colocar a placa", "coloque a placa"],
    ):
        return {
            "intencao": "INSTALAR_PLACA_AVISO",
            "agente_destino": "AGENTE_006_COMUNICACAO_VISUAL_OBRA",
            "tipo_comando": "INSTALAR_PLACA_AVISO",
            "requer_aprovacao": True,
            "confianca": 0.88,
            "justificativa": "Instalação de placa é ação sensível e exige aprovação; nenhuma instalação é executada nesta etapa.",
        }

    if menciona_placa and has_any_term(
        texto,
        [
            "prepare",
            "preparar",
            "gerar placa",
            "gerar placa de aviso",
            "crie",
            "criar",
            "monte",
            "montar",
            "elabore",
            "elaborar",
            "rascunho",
            "uso obrigatório de epi",
            "uso obrigatorio de epi",
            "aviso",
            "comunicado",
            "sinalização",
            "sinalizacao",
        ],
    ):
        return {
            "intencao": "PREPARAR_PLACA_AVISO",
            "agente_destino": "AGENTE_006_COMUNICACAO_VISUAL_OBRA",
            "tipo_comando": "PREPARAR_PLACA_AVISO",
            "requer_aprovacao": False,
            "confianca": 0.87,
            "justificativa": "Mensagem indica preparação de rascunho textual de placa de aviso sem PDF, impressão ou instalação.",
        }

    if has_any_term(
        texto,
        ["imprimir", "impressão", "impressao", "imprima", "mande imprimir", "solicite impressão", "solicite impressao"],
    ):
        return {
            "intencao": "SOLICITAR_IMPRESSAO",
            "agente_destino": "AGENTE_007_ORQUESTRADOR_EXECUTIVO",
            "tipo_comando": "SOLICITAR_IMPRESSAO",
            "requer_aprovacao": True,
            "confianca": 0.86,
            "justificativa": "Solicitação de impressão exige aprovação e não executa impressão real nesta etapa.",
        }

    if menciona_rdo and has_any_term(
        texto,
        [
            "gerar pdf",
            "gerar o pdf",
            "gere o pdf",
            "emita o pdf",
            "emitir o pdf",
            "pdf do rdo",
            "pdf do diário",
            "pdf do diario",
            "rdo em pdf",
            "diário em pdf",
            "diario em pdf",
            "feche o rdo em pdf",
            "fechar o rdo em pdf",
            "exportar rdo",
            "exporte o rdo",
            "faça o pdf",
            "faca o pdf",
            "cópia em pdf",
            "copia em pdf",
        ],
    ):
        return {
            "intencao": "GERAR_PDF_RDO",
            "agente_destino": "AGENTE_RDO",
            "tipo_comando": "GERAR_PDF_RDO",
            "requer_aprovacao": True,
            "confianca": 0.86,
            "justificativa": "Solicitação registrada apenas como comando; nenhum PDF real é gerado nesta etapa.",
        }

    if menciona_rdo and has_any_term(
        texto,
        [
            "atualizar rdo",
            "alterar rdo",
            "corrigir rdo",
            "mudar rdo",
            "atualize o rdo",
            "atualiza o rdo",
            "atualizar o rdo",
            "atualize rdo",
            "atualiza rdo",
            "altere o rdo",
            "corrija o rdo",
            "lançar rdo",
            "lancar rdo",
            "registrar rdo",
            "registre oficialmente",
            "registrar oficialmente",
            "rdo oficial",
        ],
    ):
        return {
            "intencao": "ATUALIZAR_RDO",
            "agente_destino": "AGENTE_RDO",
            "tipo_comando": "ATUALIZAR_RDO",
            "requer_aprovacao": True,
            "confianca": 0.88,
            "justificativa": "Alteração de RDO é ação sensível e não altera RDO oficial nesta etapa.",
        }

    if menciona_rdo and has_any_term(
        texto,
        [
            "prepare",
            "preparar",
            "monte",
            "montar",
            "crie",
            "criar",
            "elabore",
            "elaborar",
            "rascunho",
        ],
    ):
        return {
            "intencao": "PREPARAR_RDO",
            "agente_destino": "AGENTE_RDO",
            "tipo_comando": "PREPARAR_RDO",
            "requer_aprovacao": False,
            "confianca": 0.88,
            "justificativa": "Mensagem indica preparação de rascunho de RDO sem oficialização.",
        }

    if menciona_rdo:
        return {
            "intencao": "CONSULTAR_RDO",
            "agente_destino": "AGENTE_RDO",
            "tipo_comando": "CONSULTAR_RDO",
            "requer_aprovacao": False,
            "confianca": 0.82,
            "justificativa": "Mensagem classificada como consulta executiva de RDO.",
        }

    if has_any_term(texto, ["pendência", "pendencia", "problema", "atraso", "crítico", "critico"]):
        return {
            "intencao": "CONSULTAR_PENDENCIAS",
            "agente_destino": "AGENTE_PENDENCIAS",
            "tipo_comando": "CONSULTAR_PENDENCIAS",
            "requer_aprovacao": False,
            "confianca": 0.82,
            "justificativa": "Mensagem classificada como consulta de pendências.",
        }

    if has_any_term(texto, ["documento", "nota fiscal", "nf-e", "nfe", "contrato", "arquivo"]):
        return {
            "intencao": "CONSULTAR_DOCUMENTOS",
            "agente_destino": "AGENTE_DOCUMENTOS",
            "tipo_comando": "CONSULTAR_DOCUMENTOS",
            "requer_aprovacao": False,
            "confianca": 0.80,
            "justificativa": "Mensagem classificada como consulta documental.",
        }

    if has_any_term(texto, ["orçamento", "orcamento", "cotação", "cotacao", "proposta"]):
        return {
            "intencao": "CONSULTAR_ORCAMENTOS",
            "agente_destino": "AGENTE_ORCAMENTOS",
            "tipo_comando": "CONSULTAR_ORCAMENTOS",
            "requer_aprovacao": False,
            "confianca": 0.78,
            "justificativa": "Mensagem classificada como consulta de orçamentos.",
        }

    if has_any_term(texto, ["cronograma", "prazo", "marco", "planejamento"]):
        return {
            "intencao": "CONSULTAR_CRONOGRAMA",
            "agente_destino": "AGENTE_CRONOGRAMA",
            "tipo_comando": "CONSULTAR_CRONOGRAMA",
            "requer_aprovacao": False,
            "confianca": 0.78,
            "justificativa": "Mensagem classificada como consulta de cronograma.",
        }

    return {
        "intencao": "MENSAGEM_EXECUTIVA_GERAL",
        "agente_destino": "AGENTE_007_ORQUESTRADOR_EXECUTIVO",
        "tipo_comando": "TRIAGEM_EXECUTIVA",
        "requer_aprovacao": False,
        "confianca": 0.50,
        "justificativa": "Mensagem sem intenção executiva específica no classificador determinístico mínimo.",
    }


def table_exists(conn, qualified_name: str) -> bool:
    with conn.cursor() as cur:
        cur.execute("SELECT to_regclass(%s);", (qualified_name,))
        row = cur.fetchone()
    return bool(row and row[0])


AGENTE_008_SEGURANCA_CONSULTA = {
    "modo": "CONSULTA",
    "altera_cronograma": False,
    "executa_rpa": False,
    "sincroniza_openproject": False,
    "altera_rdo_oficial": False,
    "envia_terceiros": False,
    "envia_arquivos": False,
    "gera_links_publicos": False,
    "altera_minio": False,
    "aprova_execucao_automaticamente": False,
    "linguagem": "CONSULTIVA",
}

AGENTE_008_SEGURANCA_CONTROLE_DOCUMENTAL = {
    "modo": "CONTROLE_DOCUMENTAL_CONSULTIVO",
    "altera_cronograma": False,
    "executa_rpa": False,
    "sincroniza_openproject": False,
    "altera_rdo_oficial": False,
    "envia_terceiros": False,
    "envia_arquivos": False,
    "gera_links_publicos": False,
    "altera_minio": False,
    "move_arquivos_minio": False,
    "apaga_arquivos_minio": False,
    "libera_execucao_campo": False,
    "aprova_execucao_automaticamente": False,
    "documento_vira_vigente_automaticamente": False,
    "linguagem": "CONSULTIVA",
}

AGENTE_008_SEGURANCA_APROVACAO_DOCUMENTAL = {
    **AGENTE_008_SEGURANCA_CONTROLE_DOCUMENTAL,
    "modo": "APROVACAO_TECNICA_DOCUMENTAL_CONTROLADA",
    "documento_vira_vigente_apenas_por_decisao_explicita": True,
}

AGENTE_008_SEGURANCA_LIBERACAO_CAMPO = {
    **AGENTE_008_SEGURANCA_CONTROLE_DOCUMENTAL,
    "modo": "LIBERACAO_DOCUMENTAL_CAMPO_CONTROLADA",
    "cria_ordem_servico": False,
    "autoriza_execucao_servico": False,
    "liberacao_apenas_para_uso_documental_campo": True,
}


def extrair_decisao_revisao_documental_telegram(conteudo: Optional[str]) -> dict[str, Any]:
    texto = normalizar_texto_comparacao(conteudo)
    padroes = (
        ("SOLICITAR_AJUSTES", r"^(?:solicitar|pedir) ajustes na (?:revisao|documento|projeto)|^corrigir (?:revisao|documento|projeto)"),
        ("REJEITAR", r"^(?:rejeitar|nao aprovar) (?:a )?(?:revisao|documento|projeto)"),
        ("MARCAR_EM_ANALISE", r"^(?:marcar (?:a )?revisao .* em analise|colocar (?:a )?revisao .* em analise tecnica|analisar (?:a )?revisao)"),
        ("APROVAR_COMO_VIGENTE", r"^(?:aprovar (?:a )?(?:revisao|documento|projeto).* como vigente|aprovo (?:a )?revisao|marcar (?:a )?revisao .* como vigente)"),
    )
    decisao = next((valor for valor, padrao in padroes if re.search(padrao, texto)), None)
    if not decisao:
        return {}
    identificador = re.search(r"(?:#\s*|\b(?:revisao|documento|projeto)\s+)(\d+)\b", texto)
    extra = re.search(r"(?:\bmotivo\b|\bporque\b|:)\s*(.+)$", (conteudo or "").strip(), re.IGNORECASE)
    detalhe = extra.group(1).strip() if extra else None
    return {
        "decisao": decisao,
        "revisao_documental_id": int(identificador.group(1)) if identificador else None,
        "motivo": detalhe if decisao == "REJEITAR" else None,
        "observacao": detalhe if decisao in {"SOLICITAR_AJUSTES", "MARCAR_EM_ANALISE"} else None,
    }


def extrair_liberacao_revisao_campo_telegram(conteudo: Optional[str]) -> dict[str, Any]:
    texto = normalizar_texto_comparacao(conteudo)
    padroes = (
        ("SUSPENDER_LIBERACAO_CAMPO", r"^(?:suspender (?:liberacao da )?(?:revisao|documento|projeto)|pausar uso da (?:revisao|documento|projeto)).*(?:campo|em campo|no campo)?"),
        ("REVOGAR_LIBERACAO_CAMPO", r"^(?:revogar (?:liberacao da )?(?:revisao|documento|projeto)|bloquear uso da (?:revisao|documento|projeto)).*(?:campo|em campo|do campo)?"),
        ("LIBERAR_PARA_CAMPO", r"^(?:pode )?liberar (?:a )?(?:revisao|documento|projeto).*(?:para (?:uso em )?campo)$"),
    )
    decisao = next((valor for valor, padrao in padroes if re.search(padrao, texto)), None)
    if not decisao:
        return {}
    identificador = re.search(r"(?:#\s*|\b(?:revisao|documento|projeto)\s+)(\d+)\b", texto)
    extra = re.search(r"(?:\bmotivo\b|\bporque\b|\binstrucoes?\b|:)[\s:]*(.+)$", (conteudo or "").strip(), re.IGNORECASE)
    detalhe = extra.group(1).strip() if extra else None
    return {
        "decisao": decisao,
        "revisao_documental_id": int(identificador.group(1)) if identificador else None,
        "motivo": detalhe if decisao != "LIBERAR_PARA_CAMPO" else None,
        "observacao": detalhe,
        "instrucoes_campo": detalhe if decisao == "LIBERAR_PARA_CAMPO" else None,
    }

AGENTE_008_SEGURANCA_REGISTRO_INTERNO = {
    **AGENTE_008_SEGURANCA_CONSULTA,
    "modo": "REGISTRO_INTERNO",
}

AGENTE_008_SEGURANCA_EXPORTACAO = {
    **AGENTE_008_SEGURANCA_CONSULTA,
    "formato": "MARKDOWN",
    "gera_pdf": False,
}

AGENTE_008_SEGURANCA_PDF = {
    **AGENTE_008_SEGURANCA_CONSULTA,
    "formato": "PDF",
    "altera_minio": False,
}

AGENTE_008_SEGURANCA_APROVACAO_PDF = {
    "modo": "APROVACAO_EXECUTIVA_CONTROLADA",
    "altera_cronograma": False,
    "executa_rpa": False,
    "sincroniza_openproject": False,
    "altera_rdo_oficial": False,
    "envia_terceiros": False,
    "envia_arquivos": False,
    "gera_links_publicos": False,
    "altera_minio": False,
    "gera_pdf": False,
    "aprova_execucao_automaticamente": False,
    "aprova_apenas_documento_para_uso_interno": True,
    "linguagem": "CONSULTIVA",
}

AGENTE_008_SEGURANCA_SOLICITACAO_ENVIO = {
    "modo": "SOLICITACAO_ENVIO_CONTROLADA",
    "altera_cronograma": False,
    "executa_rpa": False,
    "sincroniza_openproject": False,
    "altera_rdo_oficial": False,
    "envia_terceiros": False,
    "envio_executado": False,
    "envia_arquivos": False,
    "anexou_arquivo": False,
    "gera_links_publicos": False,
    "gerou_presigned_url": False,
    "altera_minio": False,
    "gera_pdf": False,
    "aprova_execucao_automaticamente": False,
    "apenas_registra_solicitacao": True,
    "linguagem": "CONSULTIVA",
}


def extrair_alvo_analise_documental(conteudo: Optional[str]) -> dict[str, Any]:
    texto = re.sub(r"\s+", " ", (conteudo or "").strip().rstrip("?.!"))
    correspondencia = re.fullmatch(
        r"(?:analisar|resumir|verificar)\s+(?:documento|projeto)\s+(.+)",
        texto,
        flags=re.IGNORECASE,
    )
    if not correspondencia:
        return {}
    alvo = correspondencia.group(1).strip()
    if alvo.isdigit():
        return {"documento_id": int(alvo)}
    return {"termo": alvo}


def extrair_filtros_documentos_telegram(conteudo: Optional[str]) -> dict[str, Optional[str]]:
    texto = normalizar_texto_comparacao(conteudo)
    disciplina = None
    for termo, valor in (
        ("arquitetura", "arquitetura"),
        ("eletrica", "eletrica"),
        ("hidraulica", "hidraulica"),
        ("hidralica", "hidraulica"),
        ("hidr", "hidraulica"),
        ("estrutura", "estrutura"),
        ("luminotecnico", "luminotecnico"),
    ):
        if termo in texto:
            disciplina = valor
            break

    extensao = next((item for item in ("dwg", "pdf") if item in texto.split()), None)
    termo = "refeitório" if "refeitório" in texto else None
    return {"disciplina": disciplina, "extensao": extensao, "termo": termo}


def extrair_filtros_classificacao_telegram(conteudo: Optional[str]) -> dict[str, Any]:
    texto = normalizar_texto_comparacao(conteudo)
    filtros: dict[str, Any] = {}
    if "obsoleto" in texto:
        filtros["eh_obsoleto"] = True
    if re.search(r"\bas built\b", texto):
        filtros["eh_as_built"] = True
    if "refeitorio" in texto:
        filtros["area"] = "refeitorio"
    for termo, disciplina in (
        ("hidraulica", "hidraulica"),
        ("hidralica", "hidraulica"),
        ("hidr", "hidraulica"),
        ("eletrica", "eletrica"),
        ("luminotecnico", "luminotecnico"),
    ):
        if termo in texto:
            filtros["disciplina"] = disciplina
            break
    if texto.startswith("ultima revisao"):
        filtros["modo"] = "ULTIMA_REVISAO"
    return filtros


def extrair_ultima_revisao_telegram(conteudo: Optional[str]) -> dict[str, Any]:
    filtros = extrair_filtros_documentos_telegram(conteudo)
    return {
        "disciplina": filtros["disciplina"],
        "area": None,
        "limite_candidatos": 10,
    }


def extrair_riscos_documentais_telegram(conteudo: Optional[str]) -> dict[str, Any]:
    texto = normalizar_texto_comparacao(conteudo)
    area = None
    if "sala de jogos" in texto:
        area = "sala_de_jogos"
    elif "refeitorio" in texto:
        area = "refeitorio"

    disciplina = None
    for termo, valor in (
        ("luminotecnico", "luminotecnico"),
        ("hidraulico", "hidraulica"),
        ("hidraulica", "hidraulica"),
        ("hidralico", "hidraulica"),
        ("hidralica", "hidraulica"),
        ("eletrico", "eletrica"),
        ("eletrica", "eletrica"),
        ("estrutural", "estrutura"),
        ("estrutura", "estrutura"),
    ):
        if termo in texto:
            disciplina = valor
            break
    return {"area": area, "disciplina": disciplina, "limite_amostras": 10}


def extrair_area_diagnostico_operacional(conteudo: Optional[str]) -> Optional[str]:
    texto = normalizar_texto_comparacao(conteudo)
    if "sala de jogos" in texto:
        return "sala_de_jogos"
    if "refeitorio" in texto:
        return "refeitorio"
    return None


def extrair_area_plano_operacional(conteudo: Optional[str]) -> Optional[str]:
    return extrair_area_diagnostico_operacional(conteudo)


def extrair_acao_operacional_telegram(conteudo: Optional[str]) -> dict[str, Any]:
    texto_original = re.sub(r"\s+", " ", (conteudo or "").strip().rstrip("?.!"))
    texto = normalizar_texto_comparacao(texto_original)
    filtros = extrair_riscos_documentais_telegram(conteudo)
    titulo = re.sub(
        r"^(?:registrar\s+ação(?:\s+operacional)?|registrar\s+acao(?:\s+operacional)?|abrir\s+ação\s+para|abrir\s+acao\s+para|criar\s+ação\s+para|criar\s+acao\s+para)\s*",
        "", texto_original, flags=re.IGNORECASE,
    ).strip()
    if texto.startswith("registrar acao critica"):
        titulo = re.sub(r"^registrar\s+ação\s+crítica\s+de\s+|^registrar\s+acao\s+critica\s+de\s+", "Validar ", texto_original, flags=re.IGNORECASE).strip()
    if not titulo:
        titulo = "Ação operacional da obra"
    return {
        "area": filtros["area"],
        "disciplina": filtros["disciplina"],
        "titulo": titulo[0].upper() + titulo[1:],
        "prioridade": "ALTA" if "critica" in texto or "obsoleto" in texto else "MEDIA",
    }


def extrair_atualizacao_acao_telegram(conteudo: Optional[str]) -> dict[str, Any]:
    texto = normalizar_texto_comparacao(conteudo)
    texto_original = (conteudo or "").strip()
    correspondencia = re.search(r"\b(\d+)\b", texto)
    resultado: dict[str, Any] = {"acao_id": int(correspondencia.group(1)) if correspondencia else 0}
    if texto.startswith("cancelar"):
        resultado["status"] = "CANCELADA"
    elif "em andamento" in texto:
        resultado["status"] = "EM_ANDAMENTO"
    elif "concluida" in texto:
        resultado["status"] = "CONCLUIDA"
    prioridade = re.search(r"(?:para\s+|acao\s+\d+\s+)(baixa|media|alta|critica)$", texto)
    if "prioridade" in texto and prioridade:
        resultado["prioridade"] = prioridade.group(1).upper()
    prazo = re.search(r"(\d{2}/\d{2}/\d{4})$", texto_original)
    if "prazo" in texto and prazo:
        resultado["prazo"] = datetime.strptime(prazo.group(1), "%d/%m/%Y").date().isoformat()
    responsavel = re.search(r"(?:como|para)\s+(.+)$", texto)
    if (texto.startswith("definir responsavel") or texto.startswith("atribuir acao")) and responsavel:
        resultado["responsavel"] = responsavel.group(1).strip().title()
    observacao = re.search(r":\s*(.+)$", texto_original)
    if "observacao na acao" in texto and observacao:
        resultado["observacao"] = observacao.group(1).strip()
    return resultado


def extrair_detalhe_acao_telegram(conteudo: Optional[str]) -> dict[str, Any]:
    correspondencia = re.search(r"\b(\d+)\b", normalizar_texto_comparacao(conteudo))
    return {"acao_id": int(correspondencia.group(1)) if correspondencia else 0, "incluir_historico": True}


def extrair_pendencia_documental_telegram(conteudo: Optional[str]) -> dict[str, Any]:
    texto_original = re.sub(r"\s+", " ", (conteudo or "").strip())
    texto = normalizar_texto_comparacao(texto_original)
    filtros = extrair_riscos_documentais_telegram(conteudo)
    if "obsoleto" in texto:
        motivo = "DOCUMENTO_OBSOLETO"
    elif "revis" in texto:
        motivo = "REVISAO_NAO_IDENTIFICADA"
    else:
        motivo = "RISCO_DOCUMENTAL"
    return {
        "area": filtros["area"],
        "disciplina": filtros["disciplina"],
        "motivo": motivo,
        "descricao": texto_original or None,
    }


def extrair_documento_id_telegram(conteudo: Optional[str]) -> dict[str, int]:
    correspondencia = re.search(r"\bdocumento\s+(\d+)\b", conteudo or "", re.IGNORECASE)
    return {"documento_id": int(correspondencia.group(1))} if correspondencia else {}


def montar_resposta_documentos_resumo_telegram(obra_codigo: str) -> str:
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT COUNT(*)
                FROM documentos_minio_obra
                WHERE obra_codigo = %(obra_codigo)s;
                """,
                {"obra_codigo": obra_codigo},
            )
            total = cur.fetchone()[0]
            cur.execute(
                """
                SELECT COALESCE(disciplina_original, 'Sem disciplina'), COUNT(*)
                FROM documentos_minio_obra
                WHERE obra_codigo = %(obra_codigo)s
                GROUP BY disciplina_original
                ORDER BY COUNT(*) DESC, disciplina_original NULLS LAST
                LIMIT 8;
                """,
                {"obra_codigo": obra_codigo},
            )
            disciplinas = cur.fetchall()

    if not total:
        return "Não encontrei documentos para esse filtro."
    linhas = [f"Documentos da {obra_codigo}: {total} indexados."]
    linhas.extend(f"• {disciplina}: {quantidade}" for disciplina, quantidade in disciplinas)
    return "\n".join(linhas)


def montar_resposta_documentos_indexados_telegram(
    obra_codigo: str, conteudo: Optional[str], limite: int = 10
) -> str:
    filtros = extrair_filtros_documentos_telegram(conteudo)
    filtros_sql = ["obra_codigo = %(obra_codigo)s"]
    params: dict[str, Any] = {"obra_codigo": obra_codigo, "limite_consulta": limite + 1}
    if filtros["disciplina"]:
        filtros_sql.append(
            "TRANSLATE(LOWER(disciplina_original), "
            "'áàâãéêíóôõúüç', 'aaaaeeiooouuc') LIKE %(disciplina)s"
        )
        params["disciplina"] = f"%{filtros['disciplina']}%"
    if filtros["extensao"]:
        filtros_sql.append("LOWER(extensao) = %(extensao)s")
        params["extensao"] = filtros["extensao"]
    if filtros["termo"]:
        filtros_sql.append("(nome_arquivo ILIKE %(termo)s OR object_key ILIKE %(termo)s)")
        params["termo"] = f"%{filtros['termo']}%"

    sql = f"""
        SELECT nome_arquivo, disciplina_original, extensao, bucket, object_key
        FROM documentos_minio_obra
        WHERE {" AND ".join(filtros_sql)}
        ORDER BY atualizado_em DESC NULLS LAST, criado_em DESC NULLS LAST, id DESC
        LIMIT %(limite_consulta)s;
    """
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            documentos = cur.fetchall()

    if not documentos:
        return "Não encontrei documentos para esse filtro."
    linhas = [f"Documentos encontrados na {obra_codigo}:"]
    for nome, disciplina, extensao, bucket, object_key in documentos[:limite]:
        linhas.append(
            f"• {nome}\n  {disciplina or 'Sem disciplina'} | {extensao or 'sem extensão'}\n"
            f"  s3://{bucket}/{object_key}"
        )
    if len(documentos) > limite:
        linhas.append("Mostrando os 10 primeiros resultados.")
    return "\n".join(linhas)


def serialize_date(value: Any) -> Optional[str]:
    return value.isoformat() if value else None


def serialize_numeric(value: Any) -> float:
    return float(value or 0)


def normalizar_texto_comparacao(valor: Optional[str]) -> str:
    """Normaliza somente para comparação; o valor original nunca é modificado."""
    texto = unicodedata.normalize("NFKD", valor or "")
    texto = "".join(caractere for caractere in texto if not unicodedata.combining(caractere))
    return re.sub(r"[^a-z0-9]+", " ", texto.casefold()).strip()


def extrair_decisao_pdf_relatorio_telegram(
    conteudo: Optional[str], decisao: str,
) -> dict[str, Any]:
    texto = re.sub(r"\s+", " ", (conteudo or "").strip())
    correspondencia_id = re.search(
        r"#\s*(\d+)\b|\b(?:pdf|relat[oó]rio(?:\s+semanal)?)\s*#?\s*(\d+)\b",
        texto,
        flags=re.IGNORECASE,
    )
    pdf_relatorio_id = None
    fim_id = 0
    if correspondencia_id:
        pdf_relatorio_id = int(
            correspondencia_id.group(1) or correspondencia_id.group(2)
        )
        fim_id = correspondencia_id.end()

    motivo_match = re.search(
        r"\bmotivo\s*:\s*(.+?)(?=\s+observa[cç][aã]o\s*:|$)",
        texto,
        flags=re.IGNORECASE,
    )
    observacao_match = re.search(
        r"\bobserva[cç][aã]o\s*:\s*(.+?)(?=\s+motivo\s*:|$)",
        texto,
        flags=re.IGNORECASE,
    )
    texto_extra = texto[fim_id:].strip(" .,:;-\n\t") if fim_id else ""
    if motivo_match or observacao_match:
        motivo = motivo_match.group(1).strip() if motivo_match else None
        observacao = observacao_match.group(1).strip() if observacao_match else None
    elif decisao == "REJEITAR":
        motivo, observacao = texto_extra or None, None
    else:
        motivo, observacao = None, texto_extra or None

    return {
        "pdf_relatorio_id": pdf_relatorio_id,
        "motivo": motivo,
        "observacao": observacao,
    }


def extrair_solicitacao_envio_relatorio_telegram(
    conteudo: Optional[str],
) -> dict[str, Any]:
    texto = re.sub(r"\s+", " ", (conteudo or "").strip())
    correspondencia_id = re.search(
        r"#\s*(\d+)\b|\b(?:pdf|relat[oó]rio)\s*#?\s*(\d+)\b",
        texto,
        flags=re.IGNORECASE,
    )
    pdf_relatorio_id = None
    if correspondencia_id:
        pdf_relatorio_id = int(
            correspondencia_id.group(1) or correspondencia_id.group(2)
        )

    texto_normalizado = normalizar_texto_comparacao(texto)
    if re.search(r"\be mail\b|\bemail\b", texto_normalizado):
        canal = "EMAIL"
    elif re.search(r"\bwhatsapp\b|\bzap\b", texto_normalizado):
        canal = "WHATSAPP"
    elif re.search(r"\btelegram\b", texto_normalizado):
        canal = "TELEGRAM"
    else:
        canal = "INTERNO"

    destinatario_match = re.search(
        r"\bpara\s+(.+?)(?=\s+(?:por\s+(?:e-?mail|email|whatsapp|zap)|"
        r"no\s+telegram)\b|$)",
        texto,
        flags=re.IGNORECASE,
    )
    destinatario_nome = None
    if destinatario_match:
        destinatario_nome = destinatario_match.group(1).strip(" .,:;-\n\t") or None

    return {
        "pdf_relatorio_id": pdf_relatorio_id,
        "canal_pretendido": canal,
        "destinatario_nome": destinatario_nome,
        "destinatario_contato": None,
        "mensagem": None,
    }


def classificar_documento_tecnico(
    nome_arquivo: Optional[str],
    object_key: Optional[str],
    pasta_origem: Optional[str],
    disciplina_original: Optional[str],
    extensao: Optional[str],
) -> dict[str, Any]:
    campos = {
        "nome_arquivo": nome_arquivo or "",
        "object_key": object_key or "",
        "pasta_origem": pasta_origem or "",
        "disciplina_original": disciplina_original or "",
        "extensao": extensao or "",
    }
    texto = normalizar_texto_comparacao(" ".join(campos.values()))
    texto_nome_caminho = normalizar_texto_comparacao(
        f"{campos['nome_arquivo']} {campos['object_key']}"
    )

    def primeiro_termo(opcoes: list[tuple[str, tuple[str, ...]]]) -> Optional[str]:
        for valor, termos in opcoes:
            if any(re.search(rf"\b{re.escape(termo)}\b", texto) for termo in termos):
                return valor
        return None

    area = primeiro_termo([
        ("refeitorio", ("refeitorio",)),
        ("sala_de_jogos", ("sala de jogos",)),
        ("area_1", ("area 1",)),
        ("area_2", ("area 2",)),
        ("area_3", ("area 3",)),
        ("cozinha", ("cozinha",)),
        ("sanitario", ("sanitario", "banheiro")),
        ("geral", ("geral",)),
    ])
    disciplina = primeiro_termo([
        ("deteccao_e_alarme", ("deteccao e alarme", "incendio")),
        ("luminotecnico", ("luminotecnico",)),
        ("terraplanagem", ("terraplanagem",)),
        ("topografia", ("topografia",)),
        ("renderizacao", ("renderizacao",)),
        ("arquitetura", ("arquitetura",)),
        ("estrutura", ("estrutura",)),
        ("eletrica", ("eletrica",)),
        ("hidraulica", ("hidraulica", "hidralica", "hidr")),
        ("forros", ("forros", "forro")),
    ])

    if re.search(r"\b(?:obsoleto|obsolete)\b", texto_nome_caminho):
        status_revisao = "OBSOLETO"
    elif re.search(r"\b(?:as built|asbult|as bult|asbuilt)\b", texto_nome_caminho):
        status_revisao = "AS_BUILT"
    elif re.search(r"\b(?:alteracao|alteracoes|alterado|revisado)\b", texto_nome_caminho):
        status_revisao = "ALTERACAO"
    elif re.search(r"\bpre executivo\b", texto_nome_caminho):
        status_revisao = "PRE_EXECUTIVO"
    elif re.search(r"\bexecutivo\b", texto_nome_caminho):
        status_revisao = "EXECUTIVO"
    else:
        status_revisao = "NAO_IDENTIFICADO"

    data_revisao = None
    data_encontrada = re.search(r"(?<!\d)(\d{2})[.\-/](\d{2})[.\-/](\d{4})(?!\d)", " ".join(campos.values()))
    if data_encontrada:
        try:
            data_revisao = date(
                int(data_encontrada.group(3)),
                int(data_encontrada.group(2)),
                int(data_encontrada.group(1)),
            )
        except ValueError:
            data_revisao = None

    revisao_encontrada = re.search(
        r"\b(?:r\s*|rev(?:isao)?\s*)(\d{1,3})\b", texto, flags=re.IGNORECASE
    )
    numero_revisao = f"R{int(revisao_encontrada.group(1)):02d}" if revisao_encontrada else None
    palavras_chave = [
        valor for valor in (area, disciplina, status_revisao if status_revisao != "NAO_IDENTIFICADO" else None, numero_revisao)
        if valor
    ]
    criterios = {
        "area": area is not None,
        "disciplina": disciplina is not None,
        "status_revisao": status_revisao != "NAO_IDENTIFICADO",
        "data_revisao": data_revisao is not None,
        "numero_revisao": numero_revisao is not None,
        "campos_avaliados": list(campos),
    }
    confianca = 0.30
    confianca += 0.15 if area else 0
    confianca += 0.20 if disciplina else 0
    confianca += 0.15 if status_revisao != "NAO_IDENTIFICADO" else 0
    confianca += 0.10 if data_revisao else 0
    confianca += 0.10 if numero_revisao else 0
    return {
        "area_detectada": area,
        "disciplina_detectada": disciplina,
        "status_revisao": status_revisao,
        "data_revisao_detectada": data_revisao,
        "eh_obsoleto": status_revisao == "OBSOLETO",
        "eh_as_built": status_revisao == "AS_BUILT",
        "numero_revisao": numero_revisao,
        "palavras_chave": palavras_chave[:6],
        "criterios_detectados": criterios,
        "confianca_classificacao": min(round(confianca, 2), 1.00),
    }


def fetch_status_counts(conn, table_name: str, status_column: str, obra_codigo: str) -> dict[str, int]:
    sql = f"""
    SELECT {status_column}, COUNT(*)
    FROM {table_name}
    WHERE obra_codigo = %(obra_codigo)s
    GROUP BY {status_column}
    ORDER BY {status_column};
    """
    with conn.cursor() as cur:
        cur.execute(sql, {"obra_codigo": obra_codigo})
        return {row[0]: row[1] for row in cur.fetchall()}


def formatar_atividade_gestao_operacional(row: tuple[Any, ...]) -> dict[str, Any]:
    return {
        "codigo_atividade": row[0],
        "codigo_eap": row[1],
        "codigo_area": row[2],
        "descricao": row[3],
        "disciplina": row[4],
        "frente_servico": row[5],
        "data_inicio_planejada": serialize_date(row[6]),
        "data_fim_planejada": serialize_date(row[7]),
        "data_inicio_reprogramada": serialize_date(row[8]),
        "data_fim_reprogramada": serialize_date(row[9]),
        "percentual_planejado": serialize_numeric(row[10]),
        "percentual_real": serialize_numeric(row[11]),
        "status_atividade": row[12],
        "responsavel": row[13],
        "horizonte_planejamento": row[14],
        "criticidade": row[15],
        "restricoes_resumo": row[16] or [],
    }


def listar_atividades_gestao_operacional(
    conn,
    obra_codigo: str,
    horizonte_planejamento: str,
    inicio_offset_dias: int,
    fim_offset_dias: int,
) -> list[dict[str, Any]]:
    sql = """
    SELECT
        codigo_atividade,
        codigo_eap,
        codigo_area,
        descricao,
        disciplina,
        frente_servico,
        data_inicio_planejada,
        data_fim_planejada,
        data_inicio_reprogramada,
        data_fim_reprogramada,
        percentual_planejado,
        percentual_real,
        status_atividade,
        responsavel,
        horizonte_planejamento,
        criticidade,
        restricoes_resumo
    FROM atividades_cronograma
    WHERE obra_codigo = %(obra_codigo)s
      AND (
          horizonte_planejamento = %(horizonte_planejamento)s
          OR (
              COALESCE(data_inicio_reprogramada, data_inicio_planejada) >= CURRENT_DATE + %(inicio_offset_dias)s::int
              AND COALESCE(data_inicio_reprogramada, data_inicio_planejada) <= CURRENT_DATE + %(fim_offset_dias)s::int
          )
      )
    ORDER BY
        COALESCE(data_inicio_reprogramada, data_inicio_planejada) NULLS LAST,
        codigo_area NULLS LAST,
        codigo_atividade
    LIMIT 200;
    """
    with conn.cursor() as cur:
        cur.execute(
            sql,
            {
                "obra_codigo": obra_codigo,
                "horizonte_planejamento": horizonte_planejamento,
                "inicio_offset_dias": inicio_offset_dias,
                "fim_offset_dias": fim_offset_dias,
            },
        )
        return [formatar_atividade_gestao_operacional(row) for row in cur.fetchall()]


def listar_pendencias_rdo(conn, obra_codigo: Optional[str]) -> list[dict[str, Any]]:
    if not obra_codigo:
        return []

    if table_exists(conn, "public.pendencias_obra"):
        sql = """
        SELECT
            descricao,
            categoria,
            prioridade,
            status_pendencia,
            responsavel_acao,
            prazo_limite
        FROM public.pendencias_obra
        WHERE obra_codigo = %(obra_codigo)s
        ORDER BY criado_em DESC
        LIMIT 5;
        """
        with conn.cursor() as cur:
            cur.execute(sql, {"obra_codigo": obra_codigo})
            return [
                {
                    "descricao": row[0],
                    "categoria": row[1],
                    "prioridade": row[2],
                    "status": row[3],
                    "responsavel": row[4],
                    "prazo_limite": row[5].isoformat() if row[5] else None,
                }
                for row in cur.fetchall()
            ]

    if table_exists(conn, "core.pendencias"):
        sql = """
        SELECT
            p.titulo,
            p.descricao,
            p.categoria,
            p.prioridade,
            p.status,
            p.responsavel,
            p.data_limite
        FROM core.pendencias AS p
        JOIN core.obras AS o ON o.id = p.obra_id
        WHERE o.codigo_obra = %(obra_codigo)s
        ORDER BY p.data_abertura DESC
        LIMIT 5;
        """
        with conn.cursor() as cur:
            cur.execute(sql, {"obra_codigo": obra_codigo})
            return [
                {
                    "titulo": row[0],
                    "descricao": row[1],
                    "categoria": row[2],
                    "prioridade": row[3],
                    "status": row[4],
                    "responsavel": row[5],
                    "prazo_limite": row[6].isoformat() if row[6] else None,
                }
                for row in cur.fetchall()
            ]

    return []


def listar_eventos_rdo(conn, obra_codigo: Optional[str]) -> list[dict[str, Any]]:
    if not obra_codigo or not table_exists(conn, "core.eventos_obra"):
        return []

    sql = """
    SELECT
        e.titulo,
        e.descricao,
        e.tipo_evento,
        e.data_evento,
        e.responsavel,
        e.severidade
    FROM core.eventos_obra AS e
    JOIN core.obras AS o ON o.id = e.obra_id
    WHERE o.codigo_obra = %(obra_codigo)s
    ORDER BY e.data_evento DESC NULLS LAST, e.created_at DESC
    LIMIT 5;
    """
    with conn.cursor() as cur:
        cur.execute(sql, {"obra_codigo": obra_codigo})
        return [
            {
                "titulo": row[0],
                "descricao": row[1],
                "tipo_evento": row[2],
                "data_evento": row[3].isoformat() if row[3] else None,
                "responsavel": row[4],
                "severidade": row[5],
            }
            for row in cur.fetchall()
        ]


def montar_resultado_agente_rdo(
    comando: dict[str, Any],
    pendencias: list[dict[str, Any]],
    eventos: list[dict[str, Any]],
) -> dict[str, Any]:
    tipo_comando = comando["tipo_comando"]
    payload_comando = comando["payload_comando"] or {}
    entrada = payload_comando.get("entrada", {})
    conteudo = entrada.get("conteudo")

    tipo_resultado = "RASCUNHO_RDO" if tipo_comando == "PREPARAR_RDO" else "RESUMO_EXECUTIVO_RDO"

    return {
        "tipo_resultado": tipo_resultado,
        "agente": "AGENTE_002_GERADOR_RDO",
        "agente_alias_compatibilidade": "AGENTE_RDO",
        "modo_processamento": "DETERMINISTICO_MOCK",
        "id_comando": str(comando["id_comando"]),
        "correlation_id": str(comando["correlation_id"]),
        "obra_codigo": comando["obra_codigo"],
        "tipo_comando": tipo_comando,
        "resumo": {
            "titulo": "Resumo executivo operacional de RDO",
            "mensagem_origem": conteudo,
            "total_pendencias_consultadas": len(pendencias),
            "total_eventos_consultados": len(eventos),
            "observacao": "Resultado preliminar gerado sem LLM externo e sem ações externas.",
        },
        "rascunho_rdo": {
            "status": "RASCUNHO_NAO_OFICIAL",
            "atividades": [],
            "pendencias_referenciadas": pendencias,
            "eventos_referenciados": eventos,
            "campos_a_confirmar": [
                "data do RDO",
                "atividades executadas",
                "equipe",
                "equipamentos",
                "condicoes_climaticas",
                "evidencias",
            ],
        },
        "controles_operacionais": {
            "alterou_rdo_oficial": False,
            "gerou_pdf_real": False,
            "imprimiu": False,
            "enviou_mensagem_terceiros": False,
            "conectou_openclaw": False,
            "executou_rpa": False,
            "requer_aprovacao_para_oficializar": True,
        },
        "gerado_em": datetime.now(timezone.utc).isoformat(),
    }


def montar_mensagem_resposta_executiva_rdo(
    obra_codigo: Optional[str],
    tipo_resultado: Optional[str],
) -> Optional[str]:
    if tipo_resultado == "RASCUNHO_RDO":
        return (
            f"Rascunho de RDO preparado para {obra_codigo}. "
            "Status: rascunho não oficial. Nenhuma ação externa foi executada. "
            "Para oficializar, será necessária aprovação."
        )

    if tipo_resultado == "RESUMO_EXECUTIVO_RDO":
        return (
            f"Resumo executivo de RDO preparado para {obra_codigo}. "
            "Nenhuma ação externa foi executada."
        )

    return None


def inferir_tipo_placa_aviso(conteudo: Optional[str]) -> str:
    texto = (conteudo or "").lower()

    if has_any_term(texto, ["epi", "capacete", "bota", "luva", "óculos", "oculos"]):
        return "SEGURANCA_USO_OBRIGATORIO_EPI"
    if has_any_term(texto, ["canteiro", "obra"]):
        return "AVISO_CANTEIRO_OBRA"
    if has_any_term(texto, ["comunicado", "comunicação", "comunicacao"]):
        return "COMUNICADO_OPERACIONAL"

    return "PLACA_AVISO_GERAL"


def inferir_local_instalacao_placa(conteudo: Optional[str]) -> Optional[str]:
    texto = (conteudo or "").lower()

    if "entrada" in texto or "portaria" in texto:
        return "Entrada/portaria da obra"
    if "canteiro" in texto:
        return "Canteiro de obras"
    if "refeitório" in texto or "refeitorio" in texto:
        return "Refeitório"
    if "almoxarifado" in texto:
        return "Almoxarifado"
    if "escada" in texto or "acesso" in texto:
        return "Área de acesso/circulação"

    return None


def extrair_campos_explicitos_placa(conteudo: Optional[str]) -> dict[str, str]:
    texto = conteudo or ""
    labels: list[tuple[str, str]] = [
        ("titulo_cabecalho", r"cabe[çc]alho"),
        ("titulo_cabecalho", r"t[íi]tulo"),
        ("texto_principal", r"frase principal"),
        ("texto_principal", r"texto principal"),
        ("texto_secundario", r"frase inferior"),
        ("texto_secundario", r"texto inferior"),
        ("texto_secundario", r"texto secund[áa]rio"),
        ("descricao_pictograma", r"pictograma"),
        ("local_instalacao_sugerido", r"local"),
        ("tipo_placa", r"tipo"),
        ("tema_placa", r"tema"),
    ]
    label_union = "|".join(label for _, label in labels)
    campos: dict[str, str] = {}

    for campo, label in labels:
        match = re.search(
            rf"(?ims)^\s*{label}\s*:\s*(.+?)(?=^\s*(?:{label_union})\s*:|\Z)",
            texto,
        )
        if match:
            valor = re.sub(r"\s+", " ", match.group(1)).strip(" -\t\r\n")
            if valor and campo not in campos:
                campos[campo] = valor

    return campos


def normalizar_label_operacional(value: Optional[str]) -> Optional[str]:
    texto = re.sub(r"\s+", " ", value or "").strip()
    if not texto:
        return None
    return re.sub(r"[^0-9A-Za-zÀ-ÿ]+", "_", texto).strip("_").upper() or None


def normalizar_conteudo_variavel_placa(
    titulo_cabecalho: str,
    texto_principal: str,
    texto_secundario: str,
    descricao_pictograma: Optional[str],
    local_instalacao_sugerido: Optional[str],
    tipo_placa: str,
    tema_placa: Optional[str],
) -> dict[str, Optional[str]]:
    titulo_normalizado = re.sub(r"\s+", " ", titulo_cabecalho or "").strip()
    texto_principal_normalizado = re.sub(r"\s+", " ", texto_principal or "").strip()
    texto_secundario_normalizado = re.sub(r"\s+", " ", texto_secundario or "").strip()

    if len(titulo_normalizado) > 22:
        texto_longo = titulo_normalizado
        titulo_normalizado = "Atenção"
        if texto_principal_normalizado and texto_principal_normalizado != texto_longo:
            texto_principal_normalizado = f"{texto_longo} {texto_principal_normalizado}"
        else:
            texto_principal_normalizado = texto_longo

    return {
        "titulo_cabecalho": titulo_normalizado or "Atenção",
        "texto_principal": texto_principal_normalizado
        or "Aviso preliminar para comunicação visual da obra.",
        "texto_secundario": texto_secundario_normalizado
        or "Validar conteúdo, pictograma, dimensões e local antes de uso.",
        "descricao_pictograma": (
            re.sub(r"\s+", " ", descricao_pictograma).strip()
            if descricao_pictograma
            else None
        ),
        "local_instalacao_sugerido": (
            re.sub(r"\s+", " ", local_instalacao_sugerido).strip()
            if local_instalacao_sugerido
            else None
        ),
        "tipo_placa": tipo_placa,
        "tema_placa": (
            re.sub(r"\s+", " ", tema_placa).strip()
            if tema_placa
            else None
        ),
    }


def montar_template_fixo_placa_aviso() -> dict[str, Any]:
    return json.loads(
        json.dumps(TEMPLATE_VISUAL_OBRA_CAIO_SUM_V1, ensure_ascii=False)
    )


def montar_resultado_agente_comunicacao_obra(comando: dict[str, Any]) -> dict[str, Any]:
    payload_comando = comando["payload_comando"] or {}
    entrada = payload_comando.get("entrada", {})
    conteudo = entrada.get("conteudo")
    campos_explicitos = extrair_campos_explicitos_placa(conteudo)
    tipo_placa = (
        normalizar_label_operacional(campos_explicitos.get("tipo_placa"))
        or inferir_tipo_placa_aviso(conteudo)
    )
    tema_placa = campos_explicitos.get("tema_placa")
    cor_base = "azul-petroleo"
    estilo_visual_referencia = "docs/reference/placas/README.md"

    if tipo_placa == "SEGURANCA_USO_OBRIGATORIO_EPI":
        titulo = "USO OBRIGATÓRIO DE EPI"
        mensagem_principal = "O acesso à área de obra exige uso dos EPIs definidos para a atividade."
        mensagem_secundaria = (
            "Antes do uso oficial, validar texto, pictogramas, local de instalação "
            "e requisitos aplicáveis com o responsável técnico/segurança do trabalho."
        )
        formato_sugerido = "A3 vertical, alta legibilidade, material resistente ao ambiente da obra"
    elif tipo_placa == "AVISO_CANTEIRO_OBRA":
        titulo = "AVISO AO CANTEIRO"
        mensagem_principal = "Atenção às orientações operacionais e de segurança da obra."
        mensagem_secundaria = (
            "Conteúdo preliminar para revisão interna; validar com responsável técnico/"
            "segurança do trabalho antes de uso oficial."
        )
        formato_sugerido = "A3 vertical ou A4 horizontal, conforme ponto de instalação"
    elif tipo_placa == "COMUNICADO_OPERACIONAL":
        titulo = "COMUNICADO DA OBRA"
        mensagem_principal = "Comunicado operacional para equipes e visitantes da obra."
        mensagem_secundaria = (
            "Revisar destinatários, data de validade e responsável pela autorização "
            "antes de qualquer divulgação oficial."
        )
        formato_sugerido = "A4 vertical, linguagem objetiva e identificação da obra"
    else:
        titulo = "PLACA DE AVISO"
        mensagem_principal = "Aviso preliminar para comunicação visual da obra."
        mensagem_secundaria = (
            "Rascunho não oficial. O conteúdo deve ser validado pelo responsável "
            "técnico/segurança do trabalho antes de uso oficial."
        )
        formato_sugerido = "A3 vertical ou A4 vertical, conforme distância de leitura"

    conteudo_variavel = normalizar_conteudo_variavel_placa(
        campos_explicitos.get("titulo_cabecalho", titulo),
        campos_explicitos.get("texto_principal", mensagem_principal),
        campos_explicitos.get("texto_secundario", mensagem_secundaria),
        campos_explicitos.get("descricao_pictograma"),
        campos_explicitos.get("local_instalacao_sugerido")
        or inferir_local_instalacao_placa(conteudo),
        tipo_placa,
        tema_placa,
    )
    tipo_icone = "triangulo_amarelo_atencao"
    area_pictograma = {
        "posicao": "corpo branco central",
        "composicao": "placeholder técnico textual em círculo azul-petróleo fixo de 15 cm",
        "status": "placeholder_tecnico_sem_imagem_externa",
    }
    template_fixo = montar_template_fixo_placa_aviso()
    observacao_validacao_tecnica = (
        "Rascunho não oficial baseado na referência visual Obra-Caio/SUM. "
        "Validar texto, pictograma, local, dimensões e requisitos aplicáveis "
        "com o responsável técnico/segurança do trabalho antes de uso oficial."
    )

    campos_a_confirmar = [
        "texto final autorizado",
        "responsavel_tecnico_ou_seguranca_do_trabalho",
        "local_exato_de_instalacao",
        "dimensoes_finais",
        "material_da_placa",
        "necessidade_de_pictogramas",
    ]
    if conteudo_variavel["local_instalacao_sugerido"] is None:
        campos_a_confirmar.append("local_instalacao")

    return {
        "tipo_resultado": "RASCUNHO_PLACA_AVISO",
        "agente": "AGENTE_006_COMUNICACAO_VISUAL_OBRA",
        "modo_processamento": "DETERMINISTICO_MOCK",
        "id_comando": str(comando["id_comando"]),
        "correlation_id": str(comando["correlation_id"]),
        "obra_codigo": comando["obra_codigo"],
        "tipo_comando": comando["tipo_comando"],
        "template_visual": "OBRA_CAIO_SUM_V1",
        "template_fixo": template_fixo,
        "conteudo_variavel": conteudo_variavel,
        "titulo": conteudo_variavel["titulo_cabecalho"],
        "titulo_cabecalho": conteudo_variavel["titulo_cabecalho"],
        "mensagem_principal": conteudo_variavel["texto_principal"],
        "mensagem_secundaria": conteudo_variavel["texto_secundario"],
        "tipo_placa": tipo_placa,
        "tema_placa": conteudo_variavel["tema_placa"],
        "tipo_icone": tipo_icone,
        "cor_base": cor_base,
        "area_pictograma": area_pictograma,
        "texto_principal": conteudo_variavel["texto_principal"],
        "texto_secundario": conteudo_variavel["texto_secundario"],
        "descricao_pictograma": conteudo_variavel["descricao_pictograma"],
        "formato_sugerido": formato_sugerido,
        "estilo_visual_referencia": estilo_visual_referencia,
        "observacao_validacao_tecnica": observacao_validacao_tecnica,
        "local_instalacao_sugerido": conteudo_variavel["local_instalacao_sugerido"],
        "status": "RASCUNHO_NAO_OFICIAL",
        "campos_a_confirmar": campos_a_confirmar,
        "observacao_validacao": (
            "Este conteúdo é um rascunho textual não oficial e não afirma "
            "conformidade normativa final. Deve ser validado pelo responsável "
            "técnico/segurança do trabalho antes de uso oficial."
        ),
        "controles_operacionais": {
            "gerou_pdf_real": False,
            "imprimiu": False,
            "executou_rpa": False,
            "conectou_openclaw": False,
            "alterou_rdo_oficial": False,
            "enviou_mensagem_terceiros": False,
            "requer_aprovacao_para_pdf_ou_impressao": True,
        },
        "gerado_em": datetime.now(timezone.utc).isoformat(),
    }


def montar_mensagem_resposta_executiva_comunicacao_obra(
    obra_codigo: Optional[str],
    tipo_resultado: Optional[str],
) -> Optional[str]:
    if tipo_resultado == "RASCUNHO_PLACA_AVISO":
        return (
            f"Rascunho de placa de aviso preparado para {obra_codigo}. "
            "Status: rascunho não oficial. Nenhum PDF, impressão, RPA ou envio externo foi executado. "
            "Validar com responsável técnico/segurança do trabalho antes de uso oficial."
        )

    if tipo_resultado == "PDF_PLACA_AVISO_GERADO":
        return (
            f"PDF local de rascunho de placa de aviso gerado para {obra_codigo}. "
            "Status: rascunho não oficial. Nenhuma impressão, RPA, OpenClaw, RDO oficial "
            "ou envio a terceiros foi executado."
        )

    return None


def normalizar_nome_diretorio(value: Optional[str]) -> str:
    normalized = re.sub(r"[^A-Za-z0-9_-]+", "-", value or "obra-sem-codigo").strip("-")
    return normalized or "obra-sem-codigo"


def extrair_texto_placa_pdf(
    comando: dict[str, Any],
    rascunho_origem: Optional[dict[str, Any]] = None,
) -> dict[str, str]:
    payload_comando = comando["payload_comando"] or {}
    rascunho = (
        (rascunho_origem or {}).get("resultado")
        or payload_comando.get("rascunho")
        or payload_comando.get("resultado_origem")
        or {}
    )
    entrada = payload_comando.get("entrada", {})

    titulo = rascunho.get("titulo_cabecalho") or rascunho.get("titulo") or "Atenção"
    texto_principal = (
        rascunho.get("texto_principal")
        or rascunho.get("mensagem_principal")
        # Fallback final: texto bruto do comando de PDF só entra se não houver rascunho.
        or entrada.get("conteudo")
        or "Aviso preliminar para comunicação visual da obra."
    )
    texto_secundario = (
        rascunho.get("texto_secundario")
        or rascunho.get("mensagem_secundaria")
        or "Validar conteúdo, pictograma, dimensões e local antes de uso."
    )
    descricao_pictograma = rascunho.get("descricao_pictograma")
    local_instalacao_sugerido = rascunho.get("local_instalacao_sugerido")
    tipo_placa = rascunho.get("tipo_placa") or rascunho.get(
        "conteudo_variavel",
        {},
    ).get("tipo_placa")
    tema_placa = rascunho.get("tema_placa") or rascunho.get(
        "conteudo_variavel",
        {},
    ).get("tema_placa")

    titulo = str(titulo).strip() or "Atenção"
    texto_principal = str(texto_principal).strip()
    if len(titulo) > 22:
        texto_principal = f"{titulo} {texto_principal}".strip()
        titulo = "Atenção"

    return {
        "titulo": titulo,
        "texto_principal": texto_principal,
        "texto_secundario": str(texto_secundario).strip(),
        "descricao_pictograma": str(descricao_pictograma or "").strip(),
        "local_instalacao_sugerido": str(local_instalacao_sugerido or "").strip(),
        "tipo_icone": str(rascunho.get("tipo_icone") or "atencao").strip(),
        "cor_base": str(rascunho.get("cor_base") or "#005A64").strip(),
        "area_pictograma": str(rascunho.get("area_pictograma") or "central").strip(),
        "template_visual": str(
            rascunho.get("template_visual") or "OBRA_CAIO_SUM_V1"
        ).strip(),
        "tipo_placa": str(tipo_placa or "").strip(),
        "tema_placa": str(tema_placa or "").strip(),
    }


def gerar_pdf_local_placa_aviso(
    comando: dict[str, Any],
    rascunho_origem: Optional[dict[str, Any]] = None,
) -> str:
    try:
        from reportlab.lib import colors
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.pdfgen import canvas
        from reportlab.platypus import Paragraph
    except ImportError as exc:
        raise RuntimeError(
            "Biblioteca reportlab não instalada. Instale as dependências de requirements-api.txt."
        ) from exc

    obra_dir_name = normalizar_nome_diretorio(comando["obra_codigo"])
    output_dir = Path("outputs") / "placas" / obra_dir_name
    output_dir.mkdir(parents=True, exist_ok=True)
    arquivo_pdf = output_dir / f"placa_aviso_comando_{comando['id']}.pdf"

    template = TEMPLATE_VISUAL_OBRA_CAIO_SUM_V1
    textos = extrair_texto_placa_pdf(comando, rascunho_origem)
    cores = template["cores"]
    fontes = template["fontes"]
    pagina = template["pagina"]

    def cm_value(value: float) -> float:
        return value * cm

    def box(nome: str) -> dict[str, float]:
        return {
            chave: cm_value(valor)
            for chave, valor in template[nome].items()
            if isinstance(valor, (int, float))
        }

    def draw_centered_text(
        text: str,
        x_cm: float,
        y_cm: float,
        width_cm: float,
        font_name: str,
        font_size: int,
        color: Any,
    ) -> None:
        text_width = c.stringWidth(text, font_name, font_size)
        text_object = c.beginText()
        text_object.setTextOrigin(
            cm_value(x_cm) + (cm_value(width_cm) - text_width) / 2,
            cm_value(y_cm),
        )
        text_object.setFont(font_name, font_size)
        text_object.setCharSpace(0)
        text_object.setFillColor(color)
        text_object.textLine(text)
        c.drawText(text_object)

    def draw_sum_text(
        text: str,
        x_cm: float,
        y_cm: float,
        width_cm: float,
        font_name: str,
        font_size: int,
        color: Any,
    ) -> None:
        gap_su = fontes["sum_gap_su"]
        gap_um = fontes["sum_gap_um"]
        letter_widths = [
            c.stringWidth(letter, font_name, font_size) for letter in text
        ]
        text_width = sum(letter_widths) + gap_su + gap_um
        x_position = cm_value(x_cm) + (cm_value(width_cm) - text_width) / 2
        y_position = cm_value(y_cm)

        c.setFont(font_name, font_size)
        c.setFillColor(color)
        for index, letter in enumerate(text):
            c.drawString(x_position, y_position, letter)
            x_position += letter_widths[index]
            if index == 0:
                x_position += gap_su
            elif index == 1:
                x_position += gap_um

    width = cm_value(pagina["largura_cm"])
    height = cm_value(pagina["altura_cm"])
    azul_petroleo = colors.HexColor(cores["azul_petroleo"])
    amarelo_atencao = colors.HexColor(cores["amarelo_atencao"])
    branco = colors.HexColor(cores["branco"])
    preto = colors.HexColor(cores["preto"])

    c = canvas.Canvas(str(arquivo_pdf), pagesize=(width, height))
    c.setTitle(f"Rascunho placa aviso - {comando['obra_codigo']}")

    c.setFillColor(azul_petroleo)
    c.rect(0, 0, width, height, fill=1, stroke=0)

    cabecalho = box("cabecalho")
    triangulo = box("triangulo_atencao")
    titulo_box = box("titulo_cabecalho")
    separador = box("separador_vertical")
    corpo = box("corpo_branco")
    circulo = box("circulo_pictograma")
    texto_principal_box = box("texto_principal")
    texto_secundario_box = box("texto_secundario")

    # Cabeçalho: triângulo de atenção, título, separador e referência textual SUM.
    c.setFillColor(azul_petroleo)
    c.rect(
        cabecalho["x_cm"],
        cabecalho["y_cm"],
        cabecalho["largura_cm"],
        cabecalho["altura_cm"],
        fill=1,
        stroke=0,
    )

    path = c.beginPath()
    path.moveTo(
        triangulo["x_cm"] + triangulo["largura_cm"] / 2,
        triangulo["y_cm"] + triangulo["altura_cm"],
    )
    path.lineTo(triangulo["x_cm"], triangulo["y_cm"])
    path.lineTo(triangulo["x_cm"] + triangulo["largura_cm"], triangulo["y_cm"])
    path.close()
    c.setFillColor(amarelo_atencao)
    c.setStrokeColor(branco)
    c.setLineWidth(1.5)
    c.drawPath(path, fill=1, stroke=1)
    c.setFillColor(preto)
    c.setFont(fontes["familia_negrito"], 32)
    c.drawCentredString(
        triangulo["x_cm"] + triangulo["largura_cm"] / 2,
        triangulo["y_cm"] + triangulo["exclamacao_y_offset_cm"],
        "!",
    )

    def criar_paragrafo_ajustado(
        nome: str,
        texto: str,
        fonte: str,
        cor: Any,
        largura_caixa: float,
        altura_caixa: float,
        tamanho_max: int,
        tamanho_min: int,
        alinhamento: int = 1,
        leading_fator: float = 1.12,
        max_linhas: Optional[int] = None,
    ) -> tuple[Paragraph, float]:
        for tamanho in range(tamanho_max, tamanho_min - 1, -1):
            leading = tamanho * leading_fator
            estilo = ParagraphStyle(
                nome,
                fontName=fonte,
                fontSize=tamanho,
                leading=leading,
                alignment=alinhamento,
                textColor=cor,
            )
            paragrafo = Paragraph(html.escape(texto), estilo)
            _, altura = paragrafo.wrapOn(c, largura_caixa, altura_caixa)
            altura_limite = altura_caixa
            if max_linhas is not None:
                altura_limite = min(altura_limite, leading * max_linhas)
            if altura <= altura_limite:
                return paragrafo, altura

        leading = tamanho_min * leading_fator
        estilo = ParagraphStyle(
            nome,
            fontName=fonte,
            fontSize=tamanho_min,
            leading=leading,
            alignment=alinhamento,
            textColor=cor,
        )
        paragrafo = Paragraph(html.escape(texto), estilo)
        _, altura = paragrafo.wrapOn(c, largura_caixa, altura_caixa)
        return paragrafo, min(altura, altura_caixa)

    titulo_header = textos["titulo"].upper()
    titulo_font_size = fontes["cabecalho_max"]
    titulo_font_name = fontes["familia_negrito"]
    titulo_width = c.stringWidth(titulo_header, titulo_font_name, titulo_font_size)
    while (
        titulo_width > titulo_box["largura_cm"]
        and titulo_font_size > fontes["cabecalho_min"]
    ):
        titulo_font_size -= 1
        titulo_width = c.stringWidth(titulo_header, titulo_font_name, titulo_font_size)
    c.setFillColor(branco)
    c.setFont(titulo_font_name, titulo_font_size)
    c.drawString(
        titulo_box["x_cm"] + (titulo_box["largura_cm"] - titulo_width) / 2,
        titulo_box["y_cm"],
        titulo_header,
    )

    c.setFillColor(branco)
    c.rect(
        separador["x_cm"],
        separador["y_cm"],
        separador["largura_cm"],
        separador["altura_cm"],
        fill=1,
        stroke=0,
    )

    draw_sum_text(
        text=template["sum"]["texto"],
        x_cm=template["sum"]["x_cm"],
        y_cm=template["sum"]["y_cm"],
        width_cm=template["sum"]["largura_cm"],
        font_name=fontes["familia_sum"],
        font_size=fontes["sum"],
        color=branco,
    )

    c.setFillColor(branco)
    c.rect(
        corpo["x_cm"],
        corpo["y_cm"],
        corpo["largura_cm"],
        corpo["altura_cm"],
        fill=1,
        stroke=0,
    )

    c.setFillColor(azul_petroleo)
    c.circle(
        circulo["centro_x_cm"],
        circulo["centro_y_cm"],
        circulo["raio_cm"],
        fill=1,
        stroke=0,
    )
    c.setFillColor(branco)
    pictograma = "!"
    c.setFont(fontes["familia_negrito"], fontes["pictograma_placeholder_grande"])
    c.drawCentredString(
        circulo["centro_x_cm"],
        circulo["exclamacao_y_cm"],
        pictograma,
    )

    texto_principal = textos["texto_principal"]
    if (
        c.stringWidth(
            texto_principal,
            fontes["familia_negrito"],
            fontes["texto_principal_max"],
        )
        <= texto_principal_box["largura_cm"]
    ):
        draw_centered_text(
            text=texto_principal,
            x_cm=template["texto_principal"]["x_cm"],
            y_cm=template["texto_principal"]["y_cm"],
            width_cm=template["texto_principal"]["largura_cm"],
            font_name=fontes["familia_negrito"],
            font_size=fontes["texto_principal_max"],
            color=azul_petroleo,
        )
    else:
        principal, principal_altura = criar_paragrafo_ajustado(
            "texto_principal_placa",
            texto_principal,
            fontes["familia_negrito"],
            azul_petroleo,
            texto_principal_box["largura_cm"],
            texto_principal_box["altura_cm"],
            fontes["texto_principal_max"],
            fontes["texto_principal_min"],
            alinhamento=1,
            max_linhas=template["texto_principal"]["max_linhas"],
        )
        principal.wrapOn(
            c,
            texto_principal_box["largura_cm"],
            texto_principal_box["altura_cm"],
        )
        principal_y = texto_principal_box["y_cm"] + (
            texto_principal_box["altura_cm"] - principal_altura
        ) / 2
        principal.drawOn(c, texto_principal_box["x_cm"], principal_y)

    texto_secundario = textos["texto_secundario"]
    if (
        c.stringWidth(
            texto_secundario,
            fontes["familia_base"],
            fontes["texto_secundario_max"],
        )
        <= texto_secundario_box["largura_cm"]
    ):
        draw_centered_text(
            text=texto_secundario,
            x_cm=template["texto_secundario"]["x_cm"],
            y_cm=template["texto_secundario"]["y_cm"],
            width_cm=template["texto_secundario"]["largura_cm"],
            font_name=fontes["familia_base"],
            font_size=fontes["texto_secundario_max"],
            color=preto,
        )
    else:
        secundario, secundario_altura = criar_paragrafo_ajustado(
            "texto_secundario_placa",
            texto_secundario,
            fontes["familia_base"],
            preto,
            texto_secundario_box["largura_cm"],
            texto_secundario_box["altura_cm"],
            fontes["texto_secundario_max"],
            fontes["texto_secundario_min"],
            alinhamento=1,
            max_linhas=template["texto_secundario"]["max_linhas"],
        )
        secundario.wrapOn(
            c,
            texto_secundario_box["largura_cm"],
            texto_secundario_box["altura_cm"],
        )
        secundario_y = texto_secundario_box["y_cm"] + (
            texto_secundario_box["altura_cm"] - secundario_altura
        ) / 2
        secundario.drawOn(c, texto_secundario_box["x_cm"], secundario_y)

    c.setFillColor(branco)
    c.setFont(fontes["familia_base"], fontes["rodape"])
    c.drawCentredString(
        cm_value(template["rodape"]["x_centro_cm"]),
        cm_value(template["rodape"]["y_cm"]),
        template["rodape"]["texto"],
    )

    c.showPage()
    c.save()
    return str(arquivo_pdf)


def montar_resultado_pdf_placa_aviso(
    comando: dict[str, Any],
    arquivo_pdf: str,
    rascunho_origem: Optional[dict[str, Any]] = None,
) -> dict[str, Any]:
    textos = extrair_texto_placa_pdf(comando, rascunho_origem)
    conteudo_variavel = {
        "titulo_cabecalho": textos["titulo"],
        "texto_principal": textos["texto_principal"],
        "texto_secundario": textos["texto_secundario"],
        "descricao_pictograma": textos["descricao_pictograma"] or None,
        "local_instalacao_sugerido": textos["local_instalacao_sugerido"] or None,
        "tipo_placa": textos["tipo_placa"] or None,
        "tema_placa": textos["tema_placa"] or None,
    }

    return {
        "tipo_resultado": "PDF_PLACA_AVISO_GERADO",
        "status": "PDF_RASCUNHO_GERADO",
        "arquivo_pdf": arquivo_pdf,
        "obra_codigo": comando["obra_codigo"],
        "agente": "AGENTE_006_COMUNICACAO_VISUAL_OBRA",
        "template_visual": textos["template_visual"],
        "template_fixo": montar_template_fixo_placa_aviso(),
        "conteudo_variavel": conteudo_variavel,
        "titulo_cabecalho": conteudo_variavel["titulo_cabecalho"],
        "texto_principal": conteudo_variavel["texto_principal"],
        "texto_secundario": conteudo_variavel["texto_secundario"],
        "descricao_pictograma": conteudo_variavel["descricao_pictograma"],
        "local_instalacao_sugerido": conteudo_variavel["local_instalacao_sugerido"],
        "tipo_placa": conteudo_variavel["tipo_placa"],
        "tema_placa": conteudo_variavel["tema_placa"],
        "comando_origem_id": str(comando["id_comando"]),
        "comando_executivo_id": comando["id"],
        "comando_rascunho_origem_id": (
            rascunho_origem.get("id") if rascunho_origem else None
        ),
        "id_comando_rascunho_origem": (
            str(rascunho_origem["id_comando"]) if rascunho_origem else None
        ),
        "tipo_comando": comando["tipo_comando"],
        "gerado_em": datetime.now(timezone.utc).isoformat(),
        "observacao_validacao": (
            "PDF local de rascunho não oficial. Não afirma conformidade normativa final. "
            "Validar com responsável técnico/segurança do trabalho antes de uso."
        ),
        "controles_operacionais": {
            "gerou_pdf_real": True,
            "imprimiu": False,
            "executou_rpa": False,
            "conectou_openclaw": False,
            "alterou_rdo_oficial": False,
            "enviou_mensagem_terceiros": False,
            "rascunho_nao_oficial": True,
        },
    }


def gerar_pdf_teste_local_placa_aviso(
    conteudo: Optional[str] = None,
    obra_codigo: str = "OBRA-CAIO",
) -> dict[str, Any]:
    """Helper de desenvolvimento local: não acessa banco, Telegram ou comandos_executivos."""
    comando_rascunho = {
        "id": 0,
        "id_comando": uuid.uuid4(),
        "correlation_id": uuid.uuid4(),
        "obra_codigo": obra_codigo,
        "tipo_comando": "PREPARAR_PLACA_AVISO",
        "payload_comando": {
            "entrada": {
                "conteudo": conteudo
                or (
                    "Cabeçalho: Atenção\n"
                    "Frase principal: Obrigatório uso de EPI\n"
                    "Frase inferior: Use capacete, bota e colete nesta área.\n"
                    "Pictograma: trabalhador usando capacete, bota e colete\n"
                    "Local: entrada da obra\n"
                    "Tipo: obrigatório\n"
                    "Tema: EPI"
                )
            }
        },
    }
    resultado_rascunho = montar_resultado_agente_comunicacao_obra(comando_rascunho)
    comando_pdf = {
        **comando_rascunho,
        "id": 0,
        "id_comando": uuid.uuid4(),
        "tipo_comando": "GERAR_PDF_PLACA_AVISO",
    }
    rascunho_origem = {
        "id": None,
        "id_comando": comando_rascunho["id_comando"],
        "resultado": resultado_rascunho,
    }
    arquivo_pdf = gerar_pdf_local_placa_aviso(comando_pdf, rascunho_origem)
    resultado_pdf = montar_resultado_pdf_placa_aviso(
        comando_pdf,
        arquivo_pdf,
        rascunho_origem,
    )
    return {
        "rascunho": resultado_rascunho,
        "pdf": resultado_pdf,
    }


def ensure_core_tables() -> None:
    sql = """
    CREATE TABLE IF NOT EXISTS mensagens_recebidas (
        id SERIAL PRIMARY KEY,
        tenant_id TEXT NOT NULL,
        obra_codigo TEXT NOT NULL,
        canal TEXT NOT NULL,
        remetente_nome TEXT,
        remetente_identificador TEXT,
        tipo_mensagem TEXT,
        conteudo TEXT,
        anexos JSONB DEFAULT '[]'::jsonb,
        payload_original JSONB,
        payload_normalizado JSONB,
        correlation_id UUID NOT NULL,
        idempotency_key TEXT UNIQUE NOT NULL,
        payload_hash TEXT,
        status_processamento TEXT DEFAULT 'RECEBIDA',
        criado_em TIMESTAMP DEFAULT NOW()
    );

    CREATE TABLE IF NOT EXISTS eventos_processados (
        id SERIAL PRIMARY KEY,
        tenant_id TEXT NOT NULL,
        obra_codigo TEXT,
        idempotency_key TEXT UNIQUE NOT NULL,
        correlation_id UUID,
        agente TEXT,
        origem TEXT,
        status TEXT NOT NULL,
        recurso_tipo TEXT,
        recurso_id INTEGER,
        payload_hash TEXT,
        mensagem_erro TEXT,
        criado_em TIMESTAMP DEFAULT NOW(),
        atualizado_em TIMESTAMP DEFAULT NOW()
    );
    """
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(sql)
        conn.commit()


@app.on_event("startup")
def startup_event():
    ensure_core_tables()


@app.get("/")
def root():
    return {
        "service": APP_NAME,
        "version": APP_VERSION,
        "status": "online",
        "mvp": "Obra-Caio Control Tower",
    }


@app.get("/status/health")
def healthcheck():
    db_status = "unknown"

    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT 1;")
                cur.fetchone()
        db_status = "ok"
    except Exception as exc:
        raise HTTPException(
            status_code=503,
            detail={
                "status": "unhealthy",
                "database": "error",
                "error": str(exc),
            },
        )

    return {
        "status": "healthy",
        "service": APP_NAME,
        "version": APP_VERSION,
        "database": db_status,
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


@app.get("/agentes/gestao-operacional/status")
def status_agente_gestao_operacional():
    return {
        "ok": True,
        "agente": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
        "mvp": "0.6C",
        "descricao": "Endpoints iniciais de consulta operacional da obra.",
        **AGENTE_008_SEGURANCA_CONSULTA,
    }


@app.post("/agentes/gestao-operacional/status-obra")
def status_obra_gestao_operacional(payload: GestaoOperacionalObraRequest):
    obra_codigo = payload.obra_codigo

    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT
                        (SELECT COUNT(*) FROM areas_obra WHERE obra_codigo = %(obra_codigo)s),
                        (SELECT COUNT(*) FROM eap_obra WHERE obra_codigo = %(obra_codigo)s),
                        (SELECT COUNT(*) FROM atividades_cronograma WHERE obra_codigo = %(obra_codigo)s),
                        (
                            SELECT COUNT(*)
                            FROM restricoes_atividade
                            WHERE obra_codigo = %(obra_codigo)s
                              AND criticidade = 'CRITICA'
                              AND status_restricao IN ('ABERTA', 'EM_TRATAMENTO', 'BLOQUEANTE')
                        ),
                        (
                            SELECT resumo_executivo
                            FROM planos_operacionais_obra
                            WHERE obra_codigo = %(obra_codigo)s
                              AND resumo_executivo IS NOT NULL
                            ORDER BY data_plano DESC, criado_em DESC
                            LIMIT 1
                        );
                    """,
                    {"obra_codigo": obra_codigo},
                )
                row = cur.fetchone()

            atividades_por_status = fetch_status_counts(
                conn,
                "atividades_cronograma",
                "status_atividade",
                obra_codigo,
            )
            restricoes_por_status = fetch_status_counts(
                conn,
                "restricoes_atividade",
                "status_restricao",
                obra_codigo,
            )
            planos_por_status = fetch_status_counts(
                conn,
                "planos_operacionais_obra",
                "status_plano",
                obra_codigo,
            )

    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={
                "message": "Erro ao consultar status operacional da obra.",
                "error": str(exc),
            },
        )

    resumo_executivo = row[4] if row and row[4] else "Sem plano operacional resumido cadastrado."
    return {
        "ok": True,
        "agente": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
        "obra_codigo": obra_codigo,
        **AGENTE_008_SEGURANCA_CONSULTA,
        "areas_total": row[0] if row else 0,
        "eap_total": row[1] if row else 0,
        "atividades_total": row[2] if row else 0,
        "atividades_por_status": atividades_por_status,
        "restricoes_por_status": restricoes_por_status,
        "restricoes_criticas": row[3] if row else 0,
        "planos_por_status": planos_por_status,
        "resumo_executivo": resumo_executivo,
    }


@app.post("/agentes/gestao-operacional/cronograma-15d")
def cronograma_15d_gestao_operacional(payload: GestaoOperacionalObraRequest):
    try:
        with get_db_connection() as conn:
            atividades = listar_atividades_gestao_operacional(
                conn,
                payload.obra_codigo,
                "0_15_DIAS",
                0,
                15,
            )
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={
                "message": "Erro ao consultar cronograma operacional de 15 dias.",
                "error": str(exc),
            },
        )

    return {
        "ok": True,
        "agente": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
        "obra_codigo": payload.obra_codigo,
        **AGENTE_008_SEGURANCA_CONSULTA,
        "horizonte": "0_15_DIAS",
        "atividades_total": len(atividades),
        "atividades": atividades,
    }


@app.post("/agentes/gestao-operacional/lookahead-30d")
def lookahead_30d_gestao_operacional(payload: GestaoOperacionalObraRequest):
    try:
        with get_db_connection() as conn:
            atividades = listar_atividades_gestao_operacional(
                conn,
                payload.obra_codigo,
                "15_30_DIAS",
                15,
                30,
            )
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={
                "message": "Erro ao consultar lookahead operacional de 30 dias.",
                "error": str(exc),
            },
        )

    return {
        "ok": True,
        "agente": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
        "obra_codigo": payload.obra_codigo,
        **AGENTE_008_SEGURANCA_CONSULTA,
        "horizonte": "15_30_DIAS",
        "atividades_total": len(atividades),
        "atividades": atividades,
    }


@app.post("/agentes/gestao-operacional/programacao-mensal")
def programacao_mensal_gestao_operacional(payload: GestaoOperacionalObraRequest):
    sql = """
    SELECT
        a.codigo_area,
        a.nome_area,
        COUNT(ac.id) AS atividades_total,
        COUNT(ac.id) FILTER (WHERE ac.status_atividade = 'EM_EXECUCAO') AS em_execucao,
        COUNT(ac.id) FILTER (WHERE ac.status_atividade = 'BLOQUEADA') AS bloqueadas,
        COUNT(ac.id) FILTER (WHERE ac.status_atividade = 'CONCLUIDA') AS concluidas,
        COALESCE(ROUND(AVG(ac.percentual_real), 2), 0) AS percentual_real_medio,
        COUNT(DISTINCT r.id) FILTER (
            WHERE r.status_restricao IN ('ABERTA', 'EM_TRATAMENTO', 'BLOQUEANTE')
        ) AS restricoes_abertas
    FROM areas_obra AS a
    LEFT JOIN atividades_cronograma AS ac
        ON ac.obra_codigo = a.obra_codigo
       AND ac.codigo_area = a.codigo_area
    LEFT JOIN restricoes_atividade AS r
        ON r.obra_codigo = a.obra_codigo
       AND r.atividade_id = ac.id
    WHERE a.obra_codigo = %(obra_codigo)s
      AND a.ativo = TRUE
    GROUP BY a.codigo_area, a.nome_area, a.ordem
    ORDER BY a.ordem, a.codigo_area;
    """

    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(sql, {"obra_codigo": payload.obra_codigo})
                areas = [
                    {
                        "codigo_area": row[0],
                        "nome_area": row[1],
                        "atividades_total": row[2],
                        "em_execucao": row[3],
                        "bloqueadas": row[4],
                        "concluidas": row[5],
                        "percentual_real_medio": serialize_numeric(row[6]),
                        "restricoes_abertas": row[7],
                    }
                    for row in cur.fetchall()
                ]
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={
                "message": "Erro ao consultar programação mensal operacional.",
                "error": str(exc),
            },
        )

    return {
        "ok": True,
        "agente": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
        "obra_codigo": payload.obra_codigo,
        **AGENTE_008_SEGURANCA_CONSULTA,
        "areas_total": len(areas),
        "areas": areas,
    }


@app.post("/agentes/gestao-operacional/area-status")
def area_status_gestao_operacional(payload: GestaoOperacionalAreaRequest):
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT
                        codigo_area,
                        nome_area,
                        tipo_area,
                        descricao,
                        ordem,
                        ativo
                    FROM areas_obra
                    WHERE obra_codigo = %(obra_codigo)s
                      AND codigo_area = %(codigo_area)s
                    LIMIT 1;
                    """,
                    {
                        "obra_codigo": payload.obra_codigo,
                        "codigo_area": payload.codigo_area,
                    },
                )
                area_row = cur.fetchone()

            if area_row is None:
                raise HTTPException(
                    status_code=404,
                    detail={
                        "message": "Área operacional não encontrada.",
                        "obra_codigo": payload.obra_codigo,
                        "codigo_area": payload.codigo_area,
                    },
                )

            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT
                        COUNT(*),
                        COUNT(*) FILTER (WHERE status_atividade = 'EM_EXECUCAO'),
                        COUNT(*) FILTER (WHERE status_atividade = 'BLOQUEADA'),
                        COUNT(*) FILTER (WHERE status_atividade = 'CONCLUIDA'),
                        COALESCE(ROUND(AVG(percentual_real), 2), 0)
                    FROM atividades_cronograma
                    WHERE obra_codigo = %(obra_codigo)s
                      AND codigo_area = %(codigo_area)s;
                    """,
                    {
                        "obra_codigo": payload.obra_codigo,
                        "codigo_area": payload.codigo_area,
                    },
                )
                resumo_row = cur.fetchone()

                cur.execute(
                    """
                    SELECT r.status_restricao, COUNT(*)
                    FROM restricoes_atividade AS r
                    JOIN atividades_cronograma AS ac
                        ON ac.id = r.atividade_id
                       AND ac.obra_codigo = r.obra_codigo
                    WHERE r.obra_codigo = %(obra_codigo)s
                      AND ac.codigo_area = %(codigo_area)s
                    GROUP BY r.status_restricao
                    ORDER BY r.status_restricao;
                    """,
                    {
                        "obra_codigo": payload.obra_codigo,
                        "codigo_area": payload.codigo_area,
                    },
                )
                restricoes_por_status = {row[0]: row[1] for row in cur.fetchall()}

                cur.execute(
                    """
                    SELECT
                        codigo_atividade,
                        codigo_eap,
                        codigo_area,
                        descricao,
                        disciplina,
                        frente_servico,
                        data_inicio_planejada,
                        data_fim_planejada,
                        data_inicio_reprogramada,
                        data_fim_reprogramada,
                        percentual_planejado,
                        percentual_real,
                        status_atividade,
                        responsavel,
                        horizonte_planejamento,
                        criticidade,
                        restricoes_resumo
                    FROM atividades_cronograma
                    WHERE obra_codigo = %(obra_codigo)s
                      AND codigo_area = %(codigo_area)s
                    ORDER BY
                        COALESCE(data_inicio_reprogramada, data_inicio_planejada) NULLS LAST,
                        codigo_atividade
                    LIMIT 100;
                    """,
                    {
                        "obra_codigo": payload.obra_codigo,
                        "codigo_area": payload.codigo_area,
                    },
                )
                atividades = [
                    formatar_atividade_gestao_operacional(row) for row in cur.fetchall()
                ]

    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={
                "message": "Erro ao consultar status operacional da área.",
                "error": str(exc),
            },
        )

    return {
        "ok": True,
        "agente": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
        "obra_codigo": payload.obra_codigo,
        **AGENTE_008_SEGURANCA_CONSULTA,
        "area": {
            "codigo_area": area_row[0],
            "nome_area": area_row[1],
            "tipo_area": area_row[2],
            "descricao": area_row[3],
            "ordem": area_row[4],
            "ativo": area_row[5],
        },
        "atividades_total": resumo_row[0] if resumo_row else 0,
        "em_execucao": resumo_row[1] if resumo_row else 0,
        "bloqueadas": resumo_row[2] if resumo_row else 0,
        "concluidas": resumo_row[3] if resumo_row else 0,
        "percentual_real_medio": serialize_numeric(resumo_row[4] if resumo_row else 0),
        "restricoes_por_status": restricoes_por_status,
        "restricoes_abertas": sum(
            restricoes_por_status.get(status, 0)
            for status in ("ABERTA", "EM_TRATAMENTO", "BLOQUEANTE")
        ),
        "atividades": atividades,
    }


def _classificar_documentos_obra(
    cur: Any, obra_codigo: str, limite: int, reprocessar: bool
) -> dict[str, Any]:
    cur.execute(
        """
        SELECT COUNT(*)
        FROM classificacoes_documentais_obra AS c
        JOIN documentos_minio_obra AS d ON d.id = c.documento_id
        WHERE d.obra_codigo = %(obra_codigo)s;
        """,
        {"obra_codigo": obra_codigo},
    )
    total_ja_classificado = cur.fetchone()[0]
    cur.execute(
        """
        SELECT d.id, d.nome_arquivo, d.object_key, d.pasta_origem,
               d.disciplina_original, d.extensao,
               (c.documento_id IS NOT NULL) AS ja_classificado
        FROM documentos_minio_obra AS d
        LEFT JOIN classificacoes_documentais_obra AS c ON c.documento_id = d.id
        WHERE d.obra_codigo = %(obra_codigo)s
          AND (%(reprocessar)s OR c.documento_id IS NULL)
        ORDER BY d.id
        LIMIT %(limite)s;
        """,
        {"obra_codigo": obra_codigo, "reprocessar": reprocessar, "limite": limite},
    )
    documentos = cur.fetchall()
    resumo_disciplina: dict[str, int] = {}
    resumo_area: dict[str, int] = {}
    resumo_status: dict[str, int] = {}
    total_classificado = 0
    for documento in documentos:
        classificacao = classificar_documento_tecnico(*documento[1:6])
        cur.execute(
            """
            INSERT INTO classificacoes_documentais_obra (
                documento_id, obra_codigo, area_detectada, disciplina_detectada,
                status_revisao, data_revisao_detectada, eh_obsoleto, eh_as_built,
                numero_revisao, palavras_chave, criterios_detectados,
                confianca_classificacao
            ) VALUES (
                %(documento_id)s, %(obra_codigo)s, %(area_detectada)s,
                %(disciplina_detectada)s, %(status_revisao)s,
                %(data_revisao_detectada)s, %(eh_obsoleto)s, %(eh_as_built)s,
                %(numero_revisao)s, %(palavras_chave)s, %(criterios_detectados)s,
                %(confianca_classificacao)s
            )
            ON CONFLICT (documento_id) DO UPDATE SET
                obra_codigo = EXCLUDED.obra_codigo,
                area_detectada = EXCLUDED.area_detectada,
                disciplina_detectada = EXCLUDED.disciplina_detectada,
                status_revisao = EXCLUDED.status_revisao,
                data_revisao_detectada = EXCLUDED.data_revisao_detectada,
                eh_obsoleto = EXCLUDED.eh_obsoleto,
                eh_as_built = EXCLUDED.eh_as_built,
                numero_revisao = EXCLUDED.numero_revisao,
                palavras_chave = EXCLUDED.palavras_chave,
                criterios_detectados = EXCLUDED.criterios_detectados,
                confianca_classificacao = EXCLUDED.confianca_classificacao,
                metodo_classificacao = 'REGRA_NOME_CAMINHO',
                status = 'CLASSIFICADO', atualizado_em = NOW();
            """,
            {
                "documento_id": documento[0],
                "obra_codigo": obra_codigo,
                **classificacao,
                "palavras_chave": Json(classificacao["palavras_chave"]),
                "criterios_detectados": Json(classificacao["criterios_detectados"]),
            },
        )
        total_classificado += 1
        for resumo, chave in (
            (resumo_disciplina, classificacao["disciplina_detectada"] or "nao_identificada"),
            (resumo_area, classificacao["area_detectada"] or "nao_identificada"),
            (resumo_status, classificacao["status_revisao"]),
        ):
            resumo[chave] = resumo.get(chave, 0) + 1

    return {
        "total_processado": len(documentos),
        "total_classificado": total_classificado,
        "total_ignorados": 0 if reprocessar else total_ja_classificado,
        "resumo_por_disciplina": resumo_disciplina,
        "resumo_por_area": resumo_area,
        "resumo_por_status_revisao": resumo_status,
    }


def _consultar_documentos_classificados(
    cur: Any, obra_codigo: str, filtros: dict[str, Any], modo: Optional[str] = None
) -> list[dict[str, Any]]:
    clausulas = ["c.obra_codigo = %(obra_codigo)s"]
    params: dict[str, Any] = {"obra_codigo": obra_codigo, "limite": filtros["limite"]}
    for campo_payload, campo_sql in (
        ("area", "area_detectada"),
        ("disciplina", "disciplina_detectada"),
        ("status_revisao", "status_revisao"),
    ):
        valor = filtros.get(campo_payload)
        if valor:
            clausulas.append(f"c.{campo_sql} = %({campo_payload})s")
            params[campo_payload] = normalizar_texto_comparacao(str(valor)).replace(" ", "_")
            if campo_payload == "status_revisao":
                params[campo_payload] = str(valor).upper()
    for campo in ("eh_obsoleto", "eh_as_built"):
        if filtros.get(campo) is not None:
            clausulas.append(f"c.{campo} = %({campo})s")
            params[campo] = filtros[campo]
    if filtros.get("termo"):
        clausulas.append("(d.nome_arquivo ILIKE %(termo)s OR d.object_key ILIKE %(termo)s)")
        params["termo"] = f"%{str(filtros['termo']).strip()}%"
    ordenacao = (
        "c.data_revisao_detectada DESC NULLS LAST, c.numero_revisao DESC NULLS LAST, d.id DESC"
        if modo == "ULTIMA_REVISAO"
        else "d.id DESC"
    )
    cur.execute(
        f"""
        SELECT d.id, d.nome_arquivo, d.extensao, c.area_detectada,
               c.disciplina_detectada, c.status_revisao, c.data_revisao_detectada,
               c.eh_obsoleto, c.eh_as_built, c.numero_revisao,
               c.confianca_classificacao, d.bucket, d.object_key
        FROM classificacoes_documentais_obra AS c
        JOIN documentos_minio_obra AS d ON d.id = c.documento_id
        WHERE {" AND ".join(clausulas)}
        ORDER BY {ordenacao}
        LIMIT %(limite)s;
        """,
        params,
    )
    return [
        {
            "documento_id": row[0], "nome_arquivo": row[1], "extensao": row[2],
            "area_detectada": row[3], "disciplina_detectada": row[4],
            "status_revisao": row[5], "data_revisao_detectada": serialize_date(row[6]),
            "eh_obsoleto": row[7], "eh_as_built": row[8], "numero_revisao": row[9],
            "confianca_classificacao": serialize_numeric(row[10]),
            "minio_uri": f"s3://{row[11]}/{row[12]}",
        }
        for row in cur.fetchall()
    ]


def _formatar_documento_revisao(row: tuple[Any, ...]) -> dict[str, Any]:
    return {
        "documento_id": row[0],
        "nome_arquivo": row[1],
        "disciplina_detectada": row[2],
        "area_detectada": row[3],
        "status_revisao": row[4],
        "data_revisao_detectada": serialize_date(row[5]),
        "numero_revisao": row[6],
        "eh_obsoleto": row[7],
        "eh_as_built": row[8],
        "confianca_classificacao": serialize_numeric(row[9]),
    }


def _ultima_revisao_documental(
    cur: Any, obra_codigo: str, disciplina: str | None,
    area: str | None, limite_candidatos: int,
) -> dict[str, Any]:
    disciplina_normalizada = (
        normalizar_texto_comparacao(disciplina).replace(" ", "_")
        if disciplina else None
    )
    area_normalizada = (
        normalizar_texto_comparacao(area).replace(" ", "_") if area else None
    )
    params = {
        "obra_codigo": obra_codigo,
        "disciplina": disciplina_normalizada,
        "area": area_normalizada,
        "limite": limite_candidatos,
    }
    filtros = """
        c.obra_codigo = %(obra_codigo)s
        AND (
            %(disciplina)s::text IS NULL
            OR c.disciplina_detectada = %(disciplina)s::text
        )
        AND (%(area)s::text IS NULL OR c.area_detectada = %(area)s::text)
    """
    cur.execute(
        f"""
        SELECT COUNT(*) FILTER (WHERE c.eh_obsoleto IS TRUE),
               COUNT(*) FILTER (WHERE c.eh_obsoleto IS NOT TRUE), COUNT(*)
        FROM classificacoes_documentais_obra AS c
        JOIN documentos_minio_obra AS d ON d.id = c.documento_id
        WHERE {filtros};
        """,
        params,
    )
    total_obsoletos, total_candidatos, total_documentos = cur.fetchone()
    cur.execute(
        f"""
        SELECT d.id, d.nome_arquivo, c.disciplina_detectada, c.area_detectada,
               c.status_revisao, c.data_revisao_detectada, c.numero_revisao,
               c.eh_obsoleto, c.eh_as_built, c.confianca_classificacao
        FROM classificacoes_documentais_obra AS c
        JOIN documentos_minio_obra AS d ON d.id = c.documento_id
        WHERE {filtros} AND c.eh_obsoleto IS NOT TRUE
        ORDER BY c.data_revisao_detectada DESC NULLS LAST,
                 c.numero_revisao DESC NULLS LAST, c.documento_id DESC
        LIMIT %(limite)s::int;
        """,
        params,
    )
    candidatos = [_formatar_documento_revisao(row) for row in cur.fetchall()]
    recomendado = candidatos[0] if candidatos else None
    alertas: list[str] = []
    if total_obsoletos:
        alertas.append(
            f"Existem {total_obsoletos} documento(s) obsoleto(s) no recorte consultado."
        )
    if not recomendado:
        alertas.append("Não há documento seguro recomendado para os filtros informados.")
        if total_documentos and total_documentos == total_obsoletos:
            alertas.append("Existem documentos, mas todos estão marcados como obsoletos.")
    recomendacao = "Validar formalmente antes de liberar uso em campo."
    titulo = disciplina_normalizada or "todas as disciplinas"
    linhas = [f"📌 Última revisão — {titulo} / {obra_codigo}", ""]
    if recomendado:
        linhas.extend([
            "Documento recomendado:",
            f"ID {recomendado['documento_id']} — {recomendado['nome_arquivo']}", "",
            "Status:", recomendado["status_revisao"] or "NAO_IDENTIFICADO", "",
            "Data detectada:",
            formatar_data_telegram(recomendado["data_revisao_detectada"]), "",
        ])
    else:
        linhas.extend(["Documento recomendado:", "Nenhum documento seguro identificado.", ""])
    linhas.append("Alertas:")
    linhas.extend(f"- {alerta}" for alerta in (alertas or ["Nenhum alerta adicional."]))
    linhas.extend(["", "Recomendação:", recomendacao])
    return {
        "obra_codigo": obra_codigo,
        "disciplina": disciplina_normalizada,
        "area": area_normalizada,
        "documento_recomendado": recomendado,
        "candidatos_considerados": candidatos,
        "existem_obsoletos_na_disciplina": total_obsoletos > 0,
        "total_obsoletos_na_disciplina": total_obsoletos,
        "total_candidatos": total_candidatos,
        "alertas": alertas,
        "recomendacao": recomendacao,
        "resposta_telegram": "\n".join(linhas),
        **AGENTE_008_SEGURANCA_CONSULTA,
    }


def formatar_data_telegram(valor: str | None) -> str:
    if not valor:
        return "não identificada"
    try:
        return date.fromisoformat(valor).strftime("%d/%m/%Y")
    except ValueError:
        return valor


def _validar_documento_campo(
    cur: Any, obra_codigo: str, documento_id: int,
) -> dict[str, Any]:
    cur.execute(
        """
        SELECT d.id, d.nome_arquivo, c.disciplina_detectada, c.area_detectada,
               c.status_revisao, c.data_revisao_detectada, c.numero_revisao,
               c.eh_obsoleto, c.eh_as_built, c.confianca_classificacao
        FROM documentos_minio_obra AS d
        LEFT JOIN classificacoes_documentais_obra AS c ON c.documento_id = d.id
        WHERE d.obra_codigo = %(obra_codigo)s AND d.id = %(documento_id)s
        LIMIT 1;
        """,
        {"obra_codigo": obra_codigo, "documento_id": documento_id},
    )
    row = cur.fetchone()
    if row is None:
        raise HTTPException(status_code=404, detail={"message": "Documento não encontrado na obra."})
    documento = _formatar_documento_revisao(row)
    cur.execute(
        """
        SELECT d.id, d.nome_arquivo
        FROM classificacoes_documentais_obra AS c
        JOIN documentos_minio_obra AS d ON d.id = c.documento_id
        WHERE c.obra_codigo = %(obra_codigo)s
          AND c.disciplina_detectada IS NOT DISTINCT FROM %(disciplina)s
          AND c.area_detectada IS NOT DISTINCT FROM %(area)s
          AND c.eh_obsoleto IS NOT TRUE
        ORDER BY c.data_revisao_detectada DESC NULLS LAST,
                 c.numero_revisao DESC NULLS LAST, c.documento_id DESC
        LIMIT 1;
        """,
        {"obra_codigo": obra_codigo, "disciplina": row[2], "area": row[3]},
    )
    mais_recente = cur.fetchone()
    existe_mais_recente = bool(mais_recente and mais_recente[0] != documento_id)
    status_revisao = documento["status_revisao"] or "NAO_IDENTIFICADO"
    confianca = documento["confianca_classificacao"]
    alertas: list[str] = []
    if documento["eh_obsoleto"] or status_revisao == "OBSOLETO":
        status_uso = "LIBERACAO_NAO_RECOMENDADA"
        alertas.append("Documento marcado como obsoleto.")
    elif status_revisao == "NAO_IDENTIFICADO" or confianca is None or confianca < 0.50:
        status_uso = "REQUER_VALIDACAO_TECNICA"
        alertas.append("Revisão não identificada ou classificação com baixa confiança.")
    elif existe_mais_recente:
        status_uso = "REQUER_VALIDACAO_TECNICA"
    else:
        status_uso = "POTENCIALMENTE_UTILIZAVEL"
    if existe_mais_recente:
        alertas.append(
            f"Existe outro documento potencialmente mais recente: ID {mais_recente[0]} — {mais_recente[1]}."
        )
    recomendacao = (
        "Não usar como documento executivo sem validação do engenheiro responsável."
    )
    linhas = [
        f"⚠️ Validação documental — Documento {documento_id}", "", "Arquivo:",
        documento["nome_arquivo"], "", "Status de uso em campo:", status_uso, "",
        "Motivos:",
    ]
    linhas.extend(f"- {alerta}" for alerta in (alertas or ["Nenhum impedimento automático identificado."]))
    linhas.extend(["", "Recomendação:", recomendacao])
    return {
        **documento,
        "status_uso_campo": status_uso,
        "alertas": alertas,
        "recomendacao": recomendacao,
        "resposta_telegram": "\n".join(linhas),
        **AGENTE_008_SEGURANCA_CONSULTA,
    }


def _gerar_relatorio_documental(
    cur: Any, obra_codigo: str, incluir_amostras: bool, limite_amostras: int
) -> dict[str, Any]:
    limite_amostras = max(1, min(limite_amostras, 5))
    cur.execute(
        """
        SELECT
            (SELECT COUNT(*) FROM documentos_minio_obra WHERE obra_codigo = %(obra_codigo)s),
            COUNT(*),
            COUNT(*) FILTER (WHERE eh_obsoleto IS TRUE),
            COUNT(*) FILTER (WHERE eh_as_built IS TRUE),
            COUNT(*) FILTER (WHERE status_revisao = 'ALTERACAO'),
            COUNT(*) FILTER (WHERE status_revisao = 'NAO_IDENTIFICADO' OR status_revisao IS NULL),
            COUNT(*) FILTER (WHERE confianca_classificacao < 0.50)
        FROM classificacoes_documentais_obra
        WHERE obra_codigo = %(obra_codigo)s;
        """,
        {"obra_codigo": obra_codigo},
    )
    totais = cur.fetchone()
    total_indexado, total_classificado = totais[0], totais[1]

    def resumo(campo: str) -> dict[str, int]:
        cur.execute(
            f"""
            SELECT COALESCE(NULLIF({campo}, ''), 'NAO_IDENTIFICADO'), COUNT(*)
            FROM classificacoes_documentais_obra
            WHERE obra_codigo = %(obra_codigo)s
            GROUP BY {campo}
            ORDER BY COUNT(*) DESC, COALESCE(NULLIF({campo}, ''), 'NAO_IDENTIFICADO');
            """,
            {"obra_codigo": obra_codigo},
        )
        return {row[0]: row[1] for row in cur.fetchall()}

    resumo_disciplina = resumo("disciplina_detectada")
    resumo_area = resumo("area_detectada")
    resumo_status = resumo("status_revisao")

    def amostra(condicao: str) -> list[dict[str, Any]]:
        if not incluir_amostras:
            return []
        cur.execute(
            f"""
            SELECT d.id, d.nome_arquivo, c.area_detectada, c.disciplina_detectada,
                   c.status_revisao, c.numero_revisao, c.confianca_classificacao
            FROM classificacoes_documentais_obra AS c
            JOIN documentos_minio_obra AS d ON d.id = c.documento_id
            WHERE c.obra_codigo = %(obra_codigo)s AND ({condicao})
            ORDER BY c.confianca_classificacao, d.id
            LIMIT %(limite)s;
            """,
            {"obra_codigo": obra_codigo, "limite": limite_amostras},
        )
        return [
            {
                "documento_id": row[0],
                "nome_arquivo": row[1],
                "area": row[2],
                "disciplina": row[3],
                "status_revisao": row[4],
                "numero_revisao": row[5],
                "confianca_classificacao": serialize_numeric(row[6]),
            }
            for row in cur.fetchall()
        ]

    recomendacoes = ["Validar revisões antes de liberar uso em campo."]
    if totais[2]:
        recomendacoes.append("Segregar documentos obsoletos das referências vigentes.")
    if totais[5] or totais[6]:
        recomendacoes.append("Revisar documentos sem revisão identificada ou com baixa confiança.")

    linhas = [
        f"📁 Relatório documental — {obra_codigo}", "",
        f"Total indexado: {total_indexado}",
        f"Classificados: {total_classificado}",
        f"Sem classificação: {max(total_indexado - total_classificado, 0)}",
        "", "Por disciplina:",
    ]
    linhas.extend(
        f"- {disciplina}: {total}"
        for disciplina, total in list(resumo_disciplina.items())[:8]
    )
    if resumo_area:
        linhas.extend(["", "Por área:"])
        linhas.extend(
            f"- {area}: {total}"
            for area, total in list(resumo_area.items())[:5]
        )
    linhas.extend([
        "", "Pontos de atenção:",
        f"- {totais[2]} documentos obsoletos",
        f"- {totais[3]} documentos As Built",
        f"- {totais[4]} documentos com alteração",
        f"- {totais[5]} sem revisão identificada",
        f"- {totais[6]} com baixa confiança",
        "", "Recomendação:", recomendacoes[0],
    ])
    return {
        "obra_codigo": obra_codigo,
        "total_indexado": total_indexado,
        "total_classificado": total_classificado,
        "total_sem_classificacao": max(total_indexado - total_classificado, 0),
        "resumo_por_disciplina": resumo_disciplina,
        "resumo_por_area": resumo_area,
        "resumo_por_status_revisao": resumo_status,
        "total_obsoletos": totais[2],
        "total_as_built": totais[3],
        "total_com_alteracao": totais[4],
        "total_sem_revisao_identificada": totais[5],
        "total_baixa_confianca": totais[6],
        "documentos_obsoletos_amostra": amostra("c.eh_obsoleto IS TRUE"),
        "documentos_baixa_confianca_amostra": amostra("c.confianca_classificacao < 0.50"),
        "documentos_sem_revisao_amostra": amostra(
            "c.status_revisao = 'NAO_IDENTIFICADO' OR c.status_revisao IS NULL"
        ),
        "recomendacoes": recomendacoes,
        "resposta_telegram": "\n".join(linhas),
        **AGENTE_008_SEGURANCA_CONSULTA,
    }


def _avaliar_riscos_documentais(
    cur: Any, obra_codigo: str, area: str | None,
    disciplina: str | None, limite_amostras: int,
) -> dict[str, Any]:
    area_normalizada = normalizar_texto_comparacao(area).replace(" ", "_") if area else None
    disciplina_normalizada = (
        normalizar_texto_comparacao(disciplina).replace(" ", "_")
        if disciplina else None
    )
    limite_amostras = max(1, min(limite_amostras, 50))
    params = {
        "obra_codigo": obra_codigo, "area": area_normalizada,
        "disciplina": disciplina_normalizada, "limite": limite_amostras,
    }
    filtros = """
        c.obra_codigo = %(obra_codigo)s
        AND (%(area)s::text IS NULL OR c.area_detectada = %(area)s::text)
        AND (%(disciplina)s::text IS NULL OR c.disciplina_detectada = %(disciplina)s::text)
    """
    cur.execute(
        f"""
        SELECT COUNT(*),
               COUNT(*) FILTER (WHERE c.eh_obsoleto IS TRUE),
               COUNT(*) FILTER (WHERE c.eh_as_built IS TRUE),
               COUNT(*) FILTER (WHERE c.status_revisao = 'ALTERACAO'),
               COUNT(*) FILTER (WHERE c.status_revisao = 'NAO_IDENTIFICADO' OR c.status_revisao IS NULL),
               COUNT(*) FILTER (WHERE c.confianca_classificacao < 0.50),
               COUNT(*) FILTER (WHERE c.data_revisao_detectada IS NOT NULL),
               COUNT(*) FILTER (WHERE c.data_revisao_detectada IS NULL),
               COUNT(DISTINCT c.disciplina_detectada) FILTER (WHERE c.disciplina_detectada IS NOT NULL),
               COUNT(DISTINCT c.area_detectada) FILTER (WHERE c.area_detectada IS NOT NULL)
        FROM classificacoes_documentais_obra AS c
        JOIN documentos_minio_obra AS d ON d.id = c.documento_id
        WHERE {filtros};
        """,
        params,
    )
    valores = cur.fetchone()
    chaves = (
        "total_documentos", "total_obsoletos", "total_as_built", "total_alteracao",
        "total_sem_revisao", "total_baixa_confianca", "total_com_data_revisao",
        "total_sem_data_revisao", "total_disciplinas_envolvidas", "total_areas_envolvidas",
    )
    totais = dict(zip(chaves, valores))
    cur.execute(
        f"""
        SELECT COUNT(*)
        FROM (
            SELECT c.area_detectada, c.disciplina_detectada
            FROM classificacoes_documentais_obra AS c
            JOIN documentos_minio_obra AS d ON d.id = c.documento_id
            WHERE {filtros}
            GROUP BY c.area_detectada, c.disciplina_detectada
            HAVING BOOL_OR(c.eh_obsoleto IS TRUE)
               AND BOOL_OR(c.status_revisao = 'ALTERACAO')
        ) AS conflitos;
        """,
        params,
    )
    existe_conflito_mesma_area_disciplina = cur.fetchone()[0] > 0

    def amostra(condicao: str) -> list[dict[str, Any]]:
        cur.execute(
            f"""
            SELECT d.id, d.nome_arquivo, c.area_detectada, c.disciplina_detectada,
                   c.status_revisao, c.numero_revisao, c.data_revisao_detectada,
                   c.confianca_classificacao
            FROM classificacoes_documentais_obra AS c
            JOIN documentos_minio_obra AS d ON d.id = c.documento_id
            WHERE {filtros} AND ({condicao})
            ORDER BY c.confianca_classificacao NULLS FIRST, d.id
            LIMIT %(limite)s;
            """,
            params,
        )
        return [{
            "documento_id": row[0], "nome_arquivo": row[1], "area": row[2],
            "disciplina": row[3], "status_revisao": row[4], "numero_revisao": row[5],
            "data_revisao": serialize_date(row[6]),
            "confianca_classificacao": serialize_numeric(row[7]),
        } for row in cur.fetchall()]

    amostras = {
        "documentos_obsoletos_amostra": amostra("c.eh_obsoleto IS TRUE"),
        "documentos_com_alteracao_amostra": amostra("c.status_revisao = 'ALTERACAO'"),
        "documentos_sem_revisao_amostra": amostra(
            "c.status_revisao = 'NAO_IDENTIFICADO' OR c.status_revisao IS NULL"
        ),
        "documentos_baixa_confianca_amostra": amostra("c.confianca_classificacao < 0.50"),
    }
    alertas = []
    if existe_conflito_mesma_area_disciplina:
        alertas.append("Há documentos obsoletos e documentos com alteração na mesma área e disciplina.")
    if totais["total_obsoletos"]:
        alertas.append(f"Foram identificados {totais['total_obsoletos']} documentos obsoletos.")
    if totais["total_sem_revisao"]:
        alertas.append(f"Há {totais['total_sem_revisao']} documentos sem revisão identificada.")
    if totais["total_baixa_confianca"]:
        alertas.append(f"Há {totais['total_baixa_confianca']} classificações com baixa confiança.")
    if totais["total_sem_data_revisao"]:
        alertas.append(f"Há {totais['total_sem_data_revisao']} documentos sem data de revisão identificada.")
    alertas = alertas[:5]

    if not totais["total_documentos"]:
        nivel = "NAO_AVALIADO"
    elif existe_conflito_mesma_area_disciplina:
        nivel = "CRITICO"
    elif totais["total_obsoletos"] or totais["total_sem_revisao"] >= 10:
        nivel = "ALTO"
    elif totais["total_baixa_confianca"] or totais["total_sem_data_revisao"]:
        nivel = "MEDIO"
    else:
        nivel = "BAIXO"
    recomendacao = (
        "Não liberar execução em campo sem validação formal da revisão aplicável pelo engenheiro responsável."
        if totais["total_documentos"] else
        "Revisar o filtro e confirmar a disponibilidade da classificação documental antes de qualquer decisão em campo."
    )
    criticas = []
    ids_incluidos: set[int] = set()
    for grupo in (
        amostras["documentos_obsoletos_amostra"],
        amostras["documentos_com_alteracao_amostra"],
        amostras["documentos_sem_revisao_amostra"],
        amostras["documentos_baixa_confianca_amostra"],
    ):
        for documento in grupo:
            if documento["documento_id"] not in ids_incluidos:
                ids_incluidos.add(documento["documento_id"])
                nome = documento["nome_arquivo"] or "Sem nome"
                nome = nome if len(nome) <= 72 else f"{nome[:69]}..."
                criticas.append(f"- ID {documento['documento_id']} — {nome}")
            if len(criticas) == 5:
                break
        if len(criticas) == 5:
            break
    linhas = [
        f"⚠️ Riscos documentais — {obra_codigo}", "", "Filtro:",
        f"Área: {area_normalizada or 'todas'}",
        f"Disciplina: {disciplina_normalizada or 'todas'}", "", "Resumo:",
        f"- Total analisado: {totais['total_documentos']} documentos",
        f"- Obsoletos: {totais['total_obsoletos']}",
        f"- Com alteração: {totais['total_alteracao']}",
        f"- Sem revisão identificada: {totais['total_sem_revisao']}",
        f"- Baixa confiança: {totais['total_baixa_confianca']}", "",
        "Nível de risco:", nivel, "", "Alertas:",
    ]
    linhas.extend(f"- {alerta}" for alerta in (alertas or ["Nenhum alerta relevante identificado no recorte."]))
    if criticas:
        linhas.extend(["", "Amostras críticas:", *criticas])
    linhas.extend(["", "Recomendação:", recomendacao])
    return {
        "obra_codigo": obra_codigo, "area": area_normalizada,
        "disciplina": disciplina_normalizada, "totais": totais, "amostras": amostras,
        "alertas": alertas, "nivel_risco_documental": nivel,
        "recomendacao": recomendacao, "resposta_telegram": "\n".join(linhas),
        **AGENTE_008_SEGURANCA_CONSULTA,
    }


def _gerar_plano_saneamento_documental(
    cur: Any, obra_codigo: str, area: str | None,
    disciplina: str | None, limite_acoes: int,
) -> dict[str, Any]:
    riscos = _avaliar_riscos_documentais(
        cur, obra_codigo, area, disciplina, min(limite_acoes, 50)
    )
    area_normalizada = riscos["area"]
    disciplina_normalizada = riscos["disciplina"]
    params = {
        "obra_codigo": obra_codigo,
        "area": area_normalizada,
        "disciplina": disciplina_normalizada,
        "limite": max(1, min(limite_acoes, 50)),
    }
    cur.execute(
        """
        SELECT d.id, d.nome_arquivo, c.area_detectada, c.disciplina_detectada,
               c.status_revisao, c.data_revisao_detectada,
               c.confianca_classificacao, c.eh_obsoleto,
               (
                   c.eh_obsoleto IS TRUE AND EXISTS (
                       SELECT 1
                       FROM classificacoes_documentais_obra AS conflito
                       WHERE conflito.obra_codigo = c.obra_codigo
                         AND conflito.area_detectada IS NOT DISTINCT FROM c.area_detectada
                         AND conflito.disciplina_detectada IS NOT DISTINCT FROM c.disciplina_detectada
                         AND conflito.status_revisao = 'ALTERACAO'
                   )
               ) AS conflito_obsoleto_alteracao,
               CASE
                   WHEN c.eh_obsoleto IS TRUE AND EXISTS (
                       SELECT 1
                       FROM classificacoes_documentais_obra AS conflito
                       WHERE conflito.obra_codigo = c.obra_codigo
                         AND conflito.area_detectada IS NOT DISTINCT FROM c.area_detectada
                         AND conflito.disciplina_detectada IS NOT DISTINCT FROM c.disciplina_detectada
                         AND conflito.status_revisao = 'ALTERACAO'
                   ) THEN 1
                   WHEN c.eh_obsoleto IS TRUE THEN 1
                   WHEN c.confianca_classificacao < 0.50
                        AND COALESCE(d.categoria_documental, '') ILIKE '%%projeto%%' THEN 1
                   WHEN c.status_revisao = 'NAO_IDENTIFICADO' OR c.status_revisao IS NULL THEN 2
                   WHEN c.data_revisao_detectada IS NULL THEN 2
                   WHEN c.confianca_classificacao < 0.50 THEN 3
               END AS ordem_prioridade
        FROM classificacoes_documentais_obra AS c
        JOIN documentos_minio_obra AS d ON d.id = c.documento_id
        WHERE c.obra_codigo = %(obra_codigo)s
          AND (%(area)s::text IS NULL OR c.area_detectada = %(area)s::text)
          AND (%(disciplina)s::text IS NULL OR c.disciplina_detectada = %(disciplina)s::text)
          AND (
              c.eh_obsoleto IS TRUE
              OR c.status_revisao = 'NAO_IDENTIFICADO'
              OR c.status_revisao IS NULL
              OR c.data_revisao_detectada IS NULL
              OR c.confianca_classificacao < 0.50
          )
        ORDER BY ordem_prioridade, c.confianca_classificacao NULLS FIRST, d.id
        LIMIT %(limite)s;
        """,
        params,
    )
    acoes = []
    for row in cur.fetchall():
        motivos = []
        if row[8]:
            motivos.append("CONFLITO_OBSOLETO_ALTERACAO")
        elif row[7]:
            motivos.append("DOCUMENTO_OBSOLETO")
        if row[4] in (None, "NAO_IDENTIFICADO"):
            motivos.append("REVISAO_NAO_IDENTIFICADA")
        if row[5] is None:
            motivos.append("DATA_REVISAO_NAO_IDENTIFICADA")
        if serialize_numeric(row[6]) < 0.50:
            motivos.append("BAIXA_CONFIANCA_CLASSIFICACAO")
        acoes.append({
            "prioridade": f"PRIORIDADE_{row[9]}",
            "documento_id": row[0],
            "nome_arquivo": row[1],
            "area": row[2],
            "disciplina": row[3],
            "motivos": motivos,
            "acao_recomendada": "Validar revisão, vigência e aplicabilidade com o responsável técnico antes do uso em campo.",
        })
    pendencias_sugeridas = [{
        "motivo": item["motivos"][0],
        "descricao": f"Revisar documento {item['documento_id']} — {item['nome_arquivo'] or 'sem nome'}.",
        "requer_pedido_explicito": True,
    } for item in acoes]
    recomendacao = (
        "Priorizar as ações listadas e abrir pendência interna somente mediante pedido explícito do usuário."
        if acoes else "Nenhuma ação de saneamento foi identificada para o recorte informado."
    )
    linhas = [
        f"🧭 Plano de saneamento documental — {obra_codigo}",
        f"Documentos avaliados: {riscos['totais']['total_documentos']}",
        f"Ações recomendadas: {len(acoes)}", "",
    ]
    for item in acoes:
        linhas.append(
            f"- {item['prioridade']} | ID {item['documento_id']} | "
            f"{item['nome_arquivo'] or 'Sem nome'} | {', '.join(item['motivos'])}"
        )
    linhas.extend(["", recomendacao])
    return {
        "obra_codigo": obra_codigo,
        "area": area_normalizada,
        "disciplina": disciplina_normalizada,
        "total_documentos_avaliados": riscos["totais"]["total_documentos"],
        "acoes_recomendadas": acoes,
        "pendencias_sugeridas": pendencias_sugeridas,
        "recomendacao": recomendacao,
        "resposta_telegram": "\n".join(linhas),
        **AGENTE_008_SEGURANCA_CONSULTA,
    }


def _criar_pendencia_documental(
    cur: Any, obra_codigo: str, area: str | None, disciplina: str | None,
    motivo: str, descricao: str | None,
) -> dict[str, Any]:
    motivos_validos = {
        "DOCUMENTO_OBSOLETO", "REVISAO_NAO_IDENTIFICADA", "RISCO_DOCUMENTAL"
    }
    motivo_normalizado = normalizar_texto_comparacao(motivo).upper().replace(" ", "_")
    if motivo_normalizado not in motivos_validos:
        return _erro_pendencia_documental(
            obra_codigo, "Motivo documental inválido; a pendência não foi criada."
        )
    cur.execute("SELECT to_regclass('public.pendencias_obra');")
    if cur.fetchone()[0] is None:
        return _erro_pendencia_documental(
            obra_codigo, "A tabela public.pendencias_obra não existe; a pendência não foi criada."
        )
    cur.execute(
        """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = 'public' AND table_name = 'pendencias_obra';
        """
    )
    colunas = {row[0] for row in cur.fetchall()}
    obrigatorias = {"obra_codigo", "descricao", "categoria", "prioridade", "status_pendencia"}
    origem_coluna = next(
        (nome for nome in ("agente_origem", "origem", "criado_por") if nome in colunas), None
    )
    ausentes = sorted(obrigatorias - colunas)
    if ausentes or origem_coluna is None:
        detalhe = ", ".join(ausentes + (["agente_origem/origem/criado_por"] if origem_coluna is None else []))
        return _erro_pendencia_documental(
            obra_codigo, f"Schema de public.pendencias_obra incompatível ({detalhe}); a pendência não foi criada."
        )
    area_normalizada = normalizar_texto_comparacao(area).replace(" ", "_") if area else None
    disciplina_normalizada = normalizar_texto_comparacao(disciplina).replace(" ", "_") if disciplina else None
    descricao_final = descricao or f"Saneamento documental: {motivo_normalizado}."
    if area_normalizada:
        descricao_final += f" Área: {area_normalizada}."
    if disciplina_normalizada:
        descricao_final += f" Disciplina: {disciplina_normalizada}."
    colunas_insert = ["obra_codigo", "descricao", "categoria", "prioridade", "status_pendencia", origem_coluna]
    valores = [obra_codigo, descricao_final, "GESTAO_DOCUMENTAL", "ALTA", "ABERTA", "AGENTE_008_GESTAO_OPERACIONAL_OBRA"]
    placeholders = ", ".join(["%s"] * len(valores))
    cur.execute(
        f"INSERT INTO public.pendencias_obra ({', '.join(colunas_insert)}) VALUES ({placeholders}) RETURNING id;",
        valores,
    )
    pendencia_id = cur.fetchone()[0]
    resposta = f"Pendência documental {pendencia_id} criada internamente para {obra_codigo}."
    return {
        "ok": True, "pendencia_criada": True, "pendencia_id": pendencia_id,
        "obra_codigo": obra_codigo, "area": area_normalizada,
        "disciplina": disciplina_normalizada, "motivo": motivo_normalizado,
        "descricao": descricao_final, "categoria": "GESTAO_DOCUMENTAL",
        "origem": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
        "resposta_telegram": resposta,
        **AGENTE_008_SEGURANCA_REGISTRO_INTERNO,
    }


def _erro_pendencia_documental(obra_codigo: str, mensagem: str) -> dict[str, Any]:
    return {
        "ok": False, "pendencia_criada": False, "obra_codigo": obra_codigo,
        "erro_controlado": mensagem, "resposta_telegram": mensagem,
        **AGENTE_008_SEGURANCA_REGISTRO_INTERNO,
    }


@app.post("/agentes/gestao-operacional/relatorio-documental")
def relatorio_documental_gestao_operacional(
    payload: GestaoOperacionalRelatorioDocumentalRequest,
):
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                resultado = _gerar_relatorio_documental(
                    cur, payload.obra_codigo, payload.incluir_amostras,
                    payload.limite_amostras,
                )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={"message": "Erro ao gerar relatório documental da obra.", "error": str(exc)},
        )
    return {"ok": True, "agente": "AGENTE_008_GESTAO_OPERACIONAL_OBRA", "mvp": "0.6J", **resultado}


@app.post("/agentes/gestao-operacional/riscos-documentais")
def riscos_documentais_gestao_operacional(
    payload: GestaoOperacionalRiscosDocumentaisRequest,
):
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                resultado = _avaliar_riscos_documentais(
                    cur, payload.obra_codigo, payload.area,
                    payload.disciplina, payload.limite_amostras,
                )
    except Exception as exc:
        raise HTTPException(status_code=500, detail={
            "message": "Erro ao avaliar riscos documentais da obra.", "error": str(exc),
        })
    return {"ok": True, "agente": "AGENTE_008_GESTAO_OPERACIONAL_OBRA", "mvp": "0.6L", **resultado}


@app.post("/agentes/gestao-operacional/plano-saneamento-documental")
def plano_saneamento_documental_gestao_operacional(
    payload: GestaoOperacionalPlanoSaneamentoDocumentalRequest,
):
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                resultado = _gerar_plano_saneamento_documental(
                    cur, payload.obra_codigo, payload.area,
                    payload.disciplina, payload.limite_acoes,
                )
    except Exception as exc:
        raise HTTPException(status_code=500, detail={
            "message": "Erro ao gerar plano de saneamento documental.", "error": str(exc),
        })
    return {"ok": True, "agente": "AGENTE_008_GESTAO_OPERACIONAL_OBRA", "mvp": "0.6M", **resultado}


@app.post("/agentes/gestao-operacional/criar-pendencia-documental")
def criar_pendencia_documental_gestao_operacional(
    payload: GestaoOperacionalCriarPendenciaDocumentalRequest,
):
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                resultado = _criar_pendencia_documental(
                    cur, payload.obra_codigo, payload.area, payload.disciplina,
                    payload.motivo, payload.descricao,
                )
                if not resultado["pendencia_criada"]:
                    conn.rollback()
    except Exception as exc:
        resultado = _erro_pendencia_documental(
            payload.obra_codigo,
            f"Erro controlado ao criar pendência documental; a pendência não foi criada: {exc}",
        )
    return {"agente": "AGENTE_008_GESTAO_OPERACIONAL_OBRA", "mvp": "0.6M", **resultado}


@app.post("/agentes/gestao-operacional/ultima-revisao-documental")
def ultima_revisao_documental_gestao_operacional(
    payload: GestaoOperacionalUltimaRevisaoDocumentalRequest,
):
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                resultado = _ultima_revisao_documental(
                    cur, payload.obra_codigo, payload.disciplina,
                    payload.area, payload.limite_candidatos,
                )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={
                "message": "Erro ao consultar última revisão documental.",
                "error": str(exc),
            },
        )
    return {"ok": True, "agente": "AGENTE_008_GESTAO_OPERACIONAL_OBRA", "mvp": "0.6K", **resultado}


@app.post("/agentes/gestao-operacional/validar-documento-campo")
def validar_documento_campo_gestao_operacional(
    payload: GestaoOperacionalValidarDocumentoCampoRequest,
):
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                resultado = _validar_documento_campo(
                    cur, payload.obra_codigo, payload.documento_id
                )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail={"message": "Erro ao validar documento para uso em campo.", "error": str(exc)})
    return {"ok": True, "agente": "AGENTE_008_GESTAO_OPERACIONAL_OBRA", "mvp": "0.6K", "obra_codigo": payload.obra_codigo, **resultado}


@app.post("/agentes/gestao-operacional/classificar-documentos")
def classificar_documentos_gestao_operacional(payload: GestaoOperacionalClassificarDocumentosRequest):
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                resultado = _classificar_documentos_obra(
                    cur, payload.obra_codigo, payload.limite, payload.reprocessar
                )
    except Exception as exc:
        raise HTTPException(status_code=500, detail={"message": "Erro ao classificar documentos da obra.", "error": str(exc)})
    return {"ok": True, "agente": "AGENTE_008_GESTAO_OPERACIONAL_OBRA", "mvp": "0.6I", "obra_codigo": payload.obra_codigo, **AGENTE_008_SEGURANCA_CONSULTA, **resultado}


@app.post("/agentes/gestao-operacional/documentos-classificados")
def documentos_classificados_gestao_operacional(payload: GestaoOperacionalDocumentosClassificadosRequest):
    filtros = payload.model_dump(exclude={"obra_codigo"})
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                documentos = _consultar_documentos_classificados(cur, payload.obra_codigo, filtros)
    except Exception as exc:
        raise HTTPException(status_code=500, detail={"message": "Erro ao consultar documentos classificados.", "error": str(exc)})
    return {"ok": True, "agente": "AGENTE_008_GESTAO_OPERACIONAL_OBRA", "mvp": "0.6I", "obra_codigo": payload.obra_codigo, **AGENTE_008_SEGURANCA_CONSULTA, "total_retornado": len(documentos), "documentos": documentos}


@app.post("/agentes/gestao-operacional/documentos-resumo")
def documentos_resumo_gestao_operacional(
    payload: GestaoOperacionalDocumentosResumoRequest,
):
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT
                        disciplina_original,
                        extensao,
                        categoria_documental,
                        COUNT(*) AS total
                    FROM documentos_minio_obra
                    WHERE obra_codigo = %(obra_codigo)s
                    GROUP BY disciplina_original, extensao, categoria_documental
                    ORDER BY
                        disciplina_original NULLS LAST,
                        extensao NULLS LAST,
                        categoria_documental NULLS LAST;
                    """,
                    {"obra_codigo": payload.obra_codigo},
                )
                resumo = [
                    {
                        "disciplina_original": row[0],
                        "extensao": row[1],
                        "categoria_documental": row[2],
                        "total": row[3],
                    }
                    for row in cur.fetchall()
                ]
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={
                "message": "Erro ao consultar resumo de documentos indexados da obra.",
                "error": str(exc),
            },
        )

    return {
        "ok": True,
        "agente": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
        "mvp": "0.6E",
        "descricao": "Resumo consultivo de documentos indexados no MinIO por disciplina, extensão e categoria documental.",
        "obra_codigo": payload.obra_codigo,
        **AGENTE_008_SEGURANCA_CONSULTA,
        "total_documentos": sum(item["total"] for item in resumo),
        "resumo": resumo,
    }


@app.post("/agentes/gestao-operacional/documentos-indexados")
def documentos_indexados_gestao_operacional(
    payload: GestaoOperacionalDocumentosIndexadosRequest,
):
    filtros_sql = ["obra_codigo = %(obra_codigo)s"]
    params: dict[str, Any] = {
        "obra_codigo": payload.obra_codigo,
        "limite": payload.limite,
    }

    disciplina = payload.disciplina.strip() if payload.disciplina else None
    extensao = payload.extensao.strip().lower() if payload.extensao else None
    termo = payload.termo.strip() if payload.termo else None

    if disciplina:
        filtros_sql.append("disciplina_original ILIKE %(disciplina)s")
        params["disciplina"] = f"%{disciplina}%"

    if extensao:
        filtros_sql.append("LOWER(extensao) = %(extensao)s")
        params["extensao"] = extensao

    if termo:
        filtros_sql.append("(nome_arquivo ILIKE %(termo)s OR object_key ILIKE %(termo)s)")
        params["termo"] = f"%{termo}%"

    sql = f"""
    SELECT
        id,
        bucket,
        object_key,
        nome_arquivo,
        extensao,
        pasta_origem,
        disciplina_original,
        categoria_documental,
        status_indexacao,
        manifest_path
    FROM documentos_minio_obra
    WHERE {" AND ".join(filtros_sql)}
    ORDER BY atualizado_em DESC NULLS LAST, criado_em DESC NULLS LAST, id DESC
    LIMIT %(limite)s;
    """

    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(sql, params)
                documentos = [
                    {
                        "id": row[0],
                        "bucket": row[1],
                        "object_key": row[2],
                        "nome_arquivo": row[3],
                        "extensao": row[4],
                        "pasta_origem": row[5],
                        "disciplina_original": row[6],
                        "categoria_documental": row[7],
                        "status_indexacao": row[8],
                        "manifest_path": row[9],
                        "minio_uri": f"s3://{row[1]}/{row[2]}",
                    }
                    for row in cur.fetchall()
                ]
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={
                "message": "Erro ao consultar documentos indexados da obra.",
                "error": str(exc),
            },
        )

    return {
        "ok": True,
        "agente": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
        "mvp": "0.6E",
        "descricao": "Consulta controlada de documentos indexados no MinIO, sem download ou alteração de arquivos.",
        "obra_codigo": payload.obra_codigo,
        "filtros": {
            "disciplina": disciplina,
            "extensao": extensao,
            "termo": termo,
            "limite": payload.limite,
        },
        **AGENTE_008_SEGURANCA_CONSULTA,
        "total_retornado": len(documentos),
        "documentos": documentos,
    }


def _filtros_comando_documental(payload_comando: Any) -> dict[str, Any]:
    payload = payload_comando if isinstance(payload_comando, dict) else {}
    filtros = payload.get("filtros") if isinstance(payload.get("filtros"), dict) else {}
    entrada = payload.get("entrada") if isinstance(payload.get("entrada"), dict) else {}
    inferidos = extrair_filtros_documentos_telegram(entrada.get("conteudo"))

    def texto_filtro(nome: str) -> Optional[str]:
        valor = payload.get(nome, filtros.get(nome, inferidos.get(nome)))
        return valor.strip() if isinstance(valor, str) and valor.strip() else None

    limite_bruto = payload.get("limite", filtros.get("limite", 10))
    try:
        limite = int(limite_bruto)
    except (TypeError, ValueError):
        limite = 10

    extensao = texto_filtro("extensao")
    return {
        "disciplina": texto_filtro("disciplina"),
        "extensao": extensao.lower() if extensao else None,
        "termo": texto_filtro("termo"),
        "limite": max(1, min(limite, 10)),
    }


def _resultado_resumo_documental(cur: Any, obra_codigo: str) -> dict[str, Any]:
    cur.execute(
        """
        SELECT
            COALESCE(disciplina_original, 'Sem disciplina') AS disciplina_original,
            COALESCE(NULLIF(extensao, ''), 'sem extensão') AS extensao,
            COUNT(*) AS total
        FROM documentos_minio_obra
        WHERE obra_codigo = %(obra_codigo)s
        GROUP BY disciplina_original, extensao
        ORDER BY total DESC, disciplina_original, extensao;
        """,
        {"obra_codigo": obra_codigo},
    )
    resumo = [
        {"disciplina_original": row[0], "extensao": row[1], "total": row[2]}
        for row in cur.fetchall()
    ]
    total_documentos = sum(item["total"] for item in resumo)
    totais_disciplina: dict[str, int] = {}
    for item in resumo:
        disciplina = item["disciplina_original"]
        totais_disciplina[disciplina] = totais_disciplina.get(disciplina, 0) + item["total"]

    linhas = [
        f"📁 Documentos da {obra_codigo}",
        "",
        f"Total indexado: {total_documentos} documentos",
    ]
    if totais_disciplina:
        linhas.extend(["", "Resumo por disciplina:"])
        linhas.extend(
            f"- {disciplina}: {total}"
            for disciplina, total in sorted(
                totais_disciplina.items(), key=lambda item: (-item[1], item[0])
            )
        )

    return {
        "tipo_resultado": "RESUMO_DOCUMENTOS_OBRA",
        "resposta_telegram": "\n".join(linhas),
        "total_documentos": total_documentos,
        "resumo": resumo,
        **AGENTE_008_SEGURANCA_CONSULTA,
    }


def _resultado_documentos_indexados(
    cur: Any, obra_codigo: str, payload_comando: Any
) -> dict[str, Any]:
    filtros = _filtros_comando_documental(payload_comando)
    filtros_sql = ["obra_codigo = %(obra_codigo)s"]
    params: dict[str, Any] = {
        "obra_codigo": obra_codigo,
        "limite_consulta": 11,
    }
    if filtros["disciplina"]:
        filtros_sql.append("disciplina_original ILIKE %(disciplina)s")
        params["disciplina"] = f"%{filtros['disciplina']}%"
    if filtros["extensao"]:
        filtros_sql.append("LOWER(extensao) = %(extensao)s")
        params["extensao"] = filtros["extensao"]
    if filtros["termo"]:
        filtros_sql.append("(nome_arquivo ILIKE %(termo)s OR object_key ILIKE %(termo)s)")
        params["termo"] = f"%{filtros['termo']}%"

    cur.execute(
        f"""
        SELECT nome_arquivo, disciplina_original, extensao, bucket, object_key
        FROM documentos_minio_obra
        WHERE {" AND ".join(filtros_sql)}
        ORDER BY atualizado_em DESC NULLS LAST, criado_em DESC NULLS LAST, id DESC
        LIMIT %(limite_consulta)s;
        """,
        params,
    )
    rows = cur.fetchall()
    ha_mais_resultados = len(rows) > 10
    documentos = [
        {
            "nome_arquivo": row[0],
            "disciplina_original": row[1],
            "extensao": row[2],
            "minio_uri": f"s3://{row[3]}/{row[4]}",
        }
        for row in rows[: filtros["limite"]]
    ]

    if not documentos:
        resposta = "Não encontrei documentos para esse filtro."
    else:
        linhas = [f"📁 Documentos encontrados na {obra_codigo}", ""]
        for documento in documentos:
            linhas.extend(
                [
                    f"- {documento['nome_arquivo']}",
                    f"  {documento['disciplina_original'] or 'Sem disciplina'} | "
                    f"{documento['extensao'] or 'sem extensão'}",
                    f"  {documento['minio_uri']}",
                ]
            )
        if ha_mais_resultados:
            linhas.extend(["", "Mostrando os 10 primeiros resultados."])
        resposta = "\n".join(linhas)

    return {
        "tipo_resultado": "DOCUMENTOS_OBRA_INDEXADOS",
        "resposta_telegram": resposta,
        "total_retornado": len(documentos),
        "documentos": documentos,
        "filtros": filtros,
        "ha_mais_resultados": ha_mais_resultados,
        **AGENTE_008_SEGURANCA_CONSULTA,
    }


def _resultado_classificacao_documental(
    cur: Any, obra_codigo: str, payload_comando: Any
) -> dict[str, Any]:
    payload = payload_comando if isinstance(payload_comando, dict) else {}
    try:
        limite = max(1, min(int(payload.get("limite", 200)), 1000))
    except (TypeError, ValueError):
        limite = 200
    resultado = _classificar_documentos_obra(
        cur, obra_codigo, limite, bool(payload.get("reprocessar", False))
    )
    return {
        "tipo_resultado": "CLASSIFICACAO_DOCUMENTAL_OBRA",
        "resposta_telegram": (
            "Classificação documental concluída. "
            f"{resultado['total_classificado']} documentos classificados; "
            f"{resultado['total_ignorados']} ignorados."
        ),
        **resultado,
        **AGENTE_008_SEGURANCA_CONSULTA,
    }


def _resultado_consulta_documentos_classificados(
    cur: Any, obra_codigo: str, payload_comando: Any
) -> dict[str, Any]:
    payload = payload_comando if isinstance(payload_comando, dict) else {}
    modo = payload.get("modo")
    filtros = {
        "area": payload.get("area"), "disciplina": payload.get("disciplina"),
        "status_revisao": payload.get("status_revisao"),
        "eh_obsoleto": payload.get("eh_obsoleto"),
        "eh_as_built": payload.get("eh_as_built"),
        "termo": payload.get("termo"),
        "limite": 1 if modo == "ULTIMA_REVISAO" else 10,
    }
    documentos = _consultar_documentos_classificados(cur, obra_codigo, filtros, modo)
    if not documentos:
        resposta = "Não encontrei documentos classificados para esse filtro."
    else:
        linhas = [
            "Última revisão encontrada:"
            if modo == "ULTIMA_REVISAO"
            else "Documentos classificados:"
        ]
        for documento in documentos:
            nome = documento["nome_arquivo"] or "Sem nome"
            nome = nome if len(nome) <= 70 else f"{nome[:67]}..."
            revisao = (
                documento["numero_revisao"]
                or documento["data_revisao_detectada"]
                or "sem revisão/data"
            )
            linhas.append(
                f"• #{documento['documento_id']} {nome}\n"
                f"  {documento['disciplina_detectada'] or 'sem disciplina'} | "
                f"{documento['area_detectada'] or 'sem área'} | "
                f"{documento['status_revisao']} | {revisao}"
            )
        if modo == "ULTIMA_REVISAO" and documentos[0]["confianca_classificacao"] < 0.60:
            linhas.append("⚠️ Candidato com baixa confiança; confirme manualmente.")
        resposta = "\n".join(linhas)
    return {
        "tipo_resultado": "CONSULTA_DOCUMENTOS_CLASSIFICADOS",
        "resposta_telegram": resposta, "total_retornado": len(documentos),
        "documentos": documentos, "filtros": filtros, "modo": modo,
        **AGENTE_008_SEGURANCA_CONSULTA,
    }


LIMITE_PDF_PAGINAS = 5
LIMITE_PDF_CARACTERES = 12_000
LIMITE_DOWNLOAD_ANALISE_BYTES = 25 * 1024 * 1024
LIMITE_RESPOSTA_TELEGRAM = 3_000
PASTA_TEMPORARIA_ANALISE = Path("/tmp/agente_008_documentos")


def _avaliar_qualidade_texto(texto: str) -> dict[str, Any]:
    total_caracteres = len(texto)
    total_letras = sum(caractere.isalpha() for caractere in texto)
    total_digitos = sum(caractere.isdigit() for caractere in texto)
    total_palavras = len(re.findall(r"\b[^\W\d_]+\b", texto, flags=re.UNICODE))
    alpha_ratio = total_letras / total_caracteres if total_caracteres else 0.0
    digit_ratio = total_digitos / total_caracteres if total_caracteres else 0.0
    texto_util = (
        total_letras >= 80
        and total_palavras >= 15
        and alpha_ratio >= 0.25
    )
    return {
        "total_caracteres": total_caracteres,
        "total_letras": total_letras,
        "total_digitos": total_digitos,
        "total_palavras": total_palavras,
        "alpha_ratio": alpha_ratio,
        "digit_ratio": digit_ratio,
        "texto_util": texto_util,
    }


def _buscar_documento_para_analise(
    cur: Any, obra_codigo: str, payload_comando: Any
) -> Optional[dict[str, Any]]:
    payload = payload_comando if isinstance(payload_comando, dict) else {}
    documento_id = payload.get("documento_id")
    termo = payload.get("termo")
    if documento_id is not None:
        try:
            documento_id = int(documento_id)
        except (TypeError, ValueError) as exc:
            raise ValueError("documento_id deve ser um número inteiro.") from exc
        cur.execute(
            """
            SELECT id, bucket, object_key, nome_arquivo, extensao,
                   disciplina_original, categoria_documental
            FROM documentos_minio_obra
            WHERE id = %(documento_id)s AND obra_codigo = %(obra_codigo)s
            LIMIT 1;
            """,
            {"documento_id": documento_id, "obra_codigo": obra_codigo},
        )
    elif isinstance(termo, str) and termo.strip():
        cur.execute(
            """
            SELECT id, bucket, object_key, nome_arquivo, extensao,
                   disciplina_original, categoria_documental
            FROM documentos_minio_obra
            WHERE obra_codigo = %(obra_codigo)s
              AND (nome_arquivo ILIKE %(termo)s OR object_key ILIKE %(termo)s)
            ORDER BY CASE WHEN LOWER(COALESCE(extensao, '')) = 'pdf' THEN 0 ELSE 1 END,
                     atualizado_em DESC NULLS LAST, criado_em DESC NULLS LAST, id DESC
            LIMIT 1;
            """,
            {"obra_codigo": obra_codigo, "termo": f"%{termo.strip()}%"},
        )
    else:
        raise ValueError("Informe documento_id ou termo para a análise documental.")

    row = cur.fetchone()
    if row is None:
        return None
    return {
        "id": row[0],
        "bucket": row[1],
        "object_key": row[2],
        "nome_arquivo": row[3],
        "extensao": (row[4] or Path(row[3] or row[2]).suffix.lstrip(".")).lower(),
        "disciplina_original": row[5],
        "categoria_documental": row[6],
    }


def _baixar_objeto_minio_temporariamente(documento: dict[str, Any]) -> Path:
    PASTA_TEMPORARIA_ANALISE.mkdir(parents=True, exist_ok=True)
    sufixo = Path(documento["nome_arquivo"] or documento["object_key"]).suffix
    destino = PASTA_TEMPORARIA_ANALISE / f"{uuid.uuid4().hex}{sufixo}"
    try:
        get_minio_client().fget_object(
            documento["bucket"], documento["object_key"], str(destino)
        )
    except Exception:
        destino.unlink(missing_ok=True)
        raise
    return destino


def _resumir_texto_pdf(caminho: Path) -> tuple[str, list[str], dict[str, Any]]:
    try:
        from pypdf import PdfReader
    except ImportError:
        return (
            "PDF identificado; extração textual indisponível no ambiente.",
            ["Instale pypdf para habilitar a extração básica de texto."],
            {"extracao_texto": False},
        )

    leitor = PdfReader(str(caminho))
    partes: list[str] = []
    paginas_lidas = 0
    for pagina in leitor.pages[:LIMITE_PDF_PAGINAS]:
        partes.append(pagina.extract_text() or "")
        paginas_lidas += 1
        if sum(len(parte) for parte in partes) >= LIMITE_PDF_CARACTERES:
            break
    texto = re.sub(r"\s+", " ", " ".join(partes)).strip()[:LIMITE_PDF_CARACTERES]
    qualidade_texto = _avaliar_qualidade_texto(texto)
    observacoes = []
    if qualidade_texto["texto_util"]:
        resumo = texto[:1_500]
    else:
        resumo = (
            "PDF analisado, mas sem texto técnico útil extraível nas páginas avaliadas."
        )
        observacoes.extend(
            [
                "O arquivo pode ser prancha técnica, desenho vetorial, digitalização ou exportação CAD.",
                "A análise ficou limitada a metadados do documento.",
            ]
        )
    if len(leitor.pages) > paginas_lidas:
        observacoes.append(f"Análise limitada às primeiras {paginas_lidas} páginas.")
    return resumo, observacoes, {
        "extracao_texto": bool(texto),
        "paginas_total": len(leitor.pages),
        "paginas_analisadas": paginas_lidas,
        "caracteres_extraidos": len(texto),
        "qualidade_texto": qualidade_texto,
    }


def _resumir_planilha(caminho: Path) -> tuple[str, list[str], dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return (
            "Planilha XLSX identificada; inspeção interna indisponível no ambiente.",
            ["Instale openpyxl para extrair abas e dimensões básicas."],
            {"extracao_planilha": False},
        )

    workbook = load_workbook(caminho, read_only=True, data_only=True)
    try:
        abas = [
            {"nome": aba.title, "linhas": aba.max_row, "colunas": aba.max_column}
            for aba in workbook.worksheets
        ]
    finally:
        workbook.close()
    descricao = "; ".join(
        f"{aba['nome']} ({aba['linhas']}x{aba['colunas']})" for aba in abas
    )
    return f"Planilha com {len(abas)} aba(s): {descricao}."[:1_500], [], {
        "extracao_planilha": True,
        "abas": abas,
    }


def _analisar_documento_minio(documento: dict[str, Any]) -> dict[str, Any]:
    cliente = get_minio_client()
    stat = cliente.stat_object(documento["bucket"], documento["object_key"])
    extensao = documento["extensao"]
    metadados = {
        "tamanho_bytes": stat.size,
        "content_type": stat.content_type,
        "ultima_modificacao": (
            stat.last_modified.isoformat() if stat.last_modified else None
        ),
        "disciplina_original": documento["disciplina_original"],
        "categoria_documental": documento["categoria_documental"],
    }
    observacoes: list[str] = []
    caminho: Optional[Path] = None

    if stat.size > LIMITE_DOWNLOAD_ANALISE_BYTES:
        resumo = "Arquivo grande; análise preliminar restrita aos metadados."
        observacoes.append(
            f"Download não realizado: tamanho superior a {LIMITE_DOWNLOAD_ANALISE_BYTES} bytes."
        )
    elif extensao == "dwg":
        resumo = "Arquivo DWG identificado; análise preliminar restrita aos metadados."
        observacoes.append(
            "Conteúdo técnico DWG requer ferramenta especializada; não foi interpretado."
        )
    elif extensao in {"png", "jpg", "jpeg", "gif", "bmp", "tif", "tiff", "webp"}:
        resumo = "Imagem identificada; análise preliminar restrita aos metadados."
        observacoes.append("OCR não foi aplicado.")
    elif extensao not in {"pdf", "xlsx"}:
        resumo = f"Arquivo {extensao.upper() or 'sem extensão'}; análise restrita aos metadados."
    else:
        caminho = _baixar_objeto_minio_temporariamente(documento)
        try:
            if extensao == "pdf":
                resumo, observacoes_formato, metadados_formato = _resumir_texto_pdf(caminho)
            else:
                resumo, observacoes_formato, metadados_formato = _resumir_planilha(caminho)
            observacoes.extend(observacoes_formato)
            metadados.update(metadados_formato)
        finally:
            caminho.unlink(missing_ok=True)

    resposta = (
        f"Análise preliminar do documento {documento['id']} — {documento['nome_arquivo']}\n"
        f"{resumo}\n"
        + ("Observações: " + " ".join(observacoes) if observacoes else "")
    ).strip()[:LIMITE_RESPOSTA_TELEGRAM]
    return {
        "resumo": resumo,
        "observacoes": observacoes,
        "metadados_extraidos": metadados,
        "resposta_telegram": resposta,
    }


def _resultado_analise_documental(
    cur: Any, obra_codigo: str, payload_comando: Any
) -> dict[str, Any]:
    documento = _buscar_documento_para_analise(cur, obra_codigo, payload_comando)
    if documento is None:
        return {
            "tipo_resultado": "DOCUMENTO_NAO_ENCONTRADO",
            "documento_id": None,
            "nome_arquivo": None,
            "bucket": None,
            "object_key": None,
            "minio_uri": None,
            "extensao": None,
            "resumo": "Documento não encontrado para a obra e o filtro informados.",
            "observacoes": [],
            "metadados_extraidos": {},
            "resposta_telegram": "Não encontrei o documento solicitado nesta obra.",
            **AGENTE_008_SEGURANCA_CONSULTA,
        }

    analise = _analisar_documento_minio(documento)
    cur.execute(
        """
        INSERT INTO analises_documentais_obra (
            documento_id, obra_codigo, resumo, observacoes,
            metadados_extraidos, resposta_telegram
        ) VALUES (
            %(documento_id)s, %(obra_codigo)s, %(resumo)s, %(observacoes)s,
            %(metadados_extraidos)s, %(resposta_telegram)s
        );
        """,
        {
            "documento_id": documento["id"],
            "obra_codigo": obra_codigo,
            "resumo": analise["resumo"],
            "observacoes": Json(analise["observacoes"]),
            "metadados_extraidos": Json(analise["metadados_extraidos"]),
            "resposta_telegram": analise["resposta_telegram"],
        },
    )
    return {
        "tipo_resultado": "ANALISE_DOCUMENTAL_PRELIMINAR",
        "documento_id": documento["id"],
        "nome_arquivo": documento["nome_arquivo"],
        "bucket": documento["bucket"],
        "object_key": documento["object_key"],
        "minio_uri": f"s3://{documento['bucket']}/{documento['object_key']}",
        "extensao": documento["extensao"],
        **analise,
        **AGENTE_008_SEGURANCA_CONSULTA,
    }


def _colunas_tabela_operacional(cur, tabela: str) -> set[str]:
    cur.execute(
        """
        SELECT column_name FROM information_schema.columns
        WHERE table_schema = 'public' AND table_name = %s;
        """,
        (tabela,),
    )
    return {row[0] for row in cur.fetchall()}


def _linhas_fonte_operacional(
    cur, tabela: str, colunas_desejadas: list[str], obra_codigo: str
) -> tuple[list[dict[str, Any]], Optional[str]]:
    """Lê apenas colunas conhecidas; incompatibilidades ficam isoladas por savepoint."""
    try:
        with cur.connection.transaction():
            colunas = _colunas_tabela_operacional(cur, tabela)
            if not colunas:
                return [], "tabela ausente"
            selecionadas = [coluna for coluna in colunas_desejadas if coluna in colunas]
            if "obra_codigo" not in colunas:
                return [], "coluna obra_codigo ausente"
            if not selecionadas:
                return [], "schema sem colunas compatíveis"
            cur.execute(
                f"SELECT {', '.join(selecionadas)} FROM {tabela} WHERE obra_codigo = %s",
                (obra_codigo,),
            )
            return [dict(zip(selecionadas, row)) for row in cur.fetchall()], None
    except Exception as exc:
        return [], f"fonte indisponível: {str(exc)[:120]}"


def _texto_curto(valor: Any, limite: int = 120) -> str:
    texto = re.sub(r"\s+", " ", str(valor or "")).strip()
    return texto if len(texto) <= limite else f"{texto[: limite - 1].rstrip()}…"


def _gerar_diagnostico_operacional(
    cur, obra_codigo: str, area: Optional[str], incluir_documentos: bool,
    incluir_pendencias: bool, incluir_restricoes: bool,
) -> dict[str, Any]:
    fontes_consultadas: list[str] = []
    indisponibilidades: list[str] = []

    def ler(tabela: str, colunas: list[str], habilitada: bool = True) -> list[dict[str, Any]]:
        if not habilitada:
            return []
        linhas, erro = _linhas_fonte_operacional(cur, tabela, colunas, obra_codigo)
        fontes_consultadas.append(tabela)
        if erro:
            indisponibilidades.append(f"{tabela}: {_texto_curto(erro, 90)}")
        return linhas

    areas = ler("areas_obra", ["codigo_area", "nome_area", "ativo"])
    ler("eap_obra", ["codigo_eap", "descricao", "ativo"])
    atividades = ler(
        "atividades_cronograma",
        ["id", "codigo_area", "frente_servico", "status_atividade", "descricao"],
    )
    restricoes = ler(
        "restricoes_atividade",
        ["atividade_id", "codigo_area", "area", "status_restricao", "status", "descricao", "criticidade"],
        incluir_restricoes,
    )
    pendencias = ler(
        "pendencias_obra",
        ["codigo_area", "area", "area_detectada", "status_pendencia", "status", "descricao", "titulo", "criticidade"],
        incluir_pendencias,
    )
    ler("documentos_minio_obra", ["id", "nome_arquivo"], incluir_documentos)
    classificacoes = ler(
        "classificacoes_documentais_obra",
        ["area_detectada", "status_revisao", "eh_obsoleto", "confianca_classificacao", "status"],
        incluir_documentos,
    )

    atividade_area = {item.get("id"): item.get("codigo_area") for item in atividades}

    def area_item(item: dict[str, Any], restricao: bool = False) -> Optional[str]:
        valor = item.get("codigo_area") or item.get("area") or item.get("area_detectada")
        if not valor and restricao:
            valor = atividade_area.get(item.get("atividade_id"))
        return normalizar_texto_comparacao(str(valor)).replace(" ", "_") if valor else None

    filtro_area = normalizar_texto_comparacao(area).replace(" ", "_") if area else None
    if filtro_area:
        classificacoes = [item for item in classificacoes if area_item(item) == filtro_area]
        pendencias = [item for item in pendencias if area_item(item) == filtro_area]
        restricoes = [item for item in restricoes if area_item(item, True) == filtro_area]

    status_fechados = {"CONCLUIDO", "CONCLUIDA", "RESOLVIDO", "RESOLVIDA", "FECHADO", "FECHADA", "ENCERRADO", "ENCERRADA", "CANCELADO", "CANCELADA", "LIBERADA"}

    def esta_aberto(item: dict[str, Any], campos: tuple[str, ...]) -> bool:
        status = next((item.get(campo) for campo in campos if item.get(campo) is not None), None)
        return status is None or str(status).upper() not in status_fechados

    pendencias_abertas = [item for item in pendencias if esta_aberto(item, ("status_pendencia", "status"))]
    restricoes_abertas = [item for item in restricoes if esta_aberto(item, ("status_restricao", "status"))]
    obsoletos = [item for item in classificacoes if item.get("eh_obsoleto") is True or str(item.get("eh_obsoleto", "")).lower() in {"true", "t", "1"} or str(item.get("status_revisao", "")).upper() == "OBSOLETO"]
    sem_revisao = [item for item in classificacoes if not item.get("status_revisao") or str(item.get("status_revisao")).upper() in {"NAO_IDENTIFICADO", "SEM_REVISAO"}]
    def confianca(item: dict[str, Any]) -> float:
        try:
            return float(item.get("confianca_classificacao") or 0)
        except (TypeError, ValueError):
            return 0

    baixa_confianca = [item for item in classificacoes if confianca(item) < 0.70]

    nomes_area: dict[str, str] = {}
    for item in areas:
        codigo = area_item(item)
        if codigo and item.get("ativo") is not False:
            nomes_area[codigo] = str(item.get("nome_area") or item.get("codigo_area") or codigo)
    for colecao, eh_restricao in ((classificacoes, False), (pendencias_abertas, False), (restricoes_abertas, True)):
        for item in colecao:
            codigo = area_item(item, eh_restricao)
            if codigo:
                nomes_area.setdefault(codigo, codigo.replace("_", " ").title())
    if filtro_area:
        nomes_area = {filtro_area: nomes_area.get(filtro_area, filtro_area.replace("_", " ").title())}

    riscos_todas_areas = []
    for codigo, nome in sorted(nomes_area.items()):
        docs_area = [item for item in classificacoes if area_item(item) == codigo]
        obs_area = sum(item in obsoletos for item in docs_area)
        pend_area = sum(area_item(item) == codigo for item in pendencias_abertas)
        rest_area = sum(area_item(item, True) == codigo for item in restricoes_abertas)
        if obs_area and (pend_area or rest_area):
            risco = "BLOQUEIO_CRITICO"
        elif obs_area:
            risco = "BLOQUEIO_POTENCIAL"
        elif sum(item in sem_revisao or item in baixa_confianca for item in docs_area) >= 5:
            risco = "REQUER_VALIDACAO_TECNICA"
        else:
            risco = "OPERACIONAL_SEM_ALERTAS"
        riscos_todas_areas.append({
            "area": codigo, "nome_area": nome, "status_operacional": risco,
            "documentos_obsoletos": obs_area, "pendencias_abertas": pend_area,
            "restricoes_abertas": rest_area,
        })

    muitos_incertos = len(sem_revisao) + len(baixa_confianca) >= max(5, len(classificacoes) // 5)
    if any(item["status_operacional"] == "BLOQUEIO_CRITICO" for item in riscos_todas_areas):
        status_operacional = "BLOQUEIO_CRITICO"
    elif obsoletos:
        status_operacional = "BLOQUEIO_POTENCIAL"
    elif muitos_incertos:
        status_operacional = "REQUER_VALIDACAO_TECNICA"
    else:
        status_operacional = "OPERACIONAL_SEM_ALERTAS"

    alertas: list[str] = []
    if obsoletos:
        alertas.append(f"Há {len(obsoletos)} documento(s) obsoleto(s) no recorte consultado.")
    if pendencias_abertas:
        alertas.append(f"Há {len(pendencias_abertas)} pendência(s) aberta(s) que requerem verificação.")
    if restricoes_abertas:
        alertas.append(f"Há {len(restricoes_abertas)} restrição(ões) aberta(s) que podem afetar o avanço.")
    if sem_revisao:
        alertas.append(f"Há {len(sem_revisao)} documento(s) sem revisão identificada.")
    if baixa_confianca:
        alertas.append(f"Há {len(baixa_confianca)} classificação(ões) documentais com baixa confiança.")
    if not alertas and indisponibilidades:
        alertas.append("O diagnóstico é parcial porque uma ou mais fontes não estavam disponíveis.")
    alertas = [_texto_curto(item) for item in alertas[:5]]

    recomendacao = (
        "Validar documentação, pendências e restrições antes de liberar execução em campo. "
        "Este diagnóstico é consultivo e não confirma liberação definitiva de nenhuma frente."
    )
    area_exibicao = nomes_area.get(filtro_area, area) if filtro_area else "todas"
    linhas = [
        f"🏗️ Diagnóstico operacional — {obra_codigo}", "", "Filtro:",
        f"Área: {area_exibicao}", "", "Resumo:",
        f"- Documentos classificados: {len(classificacoes)}",
        f"- Documentos obsoletos: {len(obsoletos)}",
        f"- Sem revisão identificada: {len(sem_revisao)}",
        f"- Baixa confiança documental: {len(baixa_confianca)}",
        f"- Pendências abertas: {len(pendencias_abertas)}",
        f"- Restrições abertas: {len(restricoes_abertas)}", "",
        "Status operacional:", status_operacional,
    ]
    if alertas:
        linhas.extend(["", "Alertas:", *[f"- {item}" for item in alertas]])
    linhas.extend(["", "Recomendação:", recomendacao])
    return {
        "obra_codigo": obra_codigo, "area": area, "fontes_consultadas": fontes_consultadas,
        "total_documentos_classificados": len(classificacoes),
        "total_documentos_obsoletos": len(obsoletos),
        "total_documentos_sem_revisao": len(sem_revisao),
        "total_documentos_baixa_confianca": len(baixa_confianca),
        "total_pendencias_abertas": len(pendencias_abertas),
        "total_restricoes_abertas": len(restricoes_abertas),
        "areas_identificadas": [{"area": codigo, "nome_area": nome} for codigo, nome in list(sorted(nomes_area.items()))[:5]],
        "riscos_por_area": riscos_todas_areas[:5], "alertas": alertas,
        "fontes_indisponiveis": indisponibilidades, "status_operacional": status_operacional,
        "recomendacao": recomendacao, "resposta_telegram": "\n".join(linhas),
        **AGENTE_008_SEGURANCA_CONSULTA,
    }


def _gerar_plano_operacional(
    cur, obra_codigo: str, area: Optional[str], limite_acoes: int,
) -> dict[str, Any]:
    fontes_indisponiveis: list[str] = []

    def ler(tabela: str, colunas: list[str]) -> list[dict[str, Any]]:
        linhas, erro = _linhas_fonte_operacional(cur, tabela, colunas, obra_codigo)
        if erro:
            fontes_indisponiveis.append(f"{tabela}: {_texto_curto(erro, 90)}")
        return linhas

    classificacoes = ler("classificacoes_documentais_obra", [
        "documento_id", "area_detectada", "disciplina_detectada", "status_revisao",
        "eh_obsoleto", "confianca_classificacao", "risco_documental", "nivel_risco",
        "data_revisao_detectada", "numero_revisao", "atualizado_em", "updated_at", "status",
    ])
    documentos = ler("documentos_minio_obra", [
        "id", "nome_arquivo", "area", "codigo_area", "ultima_modificacao", "atualizado_em",
    ])
    pendencias = ler("pendencias_obra", [
        "codigo_area", "area", "area_detectada", "status_pendencia", "status", "titulo",
        "descricao", "criticidade", "prazo", "data_prazo", "responsavel", "responsavel_id",
    ])
    restricoes = ler("restricoes_atividade", [
        "atividade_id", "codigo_area", "area", "status_restricao", "status", "descricao", "criticidade",
    ])
    areas = ler("areas_obra", ["codigo_area", "nome_area", "ativo"])
    atividades = ler("atividades_cronograma", ["id", "codigo_area", "frente_servico", "descricao"])

    filtro_area = normalizar_texto_comparacao(area).replace(" ", "_") if area else None
    atividade_area = {item.get("id"): item.get("codigo_area") for item in atividades}
    def area_item(item: dict[str, Any], usar_atividade: bool = False) -> Optional[str]:
        valor = item.get("codigo_area") or item.get("area") or item.get("area_detectada")
        if not valor and usar_atividade:
            valor = atividade_area.get(item.get("atividade_id"))
        return normalizar_texto_comparacao(str(valor)).replace(" ", "_") if valor else None

    if filtro_area:
        classificacoes = [item for item in classificacoes if area_item(item) == filtro_area]
        pendencias = [item for item in pendencias if area_item(item) == filtro_area]
        restricoes = [item for item in restricoes if area_item(item, True) == filtro_area]
        documentos = [item for item in documentos if area_item(item) == filtro_area]

    nomes_area = {
        area_item(item): str(item.get("nome_area") or item.get("codigo_area"))
        for item in areas if area_item(item) and item.get("ativo") is not False
    }
    area_exibicao = nomes_area.get(filtro_area, area) if filtro_area else "todas"
    status_fechados = {
        "CONCLUIDO", "CONCLUIDA", "RESOLVIDO", "RESOLVIDA", "FECHADO", "FECHADA",
        "ENCERRADO", "ENCERRADA", "CANCELADO", "CANCELADA", "LIBERADA",
    }

    def aberto(item: dict[str, Any], campos: tuple[str, ...]) -> bool:
        status = next((item.get(campo) for campo in campos if item.get(campo) is not None), None)
        return status is None or str(status).upper() not in status_fechados

    def verdadeiro(valor: Any) -> bool:
        return valor is True or str(valor).lower() in {"true", "t", "1", "sim"}

    def baixa_confianca_documental(item: dict[str, Any]) -> bool:
        valor = item.get("confianca_classificacao")
        if valor is None:
            return False
        try:
            return float(valor) < 0.50
        except (TypeError, ValueError):
            return False

    pendencias_abertas = [item for item in pendencias if aberto(item, ("status_pendencia", "status"))]
    restricoes_abertas = [item for item in restricoes if aberto(item, ("status_restricao", "status"))]
    obsoletos = [item for item in classificacoes if verdadeiro(item.get("eh_obsoleto")) or str(item.get("status_revisao", "")).upper() == "OBSOLETO"]
    alterados = [item for item in classificacoes if str(item.get("status_revisao", "")).upper() == "ALTERACAO"]
    risco_alto = [item for item in classificacoes if str(item.get("risco_documental") or item.get("nivel_risco") or "").upper() in {"ALTO", "CRITICO", "CRÍTICO"}]
    sem_revisao = [item for item in classificacoes if not (item.get("numero_revisao") or item.get("status_revisao")) or str(item.get("status_revisao", "")).upper() in {"NAO_IDENTIFICADO", "SEM_REVISAO"}]
    sem_data_revisao = [item for item in classificacoes if not item.get("data_revisao_detectada")]
    baixa_confianca = [item for item in classificacoes if baixa_confianca_documental(item)]
    pendencias_criticas = [item for item in pendencias_abertas if str(item.get("criticidade", "")).upper() in {"ALTA", "CRITICA", "CRÍTICA"}]
    pendencias_incompletas = [item for item in pendencias_abertas if not (item.get("prazo") or item.get("data_prazo")) or not (item.get("responsavel") or item.get("responsavel_id"))]

    totais_documentais = {
        "total_obsoletos": len(obsoletos),
        "total_alteracao": len(alterados),
        "total_sem_revisao": len(sem_revisao),
        "total_baixa_confianca": len(baixa_confianca),
        "total_sem_data_revisao": len(sem_data_revisao),
    }
    try:
        with cur.connection.transaction():
            riscos_documentais = _avaliar_riscos_documentais(
                cur, obra_codigo, area, None, min(limite_acoes, 50)
            )
        totais_documentais.update(riscos_documentais["totais"])
    except Exception as exc:
        fontes_indisponiveis.append(
            f"riscos_documentais: {_texto_curto(exc, 90)}"
        )

    candidatas: dict[str, list[dict[str, Any]]] = {f"PRIORIDADE_{n}": [] for n in range(1, 4)}

    def adicionar(prioridade: str, tipo: str, quantidade: int, acao: str) -> None:
        if quantidade:
            candidatas[prioridade].append({
                "prioridade": prioridade, "tipo": tipo, "quantidade": quantidade,
                "area": area, "acao": _texto_curto(acao, 150),
            })

    total_obsoletos = totais_documentais["total_obsoletos"]
    total_alteracao = totais_documentais["total_alteracao"]
    total_sem_revisao = totais_documentais["total_sem_revisao"]
    total_baixa_confianca = totais_documentais["total_baixa_confianca"]
    total_sem_data_revisao = totais_documentais["total_sem_data_revisao"]

    adicionar("PRIORIDADE_1", "DOCUMENTO_OBSOLETO", total_obsoletos, f"Segregar e validar tecnicamente {total_obsoletos} documento(s) obsoleto(s) antes do uso em campo.")
    adicionar("PRIORIDADE_1", "ALTERACAO_DOCUMENTAL_RECENTE", total_alteracao, f"Revisar o impacto de {total_alteracao} alteração(ões) documental(is) recente(s) com a equipe técnica.")
    adicionar("PRIORIDADE_1", "RESTRICAO_ABERTA", len(restricoes_abertas), f"Tratar {len(restricoes_abertas)} restrição(ões) aberta(s) e registrar responsável e evidência de resolução.")
    adicionar("PRIORIDADE_1", "PENDENCIA_CRITICA", len(pendencias_criticas), f"Priorizar {len(pendencias_criticas)} pendência(s) crítica(s) aberta(s) para validação técnica.")
    adicionar("PRIORIDADE_1", "RISCO_DOCUMENTAL_ALTO", len(risco_alto), f"Avaliar {len(risco_alto)} documento(s) com risco alto ou crítico antes de orientar a frente.")
    adicionar("PRIORIDADE_2", "REVISAO_NAO_IDENTIFICADA", total_sem_revisao, f"Identificar a revisão de {total_sem_revisao} documento(s) e confirmar a versão aplicável.")
    adicionar("PRIORIDADE_2", "DATA_REVISAO_AUSENTE", total_sem_data_revisao, f"Complementar a data de revisão de {total_sem_data_revisao} documento(s).")
    adicionar("PRIORIDADE_2", "BAIXA_CONFIANCA_DOCUMENTAL", total_baixa_confianca, f"Validar manualmente {total_baixa_confianca} classificação(ões) documental(is) de baixa confiança.")
    adicionar("PRIORIDADE_2", "PENDENCIA_INCOMPLETA", len(pendencias_incompletas), f"Definir prazo e responsável nas {len(pendencias_incompletas)} pendência(s) incompleta(s).")
    adicionar("PRIORIDADE_3", "REVISAR_CLASSIFICACAO", len(classificacoes), "Revisar a classificação documental do recorte e corrigir inconsistências.")
    adicionar("PRIORIDADE_3", "ORGANIZAR_DOCUMENTACAO", len(documentos), "Organizar a documentação por área e disciplina para facilitar a consulta técnica.")
    adicionar("PRIORIDADE_3", "COMPLEMENTAR_METADADOS", len(documentos), "Complementar metadados ausentes dos documentos catalogados.")

    acoes: list[dict[str, Any]] = []
    for prioridade in ("PRIORIDADE_1", "PRIORIDADE_2", "PRIORIDADE_3"):
        acoes.extend(candidatas[prioridade][:3])
    acoes = acoes[:min(limite_acoes, 10)]
    pendencias_sugeridas = [
        {"tipo": item["tipo"], "area": area, "descricao": item["acao"]}
        for item in acoes if item["prioridade"] in {"PRIORIDADE_1", "PRIORIDADE_2"}
    ]
    if any(item["prioridade"] == "PRIORIDADE_1" for item in acoes):
        status_operacional = "ACAO_IMEDIATA_RECOMENDADA"
    elif acoes:
        status_operacional = "VALIDACAO_OPERACIONAL_RECOMENDADA"
    else:
        status_operacional = "SEM_ACOES_IDENTIFICADAS"
    recomendacao = (
        "Executar validação técnica antes de liberar frente em campo. "
        "O plano é consultivo e não confirma liberação definitiva."
    )
    linhas = [f"🛠️ Plano operacional — {obra_codigo}", "", "Filtro:", f"Área: {area_exibicao}"]
    for prioridade, titulo in (("PRIORIDADE_1", "Prioridade 1:"), ("PRIORIDADE_2", "Prioridade 2:"), ("PRIORIDADE_3", "Prioridade 3:")):
        itens = [item for item in acoes if item["prioridade"] == prioridade]
        linhas.extend(["", titulo])
        linhas.extend([f"- {item['acao']}" for item in itens] or ["- Nenhuma ação identificada."])
    linhas.extend(["", "Recomendação:", recomendacao])
    return {
        "obra_codigo": obra_codigo, "area": area, "total_acoes": len(acoes),
        "acoes_priorizadas": acoes, "pendencias_sugeridas": pendencias_sugeridas,
        "status_operacional": status_operacional, "recomendacao": recomendacao,
        "resposta_telegram": "\n".join(linhas),
        "fontes_indisponiveis": fontes_indisponiveis,
        **AGENTE_008_SEGURANCA_CONSULTA,
    }


def _gerar_resumo_executivo_operacional(
    cur, obra_codigo: str, area: Optional[str], limite_itens: int,
) -> dict[str, Any]:
    fontes_indisponiveis: list[str] = []

    def ler(tabela: str, colunas: list[str]) -> list[dict[str, Any]]:
        linhas, erro = _linhas_fonte_operacional(cur, tabela, colunas, obra_codigo)
        if erro:
            fontes_indisponiveis.append(f"{tabela}: {_texto_curto(erro, 90)}")
        return linhas

    classificacoes = ler("classificacoes_documentais_obra", [
        "area_detectada", "codigo_area", "area", "status_revisao", "numero_revisao",
        "eh_obsoleto", "confianca_classificacao", "status",
    ])
    documentos = ler("documentos_minio_obra", ["id", "area", "codigo_area", "nome_arquivo"])
    acoes = ler("acoes_operacionais_obra", [
        "area", "codigo_area", "status", "prioridade", "titulo", "descricao", "prazo",
    ])
    pendencias = ler("pendencias_obra", [
        "area", "codigo_area", "area_detectada", "status", "status_pendencia",
        "titulo", "descricao", "criticidade",
    ])
    restricoes = ler("restricoes_atividade", [
        "atividade_id", "area", "codigo_area", "status", "status_restricao",
        "descricao", "criticidade",
    ])
    areas = ler("areas_obra", ["codigo_area", "nome_area", "ativo"])
    atividades = ler("atividades_cronograma", ["id", "codigo_area", "area", "descricao"])

    filtro_area = normalizar_texto_comparacao(area).replace(" ", "_") if area else None
    atividade_area = {item.get("id"): item.get("codigo_area") or item.get("area") for item in atividades}

    def area_item(item: dict[str, Any], usar_atividade: bool = False) -> Optional[str]:
        valor = item.get("codigo_area") or item.get("area") or item.get("area_detectada")
        if not valor and usar_atividade:
            valor = atividade_area.get(item.get("atividade_id"))
        return normalizar_texto_comparacao(str(valor)).replace(" ", "_") if valor else None

    if filtro_area:
        classificacoes = [item for item in classificacoes if area_item(item) == filtro_area]
        documentos = [item for item in documentos if area_item(item) == filtro_area]
        acoes = [item for item in acoes if area_item(item) == filtro_area]
        pendencias = [item for item in pendencias if area_item(item) == filtro_area]
        restricoes = [item for item in restricoes if area_item(item, True) == filtro_area]

    status_fechados = {
        "CONCLUIDO", "CONCLUIDA", "RESOLVIDO", "RESOLVIDA", "FECHADO", "FECHADA",
        "ENCERRADO", "ENCERRADA", "CANCELADO", "CANCELADA", "LIBERADO", "LIBERADA",
    }

    def aberto(item: dict[str, Any], campos: tuple[str, ...]) -> bool:
        status = next((item.get(campo) for campo in campos if item.get(campo) is not None), None)
        return status is None or str(status).upper() not in status_fechados

    def verdadeiro(valor: Any) -> bool:
        return valor is True or str(valor).lower() in {"true", "t", "1", "sim"}

    def baixa_confianca(item: dict[str, Any]) -> bool:
        valor = item.get("confianca_classificacao")
        if valor is None:
            return False
        try:
            return float(valor) < 0.70
        except (TypeError, ValueError):
            return False

    obsoletos = [item for item in classificacoes if verdadeiro(item.get("eh_obsoleto")) or str(item.get("status_revisao", "")).upper() == "OBSOLETO"]
    sem_revisao = [item for item in classificacoes if not (item.get("numero_revisao") or item.get("status_revisao")) or str(item.get("status_revisao", "")).upper() in {"NAO_IDENTIFICADO", "SEM_REVISAO"}]
    baixa_confianca_itens = [item for item in classificacoes if baixa_confianca(item)]
    acoes_abertas = [item for item in acoes if str(item.get("status", "")).upper() == "ABERTA"]
    acoes_andamento = [item for item in acoes if str(item.get("status", "")).upper() == "EM_ANDAMENTO"]
    acoes_concluidas = [item for item in acoes if str(item.get("status", "")).upper() == "CONCLUIDA"]
    acoes_canceladas = [item for item in acoes if str(item.get("status", "")).upper() == "CANCELADA"]
    acoes_nao_encerradas = acoes_abertas + acoes_andamento
    criticas = [item for item in acoes_nao_encerradas if str(item.get("prioridade", "")).upper() == "CRITICA"]
    altas = [item for item in acoes_nao_encerradas if str(item.get("prioridade", "")).upper() == "ALTA"]
    pendencias_abertas = [item for item in pendencias if aberto(item, ("status_pendencia", "status"))]
    restricoes_abertas = [item for item in restricoes if aberto(item, ("status_restricao", "status"))]

    muitos_incertos = (
        len(sem_revisao) + len(baixa_confianca_itens)
        >= max(5, (len(classificacoes) + 4) // 5)
    )
    if obsoletos or criticas or altas:
        status_executivo = "BLOQUEIO_POTENCIAL"
    elif muitos_incertos:
        status_executivo = "REQUER_VALIDACAO_TECNICA"
    elif acoes_nao_encerradas or pendencias_abertas or restricoes_abertas:
        status_executivo = "REQUER_ATENCAO"
    else:
        status_executivo = "SEM_ALERTAS_RELEVANTES"

    alertas: list[str] = []
    proximos_passos: list[str] = []

    def registrar(condicao: bool, alerta: str, passo: str) -> None:
        if condicao:
            alertas.append(_texto_curto(alerta))
            proximos_passos.append(_texto_curto(passo))

    registrar(bool(obsoletos), f"Há {len(obsoletos)} documento(s) obsoleto(s) no recorte.", "Validar tecnicamente e segregar documentos obsoletos antes do uso em campo.")
    registrar(bool(criticas or altas), f"Há {len(criticas)} ação(ões) crítica(s) e {len(altas)} alta(s) não encerrada(s).", "Priorizar ações críticas e altas, confirmando responsáveis e prazos.")
    registrar(bool(sem_revisao), f"Há {len(sem_revisao)} documento(s) sem revisão identificada.", "Confirmar as revisões documentais aplicáveis com a equipe técnica.")
    registrar(bool(baixa_confianca_itens), f"Há {len(baixa_confianca_itens)} classificação(ões) com baixa confiança.", "Revisar manualmente as classificações documentais de baixa confiança.")
    registrar(bool(pendencias_abertas or restricoes_abertas), f"Há {len(pendencias_abertas)} pendência(s) e {len(restricoes_abertas)} restrição(ões) aberta(s).", "Verificar pendências e restrições, registrando tratamento e evidências.")
    if fontes_indisponiveis and len(alertas) < 5:
        alertas.append("O resumo é parcial porque uma ou mais fontes não estavam disponíveis.")
    if not proximos_passos:
        proximos_passos.append("Manter acompanhamento consultivo; qualquer liberação exige validação técnica responsável.")
    alertas = alertas[:min(limite_itens, 5)]
    proximos_passos = proximos_passos[:min(limite_itens, 5)]

    nomes_area = {
        area_item(item): str(item.get("nome_area") or item.get("codigo_area"))
        for item in areas if area_item(item) and item.get("ativo") is not False
    }
    area_exibicao = nomes_area.get(filtro_area, area) if filtro_area else "todas"
    linhas = [
        f"📊 Resumo executivo operacional — {obra_codigo}", "", "Filtro:",
        f"Área: {area_exibicao}", "", "Documentação:",
        f"- Classificados: {len(classificacoes)}", f"- Obsoletos: {len(obsoletos)}",
        f"- Sem revisão: {len(sem_revisao)}", f"- Baixa confiança: {len(baixa_confianca_itens)}",
        "", "Ações operacionais:", f"- Abertas: {len(acoes_abertas)}",
        f"- Em andamento: {len(acoes_andamento)}", f"- Concluídas: {len(acoes_concluidas)}",
        f"- Críticas/altas: {len(criticas) + len(altas)}", "", "Status executivo:",
        status_executivo, "", "Alertas:",
        *([f"- {item}" for item in alertas] or ["- Nenhum alerta relevante identificado nas fontes disponíveis."]),
        "", "Próximos passos:", *[f"- {item}" for item in proximos_passos], "",
        "Modo: CONSULTA", "Nenhum cronograma, RDO, MinIO ou OpenProject foi alterado.",
        "Este resumo não confirma liberação definitiva da obra ou de qualquer frente.",
    ]
    return {
        "obra_codigo": obra_codigo, "area": area,
        "totais_documentais": {
            "documentos_classificados": len(classificacoes), "documentos_obsoletos": len(obsoletos),
            "documentos_sem_revisao": len(sem_revisao), "documentos_baixa_confianca": len(baixa_confianca_itens),
        },
        "totais_acoes_operacionais": {
            "abertas": len(acoes_abertas), "em_andamento": len(acoes_andamento),
            "concluidas": len(acoes_concluidas), "canceladas": len(acoes_canceladas),
            "criticas": len(criticas), "altas": len(altas),
        },
        "pendencias_abertas": len(pendencias_abertas), "restricoes_abertas": len(restricoes_abertas),
        "status_executivo": status_executivo, "principais_alertas": alertas,
        "proximos_passos_recomendados": proximos_passos,
        "resposta_telegram": "\n".join(linhas),
        "fontes_indisponiveis": fontes_indisponiveis,
        "documentos_catalogados": len(documentos),
        "flags": {
            "modo": "CONSULTA", "altera_cronograma": False, "executa_rpa": False,
            "sincroniza_openproject": False, "altera_rdo_oficial": False,
            "envia_terceiros": False,
        },
        **AGENTE_008_SEGURANCA_CONSULTA,
    }


def _gerar_briefing_diario(
    cur, obra_codigo: str, area: Optional[str], limite_itens: int,
    incluir_acoes: bool, incluir_documentos: bool, incluir_historico: bool,
) -> dict[str, Any]:
    fontes_indisponiveis: list[str] = []

    def ler(tabela: str, colunas: list[str], habilitada: bool = True) -> list[dict[str, Any]]:
        if not habilitada:
            return []
        linhas, erro = _linhas_fonte_operacional(cur, tabela, colunas, obra_codigo)
        if erro:
            fontes_indisponiveis.append(f"{tabela}: {_texto_curto(erro, 90)}")
        return linhas

    acoes = ler("acoes_operacionais_obra", [
        "id", "area", "codigo_area", "status", "prioridade", "titulo", "descricao",
        "responsavel", "prazo", "criado_em", "atualizado_em", "concluido_em",
    ], incluir_acoes)
    historico = ler("historico_acoes_operacionais_obra", [
        "id", "acao_id", "tipo_evento", "status_anterior", "status_novo", "observacao",
        "criado_em",
    ], incluir_historico)
    classificacoes = ler("classificacoes_documentais_obra", [
        "documento_id", "area_detectada", "codigo_area", "area", "status_revisao",
        "numero_revisao", "eh_obsoleto", "confianca_classificacao", "risco_documental",
        "nivel_risco",
    ], incluir_documentos)
    documentos = ler("documentos_minio_obra", [
        "id", "nome_arquivo", "area", "codigo_area",
    ], incluir_documentos)
    pendencias = ler("pendencias_obra", [
        "id", "area", "codigo_area", "area_detectada", "status", "status_pendencia",
        "titulo", "descricao", "criticidade", "responsavel", "prazo", "data_prazo",
    ])
    restricoes = ler("restricoes_atividade", [
        "id", "atividade_id", "area", "codigo_area", "status", "status_restricao",
        "titulo", "descricao", "criticidade", "responsavel", "prazo",
    ])
    areas = ler("areas_obra", ["codigo_area", "nome_area", "ativo"])
    atividades = ler("atividades_cronograma", [
        "id", "codigo_area", "area", "frente_servico", "descricao", "status_atividade",
    ])

    filtro_area = normalizar_texto_comparacao(area).replace(" ", "_") if area else None
    atividade_area = {item.get("id"): item.get("codigo_area") or item.get("area") for item in atividades}

    def area_item(item: dict[str, Any], usar_atividade: bool = False) -> Optional[str]:
        valor = item.get("codigo_area") or item.get("area") or item.get("area_detectada")
        if not valor and usar_atividade:
            valor = atividade_area.get(item.get("atividade_id"))
        return normalizar_texto_comparacao(str(valor)).replace(" ", "_") if valor else None

    if filtro_area:
        acoes = [item for item in acoes if area_item(item) == filtro_area]
        classificacoes = [item for item in classificacoes if area_item(item) == filtro_area]
        documentos = [item for item in documentos if area_item(item) == filtro_area]
        pendencias = [item for item in pendencias if area_item(item) == filtro_area]
        restricoes = [item for item in restricoes if area_item(item, True) == filtro_area]
        ids_acoes = {item.get("id") for item in acoes}
        historico = [item for item in historico if item.get("acao_id") in ids_acoes]

    fechados = {"CONCLUIDO", "CONCLUIDA", "RESOLVIDO", "RESOLVIDA", "FECHADO", "FECHADA", "CANCELADO", "CANCELADA", "ENCERRADO", "ENCERRADA"}

    def status(item: dict[str, Any], *campos: str) -> str:
        valor = next((item.get(campo) for campo in campos if item.get(campo) is not None), "")
        return str(valor).upper()

    def verdadeiro(valor: Any) -> bool:
        return valor is True or str(valor).lower() in {"true", "t", "1", "sim"}

    def baixa_confianca(item: dict[str, Any]) -> bool:
        try:
            return item.get("confianca_classificacao") is not None and float(item["confianca_classificacao"]) < 0.70
        except (TypeError, ValueError):
            return False

    abertas = [item for item in acoes if status(item, "status") == "ABERTA"]
    andamento = [item for item in acoes if status(item, "status") == "EM_ANDAMENTO"]
    concluidas = [item for item in acoes if status(item, "status") in {"CONCLUIDA", "CONCLUIDO"}]
    nao_encerradas = abertas + andamento
    criticas_altas = [item for item in nao_encerradas if status(item, "prioridade") in {"CRITICA", "CRÍTICA", "ALTA"}]
    obsoletos = [item for item in classificacoes if verdadeiro(item.get("eh_obsoleto")) or status(item, "status_revisao") == "OBSOLETO"]
    sem_revisao = [item for item in classificacoes if not (item.get("numero_revisao") or item.get("status_revisao")) or status(item, "status_revisao") in {"NAO_IDENTIFICADO", "SEM_REVISAO"}]
    baixa_confianca_docs = [item for item in classificacoes if baixa_confianca(item)]
    pendencias_abertas = [item for item in pendencias if status(item, "status_pendencia", "status") not in fechados]
    restricoes_abertas = [item for item in restricoes if status(item, "status_restricao", "status") not in fechados]
    muitos_incertos = len(sem_revisao) + len(baixa_confianca_docs) >= max(5, (len(classificacoes) + 4) // 5)
    if criticas_altas or obsoletos:
        status_executivo = "BLOQUEIO_POTENCIAL"
    elif muitos_incertos:
        status_executivo = "REQUER_VALIDACAO_TECNICA"
    elif nao_encerradas or pendencias_abertas or restricoes_abertas or fontes_indisponiveis:
        status_executivo = "REQUER_ATENCAO"
    else:
        status_executivo = "SEM_ALERTAS_RELEVANTES"

    def resumir_acao(item: dict[str, Any]) -> dict[str, Any]:
        return {"id": item.get("id"), "prioridade": status(item, "prioridade") or "NAO_INFORMADA", "titulo": _texto_curto(item.get("titulo") or item.get("descricao") or "Ação sem título", 100), "status": status(item, "status") or "NAO_INFORMADO", "responsavel": item.get("responsavel"), "prazo": item.get("prazo")}

    ordem = {"CRITICA": 0, "CRÍTICA": 0, "ALTA": 1, "MEDIA": 2, "BAIXA": 3}
    nao_encerradas.sort(key=lambda item: (ordem.get(status(item, "prioridade"), 9), str(item.get("prazo") or "9999")))
    limite = min(limite_itens, 10)
    para_hoje = [resumir_acao(item) for item in nao_encerradas[:min(limite, 5)]]
    concluidas.sort(key=lambda item: str(item.get("concluido_em") or item.get("atualizado_em") or item.get("criado_em") or ""), reverse=True)
    documentos_criticos = []
    vistos = set()
    for item in obsoletos + sem_revisao + baixa_confianca_docs:
        chave = item.get("documento_id") or id(item)
        if chave in vistos:
            continue
        vistos.add(chave)
        documentos_criticos.append({"documento_id": item.get("documento_id"), "status_revisao": item.get("status_revisao"), "risco": item.get("risco_documental") or item.get("nivel_risco"), "motivo": "OBSOLETO" if item in obsoletos else "REVISAO_OU_CONFIANCA_REQUER_VALIDACAO"})
    documentos_criticos = documentos_criticos[:limite]
    decisoes = [_texto_curto(item.get("titulo") or item.get("descricao") or "Item sem descrição", 110) for item in (pendencias_abertas + restricoes_abertas)[:limite]]
    atencao = []
    if criticas_altas: atencao.append(f"{len(criticas_altas)} ação(ões) crítica(s) ou alta(s) requer(em) priorização.")
    if obsoletos: atencao.append(f"{len(obsoletos)} documento(s) obsoleto(s) exige(m) validação técnica antes do uso.")
    if muitos_incertos: atencao.append(f"{len(sem_revisao) + len(baixa_confianca_docs)} documento(s) sem revisão ou com baixa confiança.")
    if pendencias_abertas or restricoes_abertas: atencao.append(f"{len(pendencias_abertas)} pendência(s) e {len(restricoes_abertas)} restrição(ões) aberta(s).")
    if fontes_indisponiveis: atencao.append("Briefing parcial: uma ou mais fontes operacionais estão indisponíveis.")
    atencao = [_texto_curto(item, 120) for item in atencao[:5]]
    proximos = [_texto_curto(f"Priorizar #{item['id']} [{item['prioridade']}] {item['titulo']}", 120) for item in para_hoje]
    if obsoletos: proximos.append("Validar e segregar documentos obsoletos antes de qualquer uso em campo.")
    if not proximos: proximos.append("Manter acompanhamento consultivo e validar tecnicamente qualquer decisão de liberação.")
    proximos = proximos[:limite]
    nomes_area = {area_item(item): str(item.get("nome_area") or item.get("codigo_area")) for item in areas if area_item(item) and item.get("ativo") is not False}
    area_exibicao = nomes_area.get(filtro_area, area) if filtro_area else "todas"
    acoes_telegram = [resumir_acao(item) for item in nao_encerradas[:7]]
    linhas = [f"🗓️ Briefing executivo diário — {obra_codigo}", "", "Filtro:", f"Área: {area_exibicao}", "", "Status geral:", status_executivo, "", "Atenção imediata:", *([f"- {item}" for item in atencao] or ["- Nenhum alerta imediato nas fontes disponíveis."]), "", "Ações para hoje:", *([f"- #{item['id']} {item['titulo']}" for item in para_hoje] or ["- Nenhuma ação priorizada."]), "", "Ações abertas:", *([f"- #{item['id']} [{item['prioridade']}] {item['titulo']} ({item['status']})" for item in acoes_telegram] or ["- Nenhuma ação aberta ou em andamento."]), "", "Decisões pendentes:", *([f"- {item}" for item in decisoes[:5]] or ["- Nenhuma decisão pendente identificada."]), "", "Modo: CONSULTA", "Nenhum cronograma, RDO, MinIO ou OpenProject foi alterado.", "Este briefing não confirma liberação definitiva da obra ou de qualquer frente."]
    resultado = {"obra_codigo": obra_codigo, "area": area, "data_briefing": date.today(), "status_executivo": status_executivo, "resumo_geral": _texto_curto(f"{len(abertas)} ação(ões) aberta(s), {len(andamento)} em andamento, {len(pendencias_abertas)} pendência(s), {len(restricoes_abertas)} restrição(ões) e {len(documentos_criticos)} documento(s) crítico(s) no recorte consultado.", 220), "atencao_imediata": atencao, "acoes_para_hoje": para_hoje, "acoes_abertas": [resumir_acao(item) for item in abertas[:limite]], "acoes_em_andamento": [resumir_acao(item) for item in andamento[:limite]], "acoes_concluidas_recentementes": [resumir_acao(item) for item in concluidas[:limite]], "documentos_criticos": documentos_criticos, "decisoes_pendentes": decisoes, "proximos_passos": proximos, "resposta_telegram": "\n".join(linhas), "fontes_indisponiveis": fontes_indisponiveis, "total_eventos_historico": len(historico), "total_documentos_catalogados": len(documentos), "flags_seguranca": dict(AGENTE_008_SEGURANCA_CONSULTA), **AGENTE_008_SEGURANCA_CONSULTA}
    return serializar_json_seguro(resultado)


@app.post("/agentes/gestao-operacional/briefing-diario")
def briefing_diario(payload: GestaoOperacionalBriefingDiarioRequest):
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                return _gerar_briefing_diario(cur, payload.obra_codigo, payload.area, payload.limite_itens, payload.incluir_acoes, payload.incluir_documentos, payload.incluir_historico)
    except Exception as exc:
        raise HTTPException(status_code=500, detail={"message": "Erro ao gerar briefing diário executivo consultivo.", "error": str(exc)})


def _resposta_briefing_agendado(status: str, **campos: Any) -> dict[str, Any]:
    return serializar_json_seguro({
        "ok": status not in {"ERRO"},
        "status": status,
        "deve_enviar_telegram": False,
        "chat_id": None,
        "resposta_telegram": None,
        "envio_id": None,
        **campos,
        **AGENTE_008_SEGURANCA_CONSULTA,
    })


def _configuracao_briefing_diario() -> dict[str, Any]:
    enabled = os.getenv("TELEGRAM_DAILY_BRIEFING_ENABLED", "false").strip().lower() == "true"
    if not enabled:
        return {"enabled": False}
    timezone_name = os.getenv("TELEGRAM_DAILY_BRIEFING_TIMEZONE", "America/Sao_Paulo").strip()
    horario_texto = os.getenv("TELEGRAM_DAILY_BRIEFING_TIME", "07:30").strip()
    try:
        fuso = ZoneInfo(timezone_name)
    except ZoneInfoNotFoundError as exc:
        raise ValueError(f"Timezone inválido: {timezone_name}") from exc
    try:
        horario = datetime.strptime(horario_texto, "%H:%M").time()
    except ValueError as exc:
        raise ValueError("TELEGRAM_DAILY_BRIEFING_TIME deve usar o formato HH:MM.") from exc
    try:
        limite = max(1, min(int(os.getenv("TELEGRAM_DAILY_BRIEFING_LIMITE_ITENS", "10")), 50))
    except ValueError as exc:
        raise ValueError("TELEGRAM_DAILY_BRIEFING_LIMITE_ITENS deve ser inteiro.") from exc
    return {
        "enabled": enabled,
        "agora_local": datetime.now(fuso),
        "horario": horario,
        "obra_codigo": os.getenv("TELEGRAM_DAILY_BRIEFING_OBRA_CODIGO", "OBRA-CAIO").strip(),
        "area": os.getenv("TELEGRAM_DAILY_BRIEFING_AREA", "").strip() or None,
        "limite_itens": limite,
        "chat_id": (TELEGRAM_EXECUTIVE_CHAT_ID or "").strip() or None,
    }


@app.post("/agentes/gestao-operacional/briefing-diario-agendado")
def briefing_diario_agendado(
    payload: GestaoOperacionalBriefingDiarioAgendadoRequest | None = None,
):
    try:
        config = _configuracao_briefing_diario()
        if not config["enabled"]:
            return _resposta_briefing_agendado("DESABILITADO")
        if not config["chat_id"]:
            return _resposta_briefing_agendado("CHAT_EXECUTIVO_NAO_CONFIGURADO")
        if not config["obra_codigo"]:
            return _resposta_briefing_agendado("CONFIGURACAO_INVALIDA")
        agora_local = config["agora_local"]
        forcar = payload.forcar if payload else False
        if agora_local.time().replace(tzinfo=None) < config["horario"] and not forcar:
            return _resposta_briefing_agendado("FORA_DO_HORARIO")

        with get_db_connection() as conn:
            with conn.transaction():
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        SELECT id, status, resposta_telegram,
                               telegram_message_id, enviado_em
                        FROM envios_briefing_diario_obra
                        WHERE obra_codigo = %s AND COALESCE(area, '') = COALESCE(%s, '')
                          AND data_briefing = %s
                          AND tipo_briefing = 'BRIEFING_EXECUTIVO_DIARIO'
                          AND canal = 'TELEGRAM'
                        FOR UPDATE
                        """,
                        (config["obra_codigo"], config["area"], agora_local.date()),
                    )
                    existente = cur.fetchone()
                    if existente and (
                        existente[1] == "CONCLUIDO"
                        or existente[3] is not None
                        or existente[4] is not None
                    ):
                        return _resposta_briefing_agendado(
                            "JA_ENVIADO",
                            envio_id=existente[0],
                            telegram_message_id=existente[3],
                            enviado_em=existente[4],
                        )

                    briefing = _gerar_briefing_diario(
                        cur, config["obra_codigo"], config["area"], config["limite_itens"],
                        True, True, True,
                    )
                    briefing["data_briefing"] = agora_local.date().isoformat()
                    briefing = serializar_json_seguro(briefing)
                    resposta = briefing.get("resposta_telegram")
                    cur.execute(
                        """
                        INSERT INTO envios_briefing_diario_obra (
                            obra_codigo, area, data_briefing, chat_id, status,
                            payload_briefing, resposta_telegram, mensagem_erro, atualizado_em
                        ) VALUES (%s, %s, %s, %s, 'PENDENTE', %s, %s, NULL, now())
                        ON CONFLICT (
                            obra_codigo, (COALESCE(area, '')), data_briefing, tipo_briefing, canal
                        ) DO UPDATE SET
                            chat_id = EXCLUDED.chat_id,
                            status = 'PENDENTE',
                            payload_briefing = EXCLUDED.payload_briefing,
                            resposta_telegram = EXCLUDED.resposta_telegram,
                            mensagem_erro = NULL,
                            atualizado_em = now()
                        RETURNING id
                        """,
                        (
                            config["obra_codigo"], config["area"], agora_local.date(),
                            config["chat_id"], Json(briefing), resposta,
                        ),
                    )
                    envio_id = cur.fetchone()[0]
        return _resposta_briefing_agendado(
            "PENDENTE", deve_enviar_telegram=True, chat_id=config["chat_id"],
            resposta_telegram=resposta, envio_id=envio_id,
        )
    except ValueError as exc:
        return _resposta_briefing_agendado("CONFIGURACAO_INVALIDA", mensagem_erro=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail={
            "message": "Erro ao agendar briefing diário executivo consultivo.", "error": str(exc),
        })


@app.post("/agentes/gestao-operacional/briefing-diario-agendado/confirmar-envio")
def confirmar_envio_briefing_diario(payload: GestaoOperacionalConfirmarBriefingDiarioRequest):
    if payload.status == "CONCLUIDO" and not payload.telegram_message_id:
        raise HTTPException(status_code=422, detail="telegram_message_id é obrigatório para CONCLUIDO.")
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                if payload.status == "CONCLUIDO":
                    cur.execute(
                        """
                        UPDATE envios_briefing_diario_obra
                        SET status = 'CONCLUIDO',
                            telegram_message_id = %s,
                            enviado_em = now(),
                            mensagem_erro = NULL,
                            atualizado_em = now()
                        WHERE id = %s
                        RETURNING id, status, telegram_message_id, enviado_em,
                                  mensagem_erro
                        """,
                        (payload.telegram_message_id, payload.envio_id),
                    )
                else:
                    cur.execute(
                        """
                        UPDATE envios_briefing_diario_obra
                        SET status = 'ERRO',
                            mensagem_erro = %s,
                            atualizado_em = now()
                        WHERE id = %s
                        RETURNING id, status, telegram_message_id, enviado_em,
                                  mensagem_erro
                        """,
                        (payload.mensagem_erro, payload.envio_id),
                    )
                atualizado = cur.fetchone()
        if not atualizado:
            raise HTTPException(status_code=404, detail="Envio de briefing não encontrado.")
        return _resposta_briefing_agendado(
            atualizado[1],
            envio_id=atualizado[0],
            telegram_message_id=atualizado[2],
            enviado_em=atualizado[3],
            mensagem_erro=atualizado[4],
        )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail={
            "message": "Erro ao confirmar envio do briefing diário.", "error": str(exc),
        })


@app.post("/agentes/gestao-operacional/diagnostico-operacional")
def diagnostico_operacional(payload: GestaoOperacionalDiagnosticoRequest):
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                resultado = _gerar_diagnostico_operacional(
                    cur, payload.obra_codigo, payload.area, payload.incluir_documentos,
                    payload.incluir_pendencias, payload.incluir_restricoes,
                )
    except Exception as exc:
        raise HTTPException(status_code=500, detail={
            "message": "Erro ao gerar diagnóstico operacional consultivo.", "error": str(exc)
        })
    return resultado


@app.post("/agentes/gestao-operacional/plano-operacional")
def plano_operacional(payload: GestaoOperacionalPlanoOperacionalRequest):
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                return _gerar_plano_operacional(
                    cur, payload.obra_codigo, payload.area, payload.limite_acoes,
                )
    except Exception as exc:
        raise HTTPException(status_code=500, detail={
            "message": "Erro ao gerar plano operacional consultivo.", "error": str(exc),
        })


@app.post("/agentes/gestao-operacional/resumo-executivo-operacional")
def resumo_executivo_operacional(payload: GestaoOperacionalResumoExecutivoRequest):
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                return _gerar_resumo_executivo_operacional(
                    cur, payload.obra_codigo, payload.area, payload.limite_itens,
                )
    except Exception as exc:
        raise HTTPException(status_code=500, detail={
            "message": "Erro ao gerar resumo executivo operacional consultivo.", "error": str(exc),
        })


def serializar_json_seguro(valor: Any) -> Any:
    if isinstance(valor, (datetime, date)):
        return valor.isoformat()
    if isinstance(valor, Decimal):
        return float(valor)
    if isinstance(valor, dict):
        return {chave: serializar_json_seguro(item) for chave, item in valor.items()}
    if isinstance(valor, (list, tuple)):
        return [serializar_json_seguro(item) for item in valor]
    return valor


def _inferir_metadados_revisao(nome: str, object_key: str | None) -> dict[str, str | None]:
    """Infere somente padrões explícitos; na dúvida mantém o campo vazio."""
    texto = normalizar_texto_comparacao(f"{object_key or ''} {nome}")
    disciplinas = {
        "arquitetura": "ARQUITETURA", "arquitetonico": "ARQUITETURA",
        "estrutura": "ESTRUTURA", "estrutural": "ESTRUTURA",
        "eletrica": "ELETRICA", "hidraulica": "HIDRAULICA",
        "incendio": "INCENDIO", "climatizacao": "CLIMATIZACAO",
    }
    disciplina = next((valor for termo, valor in disciplinas.items() if re.search(rf"\b{termo}\b", texto)), None)
    revisao = None
    match_revisao = re.search(r"(?:^|[^a-z0-9])(?:rev|revisao)[ _.-]*([a-z0-9]{1,8})(?:[^a-z0-9]|$)", f" {nome.casefold()} ")
    if match_revisao:
        revisao = match_revisao.group(1).upper()
    codigo = None
    stem = nome.rsplit(".", 1)[0]
    match_codigo = re.match(r"^([A-Za-z]{1,6}[-_][A-Za-z0-9][A-Za-z0-9_-]{1,30})", stem)
    if match_codigo:
        codigo = match_codigo.group(1).upper().replace("_", "-")
    area = None
    match_area = re.search(r"(?:^|[/_-])(?:area|setor|pavimento|pav)[ _-]*([a-z0-9]+)", normalizar_texto_comparacao(object_key or "").replace(" ", "_"))
    if match_area:
        area = match_area.group(1).upper()
    return {"disciplina": disciplina, "area": area, "revisao_detectada": revisao, "codigo_documento": codigo}


def _registrar_revisao_documental(cur, payload: GestaoOperacionalRegistrarRevisaoDocumentalRequest) -> dict[str, Any]:
    dados = payload.model_dump()
    dados["obra_codigo"] = payload.obra_codigo.strip()
    dados["nome_arquivo_original"] = payload.nome_arquivo_original.strip()
    if not dados["obra_codigo"] or not dados["nome_arquivo_original"]:
        raise ValueError("obra_codigo e nome_arquivo_original são obrigatórios.")
    if payload.documento_minio_id:
        cur.execute("SELECT to_regclass('public.documentos_minio_obra')")
        tabela_documentos_existe = cur.fetchone()[0] is not None
    else:
        tabela_documentos_existe = False
    if payload.documento_minio_id and tabela_documentos_existe:
        cur.execute(
            """SELECT bucket, object_key, nome_arquivo, pasta_origem, disciplina_original,
                      categoria_documental
               FROM documentos_minio_obra
               WHERE id = %(id)s AND obra_codigo = %(obra_codigo)s""",
            {"id": payload.documento_minio_id, "obra_codigo": dados["obra_codigo"]},
        )
        documento = cur.fetchone()
        if documento:
            for campo, valor in zip(
                ("bucket", "object_key", "nome_arquivo_original", "caminho_origem", "disciplina", "tipo_documento"),
                documento,
            ):
                if not dados.get(campo) and valor:
                    dados[campo] = valor
    inferidos = _inferir_metadados_revisao(dados["nome_arquivo_original"], dados.get("object_key"))
    for campo, valor in inferidos.items():
        if not dados.get(campo):
            dados[campo] = valor
    dados["minio_uri"] = (
        f"s3://{dados['bucket']}/{str(dados['object_key']).lstrip('/')}"
        if dados.get("bucket") and dados.get("object_key") else None
    )
    metadados = serializar_json_seguro({"inferencia_conservadora": inferidos})
    cur.execute(
        """INSERT INTO revisoes_documentais_obra (
               obra_codigo, documento_minio_id, bucket, object_key, minio_uri,
               nome_arquivo_original, caminho_origem, disciplina, area,
               codigo_documento, titulo_documento, revisao_detectada, data_documento,
               tipo_documento, motivo_alteracao, responsavel_upload, observacao, metadados,
               status_revisao, status_vigencia, liberado_para_campo
           ) VALUES (
               %(obra_codigo)s, %(documento_minio_id)s, %(bucket)s, %(object_key)s,
               %(minio_uri)s, %(nome_arquivo_original)s, %(caminho_origem)s,
               %(disciplina)s, %(area)s, %(codigo_documento)s, %(titulo_documento)s,
               %(revisao_detectada)s, %(data_documento)s, %(tipo_documento)s,
               %(motivo_alteracao)s, %(responsavel_upload)s, %(observacao)s, %(metadados)s,
               'RECEBIDO_PARA_ANALISE', 'NAO_VIGENTE', FALSE
           ) RETURNING id, criado_em""",
        {**dados, "caminho_origem": dados.get("caminho_origem"), "metadados": Json(metadados)},
    )
    revisao_id, criado_em = cur.fetchone()
    return serializar_json_seguro({
        "ok": True, "mvp": "0.8A", "revisao_documental_id": revisao_id,
        "obra_codigo": dados["obra_codigo"], "status_revisao": "RECEBIDO_PARA_ANALISE",
        "status_vigencia": "NAO_VIGENTE", "liberado_para_campo": False,
        "minio_uri": dados["minio_uri"], "criado_em": criado_em,
        "flags_seguranca": dict(AGENTE_008_SEGURANCA_CONTROLE_DOCUMENTAL),
        **AGENTE_008_SEGURANCA_CONTROLE_DOCUMENTAL,
    })


def _revisao_documental_dict(row: tuple[Any, ...]) -> dict[str, Any]:
    campos = ("id", "obra_codigo", "disciplina", "area", "codigo_documento", "titulo_documento", "revisao_detectada", "nome_arquivo_original", "bucket", "object_key", "minio_uri", "status_revisao", "status_vigencia", "liberado_para_campo", "data_documento", "criado_em", "atualizado_em")
    return serializar_json_seguro(dict(zip(campos, row)))


def _listar_revisoes_documentais(cur, payload: GestaoOperacionalRevisoesDocumentaisRequest) -> dict[str, Any]:
    params = payload.model_dump()
    base = """ FROM revisoes_documentais_obra WHERE obra_codigo = %(obra_codigo)s
        AND (%(area)s IS NULL OR area = %(area)s)
        AND (%(disciplina)s IS NULL OR disciplina = %(disciplina)s)"""
    cur.execute("SELECT status_revisao, count(*)" + base + " GROUP BY status_revisao", params)
    totais_revisao = {status: 0 for status in ("RECEBIDO_PARA_ANALISE", "EM_ANALISE_TECNICA", "APROVADO_COMO_VIGENTE", "REJEITADO", "SUBSTITUIDO", "OBSOLETO", "AS_BUILT")}
    totais_revisao.update(dict(cur.fetchall()))
    cur.execute("SELECT status_vigencia, count(*)" + base + " GROUP BY status_vigencia", params)
    totais_vigencia = {status: 0 for status in ("NAO_VIGENTE", "VIGENTE", "SUBSTITUIDO", "OBSOLETO", "HISTORICO")}
    totais_vigencia.update(dict(cur.fetchall()))
    colunas = "id, obra_codigo, disciplina, area, codigo_documento, titulo_documento, revisao_detectada, nome_arquivo_original, bucket, object_key, minio_uri, status_revisao, status_vigencia, liberado_para_campo, data_documento, criado_em, atualizado_em"
    cur.execute("SELECT " + colunas + base + " AND (%(status_revisao)s IS NULL OR status_revisao = %(status_revisao)s) AND (%(status_vigencia)s IS NULL OR status_vigencia = %(status_vigencia)s) ORDER BY criado_em DESC, id DESC LIMIT %(limite_itens)s", params)
    ultimas = [_revisao_documental_dict(row) for row in cur.fetchall()]
    grupos = {status: [item for item in ultimas if item["status_revisao"] == status] for status in ("RECEBIDO_PARA_ANALISE", "EM_ANALISE_TECNICA", "APROVADO_COMO_VIGENTE", "REJEITADO")}
    linhas = [f"📚 Revisões documentais — {payload.obra_codigo}", "", "Resumo:", f"- Recebidas para análise: {totais_revisao['RECEBIDO_PARA_ANALISE']}", f"- Em análise técnica: {totais_revisao['EM_ANALISE_TECNICA']}", f"- Vigentes: {totais_vigencia['VIGENTE']}", f"- Rejeitadas: {totais_revisao['REJEITADO']}", f"- Obsoletas/substituídas: {totais_revisao['OBSOLETO'] + totais_revisao['SUBSTITUIDO']}", "", "Últimas revisões:"]
    linhas.extend(f"- #{item['id']} [{item['status_revisao']}] {item['disciplina'] or 'Sem disciplina'} / {item['area'] or 'Sem área'} — {item['nome_arquivo_original']}" for item in ultimas)
    if not ultimas:
        linhas.append("- Nenhuma revisão cadastrada.")
    linhas.extend(["", "Governança:", "Consulta documental. Nenhum arquivo foi substituído, nenhum projeto foi liberado para campo, nenhum RDO, cronograma, OpenProject ou RPA foi alterado."])
    return serializar_json_seguro({"ok": True, "mvp": "0.8A", "obra_codigo": payload.obra_codigo, "totais_por_status_revisao": totais_revisao, "totais_por_status_vigencia": totais_vigencia, "revisoes_recebidas_para_analise": grupos["RECEBIDO_PARA_ANALISE"], "revisoes_em_analise": grupos["EM_ANALISE_TECNICA"], "revisoes_aprovadas_como_vigente": grupos["APROVADO_COMO_VIGENTE"], "revisoes_rejeitadas": grupos["REJEITADO"], "ultimas_revisoes_cadastradas": ultimas, "resposta_telegram": "\n".join(linhas), "flags_seguranca": dict(AGENTE_008_SEGURANCA_CONTROLE_DOCUMENTAL), **AGENTE_008_SEGURANCA_CONTROLE_DOCUMENTAL})


def _importar_revisoes_documentais(cur, payload: GestaoOperacionalImportarRevisoesDocumentaisRequest) -> dict[str, Any]:
    cur.execute("SELECT to_regclass('public.documentos_minio_obra')")
    if cur.fetchone()[0] is None:
        return serializar_json_seguro({"ok": True, "mvp": "0.8A", "obra_codigo": payload.obra_codigo, "encontrados": 0, "registrados": 0, "ignorados_por_duplicidade": 0, "revisoes_documentais_ids": [], "resposta_telegram": f"📥 Importação documental — {payload.obra_codigo}\nNenhum índice documentos_minio_obra está disponível; nenhum registro foi criado.", "flags_seguranca": dict(AGENTE_008_SEGURANCA_CONTROLE_DOCUMENTAL), **AGENTE_008_SEGURANCA_CONTROLE_DOCUMENTAL})
    cur.execute("""SELECT id, bucket, object_key, nome_arquivo, pasta_origem, disciplina_original, categoria_documental
                   FROM documentos_minio_obra WHERE obra_codigo = %(obra_codigo)s AND bucket = %(bucket)s
                     AND object_key LIKE %(prefixo_like)s ORDER BY criado_em, id LIMIT %(limite_itens)s""",
                {**payload.model_dump(), "prefixo_like": payload.prefixo + "%"})
    documentos = cur.fetchall()
    registrados = 0
    duplicados = 0
    ids: list[int] = []
    for documento in documentos:
        dados = GestaoOperacionalRegistrarRevisaoDocumentalRequest(obra_codigo=payload.obra_codigo, documento_minio_id=documento[0], bucket=documento[1], object_key=documento[2], nome_arquivo_original=documento[3], disciplina=documento[5], tipo_documento=documento[6] or "PROJETO")
        cur.execute("SELECT id FROM revisoes_documentais_obra WHERE obra_codigo = %s AND bucket = %s AND object_key = %s", (payload.obra_codigo, documento[1], documento[2]))
        existente = cur.fetchone()
        if existente:
            duplicados += 1
            continue
        resultado = _registrar_revisao_documental(cur, dados)
        registrados += 1
        ids.append(resultado["revisao_documental_id"])
    return serializar_json_seguro({"ok": True, "mvp": "0.8A", "obra_codigo": payload.obra_codigo, "encontrados": len(documentos), "registrados": registrados, "ignorados_por_duplicidade": duplicados, "revisoes_documentais_ids": ids, "resposta_telegram": f"📥 Importação documental — {payload.obra_codigo}\nEncontrados: {len(documentos)} | Registrados: {registrados} | Duplicados ignorados: {duplicados}\n\nArquivos mantidos no MinIO e recebidos apenas para análise; nenhuma liberação de campo foi realizada.", "flags_seguranca": dict(AGENTE_008_SEGURANCA_CONTROLE_DOCUMENTAL), **AGENTE_008_SEGURANCA_CONTROLE_DOCUMENTAL})


def _decisor_revisao_documental(payload: GestaoOperacionalAprovarRevisaoDocumentalRequest) -> str:
    return (
        payload.decisor_nome or payload.decisor_telegram_username
        or payload.decisor_telegram_user_id or "DECISOR_NAO_IDENTIFICADO"
    )


def _resposta_decisao_revisao_documental(
    payload: GestaoOperacionalAprovarRevisaoDocumentalRequest,
    documento: str,
    status_revisao: str,
    status_vigencia: str,
    substituidas: int,
) -> str:
    cabecalhos = {
        "APROVAR_COMO_VIGENTE": "✅ Revisão documental aprovada como vigente",
        "REJEITAR": "❌ Revisão documental rejeitada",
        "SOLICITAR_AJUSTES": "🛠️ Ajustes solicitados na revisão documental",
        "MARCAR_EM_ANALISE": "🔎 Revisão documental marcada em análise técnica",
    }
    linhas = [f"{cabecalhos[payload.decisao]} — {payload.obra_codigo}", "", "Revisão:", f"#{payload.revisao_documental_id}"]
    if payload.decisao == "APROVAR_COMO_VIGENTE":
        linhas.extend(["", "Documento:", documento])
    linhas.extend(["", "Status:", f"{status_revisao} / {status_vigencia}"])
    if payload.decisao == "APROVAR_COMO_VIGENTE":
        linhas.extend(["", "Substituições lógicas:", f"{substituidas} revisão(ões) anterior(es) marcada(s) como substituída(s)."])
    elif payload.decisao == "REJEITAR":
        linhas.extend(["", "Motivo:", payload.motivo or "Não informado."])
    elif payload.decisao == "SOLICITAR_AJUSTES":
        linhas.extend(["", "Observação:", payload.observacao or "Não informada."])
    registro = {
        "APROVAR_COMO_VIGENTE": "Aprovação técnica registrada.",
        "REJEITAR": "Rejeição registrada.",
        "SOLICITAR_AJUSTES": "Solicitação de ajustes registrada.",
        "MARCAR_EM_ANALISE": "Análise técnica registrada.",
    }[payload.decisao]
    linhas.extend([
        "", "Governança:", registro,
        "Nenhum arquivo foi apagado ou movido no MinIO.",
        "Nenhum projeto foi liberado automaticamente para campo.",
        "Nenhum RDO, cronograma, OpenProject ou RPA foi alterado.",
    ])
    return "\n".join(linhas)


def _aprovar_revisao_documental(
    cur, payload: GestaoOperacionalAprovarRevisaoDocumentalRequest,
) -> dict[str, Any]:
    cur.execute(
        """SELECT id, obra_codigo, documento_minio_id, documento_substituido_id,
                  revisao_anterior_id, disciplina, area, codigo_documento,
                  revisao_detectada, titulo_documento, nome_arquivo_original,
                  status_revisao, status_vigencia
           FROM revisoes_documentais_obra WHERE id = %s FOR UPDATE""",
        (payload.revisao_documental_id,),
    )
    row = cur.fetchone()
    if not row:
        raise LookupError("Revisão documental não encontrada.")
    campos = ("id", "obra_codigo", "documento_minio_id", "documento_substituido_id", "revisao_anterior_id", "disciplina", "area", "codigo_documento", "revisao_detectada", "titulo_documento", "nome_arquivo_original", "status_revisao", "status_vigencia")
    revisao = dict(zip(campos, row))
    if revisao["obra_codigo"] != payload.obra_codigo:
        raise ValueError("obra_codigo não confere com a revisão documental.")

    mapeamento = {
        "APROVAR_COMO_VIGENTE": ("APROVADO_COMO_VIGENTE", "VIGENTE", True, False, False, "JA_APROVADA_COMO_VIGENTE"),
        "REJEITAR": ("REJEITADO", "NAO_VIGENTE", False, True, False, "JA_REJEITADA"),
        "SOLICITAR_AJUSTES": ("AJUSTES_SOLICITADOS", "NAO_VIGENTE", False, False, True, "JA_COM_AJUSTES_SOLICITADOS"),
        "MARCAR_EM_ANALISE": ("EM_ANALISE_TECNICA", "NAO_VIGENTE", False, False, False, "JA_EM_ANALISE_TECNICA"),
    }
    status_revisao, status_vigencia, aprovado, rejeitado, ajustes, status_idempotente = mapeamento[payload.decisao]
    documento = revisao["codigo_documento"] or revisao["titulo_documento"] or revisao["nome_arquivo_original"]
    if revisao["status_revisao"] == status_revisao and revisao["status_vigencia"] == status_vigencia:
        return serializar_json_seguro({
            "ok": True, "mvp": "0.8B", "status": status_idempotente,
            "obra_codigo": payload.obra_codigo, "revisao_documental_id": payload.revisao_documental_id,
            "status_revisao_resultante": status_revisao, "status_vigencia_resultante": status_vigencia,
            "revisoes_anteriores_substituidas": 0,
            "resposta_telegram": _resposta_decisao_revisao_documental(payload, documento, status_revisao, status_vigencia, 0),
            "flags_seguranca": dict(AGENTE_008_SEGURANCA_APROVACAO_DOCUMENTAL),
            **AGENTE_008_SEGURANCA_APROVACAO_DOCUMENTAL,
        })

    decisor = _decisor_revisao_documental(payload)
    campos_especificos = {
        "APROVAR_COMO_VIGENTE": ", aprovado_por = %(decisor)s, aprovado_em = NOW()",
        "REJEITAR": ", rejeitado_por = %(decisor)s, rejeitado_em = NOW(), motivo_rejeicao = %(motivo)s",
        "SOLICITAR_AJUSTES": "",
        "MARCAR_EM_ANALISE": ", responsavel_analise = %(decisor)s",
    }[payload.decisao]
    cur.execute(
        f"""UPDATE revisoes_documentais_obra
            SET status_revisao = %(status_revisao)s, status_vigencia = %(status_vigencia)s,
                decisao_tecnica = %(decisao)s, decisao_tecnica_por = %(decisor)s,
                decisao_tecnica_em = NOW(), motivo_decisao = %(motivo)s,
                observacao_decisao = %(observacao)s, liberado_para_campo = FALSE,
                liberou_execucao_campo = FALSE, atualizado_em = NOW()
                {campos_especificos}
            WHERE id = %(id)s""",
        {"id": payload.revisao_documental_id, "status_revisao": status_revisao,
         "status_vigencia": status_vigencia, "decisao": payload.decisao,
         "decisor": decisor, "motivo": payload.motivo, "observacao": payload.observacao},
    )

    substituidas = 0
    avisos: list[str] = []
    if payload.decisao == "APROVAR_COMO_VIGENTE":
        params = {"id": payload.revisao_documental_id, "obra_codigo": payload.obra_codigo,
                  "codigo": revisao["codigo_documento"], "disciplina": revisao["disciplina"],
                  "area": revisao["area"], "titulo": revisao["titulo_documento"]}
        if revisao["codigo_documento"] and revisao["codigo_documento"].strip():
            compatibilidade = "codigo_documento = %(codigo)s AND (%(disciplina)s IS NULL OR disciplina = %(disciplina)s) AND (%(area)s IS NULL OR area = %(area)s)"
        elif all(isinstance(revisao[c], str) and revisao[c].strip() for c in ("disciplina", "area", "titulo_documento")):
            compatibilidade = "disciplina = %(disciplina)s AND area = %(area)s AND LOWER(TRIM(titulo_documento)) = LOWER(TRIM(%(titulo)s))"
        else:
            compatibilidade = "FALSE"
            avisos.append("substituicao_anterior_nao_aplicada_por_baixa_confianca")
        cur.execute(
            f"""UPDATE revisoes_documentais_obra
                SET status_vigencia = 'SUBSTITUIDO', status_revisao = 'SUBSTITUIDO',
                    substituida_por_revisao_id = %(id)s, substituida_em = NOW(),
                    liberado_para_campo = FALSE, liberou_execucao_campo = FALSE,
                    atualizado_em = NOW()
                WHERE obra_codigo = %(obra_codigo)s AND id <> %(id)s
                  AND status_vigencia = 'VIGENTE' AND {compatibilidade}""",
            params,
        )
        substituidas = cur.rowcount

    metadados = serializar_json_seguro({
        "flags_seguranca": dict(AGENTE_008_SEGURANCA_APROVACAO_DOCUMENTAL),
        **AGENTE_008_SEGURANCA_APROVACAO_DOCUMENTAL,
        "avisos": avisos,
    })
    cur.execute(
        """INSERT INTO aprovacoes_revisoes_documentais_obra (
               obra_codigo, revisao_documental_id, documento_minio_id, documento_substituido_id,
               revisao_anterior_id, disciplina, area, codigo_documento, revisao_detectada,
               decisao, status_revisao_resultante, status_vigencia_resultante,
               aprovado, rejeitado, ajustes_solicitados, motivo, observacao, decisor_nome,
               decisor_telegram_user_id, decisor_telegram_username, decisor_chat_id, metadados
           ) VALUES (
               %(obra_codigo)s, %(id)s, %(documento_minio_id)s, %(documento_substituido_id)s,
               %(revisao_anterior_id)s, %(disciplina)s, %(area)s, %(codigo_documento)s,
               %(revisao_detectada)s, %(decisao)s, %(status_revisao)s, %(status_vigencia)s,
               %(aprovado)s, %(rejeitado)s, %(ajustes)s, %(motivo)s, %(observacao)s,
               %(decisor_nome)s, %(telegram_user_id)s, %(telegram_username)s, %(chat_id)s, %(metadados)s)
           RETURNING id, criado_em""",
        {**revisao, "obra_codigo": payload.obra_codigo, "decisao": payload.decisao,
         "status_revisao": status_revisao, "status_vigencia": status_vigencia,
         "aprovado": aprovado, "rejeitado": rejeitado, "ajustes": ajustes,
         "motivo": payload.motivo, "observacao": payload.observacao,
         "decisor_nome": payload.decisor_nome, "telegram_user_id": payload.decisor_telegram_user_id,
         "telegram_username": payload.decisor_telegram_username, "chat_id": payload.decisor_chat_id,
         "metadados": Json(metadados)},
    )
    aprovacao_id, criado_em = cur.fetchone()
    return serializar_json_seguro({
        "ok": True, "mvp": "0.8B", "status": "DECISAO_TECNICA_REGISTRADA",
        "aprovacao_revisao_documental_id": aprovacao_id, "criado_em": criado_em,
        "obra_codigo": payload.obra_codigo, "revisao_documental_id": payload.revisao_documental_id,
        "decisao": payload.decisao, "status_revisao_resultante": status_revisao,
        "status_vigencia_resultante": status_vigencia, "aprovado": aprovado,
        "rejeitado": rejeitado, "ajustes_solicitados": ajustes,
        "revisoes_anteriores_substituidas": substituidas, "avisos": avisos,
        "resposta_telegram": _resposta_decisao_revisao_documental(payload, documento, status_revisao, status_vigencia, substituidas),
        "flags_seguranca": dict(AGENTE_008_SEGURANCA_APROVACAO_DOCUMENTAL),
        **AGENTE_008_SEGURANCA_APROVACAO_DOCUMENTAL,
    })


def _decisor_liberacao_campo(payload: GestaoOperacionalLiberarRevisaoDocumentalCampoRequest) -> str:
    return payload.decisor_nome or payload.decisor_telegram_username or payload.decisor_telegram_user_id or "DECISOR_NAO_IDENTIFICADO"


def _resposta_liberacao_campo(payload: GestaoOperacionalLiberarRevisaoDocumentalCampoRequest, documento: str, status: str) -> str:
    if payload.decisao == "LIBERAR_PARA_CAMPO":
        linhas = [f"✅ Revisão documental liberada para uso em campo — {payload.obra_codigo}", "", "Revisão:", f"#{payload.revisao_documental_id}", "", "Documento:", documento, "", "Status:", status, "", "Instruções:", payload.instrucoes_campo or "Não informadas.", "", "Governança:", "Liberação documental registrada.", "Nenhum arquivo foi apagado ou movido no MinIO.", "Nenhum link público foi criado.", "Nenhuma ordem de serviço foi criada.", "Nenhum RDO, cronograma, OpenProject ou RPA foi alterado."]
    else:
        revogada = payload.decisao == "REVOGAR_LIBERACAO_CAMPO"
        titulo = "🚫 Liberação de campo revogada" if revogada else "⏸️ Liberação de campo suspensa"
        registro = "Revogação registrada." if revogada else "Suspensão registrada."
        linhas = [f"{titulo} — {payload.obra_codigo}", "", "Revisão:", f"#{payload.revisao_documental_id}", "", "Status:", status, "", "Motivo:", payload.motivo or "Não informado.", "", "Governança:", registro, "Nenhum arquivo foi apagado ou movido.", "Nenhuma ordem de serviço foi criada.", "Nenhum RDO, cronograma, OpenProject ou RPA foi alterado."]
    return "\n".join(linhas)


def _liberar_revisao_documental_campo(cur, payload: GestaoOperacionalLiberarRevisaoDocumentalCampoRequest) -> dict[str, Any]:
    cur.execute("""SELECT id, obra_codigo, documento_minio_id, disciplina, area, codigo_documento,
                          revisao_detectada, titulo_documento, nome_arquivo_original,
                          status_revisao, status_vigencia, liberado_para_campo, liberacao_campo_status
                   FROM revisoes_documentais_obra WHERE id = %s FOR UPDATE""", (payload.revisao_documental_id,))
    row = cur.fetchone()
    if not row:
        raise LookupError("Revisão documental não encontrada.")
    campos = ("id", "obra_codigo", "documento_minio_id", "disciplina", "area", "codigo_documento", "revisao_detectada", "titulo_documento", "nome_arquivo_original", "status_revisao", "status_vigencia", "liberado_para_campo", "liberacao_campo_status")
    revisao = dict(zip(campos, row))
    if revisao["obra_codigo"] != payload.obra_codigo:
        raise ValueError("obra_codigo não confere com a revisão documental.")
    mapa = {
        "LIBERAR_PARA_CAMPO": ("LIBERADO_PARA_USO_DOCUMENTAL_EM_CAMPO", "JA_LIBERADA_PARA_CAMPO"),
        "SUSPENDER_LIBERACAO_CAMPO": ("SUSPENSA", "JA_SUSPENSA"),
        "REVOGAR_LIBERACAO_CAMPO": ("REVOGADA", "JA_REVOGADA"),
    }
    status, status_idempotente = mapa[payload.decisao]
    documento = revisao["codigo_documento"] or revisao["titulo_documento"] or revisao["nome_arquivo_original"]
    if (payload.decisao == "LIBERAR_PARA_CAMPO" and revisao["liberado_para_campo"]) or revisao["liberacao_campo_status"] == status:
        cur.execute("""SELECT id FROM liberacoes_revisoes_documentais_campo_obra
                       WHERE revisao_documental_id = %s AND status_liberacao_resultante = %s
                       ORDER BY criado_em DESC, id DESC LIMIT 1""", (revisao["id"], status))
        liberacao_existente = cur.fetchone()
        return serializar_json_seguro({"ok": True, "mvp": "0.8C", "status": status_idempotente, "liberacao_id": liberacao_existente[0] if liberacao_existente else None, "obra_codigo": payload.obra_codigo, "revisao_documental_id": payload.revisao_documental_id, "status_liberacao_resultante": status, "resposta_telegram": _resposta_liberacao_campo(payload, documento, status), "flags_seguranca": dict(AGENTE_008_SEGURANCA_LIBERACAO_CAMPO), **AGENTE_008_SEGURANCA_LIBERACAO_CAMPO})
    if payload.decisao == "LIBERAR_PARA_CAMPO" and (revisao["status_revisao"] != "APROVADO_COMO_VIGENTE" or revisao["status_vigencia"] != "VIGENTE"):
        raise ValueError("REVISAO_DOCUMENTAL_NAO_VIGENTE")
    decisor = _decisor_liberacao_campo(payload)
    if payload.decisao == "LIBERAR_PARA_CAMPO":
        extras = "liberado_para_campo_por = %(decisor)s, liberado_para_campo_em = NOW(), liberacao_campo_instrucoes = %(instrucoes)s,"
    elif payload.decisao == "REVOGAR_LIBERACAO_CAMPO":
        extras = "liberacao_campo_revogada_por = %(decisor)s, liberacao_campo_revogada_em = NOW(), motivo_revogacao_campo = %(motivo)s,"
    else:
        extras = "motivo_revogacao_campo = %(motivo)s,"
    cur.execute(f"""UPDATE revisoes_documentais_obra SET liberado_para_campo = %(liberado)s,
                        liberacao_campo_status = %(status)s, liberacao_campo_observacao = %(observacao)s,
                        {extras} liberacao_campo_decisao_em = NOW(), liberou_execucao_campo = FALSE,
                        atualizado_em = NOW() WHERE id = %(id)s""",
                {"id": revisao["id"], "liberado": payload.decisao == "LIBERAR_PARA_CAMPO", "status": status, "observacao": payload.observacao, "instrucoes": payload.instrucoes_campo, "motivo": payload.motivo, "decisor": decisor})
    cur.execute("""SELECT id FROM aprovacoes_revisoes_documentais_obra
                   WHERE revisao_documental_id = %s ORDER BY criado_em DESC, id DESC LIMIT 1""", (revisao["id"],))
    aprovacao = cur.fetchone()
    metadados = serializar_json_seguro({"flags_seguranca": dict(AGENTE_008_SEGURANCA_LIBERACAO_CAMPO), **AGENTE_008_SEGURANCA_LIBERACAO_CAMPO})
    cur.execute("""INSERT INTO liberacoes_revisoes_documentais_campo_obra (
                       obra_codigo, revisao_documental_id, aprovacao_revisao_id, documento_minio_id,
                       disciplina, area, codigo_documento, revisao_detectada, decisao,
                       status_liberacao_resultante, liberado_para_campo, liberacao_suspensa,
                       liberacao_revogada, motivo, observacao, instrucoes_campo, decisor_nome,
                       decisor_telegram_user_id, decisor_telegram_username, decisor_chat_id, metadados)
                   VALUES (%(obra)s, %(revisao_id)s, %(aprovacao_id)s, %(documento_id)s, %(disciplina)s,
                       %(area)s, %(codigo)s, %(revisao)s, %(decisao)s, %(status)s, %(liberado)s,
                       %(suspensa)s, %(revogada)s, %(motivo)s, %(observacao)s, %(instrucoes)s,
                       %(decisor_nome)s, %(user_id)s, %(username)s, %(chat_id)s, %(metadados)s)
                   RETURNING id, criado_em""",
                {"obra": payload.obra_codigo, "revisao_id": revisao["id"], "aprovacao_id": aprovacao[0] if aprovacao else None, "documento_id": revisao["documento_minio_id"], "disciplina": revisao["disciplina"], "area": revisao["area"], "codigo": revisao["codigo_documento"], "revisao": revisao["revisao_detectada"], "decisao": payload.decisao, "status": status, "liberado": payload.decisao == "LIBERAR_PARA_CAMPO", "suspensa": payload.decisao == "SUSPENDER_LIBERACAO_CAMPO", "revogada": payload.decisao == "REVOGAR_LIBERACAO_CAMPO", "motivo": payload.motivo, "observacao": payload.observacao, "instrucoes": payload.instrucoes_campo, "decisor_nome": payload.decisor_nome, "user_id": payload.decisor_telegram_user_id, "username": payload.decisor_telegram_username, "chat_id": payload.decisor_chat_id, "metadados": Json(metadados)})
    liberacao_id, criado_em = cur.fetchone()
    return serializar_json_seguro({"ok": True, "mvp": "0.8C", "status": "DECISAO_LIBERACAO_CAMPO_REGISTRADA", "liberacao_id": liberacao_id, "criado_em": criado_em, "obra_codigo": payload.obra_codigo, "revisao_documental_id": revisao["id"], "decisao": payload.decisao, "status_liberacao_resultante": status, "liberado_para_campo": payload.decisao == "LIBERAR_PARA_CAMPO", "resposta_telegram": _resposta_liberacao_campo(payload, documento, status), "flags_seguranca": dict(AGENTE_008_SEGURANCA_LIBERACAO_CAMPO), **AGENTE_008_SEGURANCA_LIBERACAO_CAMPO})


def _listar_liberacoes_revisoes_campo(cur, payload: GestaoOperacionalLiberacoesRevisoesDocumentaisCampoRequest) -> dict[str, Any]:
    params = payload.model_dump()
    base = """ FROM liberacoes_revisoes_documentais_campo_obra WHERE obra_codigo = %(obra_codigo)s
        AND (%(area)s IS NULL OR area = %(area)s) AND (%(disciplina)s IS NULL OR disciplina = %(disciplina)s)"""
    cur.execute("SELECT status_liberacao_resultante, count(*)" + base + " GROUP BY status_liberacao_resultante", params)
    totais = {"LIBERADO_PARA_USO_DOCUMENTAL_EM_CAMPO": 0, "SUSPENSA": 0, "REVOGADA": 0}
    totais.update(dict(cur.fetchall()))
    colunas = "id, revisao_documental_id, documento_minio_id, disciplina, area, codigo_documento, revisao_detectada, decisao, status_liberacao_resultante, motivo, observacao, instrucoes_campo, criado_em"
    cur.execute("SELECT " + colunas + base + " AND (%(status_liberacao)s IS NULL OR status_liberacao_resultante = %(status_liberacao)s) ORDER BY criado_em DESC, id DESC LIMIT %(limite_itens)s", params)
    campos = tuple(item.strip() for item in colunas.split(","))
    ultimas = [serializar_json_seguro(dict(zip(campos, row))) for row in cur.fetchall()]
    cur.execute("""SELECT id, disciplina, area, codigo_documento, titulo_documento,
                          nome_arquivo_original, revisao_detectada, liberacao_campo_instrucoes,
                          liberado_para_campo_em
                   FROM revisoes_documentais_obra WHERE obra_codigo = %(obra_codigo)s
                     AND liberado_para_campo = TRUE
                     AND (%(area)s IS NULL OR area = %(area)s)
                     AND (%(disciplina)s IS NULL OR disciplina = %(disciplina)s)
                   ORDER BY liberado_para_campo_em DESC NULLS LAST, id DESC LIMIT %(limite_itens)s""", params)
    campos_atuais = ("id", "disciplina", "area", "codigo_documento", "titulo_documento", "nome_arquivo_original", "revisao_detectada", "instrucoes_campo", "liberado_para_campo_em")
    atuais = [serializar_json_seguro(dict(zip(campos_atuais, row))) for row in cur.fetchall()]
    linhas = [f"📋 Liberações documentais de campo — {payload.obra_codigo}", "", f"Liberadas: {totais['LIBERADO_PARA_USO_DOCUMENTAL_EM_CAMPO']}", f"Suspensas: {totais['SUSPENSA']}", f"Revogadas: {totais['REVOGADA']}", "", "Atualmente liberadas:"]
    linhas.extend(f"- #{item['id']} {item['codigo_documento'] or item['titulo_documento'] or item['nome_arquivo_original']}" for item in atuais)
    if not atuais:
        linhas.append("- Nenhuma revisão liberada.")
    linhas.extend(["", "Governança:", "Consulta documental; nenhuma ordem de serviço ou execução automática foi autorizada."])
    return serializar_json_seguro({"ok": True, "mvp": "0.8C", "obra_codigo": payload.obra_codigo, "total_liberadas_para_campo": totais["LIBERADO_PARA_USO_DOCUMENTAL_EM_CAMPO"], "total_suspensas": totais["SUSPENSA"], "total_revogadas": totais["REVOGADA"], "ultimas_liberacoes": ultimas, "revisoes_atualmente_liberadas": atuais, "resposta_telegram": "\n".join(linhas), "flags_seguranca": dict(AGENTE_008_SEGURANCA_LIBERACAO_CAMPO), **AGENTE_008_SEGURANCA_LIBERACAO_CAMPO})


def _periodo_relatorio_semanal(
    data_inicio: date | None, data_fim: date | None,
) -> tuple[date, date]:
    if (data_inicio is None) != (data_fim is None):
        raise ValueError("data_inicio e data_fim devem ser informadas juntas.")
    if data_inicio is not None and data_fim is not None:
        if data_inicio > data_fim:
            raise ValueError("data_inicio não pode ser posterior a data_fim.")
        return data_inicio, data_fim

    hoje_local = datetime.now(ZoneInfo("America/Sao_Paulo")).date()
    inicio = hoje_local - timedelta(days=hoje_local.weekday())
    return inicio, inicio + timedelta(days=6)


def _data_operacional(valor: Any) -> date | None:
    if isinstance(valor, datetime):
        if valor.tzinfo is not None:
            valor = valor.astimezone(ZoneInfo("America/Sao_Paulo"))
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str) and valor.strip():
        try:
            return date.fromisoformat(valor.strip()[:10])
        except ValueError:
            return None
    return None


def _gerar_relatorio_semanal_executivo(
    cur, obra_codigo: str, area: Optional[str], data_inicio: date,
    data_fim: date, limite_itens: int, salvar_relatorio: bool,
) -> dict[str, Any]:
    fontes_indisponiveis: list[str] = []

    def ler(tabela: str, colunas: list[str]) -> list[dict[str, Any]]:
        linhas, erro = _linhas_fonte_operacional(cur, tabela, colunas, obra_codigo)
        if erro:
            fontes_indisponiveis.append(f"{tabela}: {_texto_curto(erro, 90)}")
        return linhas

    acoes = ler("acoes_operacionais_obra", [
        "id", "area", "codigo_area", "status", "prioridade", "titulo", "descricao",
        "responsavel", "prazo", "criado_em", "atualizado_em", "concluido_em",
    ])
    historico = ler("historico_acoes_operacionais_obra", [
        "id", "acao_id", "tipo_evento", "status_anterior", "status_novo",
        "prioridade_nova", "responsavel_novo", "prazo_novo", "observacao", "criado_em",
    ])
    classificacoes = ler("classificacoes_documentais_obra", [
        "documento_id", "area_detectada", "codigo_area", "area", "status_revisao",
        "numero_revisao", "eh_obsoleto", "eh_as_built", "confianca_classificacao",
        "risco_documental", "nivel_risco", "criado_em", "atualizado_em", "updated_at",
    ])
    documentos = ler("documentos_minio_obra", [
        "id", "nome_arquivo", "area", "codigo_area", "criado_em", "atualizado_em",
        "ultima_modificacao",
    ])
    briefings = ler("envios_briefing_diario_obra", [
        "id", "area", "data_briefing", "status", "criado_em", "enviado_em",
    ])
    comandos = ler("comandos_executivos", [
        "id", "tipo_comando", "status", "payload_comando", "resultado", "justificativa",
        "criado_em", "atualizado_em",
    ])
    pendencias = ler("pendencias_obra", [
        "id", "area", "codigo_area", "area_detectada", "status", "status_pendencia",
        "titulo", "descricao", "criticidade", "responsavel", "prazo", "data_prazo",
        "criado_em", "atualizado_em",
    ])
    restricoes = ler("restricoes_atividade", [
        "id", "atividade_id", "area", "codigo_area", "status", "status_restricao",
        "titulo", "descricao", "criticidade", "responsavel", "prazo", "criado_em",
        "atualizado_em",
    ])
    areas = ler("areas_obra", ["codigo_area", "nome_area", "ativo"])
    atividades = ler("atividades_cronograma", [
        "id", "codigo_area", "area", "frente_servico", "descricao", "status_atividade",
    ])

    filtro_area = normalizar_texto_comparacao(area).replace(" ", "_") if area else None
    atividade_area = {
        item.get("id"): item.get("codigo_area") or item.get("area")
        for item in atividades
    }
    documento_area = {
        item.get("id"): item.get("codigo_area") or item.get("area")
        for item in documentos
    }

    def normalizar_area(valor: Any) -> str | None:
        if valor is None:
            return None
        return normalizar_texto_comparacao(str(valor)).replace(" ", "_")

    def area_item(
        item: dict[str, Any], usar_atividade: bool = False,
        usar_documento: bool = False,
    ) -> str | None:
        valor = item.get("codigo_area") or item.get("area") or item.get("area_detectada")
        if not valor and usar_atividade:
            valor = atividade_area.get(item.get("atividade_id"))
        if not valor and usar_documento:
            valor = documento_area.get(item.get("documento_id"))
        return normalizar_area(valor)

    def area_comando(item: dict[str, Any]) -> str | None:
        payload = item.get("payload_comando")
        return normalizar_area(payload.get("area")) if isinstance(payload, dict) else None

    if filtro_area:
        acoes = [item for item in acoes if area_item(item) == filtro_area]
        ids_acoes = {item.get("id") for item in acoes}
        historico = [item for item in historico if item.get("acao_id") in ids_acoes]
        classificacoes = [
            item for item in classificacoes if area_item(item, usar_documento=True) == filtro_area
        ]
        documentos = [item for item in documentos if area_item(item) == filtro_area]
        briefings = [item for item in briefings if area_item(item) == filtro_area]
        comandos = [item for item in comandos if area_comando(item) in {None, filtro_area}]
        pendencias = [item for item in pendencias if area_item(item) == filtro_area]
        restricoes = [item for item in restricoes if area_item(item, usar_atividade=True) == filtro_area]

    def no_periodo(valor: Any) -> bool:
        data_valor = _data_operacional(valor)
        return data_valor is not None and data_inicio <= data_valor <= data_fim

    def status(item: dict[str, Any], *campos: str) -> str:
        valor = next((item.get(campo) for campo in campos if item.get(campo) is not None), "")
        return normalizar_texto_comparacao(str(valor)).upper().replace(" ", "_")

    def verdadeiro(valor: Any) -> bool:
        return valor is True or str(valor).lower() in {"true", "t", "1", "sim"}

    def baixa_confianca(item: dict[str, Any]) -> bool:
        try:
            return (
                item.get("confianca_classificacao") is not None
                and float(item["confianca_classificacao"]) < 0.70
            )
        except (TypeError, ValueError):
            return False

    status_encerrados = {
        "CONCLUIDO", "CONCLUIDA", "CANCELADO", "CANCELADA", "RESOLVIDO", "RESOLVIDA",
        "FECHADO", "FECHADA", "ENCERRADO", "ENCERRADA", "REJEITADO",
        "LIBERADO", "LIBERADA",
    }
    abertas = [item for item in acoes if status(item, "status") == "ABERTA"]
    em_andamento = [item for item in acoes if status(item, "status") == "EM_ANDAMENTO"]
    nao_encerradas = abertas + em_andamento
    altas_criticas = [
        item for item in nao_encerradas
        if status(item, "prioridade") in {"ALTA", "CRITICA"}
    ]
    sem_responsavel = [item for item in nao_encerradas if not item.get("responsavel")]
    sem_prazo = [item for item in nao_encerradas if not item.get("prazo")]
    hoje_local = datetime.now(ZoneInfo("America/Sao_Paulo")).date()
    atrasadas = [
        item for item in nao_encerradas
        if _data_operacional(item.get("prazo")) is not None
        and _data_operacional(item.get("prazo")) < hoje_local
    ]
    criadas_periodo = [item for item in acoes if no_periodo(item.get("criado_em"))]
    ids_concluidos_historico = {
        item.get("acao_id") for item in historico
        if status(item, "status_novo") in {"CONCLUIDA", "CONCLUIDO"}
        and no_periodo(item.get("criado_em"))
    }
    ids_cancelados_historico = {
        item.get("acao_id") for item in historico
        if status(item, "status_novo") in {"CANCELADA", "CANCELADO"}
        and no_periodo(item.get("criado_em"))
    }
    concluidas_periodo = [
        item for item in acoes
        if item.get("id") in ids_concluidos_historico
        or (status(item, "status") in {"CONCLUIDA", "CONCLUIDO"} and no_periodo(item.get("concluido_em")))
    ]
    canceladas_periodo = [
        item for item in acoes
        if item.get("id") in ids_cancelados_historico
        or (status(item, "status") in {"CANCELADA", "CANCELADO"} and no_periodo(item.get("atualizado_em")))
    ]

    obsoletos = [
        item for item in classificacoes
        if verdadeiro(item.get("eh_obsoleto")) or status(item, "status_revisao") == "OBSOLETO"
    ]
    sem_revisao_docs = [
        item for item in classificacoes
        if not (item.get("numero_revisao") or item.get("status_revisao"))
        or status(item, "status_revisao") in {"NAO_IDENTIFICADO", "SEM_REVISAO"}
    ]
    baixa_confianca_docs = [item for item in classificacoes if baixa_confianca(item)]
    classificados_periodo = [
        item for item in classificacoes
        if no_periodo(item.get("criado_em") or item.get("atualizado_em") or item.get("updated_at"))
    ]
    alteracoes_documentais = [
        item for item in classificacoes
        if no_periodo(item.get("atualizado_em") or item.get("updated_at"))
    ]
    as_built = [item for item in classificacoes if verdadeiro(item.get("eh_as_built"))]

    briefings_periodo = [
        item for item in briefings
        if no_periodo(item.get("data_briefing") or item.get("enviado_em") or item.get("criado_em"))
        and status(item, "status") in {"CONCLUIDO", "ENVIADO"}
    ]
    envios_datas = [
        item.get("enviado_em") or item.get("criado_em") for item in briefings_periodo
        if item.get("enviado_em") or item.get("criado_em")
    ]
    ultimo_briefing = max(envios_datas, key=lambda valor: str(valor)) if envios_datas else None

    pendencias_abertas = [
        item for item in pendencias
        if status(item, "status_pendencia", "status") not in status_encerrados
    ]
    restricoes_abertas = [
        item for item in restricoes
        if status(item, "status_restricao", "status") not in status_encerrados
    ]
    comandos_pendentes = [
        item for item in comandos
        if status(item, "status") in {"PENDENTE", "AGUARDANDO_APROVACAO", "APROVADO"}
        and item.get("tipo_comando") not in {
            "GERAR_RELATORIO_SEMANAL_EXECUTIVO_OBRA",
            "EXPORTAR_RELATORIO_SEMANAL_EXECUTIVO_OBRA",
            "GERAR_PDF_RELATORIO_SEMANAL_EXECUTIVO_OBRA",
        }
    ]

    muitos_incertos = (
        len(sem_revisao_docs) + len(baixa_confianca_docs)
        >= max(5, (len(classificacoes) + 4) // 5)
    )
    if altas_criticas or obsoletos:
        status_executivo = "BLOQUEIO_POTENCIAL"
    elif muitos_incertos:
        status_executivo = "REQUER_VALIDACAO_TECNICA"
    elif nao_encerradas or pendencias_abertas or restricoes_abertas or fontes_indisponiveis:
        status_executivo = "REQUER_ATENCAO"
    else:
        status_executivo = "SEM_ALERTAS_RELEVANTES"

    conclusao_por_acao = {
        item.get("acao_id"): item.get("criado_em")
        for item in historico
        if status(item, "status_novo") in {"CONCLUIDA", "CONCLUIDO"}
        and item.get("acao_id") is not None
    }

    def resumir_acao(item: dict[str, Any]) -> dict[str, Any]:
        return {
            "id": item.get("id"),
            "prioridade": status(item, "prioridade") or "NAO_INFORMADA",
            "titulo": _texto_curto(item.get("titulo") or item.get("descricao") or "Ação sem título", 100),
            "status": status(item, "status") or "NAO_INFORMADO",
            "responsavel": _texto_curto(item.get("responsavel"), 60) or None,
            "prazo": item.get("prazo"),
            "concluido_em": item.get("concluido_em") or conclusao_por_acao.get(item.get("id")),
        }

    def resumir_documento(item: dict[str, Any]) -> dict[str, Any]:
        documento = next(
            (doc for doc in documentos if doc.get("id") == item.get("documento_id")), {}
        )
        return {
            "documento_id": item.get("documento_id"),
            "nome_arquivo": _texto_curto(documento.get("nome_arquivo") or "Documento classificado", 100),
            "status_revisao": status(item, "status_revisao") or "NAO_IDENTIFICADO",
            "obsoleto": item in obsoletos,
            "baixa_confianca": item in baixa_confianca_docs,
        }

    decisoes: list[str] = []
    for item in pendencias_abertas:
        decisoes.append(_texto_curto(item.get("titulo") or item.get("descricao") or f"Pendência #{item.get('id')}", 120))
    for item in restricoes_abertas:
        decisoes.append(_texto_curto(item.get("titulo") or item.get("descricao") or f"Restrição #{item.get('id')}", 120))
    for item in comandos_pendentes:
        decisoes.append(_texto_curto(item.get("justificativa") or f"Comando {item.get('tipo_comando')} pendente", 120))
    decisoes = list(dict.fromkeys(decisoes))[:5]

    resumo_semana: list[str] = []
    if criadas_periodo:
        resumo_semana.append(f"{len(criadas_periodo)} ação(ões) operacional(is) criada(s) no período.")
    if concluidas_periodo:
        resumo_semana.append(f"{len(concluidas_periodo)} ação(ões) concluída(s) no período.")
    if altas_criticas:
        resumo_semana.append(f"{len(altas_criticas)} ação(ões) alta(s) ou crítica(s) permanece(m) aberta(s).")
    if obsoletos or sem_revisao_docs or baixa_confianca_docs:
        resumo_semana.append(
            f"Documentação requer atenção: {len(obsoletos)} obsoleto(s), "
            f"{len(sem_revisao_docs)} sem revisão e {len(baixa_confianca_docs)} com baixa confiança."
        )
    if briefings_periodo:
        resumo_semana.append(f"{len(briefings_periodo)} briefing(s) diário(s) enviado(s) no período.")
    if fontes_indisponiveis and len(resumo_semana) < 5:
        resumo_semana.append("Relatório parcial: uma ou mais fontes operacionais não estavam disponíveis.")
    if not resumo_semana:
        resumo_semana.append("Não foram identificados eventos relevantes nas fontes disponíveis para o período.")
    resumo_semana = [_texto_curto(item, 180) for item in resumo_semana[:5]]

    recomendacoes: list[str] = []
    if altas_criticas:
        recomendacoes.append("Priorizar ações altas e críticas, confirmando responsável, prazo e evidência de tratamento.")
    if sem_responsavel:
        recomendacoes.append("Definir responsáveis para as ações abertas sem atribuição.")
    if sem_prazo or atrasadas:
        recomendacoes.append("Revisar prazos ausentes ou possivelmente atrasados com a equipe responsável.")
    if obsoletos:
        recomendacoes.append("Validar tecnicamente e segregar documentos obsoletos antes de qualquer uso em campo.")
    if sem_revisao_docs or baixa_confianca_docs:
        recomendacoes.append("Revisar documentos sem revisão identificada ou com baixa confiança de classificação.")
    if not recomendacoes:
        recomendacoes.append("Manter acompanhamento consultivo; qualquer liberação depende de validação técnica responsável.")
    recomendacoes = [_texto_curto(item, 150) for item in recomendacoes[:5]]

    nomes_area = {
        area_item(item): str(item.get("nome_area") or item.get("codigo_area"))
        for item in areas if area_item(item) and item.get("ativo") is not False
    }
    area_exibicao = nomes_area.get(filtro_area, area) if filtro_area else "todas"
    acoes_telegram = sorted(
        nao_encerradas,
        key=lambda item: ({"CRITICA": 0, "ALTA": 1, "MEDIA": 2, "BAIXA": 3}.get(status(item, "prioridade"), 4), item.get("id") or 0),
    )[:7]
    linhas = [
        f"📆 Relatório semanal executivo — {obra_codigo}", "", "Período:",
        f"{data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}", "",
        "Filtro:", f"Área: {area_exibicao}", "", "Status executivo:", status_executivo,
        "", "Resumo da semana:", *[f"- {item}" for item in resumo_semana], "",
        "Ações operacionais:", f"- Criadas: {len(criadas_periodo)}", f"- Abertas: {len(abertas)}",
        f"- Em andamento: {len(em_andamento)}", f"- Concluídas: {len(concluidas_periodo)}",
        f"- Altas/críticas abertas: {len(altas_criticas)}", "", "Documentação crítica:",
        f"- Obsoletos: {len(obsoletos)}", f"- Sem revisão: {len(sem_revisao_docs)}",
        f"- Baixa confiança: {len(baixa_confianca_docs)}", "", "Ações abertas relevantes:",
        *([f"- #{item.get('id')} [{status(item, 'prioridade') or 'N/I'}] {_texto_curto(item.get('titulo') or item.get('descricao') or 'Ação sem título', 90)}" for item in acoes_telegram] or ["- Nenhuma ação aberta identificada."]),
        "", "Decisões pendentes:", *([f"- {item}" for item in decisoes] or ["- Nenhuma decisão pendente identificada nas fontes disponíveis."]),
        "", "Recomendações para a próxima semana:", *[f"- {item}" for item in recomendacoes],
        "", "Modo: CONSULTA", "Nenhum cronograma, RDO, MinIO ou OpenProject foi alterado.",
        "Este relatório não confirma liberação definitiva da obra ou de qualquer frente.",
    ]

    limite = max(1, min(limite_itens, 50))
    resultado = {
        "ok": True,
        "mvp": "0.7H",
        "tipo_relatorio": "RELATORIO_SEMANAL_EXECUTIVO",
        "obra_codigo": obra_codigo,
        "area": area,
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "status_executivo": status_executivo,
        "resumo_semana": resumo_semana,
        "indicadores_acoes": {
            "criadas_no_periodo": len(criadas_periodo), "abertas": len(abertas),
            "em_andamento": len(em_andamento), "concluidas_no_periodo": len(concluidas_periodo),
            "canceladas_no_periodo": len(canceladas_periodo),
            "altas_ou_criticas_abertas": len(altas_criticas),
            "sem_responsavel": len(sem_responsavel), "sem_prazo": len(sem_prazo),
            "possivelmente_atrasadas": len(atrasadas),
        },
        "acoes_abertas": [resumir_acao(item) for item in nao_encerradas[:limite]],
        "acoes_concluidas_no_periodo": [resumir_acao(item) for item in concluidas_periodo[:limite]],
        "acoes_sem_responsavel": [resumir_acao(item) for item in sem_responsavel[:limite]],
        "acoes_sem_prazo": [resumir_acao(item) for item in sem_prazo[:limite]],
        "indicadores_documentais": {
            "documentos_classificados": len(classificados_periodo),
            "documentos_obsoletos": len(obsoletos),
            "documentos_sem_revisao": len(sem_revisao_docs),
            "documentos_baixa_confianca": len(baixa_confianca_docs),
            "alteracoes_documentais": len(alteracoes_documentais), "as_built": len(as_built),
        },
        "documentos_criticos": [],
        "briefing_diario": {
            "total_briefings_enviados": len(briefings_periodo),
            "ultimo_briefing_enviado_em": ultimo_briefing,
        },
        "decisoes_pendentes": decisoes,
        "recomendacoes_proxima_semana": recomendacoes,
        "resposta_telegram": "\n".join(linhas),
        "fontes_indisponiveis": fontes_indisponiveis,
        "flags_seguranca": dict(AGENTE_008_SEGURANCA_CONSULTA),
        **AGENTE_008_SEGURANCA_CONSULTA,
    }
    # Reconstrói a lista crítica sem depender de hashabilidade dos registros de origem.
    criticos_unicos: list[dict[str, Any]] = []
    vistos: set[tuple[Any, ...]] = set()
    for item in obsoletos + sem_revisao_docs + baixa_confianca_docs:
        chave = (item.get("documento_id"), item.get("status_revisao"), item.get("numero_revisao"))
        if chave not in vistos:
            vistos.add(chave)
            criticos_unicos.append(resumir_documento(item))
    resultado["documentos_criticos"] = criticos_unicos[:limite]
    resultado = serializar_json_seguro(resultado)

    relatorio_id = None
    if salvar_relatorio:
        cur.execute(
            """
            INSERT INTO relatorios_semanais_executivos_obra (
                obra_codigo, area, data_inicio, data_fim, payload_relatorio, resposta_telegram
            ) VALUES (
                %(obra_codigo)s, %(area)s, %(data_inicio)s, %(data_fim)s,
                %(payload_relatorio)s, %(resposta_telegram)s
            )
            RETURNING id;
            """,
            {
                "obra_codigo": obra_codigo, "area": area,
                "data_inicio": data_inicio, "data_fim": data_fim,
                "payload_relatorio": Json(resultado),
                "resposta_telegram": resultado["resposta_telegram"],
            },
        )
        relatorio_id = cur.fetchone()[0]
    resultado["relatorio_id"] = relatorio_id
    return serializar_json_seguro(resultado)


def _formatar_data_br(valor: Any) -> str:
    data_valor = _data_operacional(valor)
    return data_valor.strftime("%d/%m/%Y") if data_valor else "não informada"


def _texto_markdown(valor: Any, padrao: str = "Não informado") -> str:
    texto = re.sub(r"\s+", " ", str(valor or "").strip()).replace("|", "\\|")
    return texto or padrao


def _carregar_ou_gerar_relatorio_para_exportacao(
    cur,
    obra_codigo: str,
    area: Optional[str],
    data_inicio: date | None,
    data_fim: date | None,
    relatorio_semanal_id: int | None,
    limite_itens: int,
) -> tuple[dict[str, Any], int | None, bool]:
    if relatorio_semanal_id is not None:
        cur.execute(
            """
            SELECT obra_codigo, area, data_inicio, data_fim, payload_relatorio
            FROM relatorios_semanais_executivos_obra
            WHERE id = %(id)s;
            """,
            {"id": relatorio_semanal_id},
        )
        row = cur.fetchone()
        if row is None:
            raise ValueError("RELATORIO_SEMANAL_NAO_ENCONTRADO")
        if row[0] != obra_codigo:
            raise ValueError("RELATORIO_SEMANAL_NAO_PERTENCE_A_OBRA")
        if not isinstance(row[4], dict):
            raise ValueError("PAYLOAD_RELATORIO_SEMANAL_INVALIDO")
        relatorio = {
            **row[4],
            "obra_codigo": row[0],
            "area": row[1],
            "data_inicio": row[2],
            "data_fim": row[3],
            "relatorio_id": relatorio_semanal_id,
        }
        return serializar_json_seguro(relatorio), relatorio_semanal_id, True

    inicio, fim = _periodo_relatorio_semanal(data_inicio, data_fim)
    relatorio = _gerar_relatorio_semanal_executivo(
        cur, obra_codigo, area, inicio, fim, limite_itens, False,
    )
    return serializar_json_seguro(relatorio), None, False


def _montar_markdown_relatorio_semanal(relatorio: dict[str, Any]) -> tuple[str, str]:
    obra_codigo = _texto_markdown(relatorio.get("obra_codigo"))
    area = relatorio.get("area")
    area_exibicao = _texto_markdown(str(area).replace("_", " ").title()) if area else "Todas"
    data_inicio = _formatar_data_br(relatorio.get("data_inicio"))
    data_fim = _formatar_data_br(relatorio.get("data_fim"))
    status_executivo = _texto_markdown(relatorio.get("status_executivo"))

    resumo = [
        _texto_markdown(item)
        for item in (relatorio.get("resumo_semana") or [])
        if item
    ][:4]
    complementos = [
        f"O status executivo consolidado é {status_executivo} e requer revisão técnica responsável.",
        "Os indicadores abaixo refletem somente as fontes operacionais disponíveis no período.",
        "Este material é consultivo e não representa liberação definitiva da obra ou de qualquer frente.",
    ]
    for complemento in complementos:
        if len(resumo) >= 3:
            break
        resumo.append(complemento)
    resumo = resumo[:6]

    indicadores_acoes = relatorio.get("indicadores_acoes") or {}
    indicadores_documentais = relatorio.get("indicadores_documentais") or {}
    acoes_abertas = relatorio.get("acoes_abertas") or []
    acoes_concluidas = relatorio.get("acoes_concluidas_no_periodo") or []
    decisoes = relatorio.get("decisoes_pendentes") or []
    recomendacoes = relatorio.get("recomendacoes_proxima_semana") or []

    linhas = [
        f"# Relatório Semanal Executivo — {obra_codigo}",
        "",
        f"**Período:** {data_inicio} a {data_fim}  ",
        f"**Área:** {area_exibicao}  ",
        f"**Status executivo:** {status_executivo}  ",
        "**Gerado por:** AGENTE_008_GESTAO_OPERACIONAL_OBRA  ",
        "**Modo:** CONSULTA  ",
        "",
        "---",
        "",
        "## 1. Sumário executivo",
        "",
        *resumo,
        "",
        "---",
        "",
        "## 2. Indicadores operacionais da semana",
        "",
        "| Indicador | Valor |",
        "|---|---:|",
        f"| Ações criadas no período | {indicadores_acoes.get('criadas_no_periodo', 0)} |",
        f"| Ações abertas | {indicadores_acoes.get('abertas', 0)} |",
        f"| Ações em andamento | {indicadores_acoes.get('em_andamento', 0)} |",
        f"| Ações concluídas no período | {indicadores_acoes.get('concluidas_no_periodo', 0)} |",
        f"| Ações altas/críticas abertas | {indicadores_acoes.get('altas_ou_criticas_abertas', 0)} |",
        f"| Ações sem responsável | {indicadores_acoes.get('sem_responsavel', 0)} |",
        f"| Ações sem prazo | {indicadores_acoes.get('sem_prazo', 0)} |",
        "",
        "---",
        "",
        "## 3. Ações abertas relevantes",
        "",
        *(
            [
                f"- #{item.get('id')} [{_texto_markdown(item.get('prioridade'), 'N/I')}] "
                f"{_texto_markdown(item.get('titulo'), 'Ação sem título')} — "
                f"{_texto_markdown(item.get('status'), 'N/I')}"
                for item in acoes_abertas
            ]
            or ["- Nenhuma ação aberta relevante identificada."]
        ),
        "",
        "---",
        "",
        "## 4. Ações concluídas no período",
        "",
        *(
            [
                f"- #{item.get('id')} {_texto_markdown(item.get('titulo'), 'Ação sem título')} "
                f"— Concluída em {_formatar_data_br(item.get('concluido_em'))}"
                for item in acoes_concluidas
            ]
            or ["- Nenhuma ação concluída identificada no período."]
        ),
        "",
        "---",
        "",
        "## 5. Documentação crítica",
        "",
        "| Item | Quantidade |",
        "|---|---:|",
        f"| Documentos obsoletos | {indicadores_documentais.get('documentos_obsoletos', 0)} |",
        f"| Documentos sem revisão identificada | {indicadores_documentais.get('documentos_sem_revisao', 0)} |",
        f"| Documentos com baixa confiança | {indicadores_documentais.get('documentos_baixa_confianca', 0)} |",
        f"| Alterações documentais | {indicadores_documentais.get('alteracoes_documentais', 0)} |",
        f"| As Built | {indicadores_documentais.get('as_built', 0)} |",
        "",
        "---",
        "",
        "## 6. Decisões pendentes",
        "",
        *([f"- {_texto_markdown(item)}" for item in decisoes] or ["- Nenhuma decisão pendente identificada nas fontes disponíveis."]),
        "",
        "---",
        "",
        "## 7. Recomendações para a próxima semana",
        "",
        *([f"- {_texto_markdown(item)}" for item in recomendacoes] or ["- Manter acompanhamento consultivo e validação técnica responsável."]),
        "",
        "---",
        "",
        "## 8. Observações de segurança e governança",
        "",
        "Este relatório é consultivo.",
        "Nenhum cronograma oficial foi alterado.",
        "Nenhum RDO oficial foi alterado.",
        "Nenhum arquivo foi enviado a terceiros.",
        "Nenhum RPA foi acionado.",
        "Nenhuma sincronização com OpenProject foi executada.",
        "Nenhum link público foi gerado.",
    ]
    return "\n".join(linhas), "\n".join(resumo)


def _exportar_relatorio_semanal_controlado(
    cur,
    payload: GestaoOperacionalExportarRelatorioSemanalRequest,
) -> dict[str, Any]:
    formato = payload.formato.strip().upper()
    if formato != "MARKDOWN":
        raise ValueError("FORMATO_NAO_SUPORTADO_NESTE_MVP")

    relatorio, relatorio_id, reutilizou_relatorio = _carregar_ou_gerar_relatorio_para_exportacao(
        cur,
        payload.obra_codigo,
        payload.area,
        payload.data_inicio,
        payload.data_fim,
        payload.relatorio_semanal_id,
        payload.limite_itens,
    )
    conteudo_markdown, resumo_executivo = _montar_markdown_relatorio_semanal(relatorio)
    titulo = f"Relatório Semanal Executivo — {relatorio['obra_codigo']}"
    metadados = serializar_json_seguro({
        "mvp": "0.7I",
        "relatorio_reutilizado": reutilizou_relatorio,
        "flags_seguranca": dict(AGENTE_008_SEGURANCA_EXPORTACAO),
    })

    exportacao_id = None
    if payload.salvar_exportacao:
        cur.execute(
            """
            INSERT INTO exportacoes_relatorios_semanais_obra (
                obra_codigo, area, relatorio_semanal_id, data_inicio, data_fim,
                formato, status, titulo, conteudo_markdown, resumo_executivo,
                enviado_para_terceiros, alterou_rdo_oficial, alterou_cronograma,
                executou_rpa, sincronizou_openproject, gerou_pdf, gerou_link_publico,
                metadados
            ) VALUES (
                %(obra_codigo)s, %(area)s, %(relatorio_semanal_id)s,
                %(data_inicio)s, %(data_fim)s, 'MARKDOWN', 'GERADO', %(titulo)s,
                %(conteudo_markdown)s, %(resumo_executivo)s, false, false, false,
                false, false, false, false, %(metadados)s
            )
            RETURNING id;
            """,
            {
                "obra_codigo": relatorio["obra_codigo"],
                "area": relatorio.get("area"),
                "relatorio_semanal_id": relatorio_id,
                "data_inicio": _data_operacional(relatorio.get("data_inicio")),
                "data_fim": _data_operacional(relatorio.get("data_fim")),
                "titulo": titulo,
                "conteudo_markdown": conteudo_markdown,
                "resumo_executivo": resumo_executivo,
                "metadados": Json(metadados),
            },
        )
        exportacao_id = cur.fetchone()[0]

    resumo_telegram = [
        _texto_markdown(item)
        for item in (relatorio.get("resumo_semana") or [])
        if item
    ][:2]
    if not resumo_telegram:
        resumo_telegram = ["Relatório consolidado a partir das fontes disponíveis."]
    identificador_exportacao = f"#{exportacao_id}" if exportacao_id is not None else "não persistida"
    resposta_telegram = "\n".join([
        f"📄 Relatório semanal exportado — {relatorio['obra_codigo']}",
        "", "Período:",
        f"{_formatar_data_br(relatorio.get('data_inicio'))} a {_formatar_data_br(relatorio.get('data_fim'))}",
        "", "Formato:", "MARKDOWN", "", "Exportação:", identificador_exportacao,
        "", "Status executivo:", str(relatorio.get("status_executivo") or "NÃO INFORMADO"),
        "", "Resumo:", *[f"- {item}" for item in resumo_telegram],
        "", "Governança:",
        "Relatório gerado para revisão interna. Nenhum envio externo, PDF, RDO, cronograma, MinIO, OpenProject ou RPA foi alterado.",
    ])

    resultado = {
        "ok": True,
        "mvp": "0.7I",
        "tipo_exportacao": "RELATORIO_SEMANAL_EXECUTIVO",
        "exportacao_id": exportacao_id,
        "relatorio_semanal_id": relatorio_id,
        "obra_codigo": relatorio["obra_codigo"],
        "area": relatorio.get("area"),
        "data_inicio": relatorio.get("data_inicio"),
        "data_fim": relatorio.get("data_fim"),
        "formato": formato,
        "status": "GERADO",
        "status_executivo": relatorio.get("status_executivo"),
        "titulo": titulo,
        "conteudo_markdown": conteudo_markdown,
        "resumo_executivo": resumo_executivo,
        "resposta_telegram": resposta_telegram,
        "salvar_exportacao": payload.salvar_exportacao,
        "flags_seguranca": dict(AGENTE_008_SEGURANCA_EXPORTACAO),
        **AGENTE_008_SEGURANCA_EXPORTACAO,
    }
    return serializar_json_seguro(resultado)


def _carregar_ou_criar_exportacao_para_pdf(
    cur,
    payload: GestaoOperacionalGerarPdfRelatorioSemanalRequest,
) -> dict[str, Any]:
    if payload.exportacao_relatorio_id is None:
        exportacao = _exportar_relatorio_semanal_controlado(
            cur,
            GestaoOperacionalExportarRelatorioSemanalRequest(
                obra_codigo=payload.obra_codigo,
                area=payload.area,
                data_inicio=payload.data_inicio,
                data_fim=payload.data_fim,
                relatorio_semanal_id=payload.relatorio_semanal_id,
                formato="MARKDOWN",
                limite_itens=payload.limite_itens,
                salvar_exportacao=True,
            ),
        )
        return {
            "id": exportacao["exportacao_id"],
            "obra_codigo": exportacao["obra_codigo"],
            "area": exportacao.get("area"),
            "relatorio_semanal_id": exportacao.get("relatorio_semanal_id"),
            "data_inicio": _data_operacional(exportacao.get("data_inicio")),
            "data_fim": _data_operacional(exportacao.get("data_fim")),
            "titulo": exportacao["titulo"],
            "conteudo_markdown": exportacao["conteudo_markdown"],
            "exportacao_criada": True,
        }

    cur.execute(
        """
        SELECT id, obra_codigo, area, relatorio_semanal_id, data_inicio, data_fim,
               titulo, conteudo_markdown
        FROM exportacoes_relatorios_semanais_obra
        WHERE id = %(id)s;
        """,
        {"id": payload.exportacao_relatorio_id},
    )
    row = cur.fetchone()
    if row is None:
        raise ValueError("EXPORTACAO_RELATORIO_NAO_ENCONTRADA")
    if row[1] != payload.obra_codigo:
        raise ValueError("EXPORTACAO_RELATORIO_NAO_PERTENCE_A_OBRA")
    if not str(row[7] or "").strip():
        raise ValueError("CONTEUDO_MARKDOWN_EXPORTACAO_INVALIDO")
    return {
        "id": row[0],
        "obra_codigo": row[1],
        "area": row[2],
        "relatorio_semanal_id": row[3],
        "data_inicio": row[4],
        "data_fim": row[5],
        "titulo": row[6],
        "conteudo_markdown": row[7],
        "exportacao_criada": False,
    }


def _texto_pdf_markdown(valor: Any) -> str:
    texto = str(valor or "").strip()
    texto = re.sub(r"\*\*(.*?)\*\*", r"\1", texto)
    texto = re.sub(r"`(.*?)`", r"\1", texto)
    return html.escape(texto)


def _gerar_bytes_pdf_relatorio_semanal(exportacao: dict[str, Any]) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.pagesizes import A4, portrait
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            KeepTogether,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError as exc:
        raise RuntimeError(
            "Biblioteca reportlab não instalada. Instale requirements-api.txt."
        ) from exc

    conteudo_markdown = str(exportacao["conteudo_markdown"])
    status_match = re.search(
        r"\*\*Status executivo:\*\*\s*([^\n]+)", conteudo_markdown, re.IGNORECASE,
    )
    status_executivo = (
        status_match.group(1).strip().rstrip("  ") if status_match else "NÃO INFORMADO"
    )
    area = exportacao.get("area")
    area_exibicao = str(area).replace("_", " ").title() if area else "Todas"

    buffer = BytesIO()
    documento = SimpleDocTemplate(
        buffer,
        pagesize=portrait(A4),
        rightMargin=1.7 * cm,
        leftMargin=1.7 * cm,
        topMargin=1.5 * cm,
        bottomMargin=2.1 * cm,
        title=str(exportacao["titulo"]),
        author="AGENTE_008_GESTAO_OPERACIONAL_OBRA",
    )
    estilos_base = getSampleStyleSheet()
    azul = colors.HexColor("#174A5B")
    cinza = colors.HexColor("#4B5563")
    estilos = {
        "titulo": ParagraphStyle(
            "PdfTitulo", parent=estilos_base["Title"], fontName="Helvetica-Bold",
            fontSize=20, leading=24, textColor=azul, alignment=TA_CENTER,
            spaceAfter=12,
        ),
        "capa": ParagraphStyle(
            "PdfCapa", parent=estilos_base["BodyText"], fontName="Helvetica",
            fontSize=10, leading=15, textColor=cinza, alignment=TA_LEFT,
        ),
        "secao": ParagraphStyle(
            "PdfSecao", parent=estilos_base["Heading2"], fontName="Helvetica-Bold",
            fontSize=12, leading=15, textColor=azul, spaceBefore=10, spaceAfter=6,
            keepWithNext=True,
        ),
        "corpo": ParagraphStyle(
            "PdfCorpo", parent=estilos_base["BodyText"], fontName="Helvetica",
            fontSize=9, leading=13, textColor=colors.HexColor("#222222"),
            spaceAfter=4,
        ),
        "item": ParagraphStyle(
            "PdfItem", parent=estilos_base["BodyText"], fontName="Helvetica",
            fontSize=9, leading=13, leftIndent=12, firstLineIndent=-7, spaceAfter=3,
        ),
    }

    elementos: list[Any] = [
        Paragraph("Relatório Semanal Executivo", estilos["titulo"]),
        Table(
            [[Paragraph(
                "<b>Obra:</b> " + _texto_pdf_markdown(exportacao["obra_codigo"]) + "<br/>"
                "<b>Área:</b> " + _texto_pdf_markdown(area_exibicao) + "<br/>"
                "<b>Período:</b> " + _texto_pdf_markdown(_formatar_data_br(exportacao["data_inicio"]))
                + " a " + _texto_pdf_markdown(_formatar_data_br(exportacao["data_fim"])) + "<br/>"
                "<b>Status executivo:</b> " + _texto_pdf_markdown(status_executivo) + "<br/>"
                "<b>Gerado por:</b> AGENTE_008_GESTAO_OPERACIONAL_OBRA<br/>"
                "<b>Modo:</b> CONSULTA",
                estilos["capa"],
            )]],
            colWidths=[16.2 * cm],
            style=TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F2F6F7")),
                ("BOX", (0, 0), (-1, -1), 0.7, azul),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
                ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]),
        ),
        Spacer(1, 10),
    ]

    linhas = conteudo_markdown.splitlines()
    indice = 0
    encontrou_secao = False
    while indice < len(linhas):
        linha = linhas[indice].strip()
        if linha.startswith("## "):
            encontrou_secao = True
            elementos.append(Paragraph(_texto_pdf_markdown(linha[3:]), estilos["secao"]))
            indice += 1
            continue
        if not encontrou_secao or not linha or linha == "---":
            indice += 1
            continue
        if linha.startswith("|"):
            tabela_markdown: list[list[str]] = []
            while indice < len(linhas) and linhas[indice].strip().startswith("|"):
                celulas = [
                    item.replace(r"\|", "|").strip()
                    for item in re.split(
                        r"(?<!\\)\|", linhas[indice].strip().strip("|")
                    )
                ]
                if not all(re.fullmatch(r":?-{3,}:?", item) for item in celulas):
                    tabela_markdown.append(celulas)
                indice += 1
            if tabela_markdown:
                tabela_pdf = [
                    [Paragraph(_texto_pdf_markdown(celula), estilos["corpo"]) for celula in linha_tabela]
                    for linha_tabela in tabela_markdown
                ]
                tabela = Table(tabela_pdf, colWidths=[12.5 * cm, 3.7 * cm], repeatRows=1)
                tabela.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), azul),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]))
                elementos.append(KeepTogether([tabela, Spacer(1, 5)]))
            continue
        if linha.startswith("- "):
            elementos.append(Paragraph("• " + _texto_pdf_markdown(linha[2:]), estilos["item"]))
        else:
            elementos.append(Paragraph(_texto_pdf_markdown(linha), estilos["corpo"]))
        indice += 1

    def rodape(canvas, doc) -> None:
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#CBD5E1"))
        canvas.line(doc.leftMargin, 1.55 * cm, A4[0] - doc.rightMargin, 1.55 * cm)
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(cinza)
        canvas.drawString(doc.leftMargin, 1.15 * cm, f"Página {doc.page}")
        canvas.drawRightString(
            A4[0] - doc.rightMargin,
            1.15 * cm,
            "Documento gerado automaticamente para revisão interna.",
        )
        canvas.drawCentredString(
            A4[0] / 2,
            0.78 * cm,
            "Não representa liberação definitiva de frente de obra.",
        )
        canvas.restoreState()

    documento.build(elementos, onFirstPage=rodape, onLaterPages=rodape)
    return buffer.getvalue()


def _nome_arquivo_pdf_relatorio(exportacao: dict[str, Any]) -> str:
    obra = re.sub(r"[^A-Za-z0-9_-]+", "_", str(exportacao["obra_codigo"]).upper())
    inicio = _data_operacional(exportacao["data_inicio"])
    fim = _data_operacional(exportacao["data_fim"])
    if inicio is None or fim is None:
        raise ValueError("PERIODO_EXPORTACAO_INVALIDO")
    return (
        f"{obra}_RELATORIO_SEMANAL_EXECUTIVO_"
        f"{inicio.isoformat()}_a_{fim.isoformat()}.pdf"
    )


def _gerar_pdf_relatorio_semanal_privado(
    cur,
    payload: GestaoOperacionalGerarPdfRelatorioSemanalRequest,
) -> dict[str, Any]:
    if not payload.salvar_pdf and not payload.armazenar_minio:
        raise ValueError("DESTINO_PRIVADO_PDF_NAO_INFORMADO")
    exportacao = _carregar_ou_criar_exportacao_para_pdf(cur, payload)
    pdf_bytes = _gerar_bytes_pdf_relatorio_semanal(exportacao)
    nome_arquivo = _nome_arquivo_pdf_relatorio(exportacao)
    tamanho_bytes = len(pdf_bytes)
    sha256 = hashlib.sha256(pdf_bytes).hexdigest()
    inicio = _data_operacional(exportacao["data_inicio"])
    fim = _data_operacional(exportacao["data_fim"])
    if inicio is None or fim is None:
        raise ValueError("PERIODO_EXPORTACAO_INVALIDO")

    caminho_relativo = (
        Path("outputs") / "relatorios_semanais"
        / normalizar_nome_diretorio(exportacao["obra_codigo"])
        / inicio.strftime("%Y") / inicio.strftime("%m") / nome_arquivo
    )
    caminho_local: str | None = None
    bucket: str | None = None
    object_key: str | None = None
    minio_uri: str | None = None
    erro_minio: str | None = None
    alterou_minio = False

    if payload.salvar_pdf:
        caminho_relativo.parent.mkdir(parents=True, exist_ok=True)
        caminho_relativo.write_bytes(pdf_bytes)
        caminho_local = str(caminho_relativo)

    if payload.armazenar_minio:
        bucket_pdf = "obra-caio"
        object_key_pdf = (
            "12_relatorios_executivos/relatorios_semanais/"
            f"{inicio:%Y}/{inicio:%m}/{nome_arquivo}"
        )
        try:
            cliente = get_minio_client()
            if not cliente.bucket_exists(bucket_pdf):
                cliente.make_bucket(bucket_pdf)
                alterou_minio = True
            cliente.put_object(
                bucket_name=bucket_pdf,
                object_name=object_key_pdf,
                data=BytesIO(pdf_bytes),
                length=tamanho_bytes,
                content_type="application/pdf",
                metadata={"sha256": sha256, "mvp": "0.7J"},
            )
            bucket = bucket_pdf
            object_key = object_key_pdf
            minio_uri = f"s3://{bucket}/{object_key}"
            alterou_minio = True
        except Exception as exc:
            erro_minio = _texto_curto(exc, 200)
            if caminho_local is None:
                caminho_relativo.parent.mkdir(parents=True, exist_ok=True)
                caminho_relativo.write_bytes(pdf_bytes)
                caminho_local = str(caminho_relativo)

    flags_seguranca = {
        **AGENTE_008_SEGURANCA_PDF,
        "altera_minio": alterou_minio,
    }
    metadados = serializar_json_seguro({
        "mvp": "0.7J",
        "tipo": "PDF_PRIVADO_REVISAO",
        "origem": "RELATORIO_SEMANAL_EXECUTIVO",
        "sem_link_publico": True,
        "sem_envio_externo": True,
        "flags_seguranca": flags_seguranca,
    })
    status = "GERADO"
    cur.execute(
        """
        INSERT INTO pdfs_relatorios_semanais_obra (
            obra_codigo, area, exportacao_relatorio_id, relatorio_semanal_id,
            data_inicio, data_fim, titulo, status, nome_arquivo, caminho_local,
            bucket, object_key, minio_uri, tamanho_bytes, sha256,
            enviado_para_terceiros, gerou_link_publico, alterou_rdo_oficial,
            alterou_cronograma, executou_rpa, sincronizou_openproject, metadados
        ) VALUES (
            %(obra_codigo)s, %(area)s, %(exportacao_relatorio_id)s,
            %(relatorio_semanal_id)s, %(data_inicio)s, %(data_fim)s, %(titulo)s,
            %(status)s, %(nome_arquivo)s, %(caminho_local)s, %(bucket)s,
            %(object_key)s, %(minio_uri)s, %(tamanho_bytes)s, %(sha256)s,
            false, false, false, false, false, false, %(metadados)s
        )
        RETURNING id;
        """,
        {
            "obra_codigo": exportacao["obra_codigo"],
            "area": exportacao.get("area"),
            "exportacao_relatorio_id": exportacao["id"],
            "relatorio_semanal_id": exportacao.get("relatorio_semanal_id"),
            "data_inicio": inicio,
            "data_fim": fim,
            "titulo": exportacao["titulo"],
            "status": status,
            "nome_arquivo": nome_arquivo,
            "caminho_local": caminho_local,
            "bucket": bucket,
            "object_key": object_key,
            "minio_uri": minio_uri,
            "tamanho_bytes": tamanho_bytes,
            "sha256": sha256,
            "metadados": Json(metadados),
        },
    )
    pdf_id = cur.fetchone()[0]
    cur.execute(
        """
        UPDATE exportacoes_relatorios_semanais_obra
        SET gerou_pdf = true, atualizado_em = NOW()
        WHERE id = %(id)s;
        """,
        {"id": exportacao["id"]},
    )

    dados_telegram = serializar_json_seguro({
        "obra_codigo": exportacao["obra_codigo"],
        "data_inicio": inicio,
        "data_fim": fim,
        "pdf_id": pdf_id,
        "nome_arquivo": nome_arquivo,
        "armazenamento": minio_uri or caminho_local,
    })
    resposta_telegram = "\n".join([
        f"📎 PDF privado gerado — {dados_telegram['obra_codigo']}",
        "", "Período:",
        f"{_formatar_data_br(dados_telegram['data_inicio'])} a "
        f"{_formatar_data_br(dados_telegram['data_fim'])}",
        "", "PDF:", f"#{dados_telegram['pdf_id']}",
        "", "Arquivo:", dados_telegram["nome_arquivo"],
        "", "Armazenamento:", dados_telegram["armazenamento"],
        "", "Governança:",
        "PDF gerado para revisão interna.",
        "Nenhum envio externo, link público, RDO, cronograma, OpenProject ou RPA foi alterado.",
    ])
    resultado = {
        "ok": True,
        "mvp": "0.7J",
        "pdf_id": pdf_id,
        "exportacao_relatorio_id": exportacao["id"],
        "relatorio_semanal_id": exportacao.get("relatorio_semanal_id"),
        "obra_codigo": exportacao["obra_codigo"],
        "area": exportacao.get("area"),
        "data_inicio": inicio,
        "data_fim": fim,
        "titulo": exportacao["titulo"],
        "status": status,
        "nome_arquivo": nome_arquivo,
        "caminho_local": caminho_local,
        "bucket": bucket,
        "object_key": object_key,
        "minio_uri": minio_uri,
        "tamanho_bytes": tamanho_bytes,
        "sha256": sha256,
        "metadados": metadados,
        "exportacao_criada": exportacao["exportacao_criada"],
        "aviso_minio": erro_minio,
        "resposta_telegram": resposta_telegram,
        "flags_seguranca": dict(flags_seguranca),
        **flags_seguranca,
    }
    return serializar_json_seguro(resultado)


def _resposta_telegram_decisao_pdf(
    obra_codigo: str,
    pdf_relatorio_id: int,
    decisao: str,
    status_resultante: str,
    decisor: str | None,
    motivo: str | None,
    observacao: str | None,
) -> str:
    dados = serializar_json_seguro({
        "obra_codigo": obra_codigo,
        "pdf_relatorio_id": pdf_relatorio_id,
        "status_resultante": status_resultante,
        "decisor": decisor or "Não informado",
        "motivo": motivo or "Não informado",
        "observacao": observacao or "Não informada",
    })
    governanca = (
        "Nenhum envio externo, link público, RDO, cronograma, OpenProject, "
        "MinIO ou RPA foi alterado."
    )
    if decisao == "APROVAR":
        linhas = [
            f"✅ PDF do relatório semanal aprovado — {dados['obra_codigo']}",
            "", "PDF:", f"#{dados['pdf_relatorio_id']}",
            "", "Status:", dados["status_resultante"],
            "", "Decisor:", dados["decisor"],
            "", "Governança:",
            "Aprovação registrada para uso interno.", governanca,
        ]
    elif decisao == "REJEITAR":
        linhas = [
            f"❌ PDF do relatório semanal rejeitado — {dados['obra_codigo']}",
            "", "PDF:", f"#{dados['pdf_relatorio_id']}",
            "", "Status:", dados["status_resultante"],
            "", "Motivo:", dados["motivo"],
            "", "Governança:", "Rejeição registrada.", governanca,
        ]
    else:
        linhas = [
            f"🛠️ Ajustes solicitados no PDF do relatório semanal — {dados['obra_codigo']}",
            "", "PDF:", f"#{dados['pdf_relatorio_id']}",
            "", "Status:", dados["status_resultante"],
            "", "Observação:", dados["observacao"],
            "", "Governança:", "Solicitação de ajustes registrada.", governanca,
        ]
    return "\n".join(str(item) for item in linhas)


def _aprovar_pdf_relatorio_semanal_controlado(
    cur,
    payload: GestaoOperacionalAprovarPdfRelatorioSemanalRequest,
) -> dict[str, Any]:
    decisao = payload.decisao
    status_por_decisao = {
        "APROVAR": "APROVADO_PARA_USO_INTERNO",
        "REJEITAR": "REJEITADO",
        "SOLICITAR_AJUSTES": "AJUSTES_SOLICITADOS",
    }
    status_idempotente = {
        "APROVAR": "JA_APROVADO",
        "REJEITAR": "JA_REJEITADO",
        "SOLICITAR_AJUSTES": "JA_COM_AJUSTES_SOLICITADOS",
    }
    status_resultante = status_por_decisao[decisao]
    cur.execute(
        """
        SELECT id, tenant_id, obra_codigo, area, exportacao_relatorio_id,
               relatorio_semanal_id, status
        FROM pdfs_relatorios_semanais_obra
        WHERE id = %(pdf_relatorio_id)s
        FOR UPDATE;
        """,
        {"pdf_relatorio_id": payload.pdf_relatorio_id},
    )
    pdf = cur.fetchone()
    if pdf is None:
        raise ValueError("PDF_RELATORIO_SEMANAL_NAO_ENCONTRADO")
    if pdf[2] != payload.obra_codigo:
        raise ValueError("PDF_RELATORIO_SEMANAL_NAO_PERTENCE_A_OBRA")

    decisor = (
        payload.decisor_nome
        or payload.decisor_telegram_username
        or payload.decisor_telegram_user_id
    )
    resposta_telegram = _resposta_telegram_decisao_pdf(
        payload.obra_codigo,
        payload.pdf_relatorio_id,
        decisao,
        status_resultante,
        decisor,
        payload.motivo,
        payload.observacao,
    )
    if pdf[6] == status_resultante:
        return serializar_json_seguro({
            "ok": True,
            "mvp": "0.7K",
            "status": status_idempotente[decisao],
            "status_resultante": status_resultante,
            "decisao": decisao,
            "pdf_relatorio_id": payload.pdf_relatorio_id,
            "obra_codigo": payload.obra_codigo,
            "idempotente": True,
            "resposta_telegram": resposta_telegram,
            "flags_seguranca": dict(AGENTE_008_SEGURANCA_APROVACAO_PDF),
            **AGENTE_008_SEGURANCA_APROVACAO_PDF,
        })

    metadados = serializar_json_seguro({
        "mvp": "0.7K",
        "tipo": "DECISAO_EXECUTIVA_PDF_RELATORIO_SEMANAL",
        "flags_seguranca": dict(AGENTE_008_SEGURANCA_APROVACAO_PDF),
    })
    cur.execute(
        """
        INSERT INTO aprovacoes_relatorios_semanais_obra (
            tenant_id, obra_codigo, area, pdf_relatorio_id,
            exportacao_relatorio_id, relatorio_semanal_id, decisao,
            status_resultante, aprovado, rejeitado, ajustes_solicitados,
            motivo, observacao, decisor_nome, decisor_telegram_user_id,
            decisor_telegram_username, decisor_chat_id, canal, origem,
            enviado_para_terceiros, gerou_link_publico, alterou_rdo_oficial,
            alterou_cronograma, executou_rpa, sincronizou_openproject,
            alterou_minio, metadados
        ) VALUES (
            %(tenant_id)s, %(obra_codigo)s, %(area)s, %(pdf_relatorio_id)s,
            %(exportacao_relatorio_id)s, %(relatorio_semanal_id)s, %(decisao)s,
            %(status_resultante)s, %(aprovado)s, %(rejeitado)s,
            %(ajustes_solicitados)s, %(motivo)s, %(observacao)s,
            %(decisor_nome)s, %(decisor_telegram_user_id)s,
            %(decisor_telegram_username)s, %(decisor_chat_id)s, 'telegram',
            'AGENTE_007_ORQUESTRADOR_EXECUTIVO', false, false, false, false,
            false, false, false, %(metadados)s
        )
        RETURNING id, criado_em;
        """,
        {
            "tenant_id": pdf[1],
            "obra_codigo": pdf[2],
            "area": pdf[3],
            "pdf_relatorio_id": pdf[0],
            "exportacao_relatorio_id": pdf[4],
            "relatorio_semanal_id": pdf[5],
            "decisao": decisao,
            "status_resultante": status_resultante,
            "aprovado": decisao == "APROVAR",
            "rejeitado": decisao == "REJEITAR",
            "ajustes_solicitados": decisao == "SOLICITAR_AJUSTES",
            "motivo": payload.motivo,
            "observacao": payload.observacao,
            "decisor_nome": payload.decisor_nome,
            "decisor_telegram_user_id": payload.decisor_telegram_user_id,
            "decisor_telegram_username": payload.decisor_telegram_username,
            "decisor_chat_id": payload.decisor_chat_id,
            "metadados": Json(metadados),
        },
    )
    aprovacao_id, criado_em = cur.fetchone()

    if decisao == "APROVAR":
        campos_decisao = "aprovado_por = %(decisor)s, aprovado_em = NOW(),"
    elif decisao == "REJEITAR":
        campos_decisao = "rejeitado_por = %(decisor)s, rejeitado_em = NOW(),"
    else:
        campos_decisao = (
            "ajustes_solicitados_por = %(decisor)s, "
            "ajustes_solicitados_em = NOW(),"
        )
    cur.execute(
        f"""
        UPDATE pdfs_relatorios_semanais_obra
        SET status = %(status_resultante)s,
            {campos_decisao}
            motivo_decisao = %(motivo)s,
            observacao_decisao = %(observacao)s,
            atualizado_em = NOW()
        WHERE id = %(pdf_relatorio_id)s;
        """,
        {
            "pdf_relatorio_id": payload.pdf_relatorio_id,
            "status_resultante": status_resultante,
            "decisor": decisor,
            "motivo": payload.motivo,
            "observacao": payload.observacao,
        },
    )

    return serializar_json_seguro({
        "ok": True,
        "mvp": "0.7K",
        "status": status_resultante,
        "status_resultante": status_resultante,
        "decisao": decisao,
        "aprovacao_id": aprovacao_id,
        "pdf_relatorio_id": payload.pdf_relatorio_id,
        "exportacao_relatorio_id": pdf[4],
        "relatorio_semanal_id": pdf[5],
        "obra_codigo": payload.obra_codigo,
        "area": pdf[3],
        "decisor": decisor,
        "criado_em": criado_em,
        "idempotente": False,
        "resposta_telegram": resposta_telegram,
        "flags_seguranca": dict(AGENTE_008_SEGURANCA_APROVACAO_PDF),
        **AGENTE_008_SEGURANCA_APROVACAO_PDF,
    })


def _resposta_telegram_solicitacao_envio(
    obra_codigo: str,
    pdf_relatorio_id: int,
    solicitacao_envio_id: int,
    canal_pretendido: str,
    destinatario_nome: str | None,
    status: str,
) -> str:
    dados = serializar_json_seguro({
        "obra_codigo": obra_codigo,
        "pdf_relatorio_id": pdf_relatorio_id,
        "solicitacao_envio_id": solicitacao_envio_id,
        "canal_pretendido": canal_pretendido,
        "destinatario_nome": destinatario_nome or "Não informado",
        "status": status,
    })
    linhas = [
        f"📨 Solicitação de envio registrada — {dados['obra_codigo']}",
        "", "PDF:", f"#{dados['pdf_relatorio_id']}",
        "", "Solicitação:", f"#{dados['solicitacao_envio_id']}",
        "", "Canal pretendido:", dados["canal_pretendido"],
        "", "Destinatário:", dados["destinatario_nome"],
        "", "Status:", dados["status"],
        "", "Governança:",
        "A solicitação foi registrada, mas nenhum envio foi executado.",
        "Nenhum arquivo foi anexado.",
        "Nenhum link público ou presigned URL foi criado.",
        "Nenhum RDO, cronograma, MinIO, OpenProject ou RPA foi alterado.",
    ]
    return "\n".join(str(item) for item in linhas)


def _solicitar_envio_relatorio_semanal_controlado(
    cur,
    payload: GestaoOperacionalSolicitarEnvioRelatorioSemanalRequest,
) -> dict[str, Any]:
    canal_pretendido = payload.canal_pretendido.strip().upper()
    if canal_pretendido not in {"EMAIL", "WHATSAPP", "TELEGRAM", "INTERNO"}:
        raise ValueError("CANAL_NAO_SUPORTADO")

    cur.execute(
        """
        SELECT id, tenant_id, obra_codigo, area, status
        FROM pdfs_relatorios_semanais_obra
        WHERE id = %(pdf_relatorio_id)s
        FOR SHARE;
        """,
        {"pdf_relatorio_id": payload.pdf_relatorio_id},
    )
    pdf = cur.fetchone()
    if pdf is None:
        raise ValueError("PDF_RELATORIO_NAO_ENCONTRADO")
    if pdf[2] != payload.obra_codigo:
        raise ValueError("PDF_RELATORIO_NAO_PERTENCE_A_OBRA")
    if pdf[4] != "APROVADO_PARA_USO_INTERNO":
        raise ValueError("PDF_RELATORIO_NAO_APROVADO")

    cur.execute(
        """
        SELECT id
        FROM aprovacoes_relatorios_semanais_obra
        WHERE pdf_relatorio_id = %(pdf_relatorio_id)s
          AND status_resultante = 'APROVADO_PARA_USO_INTERNO'
        ORDER BY criado_em DESC, id DESC
        LIMIT 1;
        """,
        {"pdf_relatorio_id": payload.pdf_relatorio_id},
    )
    aprovacao = cur.fetchone()
    aprovacao_relatorio_id = aprovacao[0] if aprovacao else None
    status = "SOLICITADO_AGUARDANDO_EXECUCAO_CONTROLADA"

    cur.execute(
        """
        SELECT id, criado_em
        FROM solicitacoes_envio_relatorios_semanais_obra
        WHERE pdf_relatorio_id = %(pdf_relatorio_id)s
          AND canal_pretendido = %(canal_pretendido)s
          AND destinatario_nome IS NOT DISTINCT FROM %(destinatario_nome)s
          AND destinatario_contato IS NOT DISTINCT FROM %(destinatario_contato)s
          AND status = %(status)s
          AND criado_em >= NOW() - INTERVAL '24 hours'
        ORDER BY criado_em DESC, id DESC
        LIMIT 1;
        """,
        {
            "pdf_relatorio_id": payload.pdf_relatorio_id,
            "canal_pretendido": canal_pretendido,
            "destinatario_nome": payload.destinatario_nome,
            "destinatario_contato": payload.destinatario_contato,
            "status": status,
        },
    )
    existente = cur.fetchone()
    if existente:
        solicitacao_id, criado_em = existente
        status_retorno = "JA_SOLICITADO_RECENTEMENTE"
        return serializar_json_seguro({
            "ok": True,
            "mvp": "0.7L",
            "status": status_retorno,
            "status_solicitacao": status,
            "solicitacao_envio_id": solicitacao_id,
            "pdf_relatorio_id": payload.pdf_relatorio_id,
            "aprovacao_relatorio_id": aprovacao_relatorio_id,
            "obra_codigo": payload.obra_codigo,
            "canal_pretendido": canal_pretendido,
            "destinatario_nome": payload.destinatario_nome,
            "pdf_aprovado": True,
            "criado_em": criado_em,
            "idempotente": True,
            "resposta_telegram": _resposta_telegram_solicitacao_envio(
                payload.obra_codigo, payload.pdf_relatorio_id, solicitacao_id,
                canal_pretendido, payload.destinatario_nome, status_retorno,
            ),
            "flags_seguranca": dict(AGENTE_008_SEGURANCA_SOLICITACAO_ENVIO),
            **AGENTE_008_SEGURANCA_SOLICITACAO_ENVIO,
        })

    metadados = serializar_json_seguro({
        "mvp": "0.7L",
        "tipo": "SOLICITACAO_ENVIO_CONTROLADA_RELATORIO_SEMANAL",
        "flags_seguranca": dict(AGENTE_008_SEGURANCA_SOLICITACAO_ENVIO),
    })
    cur.execute(
        """
        INSERT INTO solicitacoes_envio_relatorios_semanais_obra (
            tenant_id, obra_codigo, area, pdf_relatorio_id,
            aprovacao_relatorio_id, canal_pretendido, destinatario_nome,
            destinatario_contato, assunto, mensagem, status, solicitado_por,
            solicitante_telegram_user_id, solicitante_telegram_username,
            solicitante_chat_id, canal_origem, validacao_pdf_status,
            pdf_aprovado, envio_executado, enviado_para_terceiros,
            gerou_link_publico, gerou_presigned_url, anexou_arquivo,
            alterou_rdo_oficial, alterou_cronograma, executou_rpa,
            sincronizou_openproject, alterou_minio, metadados
        ) VALUES (
            %(tenant_id)s, %(obra_codigo)s, %(area)s, %(pdf_relatorio_id)s,
            %(aprovacao_relatorio_id)s, %(canal_pretendido)s,
            %(destinatario_nome)s, %(destinatario_contato)s, %(assunto)s,
            %(mensagem)s, %(status)s, %(solicitado_por)s,
            %(solicitante_telegram_user_id)s,
            %(solicitante_telegram_username)s, %(solicitante_chat_id)s,
            'telegram', %(validacao_pdf_status)s, true, false, false, false,
            false, false, false, false, false, false, false, %(metadados)s
        )
        RETURNING id, criado_em;
        """,
        {
            "tenant_id": pdf[1],
            "obra_codigo": pdf[2],
            "area": pdf[3],
            "pdf_relatorio_id": pdf[0],
            "aprovacao_relatorio_id": aprovacao_relatorio_id,
            "canal_pretendido": canal_pretendido,
            "destinatario_nome": payload.destinatario_nome,
            "destinatario_contato": payload.destinatario_contato,
            "assunto": payload.assunto,
            "mensagem": payload.mensagem,
            "status": status,
            "solicitado_por": payload.solicitado_por,
            "solicitante_telegram_user_id": payload.solicitante_telegram_user_id,
            "solicitante_telegram_username": payload.solicitante_telegram_username,
            "solicitante_chat_id": payload.solicitante_chat_id,
            "validacao_pdf_status": pdf[4],
            "metadados": Json(metadados),
        },
    )
    solicitacao_id, criado_em = cur.fetchone()
    return serializar_json_seguro({
        "ok": True,
        "mvp": "0.7L",
        "status": status,
        "solicitacao_envio_id": solicitacao_id,
        "pdf_relatorio_id": payload.pdf_relatorio_id,
        "aprovacao_relatorio_id": aprovacao_relatorio_id,
        "obra_codigo": payload.obra_codigo,
        "area": pdf[3],
        "canal_pretendido": canal_pretendido,
        "destinatario_nome": payload.destinatario_nome,
        "pdf_aprovado": True,
        "criado_em": criado_em,
        "idempotente": False,
        "resposta_telegram": _resposta_telegram_solicitacao_envio(
            payload.obra_codigo, payload.pdf_relatorio_id, solicitacao_id,
            canal_pretendido, payload.destinatario_nome, status,
        ),
        "flags_seguranca": dict(AGENTE_008_SEGURANCA_SOLICITACAO_ENVIO),
        **AGENTE_008_SEGURANCA_SOLICITACAO_ENVIO,
    })


@app.post("/agentes/gestao-operacional/relatorio-semanal-executivo")
def relatorio_semanal_executivo(payload: GestaoOperacionalRelatorioSemanalRequest):
    try:
        data_inicio, data_fim = _periodo_relatorio_semanal(
            payload.data_inicio, payload.data_fim,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail={"message": str(exc)})

    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                return _gerar_relatorio_semanal_executivo(
                    cur, payload.obra_codigo, payload.area, data_inicio, data_fim,
                    payload.limite_itens, payload.salvar_relatorio,
                )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail={
            "message": "Erro ao gerar relatório semanal executivo consultivo.",
            "error": _texto_curto(exc, 200),
        })


@app.post("/agentes/gestao-operacional/exportar-relatorio-semanal")
def exportar_relatorio_semanal(
    payload: GestaoOperacionalExportarRelatorioSemanalRequest,
):
    if payload.formato.strip().upper() != "MARKDOWN":
        raise HTTPException(
            status_code=422,
            detail={
                "code": "FORMATO_NAO_SUPORTADO_NESTE_MVP",
                "message": "FORMATO_NAO_SUPORTADO_NESTE_MVP",
            },
        )
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                resultado = _exportar_relatorio_semanal_controlado(cur, payload)
        return serializar_json_seguro(resultado)
    except ValueError as exc:
        codigo = str(exc)
        status_code = 404 if codigo == "RELATORIO_SEMANAL_NAO_ENCONTRADO" else 422
        raise HTTPException(
            status_code=status_code,
            detail={"code": codigo, "message": codigo},
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail={
            "message": "Erro ao exportar relatório semanal executivo.",
            "error": _texto_curto(exc, 200),
        })


@app.post("/agentes/gestao-operacional/gerar-pdf-relatorio-semanal")
def gerar_pdf_relatorio_semanal(
    payload: GestaoOperacionalGerarPdfRelatorioSemanalRequest,
):
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                resultado = _gerar_pdf_relatorio_semanal_privado(cur, payload)
        return serializar_json_seguro(resultado)
    except ValueError as exc:
        codigo = str(exc)
        status_code = 404 if codigo in {
            "EXPORTACAO_RELATORIO_NAO_ENCONTRADA",
            "RELATORIO_SEMANAL_NAO_ENCONTRADO",
        } else 422
        raise HTTPException(
            status_code=status_code,
            detail={"code": codigo, "message": codigo},
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail={
            "message": "Erro ao gerar PDF privado do relatório semanal executivo.",
            "error": _texto_curto(exc, 200),
        })


@app.post("/agentes/gestao-operacional/aprovar-pdf-relatorio-semanal")
def aprovar_pdf_relatorio_semanal(
    payload: GestaoOperacionalAprovarPdfRelatorioSemanalRequest,
):
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                resultado = _aprovar_pdf_relatorio_semanal_controlado(cur, payload)
        return serializar_json_seguro(resultado)
    except ValueError as exc:
        codigo = str(exc)
        status_code = (
            404 if codigo == "PDF_RELATORIO_SEMANAL_NAO_ENCONTRADO" else 422
        )
        raise HTTPException(
            status_code=status_code,
            detail={"code": codigo, "message": codigo},
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail={
            "message": "Erro ao registrar decisão executiva sobre o PDF semanal.",
            "error": _texto_curto(exc, 200),
        })


@app.post("/agentes/gestao-operacional/solicitar-envio-relatorio-semanal")
def solicitar_envio_relatorio_semanal(
    payload: GestaoOperacionalSolicitarEnvioRelatorioSemanalRequest,
):
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                resultado = _solicitar_envio_relatorio_semanal_controlado(cur, payload)
        return serializar_json_seguro(resultado)
    except ValueError as exc:
        codigo = str(exc)
        status_code = 404 if codigo == "PDF_RELATORIO_NAO_ENCONTRADO" else 422
        raise HTTPException(
            status_code=status_code,
            detail={"code": codigo, "message": codigo},
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail={
            "message": "Erro ao registrar solicitação controlada de envio.",
            "error": _texto_curto(exc, 200),
        })


def _acao_operacional_dict(row: tuple[Any, ...]) -> dict[str, Any]:
    campos = (
        "id", "tenant_id", "obra_codigo", "area", "disciplina", "titulo",
        "descricao", "origem", "tipo_acao", "prioridade", "status",
        "responsavel", "prazo", "referencia_documento_id",
        "referencia_comando_id", "metadados", "criado_em", "atualizado_em",
        "concluido_em",
    )
    return serializar_json_seguro(dict(zip(campos, row)))


def _criar_acao_operacional(
    cur, payload: GestaoOperacionalCriarAcaoRequest, referencia_comando_id: int | None = None,
) -> dict[str, Any]:
    cur.execute(
        """
        INSERT INTO acoes_operacionais_obra (
            obra_codigo, area, disciplina, titulo, descricao, prioridade,
            responsavel, prazo, referencia_documento_id, referencia_comando_id
        ) VALUES (
            %(obra_codigo)s::text, %(area)s::text, %(disciplina)s::text,
            %(titulo)s::text, %(descricao)s::text, %(prioridade)s::text,
            %(responsavel)s::text, %(prazo)s::date,
            %(referencia_documento_id)s::bigint, %(referencia_comando_id)s::bigint
        )
        RETURNING id, tenant_id, obra_codigo, area, disciplina, titulo, descricao,
                  origem, tipo_acao, prioridade, status, responsavel, prazo,
                  referencia_documento_id, referencia_comando_id, metadados,
                  criado_em, atualizado_em, concluido_em;
        """,
        {**payload.model_dump(), "referencia_comando_id": referencia_comando_id},
    )
    acao = _acao_operacional_dict(cur.fetchone())
    return {
        "ok": True, "acao": acao,
        "resposta_telegram": (
            f"✅ Ação operacional #{acao['id']} registrada — {acao['obra_codigo']}\n"
            f"Título: {acao['titulo']}\nPrioridade: {acao['prioridade']}\nStatus: {acao['status']}\n\n"
            "Registro interno. Valide responsável e prazo com a equipe da obra."
        ),
        **AGENTE_008_SEGURANCA_REGISTRO_INTERNO,
    }


def _listar_acoes_operacionais(cur, payload: GestaoOperacionalListarAcoesRequest) -> dict[str, Any]:
    filtros = {
        "obra_codigo": payload.obra_codigo, "area": payload.area,
        "disciplina": payload.disciplina, "status": payload.status,
        "limite": payload.limite,
    }
    base = """
        FROM acoes_operacionais_obra
        WHERE obra_codigo = %(obra_codigo)s::text
          AND (%(area)s::text IS NULL OR area = %(area)s::text)
          AND (%(disciplina)s::text IS NULL OR disciplina = %(disciplina)s::text)
    """
    cur.execute("SELECT status, count(*) " + base + " GROUP BY status", filtros)
    totais_status = {status: 0 for status in ("ABERTA", "EM_ANDAMENTO", "CONCLUIDA", "CANCELADA")}
    totais_status.update(dict(cur.fetchall()))
    cur.execute("SELECT prioridade, count(*) " + base + " GROUP BY prioridade", filtros)
    totais_prioridade = {prioridade: 0 for prioridade in ("BAIXA", "MEDIA", "ALTA", "CRITICA")}
    totais_prioridade.update(dict(cur.fetchall()))
    cur.execute(
        """SELECT id, tenant_id, obra_codigo, area, disciplina, titulo, descricao,
                  origem, tipo_acao, prioridade, status, responsavel, prazo,
                  referencia_documento_id, referencia_comando_id, metadados,
                  criado_em, atualizado_em, concluido_em """ + base + """
          AND (%(status)s::text IS NULL OR status = %(status)s::text)
        ORDER BY CASE prioridade WHEN 'CRITICA' THEN 1 WHEN 'ALTA' THEN 2
                                 WHEN 'MEDIA' THEN 3 ELSE 4 END,
                 prazo NULLS LAST, criado_em, id
        LIMIT %(limite)s::int
        """,
        filtros,
    )
    acoes = [_acao_operacional_dict(row) for row in cur.fetchall()]
    linhas = [f"📋 Ações operacionais — {payload.obra_codigo}", ""]
    linhas.extend([
        f"Abertas: {totais_status['ABERTA']} | Em andamento: {totais_status['EM_ANDAMENTO']}",
        f"Concluídas: {totais_status['CONCLUIDA']} | Canceladas: {totais_status['CANCELADA']}", "",
    ])
    linhas.extend(
        f"- #{acao['id']} [{acao['prioridade']}] {acao['titulo']} ({acao['status']})"
        for acao in acoes
    )
    if not acoes:
        linhas.append("Nenhuma ação encontrada para os filtros informados.")
    return {
        "ok": True, "obra_codigo": payload.obra_codigo,
        "totais_por_status": totais_status, "totais_por_prioridade": totais_prioridade,
        "total_retornado": len(acoes), "acoes": acoes,
        "resposta_telegram": "\n".join(linhas), **AGENTE_008_SEGURANCA_CONSULTA,
    }


def _atualizar_acao_operacional(cur, payload: GestaoOperacionalAtualizarAcaoRequest) -> dict[str, Any]:
    observacao = payload.observacao.strip() if payload.observacao else None
    cur.execute(
        """SELECT id, tenant_id, obra_codigo, area, disciplina, titulo, descricao,
                  origem, tipo_acao, prioridade, status, responsavel, prazo,
                  referencia_documento_id, referencia_comando_id, metadados,
                  criado_em, atualizado_em, concluido_em
           FROM acoes_operacionais_obra
           WHERE id = %(acao_id)s::bigint AND obra_codigo = %(obra_codigo)s::text
           FOR UPDATE""",
        payload.model_dump(),
    )
    row_anterior = cur.fetchone()
    if row_anterior is None:
        raise HTTPException(status_code=404, detail="Ação operacional não encontrada na obra informada.")
    anterior = _acao_operacional_dict(row_anterior)
    campos_informados = payload.model_fields_set
    novo_status = payload.status if "status" in campos_informados else anterior["status"]
    nova_prioridade = payload.prioridade if "prioridade" in campos_informados else anterior["prioridade"]
    novo_responsavel = payload.responsavel if "responsavel" in campos_informados else anterior["responsavel"]
    novo_prazo = payload.prazo if "prazo" in campos_informados else anterior["prazo"]
    alteracoes = {
        "status": anterior["status"] != novo_status,
        "prioridade": anterior["prioridade"] != nova_prioridade,
        "responsavel": anterior["responsavel"] != novo_responsavel,
        "prazo": anterior["prazo"] != serializar_json_seguro(novo_prazo),
        "observacao": observacao is not None,
    }
    if not any(alteracoes.values()):
        return {
            "ok": True, "alterado": False, "acao": anterior,
            "resposta_telegram": f"ℹ️ Ação operacional #{anterior['id']} não teve alterações relevantes.",
            **AGENTE_008_SEGURANCA_REGISTRO_INTERNO,
        }
    cur.execute(
        """
        UPDATE acoes_operacionais_obra
        SET status = %(status)s::text,
            prioridade = %(prioridade)s::text,
            responsavel = %(responsavel)s::text,
            prazo = %(prazo)s::date,
            concluido_em = CASE
                WHEN %(status)s::text = 'CONCLUIDA' AND status <> 'CONCLUIDA' THEN now()
                WHEN %(status)s::text IN ('ABERTA', 'EM_ANDAMENTO') THEN NULL
                ELSE concluido_em END,
            metadados = CASE WHEN %(observacao)s::text IS NULL THEN metadados ELSE
                jsonb_set(
                    metadados || jsonb_build_object('ultima_observacao', %(observacao)s::text),
                    '{observacoes}',
                    COALESCE(metadados->'observacoes', '[]'::jsonb) ||
                        jsonb_build_array(jsonb_build_object(
                            'texto', %(observacao)s::text, 'registrado_em', now()
                        ))
                ) END
        WHERE id = %(acao_id)s::bigint AND obra_codigo = %(obra_codigo)s::text
        RETURNING id, tenant_id, obra_codigo, area, disciplina, titulo, descricao,
                  origem, tipo_acao, prioridade, status, responsavel, prazo,
                  referencia_documento_id, referencia_comando_id, metadados,
                  criado_em, atualizado_em, concluido_em;
        """,
        {
            **payload.model_dump(), "status": novo_status, "prioridade": nova_prioridade,
            "responsavel": novo_responsavel, "prazo": novo_prazo, "observacao": observacao,
        },
    )
    row = cur.fetchone()
    if row is None:
        raise HTTPException(status_code=404, detail="Ação operacional não encontrada na obra informada.")
    acao = _acao_operacional_dict(row)
    cur.execute(
        """INSERT INTO historico_acoes_operacionais_obra (
               tenant_id, obra_codigo, acao_id, tipo_evento,
               status_anterior, status_novo, prioridade_anterior, prioridade_nova,
               responsavel_anterior, responsavel_novo, prazo_anterior, prazo_novo,
               observacao, registrado_por, metadados
           ) VALUES (
               %(tenant_id)s, %(obra_codigo)s, %(acao_id)s, 'ATUALIZACAO',
               %(status_anterior)s, %(status_novo)s, %(prioridade_anterior)s, %(prioridade_nova)s,
               %(responsavel_anterior)s, %(responsavel_novo)s, %(prazo_anterior)s::date,
               %(prazo_novo)s::date, %(observacao)s, 'AGENTE_008_GESTAO_OPERACIONAL_OBRA',
               %(metadados)s
           )""",
        {
            "tenant_id": acao["tenant_id"], "obra_codigo": acao["obra_codigo"],
            "acao_id": acao["id"], "status_anterior": anterior["status"],
            "status_novo": acao["status"], "prioridade_anterior": anterior["prioridade"],
            "prioridade_nova": acao["prioridade"], "responsavel_anterior": anterior["responsavel"],
            "responsavel_novo": acao["responsavel"], "prazo_anterior": anterior["prazo"],
            "prazo_novo": acao["prazo"], "observacao": observacao,
            "metadados": Json({"campos_alterados": [campo for campo, mudou in alteracoes.items() if mudou]}),
        },
    )
    return {
        "ok": True, "alterado": True, "acao": acao,
        "resposta_telegram": (
            f"✅ Ação operacional #{acao['id']} atualizada.\n"
            f"Status: {acao['status']} | Prioridade: {acao['prioridade']}\n"
            f"Responsável: {acao['responsavel'] or 'não definido'} | Prazo: {acao['prazo'] or 'não definido'}\n"
            "Registro interno; nenhuma alteração foi feita em cronograma, RDO ou OpenProject."
        ),
        **AGENTE_008_SEGURANCA_REGISTRO_INTERNO,
    }


def _detalhar_acao_operacional(cur, payload: GestaoOperacionalDetalheAcaoRequest) -> dict[str, Any]:
    cur.execute(
        """SELECT id, tenant_id, obra_codigo, area, disciplina, titulo, descricao,
                  origem, tipo_acao, prioridade, status, responsavel, prazo,
                  referencia_documento_id, referencia_comando_id, metadados,
                  criado_em, atualizado_em, concluido_em
           FROM acoes_operacionais_obra
           WHERE id = %(acao_id)s::bigint AND obra_codigo = %(obra_codigo)s::text""",
        payload.model_dump(),
    )
    row = cur.fetchone()
    if row is None:
        raise HTTPException(status_code=404, detail="Ação operacional não encontrada na obra informada.")
    acao = _acao_operacional_dict(row)
    historico: list[dict[str, Any]] = []
    if payload.incluir_historico:
        cur.execute(
            """SELECT id, tenant_id, obra_codigo, acao_id, tipo_evento,
                      status_anterior, status_novo, prioridade_anterior, prioridade_nova,
                      responsavel_anterior, responsavel_novo, prazo_anterior, prazo_novo,
                      observacao, origem, registrado_por, metadados, criado_em
               FROM historico_acoes_operacionais_obra
               WHERE acao_id = %(acao_id)s::bigint AND obra_codigo = %(obra_codigo)s::text
               ORDER BY criado_em, id""",
            payload.model_dump(),
        )
        campos = ("id", "tenant_id", "obra_codigo", "acao_id", "tipo_evento",
                  "status_anterior", "status_novo", "prioridade_anterior", "prioridade_nova",
                  "responsavel_anterior", "responsavel_novo", "prazo_anterior", "prazo_novo",
                  "observacao", "origem", "registrado_por", "metadados", "criado_em")
        historico = [serializar_json_seguro(dict(zip(campos, item))) for item in cur.fetchall()]
    return {
        "ok": True, "acao": acao, "historico": historico,
        "resposta_telegram": (
            f"🔎 Ação #{acao['id']} — {acao['titulo']}\n"
            f"Status: {acao['status']} | Prioridade: {acao['prioridade']}\n"
            f"Responsável: {acao['responsavel'] or 'não definido'} | Prazo: {acao['prazo'] or 'não definido'}\n"
            f"Eventos no histórico: {len(historico)}"
        ),
        **AGENTE_008_SEGURANCA_CONSULTA,
    }


@app.post("/agentes/gestao-operacional/criar-acao-operacional")
def criar_acao_operacional(payload: GestaoOperacionalCriarAcaoRequest):
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            return _criar_acao_operacional(cur, payload)


@app.post("/agentes/gestao-operacional/registrar-revisao-documental")
def registrar_revisao_documental(payload: GestaoOperacionalRegistrarRevisaoDocumentalRequest):
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                return _registrar_revisao_documental(cur, payload)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=409 if "uq_revisoes_documentais_objeto_minio" in str(exc) else 500, detail={"message": "Erro ao registrar revisão documental.", "error": _texto_curto(exc, 200)})


@app.post("/agentes/gestao-operacional/revisoes-documentais")
def revisoes_documentais(payload: GestaoOperacionalRevisoesDocumentaisRequest):
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                return _listar_revisoes_documentais(cur, payload)
    except Exception as exc:
        raise HTTPException(status_code=500, detail={"message": "Erro ao consultar revisões documentais.", "error": _texto_curto(exc, 200)})


@app.post("/agentes/gestao-operacional/importar-revisoes-documentais-pendentes")
def importar_revisoes_documentais_pendentes(payload: GestaoOperacionalImportarRevisoesDocumentaisRequest):
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                return _importar_revisoes_documentais(cur, payload)
    except Exception as exc:
        raise HTTPException(status_code=500, detail={"message": "Erro ao importar revisões documentais pendentes.", "error": _texto_curto(exc, 200)})


@app.post("/agentes/gestao-operacional/aprovar-revisao-documental")
def aprovar_revisao_documental(payload: GestaoOperacionalAprovarRevisaoDocumentalRequest):
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                return _aprovar_revisao_documental(cur, payload)
    except LookupError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail={
            "message": "Erro ao registrar decisão técnica documental.",
            "error": _texto_curto(exc, 200),
        }) from exc


@app.post("/agentes/gestao-operacional/liberar-revisao-documental-campo")
def liberar_revisao_documental_campo(payload: GestaoOperacionalLiberarRevisaoDocumentalCampoRequest):
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                return _liberar_revisao_documental_campo(cur, payload)
    except LookupError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=409, detail={"codigo": str(exc), "message": str(exc)}) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail={"message": "Erro ao registrar liberação documental de campo.", "error": _texto_curto(exc, 200)}) from exc


@app.post("/agentes/gestao-operacional/liberacoes-revisoes-documentais-campo")
def liberacoes_revisoes_documentais_campo(payload: GestaoOperacionalLiberacoesRevisoesDocumentaisCampoRequest):
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                return _listar_liberacoes_revisoes_campo(cur, payload)
    except Exception as exc:
        raise HTTPException(status_code=500, detail={"message": "Erro ao consultar liberações documentais de campo.", "error": _texto_curto(exc, 200)}) from exc


@app.post("/agentes/gestao-operacional/acoes-operacionais")
def listar_acoes_operacionais(payload: GestaoOperacionalListarAcoesRequest):
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            return _listar_acoes_operacionais(cur, payload)


@app.post("/agentes/gestao-operacional/atualizar-acao-operacional")
def atualizar_acao_operacional(payload: GestaoOperacionalAtualizarAcaoRequest):
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            return _atualizar_acao_operacional(cur, payload)


@app.post("/agentes/gestao-operacional/detalhe-acao-operacional")
def detalhe_acao_operacional(payload: GestaoOperacionalDetalheAcaoRequest):
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            return _detalhar_acao_operacional(cur, payload)


@app.post("/agentes/gestao-operacional/processar-comando")
def processar_comando_agente_gestao_operacional(
    payload: Optional[ProcessarComandoGestaoOperacionalRequest] = None,
):
    id_comando = payload.id_comando if payload else None
    if id_comando:
        try:
            id_comando = str(uuid.UUID(id_comando))
        except ValueError as exc:
            raise HTTPException(
                status_code=422,
                detail={"message": "id_comando inválido. Use um UUID válido.", "error": str(exc)},
            )

    comando: Optional[dict[str, Any]] = None
    try:
        conn = get_db_connection()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT ce.id, ce.id_comando, ce.obra_codigo, ce.tipo_comando,
                           ce.payload_comando, et.chat_id
                    FROM comandos_executivos AS ce
                    LEFT JOIN eventos_telegram AS et ON et.id = ce.evento_telegram_id
                    WHERE ce.agente_destino = 'AGENTE_008_GESTAO_OPERACIONAL_OBRA'
                      AND ce.tipo_comando IN (
                          'CONSULTAR_DOCUMENTOS_OBRA_RESUMO',
                          'CONSULTAR_DOCUMENTOS_OBRA_INDEXADOS',
                          'ANALISAR_DOCUMENTO_OBRA',
                          'CLASSIFICAR_DOCUMENTOS_OBRA',
                          'CONSULTAR_DOCUMENTOS_CLASSIFICADOS',
                          'GERAR_RELATORIO_DOCUMENTAL_OBRA',
                          'AVALIAR_RISCOS_DOCUMENTAIS_OBRA',
                          'CONSULTAR_ULTIMA_REVISAO_DOCUMENTAL',
                          'VALIDAR_DOCUMENTO_CAMPO',
                          'GERAR_PLANO_SANEAMENTO_DOCUMENTAL',
                          'CRIAR_PENDENCIA_DOCUMENTAL',
                          'GERAR_DIAGNOSTICO_OPERACIONAL_OBRA',
                          'GERAR_PLANO_OPERACIONAL_OBRA',
                          'GERAR_RESUMO_EXECUTIVO_OPERACIONAL_OBRA',
                          'GERAR_BRIEFING_DIARIO_OBRA',
                          'GERAR_RELATORIO_SEMANAL_EXECUTIVO_OBRA',
                          'EXPORTAR_RELATORIO_SEMANAL_EXECUTIVO_OBRA',
                          'GERAR_PDF_RELATORIO_SEMANAL_EXECUTIVO_OBRA',
                          'APROVAR_PDF_RELATORIO_SEMANAL_EXECUTIVO_OBRA',
                          'SOLICITAR_ENVIO_RELATORIO_SEMANAL_EXECUTIVO_OBRA',
                          'CRIAR_ACAO_OPERACIONAL_OBRA',
                          'LISTAR_ACOES_OPERACIONAIS_OBRA',
                          'ATUALIZAR_ACAO_OPERACIONAL_OBRA',
                          'DETALHAR_ACAO_OPERACIONAL_OBRA',
                          'CONSULTAR_REVISOES_DOCUMENTAIS_OBRA',
                          'IMPORTAR_REVISOES_DOCUMENTAIS_PENDENTES_OBRA',
                          'APROVAR_REVISAO_DOCUMENTAL_OBRA',
                          'LIBERAR_REVISAO_DOCUMENTAL_CAMPO_OBRA',
                          'CONSULTAR_LIBERACOES_REVISOES_DOCUMENTAIS_CAMPO_OBRA'
                      )
                      AND ce.status = 'PENDENTE'
                      AND (%(id_comando)s::uuid IS NULL OR ce.id_comando = %(id_comando)s::uuid)
                    ORDER BY ce.criado_em, ce.id
                    LIMIT 1
                    FOR UPDATE OF ce;
                    """,
                    {"id_comando": id_comando},
                )
                row = cur.fetchone()
                if row is None:
                    conn.rollback()
                    return {
                        "ok": True,
                        "agente": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
                        "mvp": "0.7K",
                        "status": "SEM_COMANDO_PENDENTE",
                        "message": (
                            "Nenhum comando PENDENTE para "
                            "AGENTE_008_GESTAO_OPERACIONAL_OBRA encontrado."
                        ),
                    }

                comando = {
                    "id": row[0],
                    "id_comando": row[1],
                    "obra_codigo": row[2],
                    "tipo_comando": row[3],
                    "payload_comando": row[4],
                    "telegram_chat_id": row[5],
                }
                if comando["tipo_comando"] == "CONSULTAR_DOCUMENTOS_OBRA_RESUMO":
                    resultado = _resultado_resumo_documental(cur, comando["obra_codigo"])
                elif comando["tipo_comando"] == "CONSULTAR_DOCUMENTOS_OBRA_INDEXADOS":
                    resultado = _resultado_documentos_indexados(
                        cur, comando["obra_codigo"], comando["payload_comando"]
                    )
                elif comando["tipo_comando"] == "ANALISAR_DOCUMENTO_OBRA":
                    resultado = _resultado_analise_documental(
                        cur, comando["obra_codigo"], comando["payload_comando"]
                    )
                elif comando["tipo_comando"] == "CLASSIFICAR_DOCUMENTOS_OBRA":
                    resultado = _resultado_classificacao_documental(
                        cur, comando["obra_codigo"], comando["payload_comando"]
                    )
                elif comando["tipo_comando"] == "GERAR_RELATORIO_DOCUMENTAL_OBRA":
                    payload_relatorio = comando["payload_comando"] or {}
                    resultado = _gerar_relatorio_documental(
                        cur,
                        comando["obra_codigo"],
                        bool(payload_relatorio.get("incluir_amostras", True)),
                        int(payload_relatorio.get("limite_amostras", 5)),
                    )
                elif comando["tipo_comando"] == "AVALIAR_RISCOS_DOCUMENTAIS_OBRA":
                    payload_riscos = comando["payload_comando"] or {}
                    resultado = _avaliar_riscos_documentais(
                        cur, comando["obra_codigo"], payload_riscos.get("area"),
                        payload_riscos.get("disciplina"),
                        int(payload_riscos.get("limite_amostras", 10)),
                    )
                elif comando["tipo_comando"] == "CONSULTAR_ULTIMA_REVISAO_DOCUMENTAL":
                    payload_revisao = comando["payload_comando"] or {}
                    resultado = _ultima_revisao_documental(
                        cur, comando["obra_codigo"], payload_revisao.get("disciplina"),
                        payload_revisao.get("area"),
                        int(payload_revisao.get("limite_candidatos", 10)),
                    )
                elif comando["tipo_comando"] == "VALIDAR_DOCUMENTO_CAMPO":
                    resultado = _validar_documento_campo(
                        cur, comando["obra_codigo"],
                        int((comando["payload_comando"] or {})["documento_id"]),
                    )
                elif comando["tipo_comando"] == "GERAR_PLANO_SANEAMENTO_DOCUMENTAL":
                    payload_plano = comando["payload_comando"] or {}
                    try:
                        with conn.transaction():
                            resultado = _gerar_plano_saneamento_documental(
                                cur, comando["obra_codigo"], payload_plano.get("area"),
                                payload_plano.get("disciplina"),
                                int(payload_plano.get("limite_acoes", 10)),
                            )
                    except Exception as exc:
                        resultado = {
                            "ok": False,
                            "erro_controlado": f"Plano de saneamento não gerado: {exc}",
                            "resposta_telegram": "Não foi possível gerar o plano de saneamento documental.",
                            **AGENTE_008_SEGURANCA_CONSULTA,
                        }
                elif comando["tipo_comando"] == "CRIAR_PENDENCIA_DOCUMENTAL":
                    payload_pendencia = comando["payload_comando"] or {}
                    try:
                        with conn.transaction():
                            resultado = _criar_pendencia_documental(
                                cur, comando["obra_codigo"], payload_pendencia.get("area"),
                                payload_pendencia.get("disciplina"),
                                str(payload_pendencia.get("motivo", "RISCO_DOCUMENTAL")),
                                payload_pendencia.get("descricao"),
                            )
                    except Exception as exc:
                        resultado = _erro_pendencia_documental(
                            comando["obra_codigo"],
                            f"Erro controlado ao criar pendência; a pendência não foi criada: {exc}",
                        )
                elif comando["tipo_comando"] == "GERAR_DIAGNOSTICO_OPERACIONAL_OBRA":
                    payload_diagnostico = comando["payload_comando"] or {}
                    resultado = _gerar_diagnostico_operacional(
                        cur, comando["obra_codigo"], payload_diagnostico.get("area"),
                        bool(payload_diagnostico.get("incluir_documentos", True)),
                        bool(payload_diagnostico.get("incluir_pendencias", True)),
                        bool(payload_diagnostico.get("incluir_restricoes", True)),
                    )
                elif comando["tipo_comando"] == "GERAR_PLANO_OPERACIONAL_OBRA":
                    payload_plano_operacional = comando["payload_comando"] or {}
                    try:
                        resultado = _gerar_plano_operacional(
                            cur, comando["obra_codigo"], payload_plano_operacional.get("area"),
                            max(1, min(int(payload_plano_operacional.get("limite_acoes", 10)), 10)),
                        )
                    except Exception as exc:
                        resultado = {
                            "ok": False,
                            "erro_controlado": f"Plano operacional não gerado: {_texto_curto(exc)}",
                            "resposta_telegram": "Não foi possível gerar o plano operacional consultivo.",
                            **AGENTE_008_SEGURANCA_CONSULTA,
                        }
                elif comando["tipo_comando"] == "GERAR_RESUMO_EXECUTIVO_OPERACIONAL_OBRA":
                    payload_resumo = comando["payload_comando"] or {}
                    try:
                        resultado = _gerar_resumo_executivo_operacional(
                            cur, comando["obra_codigo"], payload_resumo.get("area"),
                            max(1, min(int(payload_resumo.get("limite_itens", 10)), 50)),
                        )
                    except Exception as exc:
                        resultado = {
                            "ok": False,
                            "erro_controlado": f"Resumo executivo não gerado: {_texto_curto(exc)}",
                            "resposta_telegram": "Não foi possível gerar o resumo executivo operacional consultivo.",
                            **AGENTE_008_SEGURANCA_CONSULTA,
                        }
                elif comando["tipo_comando"] == "GERAR_BRIEFING_DIARIO_OBRA":
                    payload_briefing = comando["payload_comando"] or {}
                    try:
                        resultado = _gerar_briefing_diario(
                            cur, comando["obra_codigo"], payload_briefing.get("area"),
                            max(1, min(int(payload_briefing.get("limite_itens", 10)), 50)),
                            bool(payload_briefing.get("incluir_acoes", True)),
                            bool(payload_briefing.get("incluir_documentos", True)),
                            bool(payload_briefing.get("incluir_historico", True)),
                        )
                    except Exception as exc:
                        resultado = {
                            "ok": False,
                            "erro_controlado": f"Briefing diário não gerado: {_texto_curto(exc)}",
                            "resposta_telegram": "Não foi possível gerar o briefing diário executivo consultivo.",
                            **AGENTE_008_SEGURANCA_CONSULTA,
                        }
                elif comando["tipo_comando"] == "GERAR_RELATORIO_SEMANAL_EXECUTIVO_OBRA":
                    payload_relatorio_semanal = comando["payload_comando"] or {}
                    try:
                        data_inicio_informada = payload_relatorio_semanal.get("data_inicio")
                        data_fim_informada = payload_relatorio_semanal.get("data_fim")
                        for nome_data, valor_data in (
                            ("data_inicio", data_inicio_informada),
                            ("data_fim", data_fim_informada),
                        ):
                            if isinstance(valor_data, str) and not re.fullmatch(
                                r"\d{4}-\d{2}-\d{2}", valor_data,
                            ):
                                raise ValueError(f"{nome_data} inválida. Use ISO YYYY-MM-DD.")
                        data_inicio_convertida = _data_operacional(data_inicio_informada)
                        data_fim_convertida = _data_operacional(data_fim_informada)
                        if data_inicio_informada is not None and data_inicio_convertida is None:
                            raise ValueError("data_inicio inválida. Use ISO YYYY-MM-DD.")
                        if data_fim_informada is not None and data_fim_convertida is None:
                            raise ValueError("data_fim inválida. Use ISO YYYY-MM-DD.")
                        data_inicio, data_fim = _periodo_relatorio_semanal(
                            data_inicio_convertida, data_fim_convertida,
                        )
                        resultado = _gerar_relatorio_semanal_executivo(
                            cur, comando["obra_codigo"], payload_relatorio_semanal.get("area"),
                            data_inicio, data_fim,
                            max(1, min(int(payload_relatorio_semanal.get("limite_itens", 10)), 50)),
                            bool(payload_relatorio_semanal.get("salvar_relatorio", True)),
                        )
                    except Exception as exc:
                        resultado = {
                            "ok": False,
                            "erro_controlado": f"Relatório semanal não gerado: {_texto_curto(exc)}",
                            "resposta_telegram": "Não foi possível gerar o relatório semanal executivo consultivo.",
                            **AGENTE_008_SEGURANCA_CONSULTA,
                        }
                elif comando["tipo_comando"] == "EXPORTAR_RELATORIO_SEMANAL_EXECUTIVO_OBRA":
                    payload_exportacao = {
                        **(comando["payload_comando"] or {}),
                        "obra_codigo": comando["obra_codigo"],
                    }
                    try:
                        resultado = _exportar_relatorio_semanal_controlado(
                            cur,
                            GestaoOperacionalExportarRelatorioSemanalRequest(
                                **payload_exportacao
                            ),
                        )
                    except Exception as exc:
                        codigo_erro = str(exc)
                        resultado = {
                            "ok": False,
                            "codigo_erro": codigo_erro,
                            "erro_controlado": (
                                "Exportação do relatório semanal não gerada: "
                                f"{_texto_curto(exc)}"
                            ),
                            "resposta_telegram": (
                                "Não foi possível exportar o relatório semanal em Markdown. "
                                f"Código: {codigo_erro}. Nenhuma ação externa foi realizada."
                            ),
                            "flags_seguranca": dict(AGENTE_008_SEGURANCA_EXPORTACAO),
                            **AGENTE_008_SEGURANCA_EXPORTACAO,
                        }
                elif comando["tipo_comando"] == "GERAR_PDF_RELATORIO_SEMANAL_EXECUTIVO_OBRA":
                    payload_pdf = {
                        **(comando["payload_comando"] or {}),
                        "obra_codigo": comando["obra_codigo"],
                    }
                    try:
                        resultado = _gerar_pdf_relatorio_semanal_privado(
                            cur,
                            GestaoOperacionalGerarPdfRelatorioSemanalRequest(**payload_pdf),
                        )
                    except Exception as exc:
                        codigo_erro = str(exc)
                        resultado = {
                            "ok": False,
                            "codigo_erro": codigo_erro,
                            "erro_controlado": (
                                "PDF privado do relatório semanal não gerado: "
                                f"{_texto_curto(exc)}"
                            ),
                            "resposta_telegram": (
                                "Não foi possível gerar o PDF privado do relatório semanal. "
                                f"Código: {codigo_erro}. Nenhuma ação externa foi realizada."
                            ),
                            "flags_seguranca": dict(AGENTE_008_SEGURANCA_PDF),
                            **AGENTE_008_SEGURANCA_PDF,
                        }
                elif comando["tipo_comando"] == "APROVAR_PDF_RELATORIO_SEMANAL_EXECUTIVO_OBRA":
                    payload_aprovacao = {
                        **(comando["payload_comando"] or {}),
                        "obra_codigo": comando["obra_codigo"],
                    }
                    try:
                        if payload_aprovacao.get("pdf_relatorio_id") is None:
                            raise ValueError(
                                "Informe o ID do PDF. Exemplo: aprovar pdf #5"
                            )
                        with conn.transaction():
                            resultado = _aprovar_pdf_relatorio_semanal_controlado(
                                cur,
                                GestaoOperacionalAprovarPdfRelatorioSemanalRequest(
                                    **payload_aprovacao
                                ),
                            )
                    except Exception as exc:
                        codigo_erro = str(exc)
                        resultado = {
                            "ok": False,
                            "codigo_erro": codigo_erro,
                            "erro_controlado": (
                                "Decisão sobre o PDF semanal não registrada: "
                                f"{_texto_curto(exc)}"
                            ),
                            "resposta_telegram": (
                                codigo_erro
                                if codigo_erro.startswith("Informe o ID do PDF.")
                                else (
                                    "Não foi possível registrar a decisão sobre o PDF "
                                    f"semanal. Código: {codigo_erro}. Nenhuma ação "
                                    "externa foi realizada."
                                )
                            ),
                            "flags_seguranca": dict(
                                AGENTE_008_SEGURANCA_APROVACAO_PDF
                            ),
                            **AGENTE_008_SEGURANCA_APROVACAO_PDF,
                        }
                elif comando["tipo_comando"] == "SOLICITAR_ENVIO_RELATORIO_SEMANAL_EXECUTIVO_OBRA":
                    payload_solicitacao = {
                        **(comando["payload_comando"] or {}),
                        "obra_codigo": comando["obra_codigo"],
                    }
                    try:
                        if payload_solicitacao.get("pdf_relatorio_id") is None:
                            raise ValueError(
                                "Informe o ID do PDF. Exemplo: solicitar envio do pdf "
                                "#5 para Sr. Charles por email"
                            )
                        with conn.transaction():
                            resultado = _solicitar_envio_relatorio_semanal_controlado(
                                cur,
                                GestaoOperacionalSolicitarEnvioRelatorioSemanalRequest(
                                    **payload_solicitacao
                                ),
                            )
                    except Exception as exc:
                        codigo_erro = str(exc)
                        resultado = {
                            "ok": False,
                            "codigo_erro": codigo_erro,
                            "erro_controlado": (
                                "Solicitação controlada de envio não registrada: "
                                f"{_texto_curto(exc)}"
                            ),
                            "resposta_telegram": (
                                codigo_erro
                                if codigo_erro.startswith("Informe o ID do PDF.")
                                else (
                                    "Não foi possível registrar a solicitação de envio. "
                                    f"Código: {codigo_erro}. Nenhum envio foi executado."
                                )
                            ),
                            "flags_seguranca": dict(
                                AGENTE_008_SEGURANCA_SOLICITACAO_ENVIO
                            ),
                            **AGENTE_008_SEGURANCA_SOLICITACAO_ENVIO,
                        }
                elif comando["tipo_comando"] == "CRIAR_ACAO_OPERACIONAL_OBRA":
                    dados_acao = {
                        **(comando["payload_comando"] or {}),
                        "obra_codigo": comando["obra_codigo"],
                    }
                    resultado = _criar_acao_operacional(
                        cur, GestaoOperacionalCriarAcaoRequest(**dados_acao), comando["id"]
                    )
                elif comando["tipo_comando"] == "LISTAR_ACOES_OPERACIONAIS_OBRA":
                    dados_lista = {
                        **(comando["payload_comando"] or {}),
                        "obra_codigo": comando["obra_codigo"],
                    }
                    resultado = _listar_acoes_operacionais(
                        cur, GestaoOperacionalListarAcoesRequest(**dados_lista)
                    )
                elif comando["tipo_comando"] == "ATUALIZAR_ACAO_OPERACIONAL_OBRA":
                    dados_atualizacao = {
                        **(comando["payload_comando"] or {}),
                        "obra_codigo": comando["obra_codigo"],
                    }
                    resultado = _atualizar_acao_operacional(
                        cur, GestaoOperacionalAtualizarAcaoRequest(**dados_atualizacao)
                    )
                elif comando["tipo_comando"] == "DETALHAR_ACAO_OPERACIONAL_OBRA":
                    dados_detalhe = {
                        **(comando["payload_comando"] or {}),
                        "obra_codigo": comando["obra_codigo"],
                    }
                    resultado = _detalhar_acao_operacional(
                        cur, GestaoOperacionalDetalheAcaoRequest(**dados_detalhe)
                    )
                elif comando["tipo_comando"] == "CONSULTAR_REVISOES_DOCUMENTAIS_OBRA":
                    dados_revisoes = {**(comando["payload_comando"] or {}), "obra_codigo": comando["obra_codigo"]}
                    try:
                        with conn.transaction():
                            resultado = _listar_revisoes_documentais(
                                cur, GestaoOperacionalRevisoesDocumentaisRequest(**dados_revisoes)
                            )
                    except Exception as exc:
                        resultado = {"ok": False, "erro_controlado": _texto_curto(exc), "resposta_telegram": "Não foi possível consultar as revisões documentais. Nenhuma alteração externa foi realizada.", "flags_seguranca": dict(AGENTE_008_SEGURANCA_CONTROLE_DOCUMENTAL), **AGENTE_008_SEGURANCA_CONTROLE_DOCUMENTAL}
                elif comando["tipo_comando"] == "IMPORTAR_REVISOES_DOCUMENTAIS_PENDENTES_OBRA":
                    dados_importacao = {**(comando["payload_comando"] or {}), "obra_codigo": comando["obra_codigo"]}
                    try:
                        with conn.transaction():
                            resultado = _importar_revisoes_documentais(
                                cur, GestaoOperacionalImportarRevisoesDocumentaisRequest(**dados_importacao)
                            )
                    except Exception as exc:
                        resultado = {"ok": False, "erro_controlado": _texto_curto(exc), "resposta_telegram": "Não foi possível importar as revisões documentais pendentes. Nenhum arquivo foi alterado.", "flags_seguranca": dict(AGENTE_008_SEGURANCA_CONTROLE_DOCUMENTAL), **AGENTE_008_SEGURANCA_CONTROLE_DOCUMENTAL}
                elif comando["tipo_comando"] == "APROVAR_REVISAO_DOCUMENTAL_OBRA":
                    dados_decisao = {**(comando["payload_comando"] or {}), "obra_codigo": comando["obra_codigo"]}
                    campos_payload = set(GestaoOperacionalAprovarRevisaoDocumentalRequest.model_fields)
                    dados_decisao = {chave: valor for chave, valor in dados_decisao.items() if chave in campos_payload}
                    try:
                        with conn.transaction():
                            resultado = _aprovar_revisao_documental(
                                cur, GestaoOperacionalAprovarRevisaoDocumentalRequest(**dados_decisao)
                            )
                    except Exception as exc:
                        resultado = {
                            "ok": False, "erro_controlado": _texto_curto(exc),
                            "resposta_telegram": "Não foi possível registrar a decisão técnica documental. Nenhuma alteração externa foi realizada.",
                            "flags_seguranca": dict(AGENTE_008_SEGURANCA_APROVACAO_DOCUMENTAL),
                            **AGENTE_008_SEGURANCA_APROVACAO_DOCUMENTAL,
                        }
                elif comando["tipo_comando"] == "LIBERAR_REVISAO_DOCUMENTAL_CAMPO_OBRA":
                    dados_liberacao = {**(comando["payload_comando"] or {}), "obra_codigo": comando["obra_codigo"]}
                    campos_payload = set(GestaoOperacionalLiberarRevisaoDocumentalCampoRequest.model_fields)
                    dados_liberacao = {chave: valor for chave, valor in dados_liberacao.items() if chave in campos_payload}
                    try:
                        with conn.transaction():
                            resultado = _liberar_revisao_documental_campo(cur, GestaoOperacionalLiberarRevisaoDocumentalCampoRequest(**dados_liberacao))
                    except Exception as exc:
                        resultado = {"ok": False, "erro_controlado": _texto_curto(exc), "resposta_telegram": "Não foi possível registrar a liberação documental de campo. Nenhuma alteração externa foi realizada.", "flags_seguranca": dict(AGENTE_008_SEGURANCA_LIBERACAO_CAMPO), **AGENTE_008_SEGURANCA_LIBERACAO_CAMPO}
                elif comando["tipo_comando"] == "CONSULTAR_LIBERACOES_REVISOES_DOCUMENTAIS_CAMPO_OBRA":
                    dados_consulta = {**(comando["payload_comando"] or {}), "obra_codigo": comando["obra_codigo"]}
                    campos_payload = set(GestaoOperacionalLiberacoesRevisoesDocumentaisCampoRequest.model_fields)
                    dados_consulta = {chave: valor for chave, valor in dados_consulta.items() if chave in campos_payload}
                    try:
                        with conn.transaction():
                            resultado = _listar_liberacoes_revisoes_campo(cur, GestaoOperacionalLiberacoesRevisoesDocumentaisCampoRequest(**dados_consulta))
                    except Exception as exc:
                        resultado = {"ok": False, "erro_controlado": _texto_curto(exc), "resposta_telegram": "Não foi possível consultar as liberações documentais de campo.", "flags_seguranca": dict(AGENTE_008_SEGURANCA_LIBERACAO_CAMPO), **AGENTE_008_SEGURANCA_LIBERACAO_CAMPO}
                else:
                    resultado = _resultado_consulta_documentos_classificados(
                        cur, comando["obra_codigo"], comando["payload_comando"]
                    )

                status_resultado = (
                    "ERRO"
                    if comando["tipo_comando"] in {
                        "CRIAR_PENDENCIA_DOCUMENTAL",
                        "GERAR_PLANO_SANEAMENTO_DOCUMENTAL",
                        "GERAR_PLANO_OPERACIONAL_OBRA",
                        "GERAR_RESUMO_EXECUTIVO_OPERACIONAL_OBRA",
                        "GERAR_BRIEFING_DIARIO_OBRA",
                        "GERAR_RELATORIO_SEMANAL_EXECUTIVO_OBRA",
                        "EXPORTAR_RELATORIO_SEMANAL_EXECUTIVO_OBRA",
                        "GERAR_PDF_RELATORIO_SEMANAL_EXECUTIVO_OBRA",
                        "APROVAR_PDF_RELATORIO_SEMANAL_EXECUTIVO_OBRA",
                        "SOLICITAR_ENVIO_RELATORIO_SEMANAL_EXECUTIVO_OBRA",
                        "CONSULTAR_REVISOES_DOCUMENTAIS_OBRA",
                        "IMPORTAR_REVISOES_DOCUMENTAIS_PENDENTES_OBRA",
                        "APROVAR_REVISAO_DOCUMENTAL_OBRA",
                        "LIBERAR_REVISAO_DOCUMENTAL_CAMPO_OBRA",
                        "CONSULTAR_LIBERACOES_REVISOES_DOCUMENTAIS_CAMPO_OBRA",
                    }
                    and resultado.get("ok") is False
                    else "CONCLUIDO"
                )
                cur.execute(
                    """
                    UPDATE comandos_executivos
                    SET status = %(status)s,
                        executado_por = 'AGENTE_008_GESTAO_OPERACIONAL_OBRA',
                        executado_em = NOW(),
                        resultado = %(resultado)s,
                        mensagem_erro = %(mensagem_erro)s,
                        atualizado_em = NOW()
                    WHERE id = %(id)s AND status = 'PENDENTE'
                    RETURNING status;
                    """,
                    {
                        "id": comando["id"],
                        "resultado": Json(serializar_json_seguro(resultado)),
                        "status": status_resultado,
                        "mensagem_erro": resultado.get("erro_controlado"),
                    },
                )
                atualizado = cur.fetchone()
                if atualizado is None:
                    raise RuntimeError("Comando não pôde ser atualizado para o status final.")
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()
    except Exception as exc:
        if comando is not None:
            if comando["tipo_comando"] == "EXPORTAR_RELATORIO_SEMANAL_EXECUTIVO_OBRA":
                flags_erro = AGENTE_008_SEGURANCA_EXPORTACAO
            elif comando["tipo_comando"] == "GERAR_PDF_RELATORIO_SEMANAL_EXECUTIVO_OBRA":
                flags_erro = AGENTE_008_SEGURANCA_PDF
            elif comando["tipo_comando"] == "APROVAR_PDF_RELATORIO_SEMANAL_EXECUTIVO_OBRA":
                flags_erro = AGENTE_008_SEGURANCA_APROVACAO_PDF
            elif comando["tipo_comando"] == "SOLICITAR_ENVIO_RELATORIO_SEMANAL_EXECUTIVO_OBRA":
                flags_erro = AGENTE_008_SEGURANCA_SOLICITACAO_ENVIO
            elif comando["tipo_comando"] == "APROVAR_REVISAO_DOCUMENTAL_OBRA":
                flags_erro = AGENTE_008_SEGURANCA_APROVACAO_DOCUMENTAL
            elif comando["tipo_comando"] in {"LIBERAR_REVISAO_DOCUMENTAL_CAMPO_OBRA", "CONSULTAR_LIBERACOES_REVISOES_DOCUMENTAIS_CAMPO_OBRA"}:
                flags_erro = AGENTE_008_SEGURANCA_LIBERACAO_CAMPO
            else:
                flags_erro = AGENTE_008_SEGURANCA_CONSULTA
            resultado_erro = {
                "ok": False,
                "erro_controlado": str(exc),
                "resposta_telegram": (
                    "Não foi possível processar o comando de gestão operacional. "
                    "Nenhuma alteração externa foi realizada."
                ),
                "flags_seguranca": dict(flags_erro),
                **flags_erro,
            }
            try:
                with get_db_connection() as error_conn:
                    with error_conn.cursor() as error_cur:
                        error_cur.execute(
                            """
                            UPDATE comandos_executivos
                            SET status = 'ERRO', resultado = %(resultado)s,
                                mensagem_erro = %(erro)s,
                                executado_por = 'AGENTE_008_GESTAO_OPERACIONAL_OBRA',
                                executado_em = NOW(), atualizado_em = NOW()
                            WHERE id = %(id)s AND status = 'PENDENTE';
                            """,
                            {
                                "id": comando["id"], "erro": str(exc),
                                "resultado": Json(serializar_json_seguro(resultado_erro)),
                            },
                        )
            except Exception:
                pass
        raise HTTPException(
            status_code=500,
            detail={"message": "Erro ao processar comando de gestão operacional.", "error": str(exc)},
        )

    return serializar_json_seguro({
        "ok": True,
        "agente": "AGENTE_008_GESTAO_OPERACIONAL_OBRA",
        "mvp": (
            "0.8A"
            if comando["tipo_comando"] in {
                "CONSULTAR_REVISOES_DOCUMENTAIS_OBRA",
                "IMPORTAR_REVISOES_DOCUMENTAIS_PENDENTES_OBRA",
                "APROVAR_REVISAO_DOCUMENTAL_OBRA",
            }
            else (
                "0.7L"
                if comando["tipo_comando"]
                == "SOLICITAR_ENVIO_RELATORIO_SEMANAL_EXECUTIVO_OBRA"
                else "0.7K"
            )
        ),
        "status_comando": status_resultado,
        "id_comando": str(comando["id_comando"]),
        "tipo_comando": comando["tipo_comando"],
        "obra_codigo": comando["obra_codigo"],
        "telegram_chat_id": comando["telegram_chat_id"],
        "resposta_telegram": resultado["resposta_telegram"],
        "resultado": resultado,
    })


@app.post("/webhooks/entrada")
async def receber_entrada(
    request: Request,
    x_correlation_id: Optional[str] = Header(default=None),
):
    raw_payload = await request.json()

    try:
        payload = EntradaPayload(**raw_payload)
    except Exception as exc:
        raise HTTPException(
            status_code=422,
            detail={
                "message": "Payload inválido para entrada.",
                "error": str(exc),
            },
        )

    correlation_id = x_correlation_id or str(uuid.uuid4())

    normalized = {
        "tenant_id": payload.tenant_id or "construtora-piloto",
        "obra_codigo": payload.obra_codigo or "OBRA-CAIO",
        "canal": payload.canal or "desconhecido",
        "remetente_nome": payload.remetente_nome,
        "remetente_identificador": payload.remetente_identificador,
        "tipo_mensagem": payload.tipo_mensagem or "texto",
        "conteudo": payload.conteudo,
        "anexos": payload.anexos or [],
        "correlation_id": correlation_id,
    }

    payload_hash = stable_hash(raw_payload)
    idempotency_key = f"{normalized['obra_codigo']}:{normalized['canal']}:{payload_hash}"

    insert_sql = """
    INSERT INTO mensagens_recebidas (
        tenant_id,
        obra_codigo,
        canal,
        remetente_nome,
        remetente_identificador,
        tipo_mensagem,
        conteudo,
        anexos,
        payload_original,
        payload_normalizado,
        correlation_id,
        idempotency_key,
        payload_hash,
        status_processamento
    )
    VALUES (
        %(tenant_id)s,
        %(obra_codigo)s,
        %(canal)s,
        %(remetente_nome)s,
        %(remetente_identificador)s,
        %(tipo_mensagem)s,
        %(conteudo)s,
        %(anexos)s,
        %(payload_original)s,
        %(payload_normalizado)s,
        %(correlation_id)s,
        %(idempotency_key)s,
        %(payload_hash)s,
        'RECEBIDA'
    )
    ON CONFLICT (idempotency_key)
    DO NOTHING
    RETURNING id, status_processamento;
    """

    row = None

    try:
        conn = get_db_connection()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    insert_sql,
                    {
                        "tenant_id": normalized["tenant_id"],
                        "obra_codigo": normalized["obra_codigo"],
                        "canal": normalized["canal"],
                        "remetente_nome": normalized["remetente_nome"],
                        "remetente_identificador": normalized["remetente_identificador"],
                        "tipo_mensagem": normalized["tipo_mensagem"],
                        "conteudo": normalized["conteudo"],
                        "anexos": Json(normalized["anexos"]),
                        "payload_original": Json(raw_payload),
                        "payload_normalizado": Json(normalized),
                        "correlation_id": correlation_id,
                        "idempotency_key": idempotency_key,
                        "payload_hash": payload_hash,
                    },
                )

                row = cur.fetchone()

                if row is None:
                    cur.execute(
                        """
                        SELECT id, 'DUPLICADA' AS status_processamento
                        FROM mensagens_recebidas
                        WHERE idempotency_key = %(idempotency_key)s;
                        """,
                        {"idempotency_key": idempotency_key},
                    )
                    row = cur.fetchone()

            conn.commit()

        except Exception:
            conn.rollback()
            raise

        finally:
            conn.close()

    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={
                "message": "Erro ao registrar entrada no banco.",
                "error": str(exc),
                "idempotency_key": idempotency_key,
            },
        )

    if row is None:
        raise HTTPException(
            status_code=500,
            detail={
                "message": "Falha ao registrar ou recuperar mensagem recebida.",
                "idempotency_key": idempotency_key,
            },
        )

    return {
        "ok": True,
        "mensagem_id": row[0],
        "status_processamento": row[1],
        "correlation_id": correlation_id,
        "idempotency_key": idempotency_key,
        "next_action": "CLASSIFICAR",
        "message": "Entrada registrada com sucesso.",
    }


@app.post("/telegram/entrada")
async def receber_entrada_telegram(
    request: Request,
    x_correlation_id: Optional[str] = Header(default=None),
):
    raw_payload = await request.json()

    try:
        payload = TelegramEntradaPayload(**raw_payload)
    except Exception as exc:
        raise HTTPException(
            status_code=422,
            detail={
                "message": "Payload inválido para entrada Telegram.",
                "error": str(exc),
            },
        )

    correlation_id = parse_correlation_id(x_correlation_id)
    usuario_autorizado, motivo_autorizacao = avaliar_autorizacao_telegram(payload)
    instrucao_aprovacao = parse_instrucao_aprovacao_comando(payload.conteudo)
    classificacao = (
        classificar_instrucao_aprovacao(instrucao_aprovacao)
        if instrucao_aprovacao
        else classificar_intencao_executiva(payload.conteudo)
    )
    if not usuario_autorizado:
        classificacao = {
            "intencao": "USUARIO_NAO_AUTORIZADO",
            "agente_destino": "AGENTE_007_ORQUESTRADOR_EXECUTIVO",
            "tipo_comando": "BLOQUEAR_ENTRADA_NAO_AUTORIZADA",
            "requer_aprovacao": True,
            "confianca": 1.00,
            "justificativa": "Usuário não autorizado; comando registrado apenas para auditoria.",
        }

    intencao_documental = classificacao["intencao"] in {
        "CONSULTAR_DOCUMENTOS_OBRA_RESUMO",
        "CONSULTAR_DOCUMENTOS_OBRA_INDEXADOS",
        "ANALISAR_DOCUMENTO_OBRA",
        "CLASSIFICAR_DOCUMENTOS_OBRA",
        "CONSULTAR_DOCUMENTOS_CLASSIFICADOS",
        "GERAR_RELATORIO_DOCUMENTAL_OBRA",
        "AVALIAR_RISCOS_DOCUMENTAIS_OBRA",
        "CONSULTAR_ULTIMA_REVISAO_DOCUMENTAL",
        "VALIDAR_DOCUMENTO_CAMPO",
        "GERAR_PLANO_SANEAMENTO_DOCUMENTAL",
        "CRIAR_PENDENCIA_DOCUMENTAL",
        "GERAR_DIAGNOSTICO_OPERACIONAL_OBRA",
        "GERAR_PLANO_OPERACIONAL_OBRA",
        "GERAR_RESUMO_EXECUTIVO_OPERACIONAL_OBRA",
        "GERAR_BRIEFING_DIARIO_OBRA",
        "GERAR_RELATORIO_SEMANAL_EXECUTIVO_OBRA",
        "EXPORTAR_RELATORIO_SEMANAL_EXECUTIVO_OBRA",
        "GERAR_PDF_RELATORIO_SEMANAL_EXECUTIVO_OBRA",
        "APROVAR_PDF_RELATORIO_SEMANAL_EXECUTIVO_OBRA",
        "SOLICITAR_ENVIO_RELATORIO_SEMANAL_EXECUTIVO_OBRA",
        "CRIAR_ACAO_OPERACIONAL_OBRA",
        "LISTAR_ACOES_OPERACIONAIS_OBRA",
        "ATUALIZAR_ACAO_OPERACIONAL_OBRA",
        "DETALHAR_ACAO_OPERACIONAL_OBRA",
        "CONSULTAR_REVISOES_DOCUMENTAIS_OBRA",
        "IMPORTAR_REVISOES_DOCUMENTAIS_PENDENTES_OBRA",
        "APROVAR_REVISAO_DOCUMENTAL_OBRA",
        "LIBERAR_REVISAO_DOCUMENTAL_CAMPO_OBRA",
        "CONSULTAR_LIBERACOES_REVISOES_DOCUMENTAIS_CAMPO_OBRA",
    }
    obra_codigo = payload.obra_codigo or "OBRA-CAIO"
    if (
        usuario_autorizado
        and classificacao["tipo_comando"] == "APROVAR_REVISAO_DOCUMENTAL_OBRA"
        and classificacao.get("revisao_documental_id") is None
    ):
        mensagem = "Informe o ID da revisão documental. Exemplo: aprovar revisão #12 como vigente"
        return {
            "ok": True, "status": "ID_REVISAO_DOCUMENTAL_NAO_INFORMADO",
            "obra_codigo": obra_codigo, "tipo_comando": None,
            "mensagem_resposta_executiva": mensagem, "resposta_telegram": mensagem,
            "acoes_externas_executadas": False,
        }
    if (
        usuario_autorizado
        and classificacao["tipo_comando"] == "LIBERAR_REVISAO_DOCUMENTAL_CAMPO_OBRA"
        and classificacao.get("revisao_documental_id") is None
    ):
        mensagem = "Informe o ID da revisão documental. Exemplo: liberar revisão #12 para campo"
        return {"ok": True, "status": "ID_REVISAO_DOCUMENTAL_NAO_INFORMADO", "obra_codigo": obra_codigo, "tipo_comando": None, "mensagem_resposta_executiva": mensagem, "resposta_telegram": mensagem, "acoes_externas_executadas": False}
    normalized = {
        "tenant_id": payload.tenant_id or "construtora-piloto",
        "obra_codigo": obra_codigo,
        "canal": "telegram",
        "telegram_update_id": payload.telegram_update_id,
        "telegram_message_id": payload.telegram_message_id,
        "telegram_user_id": payload.telegram_user_id,
        "telegram_username": payload.telegram_username,
        "chat_id": payload.chat_id,
        "chat_type": payload.chat_type,
        "remetente_nome": payload.remetente_nome,
        "remetente_identificador": payload.remetente_identificador,
        "usuario_autorizado": usuario_autorizado,
        "motivo_autorizacao": motivo_autorizacao,
        "tipo_mensagem": payload.tipo_mensagem or "texto",
        "conteudo": payload.conteudo,
        "anexos": payload.anexos or [],
        "correlation_id": correlation_id,
        "intencao": classificacao["intencao"],
        "agente_destino": classificacao["agente_destino"],
    }
    if instrucao_aprovacao:
        normalized["instrucao_aprovacao"] = instrucao_aprovacao

    payload_hash = stable_hash(raw_payload)
    telegram_idempotency_parts = [
        normalized["telegram_update_id"],
        normalized["telegram_message_id"],
        normalized["chat_id"],
    ]
    if all(telegram_idempotency_parts):
        idempotency_parts = [
            normalized["obra_codigo"],
            normalized["canal"],
            *telegram_idempotency_parts,
        ]
    else:
        idempotency_parts = [
            normalized["obra_codigo"],
            normalized["canal"],
            payload_hash,
        ]
    idempotency_key = ":".join(str(part) for part in idempotency_parts)

    status_comando = (
        "AGUARDANDO_APROVACAO"
        if classificacao["requer_aprovacao"]
        else "PENDENTE"
    )

    insert_evento_sql = """
    INSERT INTO eventos_telegram (
        tenant_id,
        obra_codigo,
        canal,
        telegram_update_id,
        telegram_message_id,
        telegram_user_id,
        telegram_username,
        chat_id,
        chat_type,
        remetente_nome,
        remetente_identificador,
        usuario_autorizado,
        motivo_autorizacao,
        tipo_mensagem,
        conteudo,
        anexos,
        payload_original,
        payload_normalizado,
        payload_hash,
        idempotency_key,
        correlation_id,
        intencao,
        confianca,
        agente_destino,
        status_processamento
    )
    VALUES (
        %(tenant_id)s,
        %(obra_codigo)s,
        'telegram',
        %(telegram_update_id)s,
        %(telegram_message_id)s,
        %(telegram_user_id)s,
        %(telegram_username)s,
        %(chat_id)s,
        %(chat_type)s,
        %(remetente_nome)s,
        %(remetente_identificador)s,
        %(usuario_autorizado)s,
        %(motivo_autorizacao)s,
        %(tipo_mensagem)s,
        %(conteudo)s,
        %(anexos)s,
        %(payload_original)s,
        %(payload_normalizado)s,
        %(payload_hash)s,
        %(idempotency_key)s,
        %(correlation_id)s,
        %(intencao)s,
        %(confianca)s,
        %(agente_destino)s,
        'CLASSIFICADO'
    )
    ON CONFLICT (idempotency_key)
    DO NOTHING
    RETURNING id, status_processamento;
    """

    select_evento_sql = """
    SELECT id, 'DUPLICADO' AS status_processamento
    FROM eventos_telegram
    WHERE idempotency_key = %(idempotency_key)s;
    """

    insert_comando_sql = """
    INSERT INTO comandos_executivos (
        id_comando,
        evento_telegram_id,
        tenant_id,
        obra_codigo,
        correlation_id,
        agente_origem,
        agente_destino,
        tipo_comando,
        payload_comando,
        justificativa,
        status,
        requer_aprovacao,
        evidencias
    )
    VALUES (
        %(id_comando)s,
        %(evento_telegram_id)s,
        %(tenant_id)s,
        %(obra_codigo)s,
        %(correlation_id)s,
        'AGENTE_007_ORQUESTRADOR_EXECUTIVO',
        %(agente_destino)s,
        %(tipo_comando)s,
        %(payload_comando)s,
        %(justificativa)s,
        %(status)s,
        %(requer_aprovacao)s,
        '[]'::jsonb
    )
    ON CONFLICT (evento_telegram_id)
    WHERE evento_telegram_id IS NOT NULL
    DO NOTHING
    RETURNING id, id_comando, status;
    """

    select_comando_sql = """
    SELECT id, id_comando, status
    FROM comandos_executivos
    WHERE evento_telegram_id = %(evento_telegram_id)s
    ORDER BY id
    LIMIT 1;
    """

    update_evento_sql = """
    UPDATE eventos_telegram
    SET status_processamento = 'COMANDO_GERADO',
        atualizado_em = NOW()
    WHERE id = %(evento_telegram_id)s
      AND status_processamento <> 'COMANDO_GERADO';
    """

    update_evento_aprovacao_sql = """
    UPDATE eventos_telegram
    SET status_processamento = 'COMANDO_GERADO',
        atualizado_em = NOW()
    WHERE id = %(evento_telegram_id)s
      AND status_processamento <> 'COMANDO_GERADO';
    """

    select_comando_alvo_sql = """
    SELECT id, id_comando, status, tipo_comando, agente_destino
    FROM comandos_executivos
    WHERE (
            %(comando_id)s::bigint IS NOT NULL
            AND id = %(comando_id)s::bigint
        )
       OR (
            %(id_comando)s::uuid IS NOT NULL
            AND id_comando = %(id_comando)s::uuid
        )
    ORDER BY id
    LIMIT 1
    FOR UPDATE;
    """

    update_aprovar_comando_sql = """
    UPDATE comandos_executivos
    SET status = 'APROVADO',
        aprovado_por = %(aprovado_por)s,
        aprovado_em = NOW(),
        resultado = %(resultado)s,
        evidencias = COALESCE(evidencias, '[]'::jsonb) || %(evidencia)s::jsonb,
        mensagem_erro = NULL,
        atualizado_em = NOW()
    WHERE id = %(id)s
      AND status = 'AGUARDANDO_APROVACAO'
    RETURNING id, id_comando, status, tipo_comando, agente_destino;
    """

    update_cancelar_comando_sql = """
    UPDATE comandos_executivos
    SET status = 'CANCELADO',
        resultado = %(resultado)s,
        evidencias = COALESCE(evidencias, '[]'::jsonb) || %(evidencia)s::jsonb,
        mensagem_erro = NULL,
        atualizado_em = NOW()
    WHERE id = %(id)s
      AND status = 'AGUARDANDO_APROVACAO'
    RETURNING id, id_comando, status, tipo_comando, agente_destino;
    """

    evento_row = None
    comando_row = None
    evento_duplicado = False

    try:
        conn = get_db_connection()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    insert_evento_sql,
                    {
                        "tenant_id": normalized["tenant_id"],
                        "obra_codigo": normalized["obra_codigo"],
                        "telegram_update_id": normalized["telegram_update_id"],
                        "telegram_message_id": normalized["telegram_message_id"],
                        "telegram_user_id": normalized["telegram_user_id"],
                        "telegram_username": normalized["telegram_username"],
                        "chat_id": normalized["chat_id"],
                        "chat_type": normalized["chat_type"],
                        "remetente_nome": normalized["remetente_nome"],
                        "remetente_identificador": normalized["remetente_identificador"],
                        "usuario_autorizado": normalized["usuario_autorizado"],
                        "motivo_autorizacao": normalized["motivo_autorizacao"],
                        "tipo_mensagem": normalized["tipo_mensagem"],
                        "conteudo": normalized["conteudo"],
                        "anexos": Json(normalized["anexos"]),
                        "payload_original": Json(raw_payload),
                        "payload_normalizado": Json(normalized),
                        "payload_hash": payload_hash,
                        "idempotency_key": idempotency_key,
                        "correlation_id": correlation_id,
                        "intencao": classificacao["intencao"],
                        "confianca": classificacao["confianca"],
                        "agente_destino": classificacao["agente_destino"],
                    },
                )
                evento_row = cur.fetchone()

                if evento_row is None:
                    cur.execute(
                        select_evento_sql,
                        {"idempotency_key": idempotency_key},
                    )
                    evento_row = cur.fetchone()
                    evento_duplicado = True

                if evento_row is None:
                    raise RuntimeError("Evento Telegram não foi registrado nem recuperado.")

                evento_id = evento_row[0]

                if instrucao_aprovacao and usuario_autorizado:
                    cur.execute(
                        select_comando_alvo_sql,
                        {
                            "comando_id": instrucao_aprovacao["comando_id"],
                            "id_comando": instrucao_aprovacao["id_comando"],
                        },
                    )
                    comando_alvo_row = cur.fetchone()

                    if comando_alvo_row is None:
                        mensagem_resposta_executiva = (
                            f"Comando {instrucao_aprovacao['identificador']} não encontrado. "
                            "Nenhuma ação externa foi executada."
                        )
                        cur.execute(
                            update_evento_aprovacao_sql,
                            {"evento_telegram_id": evento_id},
                        )
                        conn.commit()
                        return {
                            "ok": True,
                            "evento_telegram_id": evento_row[0],
                            "status_recebimento": evento_row[1],
                            "status_evento": "COMANDO_GERADO",
                            "evento_duplicado": evento_duplicado,
                            "comando_executivo_id": None,
                            "id_comando": None,
                            "status_comando": None,
                            "tenant_id": normalized["tenant_id"],
                            "obra_codigo": normalized["obra_codigo"],
                            "correlation_id": correlation_id,
                            "idempotency_key": idempotency_key,
                            "usuario_autorizado": usuario_autorizado,
                            "motivo_autorizacao": motivo_autorizacao,
                            "intencao": classificacao["intencao"],
                            "confianca": classificacao["confianca"],
                            "agente_origem": "AGENTE_007_ORQUESTRADOR_EXECUTIVO",
                            "agente_destino": classificacao["agente_destino"],
                            "tipo_comando": classificacao["tipo_comando"],
                            "requer_aprovacao": classificacao["requer_aprovacao"],
                            "telegram_chat_id": normalized["chat_id"],
                            "telegram_user_id": normalized["telegram_user_id"],
                            "telegram_message_id": normalized["telegram_message_id"],
                            "mensagem_resposta_executiva": mensagem_resposta_executiva,
                            "acoes_externas_executadas": False,
                            "next_action": "NENHUMA_ACAO_EXTERNA_EXECUTADA",
                            "message": mensagem_resposta_executiva,
                        }

                    status_atual = comando_alvo_row[2]
                    if status_atual != "AGUARDANDO_APROVACAO":
                        mensagem_resposta_executiva = (
                            f"Comando {comando_alvo_row[0]} está com status {status_atual}. "
                            "Nenhuma alteração foi feita."
                        )
                        cur.execute(
                            update_evento_aprovacao_sql,
                            {"evento_telegram_id": evento_id},
                        )
                        conn.commit()
                        return {
                            "ok": True,
                            "evento_telegram_id": evento_row[0],
                            "status_recebimento": evento_row[1],
                            "status_evento": "COMANDO_GERADO",
                            "evento_duplicado": evento_duplicado,
                            "comando_executivo_id": comando_alvo_row[0],
                            "id_comando": str(comando_alvo_row[1]),
                            "status_comando": status_atual,
                            "tenant_id": normalized["tenant_id"],
                            "obra_codigo": normalized["obra_codigo"],
                            "correlation_id": correlation_id,
                            "idempotency_key": idempotency_key,
                            "usuario_autorizado": usuario_autorizado,
                            "motivo_autorizacao": motivo_autorizacao,
                            "intencao": classificacao["intencao"],
                            "confianca": classificacao["confianca"],
                            "agente_origem": "AGENTE_007_ORQUESTRADOR_EXECUTIVO",
                            "agente_destino": comando_alvo_row[4],
                            "tipo_comando": comando_alvo_row[3],
                            "requer_aprovacao": False,
                            "telegram_chat_id": normalized["chat_id"],
                            "telegram_user_id": normalized["telegram_user_id"],
                            "telegram_message_id": normalized["telegram_message_id"],
                            "mensagem_resposta_executiva": mensagem_resposta_executiva,
                            "acoes_externas_executadas": False,
                            "next_action": "NENHUMA_ACAO_EXTERNA_EXECUTADA",
                            "message": mensagem_resposta_executiva,
                        }

                    auditoria = {
                        "tipo_evento": "APROVACAO_EXECUTIVA_TELEGRAM"
                        if instrucao_aprovacao["acao"] == "APROVAR"
                        else "CANCELAMENTO_EXECUTIVO_TELEGRAM",
                        "acao": instrucao_aprovacao["acao"],
                        "agente": "AGENTE_007_ORQUESTRADOR_EXECUTIVO",
                        "evento_telegram_id": evento_id,
                        "telegram_user_id": normalized["telegram_user_id"],
                        "telegram_chat_id": normalized["chat_id"],
                        "telegram_message_id": normalized["telegram_message_id"],
                        "correlation_id": correlation_id,
                        "registrado_em": datetime.now(timezone.utc).isoformat(),
                        "acoes_externas_executadas": False,
                    }
                    resultado_aprovacao = {
                        "tipo_resultado": "COMANDO_EXECUTIVO_APROVADO"
                        if instrucao_aprovacao["acao"] == "APROVAR"
                        else "COMANDO_EXECUTIVO_CANCELADO",
                        "auditoria": auditoria,
                        "controles_operacionais": {
                            "executou_acao_externa": False,
                            "gerou_pdf_real": False,
                            "imprimiu": False,
                            "alterou_rdo_oficial": False,
                            "executou_rpa": False,
                            "conectou_openclaw": False,
                            "enviou_mensagem_terceiros": False,
                        },
                    }
                    update_sql = (
                        update_aprovar_comando_sql
                        if instrucao_aprovacao["acao"] == "APROVAR"
                        else update_cancelar_comando_sql
                    )
                    cur.execute(
                        update_sql,
                        {
                            "id": comando_alvo_row[0],
                            "aprovado_por": normalized["telegram_user_id"]
                            or normalized["remetente_identificador"]
                            or normalized["chat_id"],
                            "resultado": Json(resultado_aprovacao),
                            "evidencia": Json([auditoria]),
                        },
                    )
                    comando_atualizado_row = cur.fetchone()
                    if comando_atualizado_row is None:
                        raise RuntimeError("Comando alvo não foi atualizado.")

                    cur.execute(
                        update_evento_aprovacao_sql,
                        {"evento_telegram_id": evento_id},
                    )
                    mensagem_resposta_executiva = (
                        f"Comando {comando_atualizado_row[0]} aprovado. "
                        "Nenhuma ação externa foi executada."
                        if instrucao_aprovacao["acao"] == "APROVAR"
                        else f"Comando {comando_atualizado_row[0]} cancelado. "
                        "Nenhuma ação externa foi executada."
                    )
                    conn.commit()
                    return {
                        "ok": True,
                        "evento_telegram_id": evento_row[0],
                        "status_recebimento": evento_row[1],
                        "status_evento": "COMANDO_GERADO",
                        "evento_duplicado": evento_duplicado,
                        "comando_executivo_id": comando_atualizado_row[0],
                        "id_comando": str(comando_atualizado_row[1]),
                        "status_comando": comando_atualizado_row[2],
                        "tenant_id": normalized["tenant_id"],
                        "obra_codigo": normalized["obra_codigo"],
                        "correlation_id": correlation_id,
                        "idempotency_key": idempotency_key,
                        "usuario_autorizado": usuario_autorizado,
                        "motivo_autorizacao": motivo_autorizacao,
                        "intencao": classificacao["intencao"],
                        "confianca": classificacao["confianca"],
                        "agente_origem": "AGENTE_007_ORQUESTRADOR_EXECUTIVO",
                        "agente_destino": comando_atualizado_row[4],
                        "tipo_comando": comando_atualizado_row[3],
                        "requer_aprovacao": False,
                        "telegram_chat_id": normalized["chat_id"],
                        "telegram_user_id": normalized["telegram_user_id"],
                        "telegram_message_id": normalized["telegram_message_id"],
                        "mensagem_resposta_executiva": mensagem_resposta_executiva,
                        "acoes_externas_executadas": False,
                        "next_action": "COMANDO_ATUALIZADO_SEM_EXECUCAO_EXTERNA",
                        "message": mensagem_resposta_executiva,
                    }

                payload_comando = {
                    "modo": "curl_local_sem_integracoes_externas",
                    "proibicoes": [
                        "nao_conectar_telegram_real",
                        "nao_usar_openclaw",
                        "nao_executar_rpa",
                        "nao_imprimir",
                        "nao_gerar_pdf_real",
                        "nao_alterar_rdo_oficial",
                    ],
                    "entrada": normalized,
                    "classificacao": classificacao,
                }
                if intencao_documental:
                    payload_comando.update(AGENTE_008_SEGURANCA_CONSULTA)
                    if classificacao["tipo_comando"] == "CONSULTAR_DOCUMENTOS_OBRA_INDEXADOS":
                        payload_comando.update(
                            extrair_filtros_documentos_telegram(normalized["conteudo"])
                        )
                        payload_comando["limite"] = 10
                    elif classificacao["tipo_comando"] == "ANALISAR_DOCUMENTO_OBRA":
                        payload_comando.update(
                            extrair_alvo_analise_documental(normalized["conteudo"])
                        )
                    elif classificacao["tipo_comando"] == "CLASSIFICAR_DOCUMENTOS_OBRA":
                        payload_comando.update({
                            "obra_codigo": obra_codigo,
                            "reprocessar": normalizar_texto_comparacao(
                                normalized["conteudo"]
                            ).startswith("reclassificar"),
                        })
                    elif classificacao["tipo_comando"] == "CONSULTAR_DOCUMENTOS_CLASSIFICADOS":
                        payload_comando.update(
                            extrair_filtros_classificacao_telegram(normalized["conteudo"])
                        )
                    elif classificacao["tipo_comando"] == "GERAR_RELATORIO_DOCUMENTAL_OBRA":
                        payload_comando = {
                            "obra_codigo": obra_codigo,
                            "incluir_amostras": True,
                            "limite_amostras": 5,
                        }
                    elif classificacao["tipo_comando"] == "AVALIAR_RISCOS_DOCUMENTAIS_OBRA":
                        payload_comando = {
                            "obra_codigo": obra_codigo,
                            **extrair_riscos_documentais_telegram(normalized["conteudo"]),
                            **AGENTE_008_SEGURANCA_CONSULTA,
                        }
                    elif classificacao["tipo_comando"] == "CONSULTAR_ULTIMA_REVISAO_DOCUMENTAL":
                        payload_comando = {
                            "obra_codigo": obra_codigo,
                            **extrair_ultima_revisao_telegram(normalized["conteudo"]),
                            **AGENTE_008_SEGURANCA_CONSULTA,
                        }
                    elif classificacao["tipo_comando"] == "VALIDAR_DOCUMENTO_CAMPO":
                        payload_comando = {
                            "obra_codigo": obra_codigo,
                            **extrair_documento_id_telegram(normalized["conteudo"]),
                            **AGENTE_008_SEGURANCA_CONSULTA,
                        }
                    elif classificacao["tipo_comando"] == "GERAR_PLANO_SANEAMENTO_DOCUMENTAL":
                        filtros = extrair_riscos_documentais_telegram(normalized["conteudo"])
                        payload_comando = {
                            "obra_codigo": obra_codigo,
                            "area": filtros["area"],
                            "disciplina": filtros["disciplina"],
                            "limite_acoes": 10,
                            **AGENTE_008_SEGURANCA_CONSULTA,
                        }
                    elif classificacao["tipo_comando"] == "CRIAR_PENDENCIA_DOCUMENTAL":
                        payload_comando = {
                            "obra_codigo": obra_codigo,
                            **extrair_pendencia_documental_telegram(normalized["conteudo"]),
                            **AGENTE_008_SEGURANCA_REGISTRO_INTERNO,
                        }
                    elif classificacao["tipo_comando"] == "GERAR_DIAGNOSTICO_OPERACIONAL_OBRA":
                        payload_comando = {
                            "obra_codigo": obra_codigo,
                            "area": extrair_area_diagnostico_operacional(normalized["conteudo"]),
                            "incluir_documentos": True,
                            "incluir_pendencias": True,
                            "incluir_restricoes": True,
                            **AGENTE_008_SEGURANCA_CONSULTA,
                        }
                    elif classificacao["tipo_comando"] == "GERAR_PLANO_OPERACIONAL_OBRA":
                        payload_comando = {
                            "obra_codigo": obra_codigo,
                            "area": extrair_area_plano_operacional(normalized["conteudo"]),
                            "limite_acoes": 10,
                            **AGENTE_008_SEGURANCA_CONSULTA,
                        }
                    elif classificacao["tipo_comando"] == "GERAR_RESUMO_EXECUTIVO_OPERACIONAL_OBRA":
                        payload_comando = {
                            "obra_codigo": obra_codigo,
                            "area": extrair_area_diagnostico_operacional(normalized["conteudo"]),
                            "limite_itens": 10,
                            **AGENTE_008_SEGURANCA_CONSULTA,
                        }
                    elif classificacao["tipo_comando"] == "GERAR_BRIEFING_DIARIO_OBRA":
                        payload_comando = {
                            "obra_codigo": obra_codigo,
                            "area": extrair_area_diagnostico_operacional(normalized["conteudo"]),
                            "limite_itens": 10,
                            "incluir_acoes": True,
                            "incluir_documentos": True,
                            "incluir_historico": True,
                            **AGENTE_008_SEGURANCA_CONSULTA,
                        }
                    elif classificacao["tipo_comando"] == "GERAR_RELATORIO_SEMANAL_EXECUTIVO_OBRA":
                        payload_comando = {
                            "obra_codigo": obra_codigo,
                            "area": extrair_area_diagnostico_operacional(normalized["conteudo"]),
                            "data_inicio": None,
                            "data_fim": None,
                            "limite_itens": 10,
                            "salvar_relatorio": True,
                            **AGENTE_008_SEGURANCA_CONSULTA,
                        }
                    elif classificacao["tipo_comando"] == "EXPORTAR_RELATORIO_SEMANAL_EXECUTIVO_OBRA":
                        payload_comando = {
                            "obra_codigo": obra_codigo,
                            "area": extrair_area_diagnostico_operacional(normalized["conteudo"]),
                            "data_inicio": None,
                            "data_fim": None,
                            "relatorio_semanal_id": None,
                            "formato": "MARKDOWN",
                            "limite_itens": 10,
                            "salvar_exportacao": True,
                            **AGENTE_008_SEGURANCA_EXPORTACAO,
                        }
                    elif classificacao["tipo_comando"] == "GERAR_PDF_RELATORIO_SEMANAL_EXECUTIVO_OBRA":
                        payload_comando = {
                            "obra_codigo": obra_codigo,
                            "area": extrair_area_diagnostico_operacional(normalized["conteudo"]),
                            "data_inicio": None,
                            "data_fim": None,
                            "exportacao_relatorio_id": None,
                            "relatorio_semanal_id": None,
                            "limite_itens": 10,
                            "salvar_pdf": True,
                            "armazenar_minio": True,
                            **AGENTE_008_SEGURANCA_PDF,
                        }
                    elif classificacao["tipo_comando"] == "APROVAR_PDF_RELATORIO_SEMANAL_EXECUTIVO_OBRA":
                        dados_decisao = extrair_decisao_pdf_relatorio_telegram(
                            normalized["conteudo"], classificacao["decisao"],
                        )
                        payload_comando = {
                            "obra_codigo": obra_codigo,
                            **dados_decisao,
                            "decisao": classificacao["decisao"],
                            "decisor_nome": None,
                            "decisor_telegram_user_id": normalized["telegram_user_id"],
                            "decisor_telegram_username": normalized["telegram_username"],
                            "decisor_chat_id": normalized["chat_id"],
                            **AGENTE_008_SEGURANCA_APROVACAO_PDF,
                        }
                    elif classificacao["tipo_comando"] == "SOLICITAR_ENVIO_RELATORIO_SEMANAL_EXECUTIVO_OBRA":
                        dados_solicitacao = extrair_solicitacao_envio_relatorio_telegram(
                            normalized["conteudo"]
                        )
                        payload_comando = {
                            "obra_codigo": obra_codigo,
                            **dados_solicitacao,
                            "assunto": (
                                f"Relatório semanal executivo — {obra_codigo}"
                            ),
                            "solicitado_por": None,
                            "solicitante_telegram_user_id": normalized["telegram_user_id"],
                            "solicitante_telegram_username": normalized["telegram_username"],
                            "solicitante_chat_id": normalized["chat_id"],
                            **AGENTE_008_SEGURANCA_SOLICITACAO_ENVIO,
                        }
                    elif classificacao["tipo_comando"] == "CRIAR_ACAO_OPERACIONAL_OBRA":
                        payload_comando = {
                            "obra_codigo": obra_codigo,
                            **extrair_acao_operacional_telegram(normalized["conteudo"]),
                            **AGENTE_008_SEGURANCA_REGISTRO_INTERNO,
                        }
                    elif classificacao["tipo_comando"] == "LISTAR_ACOES_OPERACIONAIS_OBRA":
                        payload_comando = {
                            "obra_codigo": obra_codigo,
                            "area": extrair_area_diagnostico_operacional(normalized["conteudo"]),
                            "disciplina": None, "status": "ABERTA", "limite": 20,
                            **AGENTE_008_SEGURANCA_CONSULTA,
                        }
                    elif classificacao["tipo_comando"] == "ATUALIZAR_ACAO_OPERACIONAL_OBRA":
                        payload_comando = {
                            "obra_codigo": obra_codigo,
                            **extrair_atualizacao_acao_telegram(normalized["conteudo"]),
                            **AGENTE_008_SEGURANCA_REGISTRO_INTERNO,
                        }
                    elif classificacao["tipo_comando"] == "DETALHAR_ACAO_OPERACIONAL_OBRA":
                        payload_comando = {
                            "obra_codigo": obra_codigo,
                            **extrair_detalhe_acao_telegram(normalized["conteudo"]),
                            **AGENTE_008_SEGURANCA_CONSULTA,
                        }
                    elif classificacao["tipo_comando"] == "CONSULTAR_REVISOES_DOCUMENTAIS_OBRA":
                        payload_comando = {
                            "obra_codigo": obra_codigo, "area": None, "disciplina": None,
                            "status_revisao": None, "status_vigencia": None,
                            "limite_itens": 20,
                            **AGENTE_008_SEGURANCA_CONTROLE_DOCUMENTAL,
                        }
                    elif classificacao["tipo_comando"] == "IMPORTAR_REVISOES_DOCUMENTAIS_PENDENTES_OBRA":
                        payload_comando = {
                            "obra_codigo": obra_codigo, "bucket": "obra-caio",
                            "prefixo": "01_projetos/00_recebidos_para_analise/",
                            "limite_itens": 50,
                            **AGENTE_008_SEGURANCA_CONTROLE_DOCUMENTAL,
                        }
                    elif classificacao["tipo_comando"] == "APROVAR_REVISAO_DOCUMENTAL_OBRA":
                        dados_decisao = extrair_decisao_revisao_documental_telegram(normalized["conteudo"])
                        payload_comando = {
                            "obra_codigo": obra_codigo,
                            "revisao_documental_id": dados_decisao["revisao_documental_id"],
                            "decisao": dados_decisao["decisao"],
                            "motivo": dados_decisao["motivo"],
                            "observacao": dados_decisao["observacao"],
                            "decisor_nome": normalized["remetente_nome"],
                            "decisor_telegram_user_id": normalized["telegram_user_id"],
                            "decisor_telegram_username": normalized["telegram_username"],
                            "decisor_chat_id": normalized["chat_id"],
                            **AGENTE_008_SEGURANCA_APROVACAO_DOCUMENTAL,
                        }
                    elif classificacao["tipo_comando"] == "LIBERAR_REVISAO_DOCUMENTAL_CAMPO_OBRA":
                        dados_liberacao = extrair_liberacao_revisao_campo_telegram(normalized["conteudo"])
                        payload_comando = {
                            "obra_codigo": obra_codigo,
                            **dados_liberacao,
                            "decisor_nome": None,
                            "decisor_telegram_user_id": normalized["telegram_user_id"],
                            "decisor_telegram_username": normalized["telegram_username"],
                            "decisor_chat_id": normalized["chat_id"],
                            **AGENTE_008_SEGURANCA_LIBERACAO_CAMPO,
                        }
                    elif classificacao["tipo_comando"] == "CONSULTAR_LIBERACOES_REVISOES_DOCUMENTAIS_CAMPO_OBRA":
                        payload_comando = {
                            "obra_codigo": obra_codigo, "area": None, "disciplina": None,
                            "status_liberacao": None, "limite_itens": 20,
                            **AGENTE_008_SEGURANCA_LIBERACAO_CAMPO,
                        }

                cur.execute(
                    insert_comando_sql,
                    {
                        "id_comando": str(uuid.uuid4()),
                        "evento_telegram_id": evento_id,
                        "tenant_id": normalized["tenant_id"],
                        "obra_codigo": normalized["obra_codigo"],
                        "correlation_id": correlation_id,
                        "agente_destino": classificacao["agente_destino"],
                        "tipo_comando": classificacao["tipo_comando"],
                        "payload_comando": Json(payload_comando),
                        "justificativa": classificacao["justificativa"],
                        "status": status_comando,
                        "requer_aprovacao": classificacao["requer_aprovacao"],
                    },
                )
                comando_row = cur.fetchone()

                if comando_row is None:
                    cur.execute(
                        select_comando_sql,
                        {"evento_telegram_id": evento_id},
                    )
                    comando_row = cur.fetchone()

                cur.execute(
                    update_evento_sql,
                    {"evento_telegram_id": evento_id},
                )

            conn.commit()

        except Exception:
            conn.rollback()
            raise

        finally:
            conn.close()

    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={
                "message": "Erro ao registrar entrada Telegram e comando executivo.",
                "error": str(exc),
                "idempotency_key": idempotency_key,
            },
        )

    if comando_row is None:
        raise HTTPException(
            status_code=500,
            detail={
                "message": "Falha ao registrar ou recuperar comando executivo.",
                "idempotency_key": idempotency_key,
            },
        )

    mensagem_resposta_executiva = None
    if not usuario_autorizado:
        mensagem_resposta_executiva = (
            "Usuário não autorizado para comandos executivos. "
            "Solicitação registrada para auditoria; nenhum comando foi alterado."
        )
    elif classificacao["intencao"] == "CONSULTAR_DOCUMENTOS_OBRA_RESUMO":
        mensagem_resposta_executiva = montar_resposta_documentos_resumo_telegram(
            normalized["obra_codigo"]
        )
    elif classificacao["intencao"] == "CONSULTAR_DOCUMENTOS_OBRA_INDEXADOS":
        mensagem_resposta_executiva = montar_resposta_documentos_indexados_telegram(
            normalized["obra_codigo"], normalized["conteudo"]
        )

    return {
        "ok": True,
        "evento_telegram_id": evento_row[0],
        "status_recebimento": evento_row[1],
        "status_evento": "COMANDO_GERADO",
        "evento_duplicado": evento_duplicado,
        "comando_executivo_id": comando_row[0],
        "id_comando": str(comando_row[1]),
        "status_comando": comando_row[2],
        "tenant_id": normalized["tenant_id"],
        "obra_codigo": normalized["obra_codigo"],
        "correlation_id": correlation_id,
        "idempotency_key": idempotency_key,
        "usuario_autorizado": usuario_autorizado,
        "motivo_autorizacao": motivo_autorizacao,
        "intencao": classificacao["intencao"],
        "confianca": classificacao["confianca"],
        "agente_origem": "AGENTE_007_ORQUESTRADOR_EXECUTIVO",
        "agente_destino": classificacao["agente_destino"],
        "tipo_comando": classificacao["tipo_comando"],
        "requer_aprovacao": classificacao["requer_aprovacao"],
        "telegram_chat_id": normalized["chat_id"],
        "telegram_user_id": normalized["telegram_user_id"],
        "telegram_message_id": normalized["telegram_message_id"],
        "mensagem_resposta_executiva": mensagem_resposta_executiva,
        **(AGENTE_008_SEGURANCA_CONSULTA if intencao_documental else {}),
        "acoes_externas_executadas": False,
        "next_action": "COMANDO_AUDITAVEL_REGISTRADO_SEM_EXECUCAO_EXTERNA",
        "message": "Entrada Telegram registrada e classificada sem integrações externas.",
    }


@app.post("/agentes/rdo/processar-comando")
async def processar_comando_agente_rdo(
    payload: Optional[ProcessarComandoRDORequest] = None,
):
    id_comando = payload.id_comando if payload else None
    if id_comando:
        try:
            id_comando = str(uuid.UUID(id_comando))
        except ValueError as exc:
            raise HTTPException(
                status_code=422,
                detail={
                    "message": "id_comando inválido. Use um UUID válido.",
                    "error": str(exc),
                },
            )

    select_comando_sql = """
    SELECT
        ce.id,
        ce.id_comando,
        ce.tenant_id,
        ce.obra_codigo,
        ce.correlation_id,
        ce.agente_origem,
        ce.agente_destino,
        ce.tipo_comando,
        ce.payload_comando,
        ce.status,
        ce.requer_aprovacao,
        et.chat_id,
        et.telegram_user_id,
        et.telegram_message_id
    FROM comandos_executivos AS ce
    LEFT JOIN eventos_telegram AS et
        ON et.id = ce.evento_telegram_id
    WHERE ce.agente_destino = 'AGENTE_RDO'
      AND ce.status = 'PENDENTE'
      AND (%(id_comando)s::uuid IS NULL OR ce.id_comando = %(id_comando)s::uuid)
    ORDER BY ce.criado_em
    LIMIT 1
    FOR UPDATE OF ce;
    """

    update_resultado_sql = """
    UPDATE comandos_executivos
    SET resultado = %(resultado)s,
        executado_por = 'AGENTE_002_GERADOR_RDO',
        executado_em = NOW(),
        mensagem_erro = NULL,
        atualizado_em = NOW()
    WHERE id = %(id)s
      AND status = 'PENDENTE'
    RETURNING id;
    """

    update_status_sql = """
    UPDATE comandos_executivos
    SET status = 'CONCLUIDO',
        atualizado_em = NOW()
    WHERE id = %(id)s
      AND status = 'PENDENTE'
      AND resultado IS NOT NULL
    RETURNING id, id_comando, status, resultado;
    """

    try:
        conn = get_db_connection()
        try:
            with conn.cursor() as cur:
                cur.execute(select_comando_sql, {"id_comando": id_comando})
                row = cur.fetchone()

                if row is None:
                    conn.rollback()
                    return {
                        "ok": True,
                        "processado": False,
                        "agente_destino": "AGENTE_RDO",
                        "message": "Nenhum comando PENDENTE para AGENTE_RDO encontrado.",
                    }

                comando = {
                    "id": row[0],
                    "id_comando": row[1],
                    "tenant_id": row[2],
                    "obra_codigo": row[3],
                    "correlation_id": row[4],
                    "agente_origem": row[5],
                    "agente_destino": row[6],
                    "tipo_comando": row[7],
                    "payload_comando": row[8],
                    "status": row[9],
                    "requer_aprovacao": row[10],
                    "telegram_chat_id": row[11],
                    "telegram_user_id": row[12],
                    "telegram_message_id": row[13],
                }

                pendencias = listar_pendencias_rdo(conn, comando["obra_codigo"])
                eventos = listar_eventos_rdo(conn, comando["obra_codigo"])
                resultado = montar_resultado_agente_rdo(comando, pendencias, eventos)

                cur.execute(
                    update_resultado_sql,
                    {
                        "id": comando["id"],
                        "resultado": Json(resultado),
                    },
                )
                resultado_row = cur.fetchone()
                if resultado_row is None:
                    raise RuntimeError("Resultado do comando RDO não foi salvo.")

                cur.execute(update_status_sql, {"id": comando["id"]})
                concluido_row = cur.fetchone()
                if concluido_row is None:
                    raise RuntimeError("Status do comando RDO não foi atualizado para CONCLUIDO.")

            conn.commit()

        except Exception:
            conn.rollback()
            raise

        finally:
            conn.close()

    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={
                "message": "Erro ao processar comando executivo do Agente RDO.",
                "error": str(exc),
            },
        )

    return {
        "ok": True,
        "processado": True,
        "comando_executivo_id": concluido_row[0],
        "id_comando": str(concluido_row[1]),
        "status_comando": concluido_row[2],
        "tipo_resultado": concluido_row[3].get("tipo_resultado"),
        "agente_destino": "AGENTE_RDO",
        "agente_processador": "AGENTE_002_GERADOR_RDO",
        "telegram_chat_id": comando["telegram_chat_id"],
        "telegram_user_id": comando["telegram_user_id"],
        "telegram_message_id": comando["telegram_message_id"],
        "mensagem_resposta_executiva": montar_mensagem_resposta_executiva_rdo(
            comando["obra_codigo"],
            concluido_row[3].get("tipo_resultado"),
        ),
        "acoes_externas_executadas": False,
        "message": "Comando AGENTE_RDO processado sem ações externas.",
    }


@app.post("/agentes/comunicacao-obra/processar-comando")
async def processar_comando_agente_comunicacao_obra(
    payload: Optional[ProcessarComandoComunicacaoObraRequest] = None,
):
    id_comando = payload.id_comando if payload else None
    if id_comando:
        try:
            id_comando = str(uuid.UUID(id_comando))
        except ValueError as exc:
            raise HTTPException(
                status_code=422,
                detail={
                    "message": "id_comando inválido. Use um UUID válido.",
                    "error": str(exc),
                },
            )

    select_comando_sql = """
    SELECT
        ce.id,
        ce.id_comando,
        ce.tenant_id,
        ce.obra_codigo,
        ce.correlation_id,
        ce.agente_origem,
        ce.agente_destino,
        ce.tipo_comando,
        ce.payload_comando,
        ce.status,
        ce.requer_aprovacao,
        et.chat_id,
        et.telegram_user_id,
        et.telegram_message_id
    FROM comandos_executivos AS ce
    LEFT JOIN eventos_telegram AS et
        ON et.id = ce.evento_telegram_id
    WHERE ce.agente_destino = 'AGENTE_006_COMUNICACAO_VISUAL_OBRA'
      AND ce.tipo_comando = 'PREPARAR_PLACA_AVISO'
      AND ce.status = 'PENDENTE'
      AND (%(id_comando)s::uuid IS NULL OR ce.id_comando = %(id_comando)s::uuid)
    ORDER BY ce.criado_em
    LIMIT 1
    FOR UPDATE OF ce;
    """

    update_resultado_sql = """
    UPDATE comandos_executivos
    SET resultado = %(resultado)s,
        executado_por = 'AGENTE_006_COMUNICACAO_VISUAL_OBRA',
        executado_em = NOW(),
        mensagem_erro = NULL,
        atualizado_em = NOW()
    WHERE id = %(id)s
      AND status = 'PENDENTE'
      AND tipo_comando = 'PREPARAR_PLACA_AVISO'
    RETURNING id;
    """

    update_status_sql = """
    UPDATE comandos_executivos
    SET status = 'CONCLUIDO',
        atualizado_em = NOW()
    WHERE id = %(id)s
      AND status = 'PENDENTE'
      AND tipo_comando = 'PREPARAR_PLACA_AVISO'
      AND resultado IS NOT NULL
    RETURNING id, id_comando, status, resultado;
    """

    try:
        conn = get_db_connection()
        try:
            with conn.cursor() as cur:
                cur.execute(select_comando_sql, {"id_comando": id_comando})
                row = cur.fetchone()

                if row is None:
                    conn.rollback()
                    return {
                        "ok": True,
                        "processado": False,
                        "agente_destino": "AGENTE_006_COMUNICACAO_VISUAL_OBRA",
                        "message": (
                            "Nenhum comando PENDENTE PREPARAR_PLACA_AVISO para "
                            "AGENTE_006_COMUNICACAO_VISUAL_OBRA encontrado."
                        ),
                    }

                comando = {
                    "id": row[0],
                    "id_comando": row[1],
                    "tenant_id": row[2],
                    "obra_codigo": row[3],
                    "correlation_id": row[4],
                    "agente_origem": row[5],
                    "agente_destino": row[6],
                    "tipo_comando": row[7],
                    "payload_comando": row[8],
                    "status": row[9],
                    "requer_aprovacao": row[10],
                    "telegram_chat_id": row[11],
                    "telegram_user_id": row[12],
                    "telegram_message_id": row[13],
                }

                resultado = montar_resultado_agente_comunicacao_obra(comando)

                cur.execute(
                    update_resultado_sql,
                    {
                        "id": comando["id"],
                        "resultado": Json(resultado),
                    },
                )
                resultado_row = cur.fetchone()
                if resultado_row is None:
                    raise RuntimeError("Resultado do comando de comunicação visual não foi salvo.")

                cur.execute(update_status_sql, {"id": comando["id"]})
                concluido_row = cur.fetchone()
                if concluido_row is None:
                    raise RuntimeError(
                        "Status do comando de comunicação visual não foi atualizado para CONCLUIDO."
                    )

            conn.commit()

        except Exception:
            conn.rollback()
            raise

        finally:
            conn.close()

    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={
                "message": "Erro ao processar comando executivo do Agente 006 Comunicação Visual.",
                "error": str(exc),
            },
        )

    return {
        "ok": True,
        "processado": True,
        "comando_executivo_id": concluido_row[0],
        "id_comando": str(concluido_row[1]),
        "status_comando": concluido_row[2],
        "tipo_resultado": concluido_row[3].get("tipo_resultado"),
        "agente_destino": "AGENTE_006_COMUNICACAO_VISUAL_OBRA",
        "agente_processador": "AGENTE_006_COMUNICACAO_VISUAL_OBRA",
        "telegram_chat_id": comando["telegram_chat_id"],
        "telegram_user_id": comando["telegram_user_id"],
        "telegram_message_id": comando["telegram_message_id"],
        "mensagem_resposta_executiva": montar_mensagem_resposta_executiva_comunicacao_obra(
            comando["obra_codigo"],
            concluido_row[3].get("tipo_resultado"),
        ),
        "acoes_externas_executadas": False,
        "message": "Comando AGENTE_006_COMUNICACAO_VISUAL_OBRA processado sem ações externas.",
    }


@app.post("/agentes/comunicacao-obra/gerar-pdf-placa")
async def gerar_pdf_placa_agente_comunicacao_obra(
    payload: Optional[ProcessarComandoComunicacaoObraRequest] = None,
):
    id_comando = payload.id_comando if payload else None
    if id_comando:
        try:
            id_comando = str(uuid.UUID(id_comando))
        except ValueError as exc:
            raise HTTPException(
                status_code=422,
                detail={
                    "message": "id_comando inválido. Use um UUID válido.",
                    "error": str(exc),
                },
            )

    select_comando_sql = """
    SELECT
        ce.id,
        ce.id_comando,
        ce.tenant_id,
        ce.obra_codigo,
        ce.correlation_id,
        ce.agente_origem,
        ce.agente_destino,
        ce.tipo_comando,
        ce.payload_comando,
        ce.status,
        ce.requer_aprovacao,
        ce.criado_em,
        ce.atualizado_em,
        et.chat_id,
        et.telegram_user_id,
        et.telegram_message_id
    FROM comandos_executivos AS ce
    LEFT JOIN eventos_telegram AS et
        ON et.id = ce.evento_telegram_id
    WHERE ce.agente_destino = 'AGENTE_006_COMUNICACAO_VISUAL_OBRA'
      AND ce.tipo_comando = 'GERAR_PDF_PLACA_AVISO'
      AND ce.status = 'APROVADO'
      AND (%(id_comando)s::uuid IS NULL OR ce.id_comando = %(id_comando)s::uuid)
    ORDER BY ce.aprovado_em NULLS LAST, ce.criado_em
    LIMIT 1
    FOR UPDATE OF ce;
    """

    select_rascunho_origem_sql = """
    SELECT
        ce.id,
        ce.id_comando,
        ce.resultado,
        ce.criado_em,
        ce.atualizado_em
    FROM comandos_executivos AS ce
    WHERE ce.agente_destino = 'AGENTE_006_COMUNICACAO_VISUAL_OBRA'
      AND ce.tipo_comando = 'PREPARAR_PLACA_AVISO'
      AND ce.status = 'CONCLUIDO'
      AND ce.resultado->>'tipo_resultado' = 'RASCUNHO_PLACA_AVISO'
      AND ce.obra_codigo = %(obra_codigo)s
      AND (
          ce.criado_em <= %(comando_pdf_criado_em)s
          OR ce.atualizado_em <= %(comando_pdf_criado_em)s
      )
    ORDER BY
        CASE
            WHEN ce.atualizado_em <= %(comando_pdf_criado_em)s THEN ce.atualizado_em
            ELSE ce.criado_em
        END DESC,
        ce.criado_em DESC
    LIMIT 1;
    """

    update_resultado_evidencias_sql = """
    UPDATE comandos_executivos
    SET resultado = %(resultado)s,
        evidencias = COALESCE(evidencias, '[]'::jsonb) || %(evidencia)s::jsonb,
        executado_por = 'AGENTE_006_COMUNICACAO_VISUAL_OBRA',
        executado_em = NOW(),
        mensagem_erro = NULL,
        atualizado_em = NOW()
    WHERE id = %(id)s
      AND status = 'APROVADO'
      AND tipo_comando = 'GERAR_PDF_PLACA_AVISO'
    RETURNING id;
    """

    update_status_sql = """
    UPDATE comandos_executivos
    SET status = 'CONCLUIDO',
        atualizado_em = NOW()
    WHERE id = %(id)s
      AND status = 'APROVADO'
      AND tipo_comando = 'GERAR_PDF_PLACA_AVISO'
      AND resultado IS NOT NULL
    RETURNING id, id_comando, status, resultado;
    """

    try:
        conn = get_db_connection()
        try:
            with conn.cursor() as cur:
                cur.execute(select_comando_sql, {"id_comando": id_comando})
                row = cur.fetchone()

                if row is None:
                    conn.rollback()
                    return {
                        "ok": True,
                        "processado": False,
                        "comando_executivo_id": None,
                        "arquivo_pdf": None,
                        "tipo_resultado": None,
                        "mensagem_resposta_executiva": (
                            "Nenhum comando APROVADO GERAR_PDF_PLACA_AVISO para "
                            "AGENTE_006_COMUNICACAO_VISUAL_OBRA encontrado."
                        ),
                        "acoes_externas_executadas": False,
                    }

                comando = {
                    "id": row[0],
                    "id_comando": row[1],
                    "tenant_id": row[2],
                    "obra_codigo": row[3],
                    "correlation_id": row[4],
                    "agente_origem": row[5],
                    "agente_destino": row[6],
                    "tipo_comando": row[7],
                    "payload_comando": row[8],
                    "status": row[9],
                    "requer_aprovacao": row[10],
                    "criado_em": row[11],
                    "atualizado_em": row[12],
                    "telegram_chat_id": row[13],
                    "telegram_user_id": row[14],
                    "telegram_message_id": row[15],
                }

                cur.execute(
                    select_rascunho_origem_sql,
                    {
                        "obra_codigo": comando["obra_codigo"],
                        "comando_pdf_criado_em": comando["criado_em"],
                    },
                )
                rascunho_row = cur.fetchone()
                rascunho_origem = None
                if rascunho_row is not None:
                    rascunho_origem = {
                        "id": rascunho_row[0],
                        "id_comando": rascunho_row[1],
                        "resultado": rascunho_row[2],
                        "criado_em": rascunho_row[3],
                        "atualizado_em": rascunho_row[4],
                    }

                arquivo_pdf = gerar_pdf_local_placa_aviso(comando, rascunho_origem)
                resultado = montar_resultado_pdf_placa_aviso(
                    comando,
                    arquivo_pdf,
                    rascunho_origem,
                )
                evidencia = [
                    {
                        "tipo": "ARQUIVO_PDF_PLACA_AVISO",
                        "arquivo_pdf": arquivo_pdf,
                        "gerado_em": resultado["gerado_em"],
                        "rascunho_nao_oficial": True,
                        "comando_rascunho_origem_id": resultado[
                            "comando_rascunho_origem_id"
                        ],
                        "id_comando_rascunho_origem": resultado[
                            "id_comando_rascunho_origem"
                        ],
                    }
                ]

                cur.execute(
                    update_resultado_evidencias_sql,
                    {
                        "id": comando["id"],
                        "resultado": Json(resultado),
                        "evidencia": Json(evidencia),
                    },
                )
                resultado_row = cur.fetchone()
                if resultado_row is None:
                    raise RuntimeError("Resultado/evidências do PDF da placa não foram salvos.")

                cur.execute(update_status_sql, {"id": comando["id"]})
                concluido_row = cur.fetchone()
                if concluido_row is None:
                    raise RuntimeError(
                        "Status do comando GERAR_PDF_PLACA_AVISO não foi atualizado para CONCLUIDO."
                    )

            conn.commit()

        except Exception:
            conn.rollback()
            raise

        finally:
            conn.close()

    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={
                "message": "Erro ao gerar PDF local de placa de aviso pelo Agente 006.",
                "error": str(exc),
            },
        )

    tipo_resultado = concluido_row[3].get("tipo_resultado")
    return {
        "ok": True,
        "processado": True,
        "comando_executivo_id": concluido_row[0],
        "id_comando": str(concluido_row[1]),
        "status_comando": concluido_row[2],
        "arquivo_pdf": concluido_row[3].get("arquivo_pdf"),
        "tipo_resultado": tipo_resultado,
        "mensagem_resposta_executiva": montar_mensagem_resposta_executiva_comunicacao_obra(
            comando["obra_codigo"],
            tipo_resultado,
        ),
        "acoes_externas_executadas": False,
    }


@app.post("/agentes/classificar")
async def classificar_payload(request: Request):
    payload = await request.json()
    texto = json.dumps(payload, ensure_ascii=False).lower()

    if any(term in texto for term in ["nota fiscal", "nf-e", "nfe", "danfe"]):
        categoria = "NOTA_FISCAL"
    elif any(term in texto for term in ["orçamento", "cotação", "proposta"]):
        categoria = "ORCAMENTO_RECEBIDO"
    elif any(term in texto for term in ["comprar", "compra", "material", "faltou", "precisa cotar"]):
        categoria = "SOLICITACAO_COMPRA"
    elif any(term in texto for term in ["pendência", "problema", "atraso", "não executado"]):
        categoria = "PENDENCIA"
    elif any(term in texto for term in ["rdo", "hoje executamos", "equipe", "atividade"]):
        categoria = "RELATO_RDO"
    elif any(term in texto for term in ["foto", "imagem"]):
        categoria = "FOTO_OBRA"
    else:
        categoria = "MENSAGEM_GERAL"

    return {
        "ok": True,
        "categoria": categoria,
        "metodo": "rule_based_mvp",
        "observacao": "Classificador provisório sem LLM para teste inicial do api_core.",
    }


@app.get("/obras/{obra_codigo}/resumo-dia")
def resumo_dia(obra_codigo: str):
    sql = """
    SELECT
        COUNT(*) FILTER (WHERE status_processamento = 'RECEBIDA') AS recebidas,
        COUNT(*) FILTER (WHERE status_processamento = 'DUPLICADA') AS duplicadas,
        COUNT(*) AS total
    FROM mensagens_recebidas
    WHERE obra_codigo = %(obra_codigo)s
      AND criado_em::date = CURRENT_DATE;
    """

    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, {"obra_codigo": obra_codigo})
            row = cur.fetchone()

    return {
        "obra_codigo": obra_codigo,
        "data": datetime.now().date().isoformat(),
        "mensagens_recebidas": row[0] or 0,
        "mensagens_duplicadas": row[1] or 0,
        "total_mensagens": row[2] or 0,
    }


# -----------------------------------------------------------------------------
# Document upload / MinIO
# -----------------------------------------------------------------------------

MINIO_ENDPOINT = os.getenv("MINIO_ENDPOINT", "http://minio:9000")
MINIO_ACCESS_KEY = os.getenv("MINIO_ACCESS_KEY") or os.getenv("MINIO_ROOT_USER")
MINIO_SECRET_KEY = os.getenv("MINIO_SECRET_KEY") or os.getenv("MINIO_ROOT_PASSWORD")
MINIO_DEFAULT_BUCKET = os.getenv("MINIO_DEFAULT_BUCKET", "obra-caio")


def get_minio_client() -> Minio:
    if not MINIO_ACCESS_KEY or not MINIO_SECRET_KEY:
        raise RuntimeError("Credenciais MinIO não definidas.")

    parsed = urlparse(MINIO_ENDPOINT)

    if parsed.scheme:
        endpoint = parsed.netloc
        secure = parsed.scheme == "https"
    else:
        endpoint = MINIO_ENDPOINT.replace("http://", "").replace("https://", "")
        secure = False

    return Minio(
        endpoint,
        access_key=MINIO_ACCESS_KEY,
        secret_key=MINIO_SECRET_KEY,
        secure=secure,
    )


def sanitize_filename(filename: str | None) -> str:
    name = Path(filename or "anexo.bin").name
    name = re.sub(r"[^A-Za-z0-9._-]+", "_", name)
    return name or "anexo.bin"


def ensure_documentos_table() -> None:
    sql = """
    CREATE TABLE IF NOT EXISTS arquivos_obra (
        id SERIAL PRIMARY KEY,
        tenant_id TEXT NOT NULL,
        obra_codigo TEXT NOT NULL,
        mensagem_id INTEGER,
        canal_origem TEXT,
        remetente_identificador TEXT,
        origem_id TEXT,
        tipo_documento TEXT,
        file_name TEXT,
        mimetype TEXT,
        storage_provider TEXT DEFAULT 'minio',
        bucket TEXT NOT NULL,
        object_key TEXT NOT NULL,
        tamanho_bytes INTEGER,
        hash_arquivo TEXT,
        status_processamento TEXT DEFAULT 'ARMAZENADO',
        payload_metadata JSONB,
        criado_em TIMESTAMP DEFAULT NOW()
    );

    CREATE INDEX IF NOT EXISTS idx_arquivos_obra_codigo
        ON arquivos_obra (obra_codigo);

    CREATE INDEX IF NOT EXISTS idx_arquivos_mensagem_id
        ON arquivos_obra (mensagem_id);

    CREATE INDEX IF NOT EXISTS idx_arquivos_hash
        ON arquivos_obra (hash_arquivo);
    """

    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(sql)
        conn.commit()


@app.post("/documentos/upload")
async def upload_documento(
    file: UploadFile = File(...),
    tenant_id: str = Form(default="construtora-piloto"),
    obra_codigo: str = Form(default="OBRA-CAIO"),
    canal_origem: str = Form(default="desconhecido"),
    mensagem_id: Optional[int] = Form(default=None),
    remetente_identificador: Optional[str] = Form(default=None),
    origem_id: Optional[str] = Form(default=None),
    tipo_documento: Optional[str] = Form(default=None),
):
    ensure_documentos_table()

    content = await file.read()

    if not content:
        raise HTTPException(
            status_code=400,
            detail={"message": "Arquivo vazio ou não recebido."},
        )

    file_hash = hashlib.sha256(content).hexdigest()
    file_size = len(content)
    safe_name = sanitize_filename(file.filename)
    mimetype = file.content_type or "application/octet-stream"

    bucket = MINIO_DEFAULT_BUCKET
    date_path = datetime.now(timezone.utc).strftime("%Y/%m/%d")
    object_key = (
        f"{tenant_id}/{obra_codigo}/{canal_origem}/"
        f"{date_path}/{file_hash[:12]}_{safe_name}"
    )

    try:
        client = get_minio_client()

        if not client.bucket_exists(bucket):
            client.make_bucket(bucket)

        client.put_object(
            bucket_name=bucket,
            object_name=object_key,
            data=BytesIO(content),
            length=file_size,
            content_type=mimetype,
        )

    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={
                "message": "Erro ao salvar arquivo no MinIO.",
                "error": str(exc),
                "bucket": bucket,
                "object_key": object_key,
            },
        )

    insert_sql = """
    INSERT INTO arquivos_obra (
        tenant_id,
        obra_codigo,
        mensagem_id,
        canal_origem,
        remetente_identificador,
        origem_id,
        tipo_documento,
        file_name,
        mimetype,
        storage_provider,
        bucket,
        object_key,
        tamanho_bytes,
        hash_arquivo,
        status_processamento,
        payload_metadata
    )
    VALUES (
        %(tenant_id)s,
        %(obra_codigo)s,
        %(mensagem_id)s,
        %(canal_origem)s,
        %(remetente_identificador)s,
        %(origem_id)s,
        %(tipo_documento)s,
        %(file_name)s,
        %(mimetype)s,
        'minio',
        %(bucket)s,
        %(object_key)s,
        %(tamanho_bytes)s,
        %(hash_arquivo)s,
        'ARMAZENADO',
        %(payload_metadata)s
    )
    RETURNING id;
    """

    metadata = {
        "original_filename": file.filename,
        "content_type": mimetype,
        "hash_sha256": file_hash,
    }

    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    insert_sql,
                    {
                        "tenant_id": tenant_id,
                        "obra_codigo": obra_codigo,
                        "mensagem_id": mensagem_id,
                        "canal_origem": canal_origem,
                        "remetente_identificador": remetente_identificador,
                        "origem_id": origem_id,
                        "tipo_documento": tipo_documento,
                        "file_name": safe_name,
                        "mimetype": mimetype,
                        "bucket": bucket,
                        "object_key": object_key,
                        "tamanho_bytes": file_size,
                        "hash_arquivo": file_hash,
                        "payload_metadata": Json(metadata),
                    },
                )
                row = cur.fetchone()
            conn.commit()

    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={
                "message": "Arquivo salvo no MinIO, mas houve erro ao registrar no banco.",
                "error": str(exc),
                "bucket": bucket,
                "object_key": object_key,
            },
        )

    documento_id = row[0]

    return {
        "ok": True,
        "documento_id": documento_id,
        "tenant_id": tenant_id,
        "obra_codigo": obra_codigo,
        "mensagem_id": mensagem_id,
        "anexo": {
            "documento_id": documento_id,
            "tipo": tipo_documento or "documento",
            "file_name": safe_name,
            "mimetype": mimetype,
            "storage_provider": "minio",
            "bucket": bucket,
            "object_key": object_key,
            "tamanho_bytes": file_size,
            "hash_arquivo": file_hash,
            "status_processamento": "ARMAZENADO",
        },
    }
