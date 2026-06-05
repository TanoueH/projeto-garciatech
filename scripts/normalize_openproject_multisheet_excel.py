#!/usr/bin/env python3
"""
Normaliza cronograma Excel multiabas para importação segura no OpenProject.

Uso:
    python scripts/normalize_openproject_multisheet_excel.py \
      --arquivo "data/import/openproject/atualizacoes_excel/ANEXO IV - Cronograma_SUM_CAIO_Rev_09 - OBRA.xlsx" \
      --obra-codigo OBRA-001 \
      --saida data/staging/openproject \
      --relatorios outputs/reports/openproject

Saídas:
    - CSV consolidado normalizado
    - Excel de relatório dry-run
    - JSON de validação
"""

import argparse
import hashlib
import json
import re
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


SHEET_CONFIG = {
    "CRONOGRAMA META": {
        "area_codigo": "META",
        "area_cronograma": "Cronograma Meta",
    },
    "CRONOGRAMA FASEADO - REFEITÓRIO": {
        "area_codigo": "REFEITORIO",
        "area_cronograma": "Refeitório",
    },
    "CRONOGRAMA FASEADO - S. JOGOS": {
        "area_codigo": "SALA_JOGOS",
        "area_cronograma": "Sala de Jogos",
    },
}

# A tabela principal está em B:I, com cabeçalho na linha 10 e dados a partir da linha 13.
FIRST_DATA_ROW = 13
COL_NO = 2          # B
COL_TIPO = 3        # C
COL_DESCRICAO = 4   # D
COL_STATUS = 5      # E
COL_RESP = 6        # F
COL_INICIO = 7      # G
COL_FIM = 8         # H
COL_PRAZO = 9       # I

STATUS_MAP = {
    "planejado": "Planejado",
    "em andamento": "Em andamento",
    "andamento": "Em andamento",
    "concluido": "Concluído",
    "concluído": "Concluído",
    "finalizado": "Concluído",
    "atrasado": "Atrasado",
    "atrasada": "Atrasado",
    "paralisado": "Paralisado",
    "paralisada": "Paralisado",
    "cancelado": "Cancelado",
    "cancelada": "Cancelado",
}

TIPO_MAP = {
    "etapa": "Etapa",
    "servico": "Serviço",
    "serviço": "Serviço",
}

RESPONSAVEL_MAP = {
    "CAO": "CAIO",  # correção observada na planilha
}


def clean_text(value: Any) -> str:
    """Converte valores de Excel para texto limpo."""
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def slugify(value: str, max_len: int = 40) -> str:
    """Cria slug simples, sem acentos especiais, adequado para código interno."""
    value = clean_text(value).upper()
    replacements = {
        "Á": "A", "À": "A", "Â": "A", "Ã": "A", "Ä": "A",
        "É": "E", "È": "E", "Ê": "E", "Ë": "E",
        "Í": "I", "Ì": "I", "Î": "I", "Ï": "I",
        "Ó": "O", "Ò": "O", "Ô": "O", "Õ": "O", "Ö": "O",
        "Ú": "U", "Ù": "U", "Û": "U", "Ü": "U",
        "Ç": "C",
    }
    for old, new in replacements.items():
        value = value.replace(old, new)
    value = re.sub(r"[^A-Z0-9]+", "-", value)
    value = value.strip("-")
    return value[:max_len].strip("-") or "SEM-DESCRICAO"


def short_hash(*values: Any, size: int = 8) -> str:
    raw = "|".join(clean_text(v) for v in values)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:size].upper()


def normalize_status(value: Any) -> Tuple[str, Optional[str]]:
    original = clean_text(value)
    if not original:
        return "", "STATUS_VAZIO"
    normalized = STATUS_MAP.get(original.lower())
    if normalized:
        return normalized, None
    return original, "STATUS_NAO_MAPEADO"


def normalize_tipo(value: Any, descricao: str) -> Tuple[str, Optional[str]]:
    original = clean_text(value)
    if not original:
        # Se existe descrição, assumimos Serviço, mas registramos aviso.
        if descricao:
            return "Serviço", "TIPO_INFERIDO_COMO_SERVICO"
        return "", "TIPO_VAZIO"

    normalized = TIPO_MAP.get(original.lower())
    if normalized:
        return normalized, None

    # Erro comum observado: coluna Tipo preenchida acidentalmente com número da atividade.
    # Ex.: No. = 14.1 e Tipo = 14.2. Neste caso, mantemos a linha como Serviço e registramos aviso.
    if descricao and re.fullmatch(r"\d+(\.\d+)*", original):
        return "Serviço", f"TIPO_CORRIGIDO_DE_{original}_PARA_SERVICO"

    return original, "TIPO_NAO_MAPEADO"


def normalize_responsavel(value: Any) -> Tuple[str, Optional[str]]:
    original = clean_text(value)
    if not original:
        return "", None
    normalized = RESPONSAVEL_MAP.get(original.upper(), original)
    if normalized != original:
        return normalized, f"RESPONSAVEL_CORRIGIDO_DE_{original}_PARA_{normalized}"
    return normalized, None


def excel_date_to_iso(value: Any) -> Tuple[str, Optional[str]]:
    """Converte data vinda do Excel para YYYY-MM-DD."""
    if value is None or clean_text(value) == "":
        return "", "DATA_VAZIA"

    if isinstance(value, datetime):
        return value.date().isoformat(), None

    if isinstance(value, date):
        return value.isoformat(), None

    parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return "", f"DATA_INVALIDA: {clean_text(value)}"

    return parsed.date().isoformat(), None


def safe_int(value: Any) -> Tuple[Optional[int], Optional[str]]:
    if value is None or clean_text(value) == "":
        return None, None
    try:
        return int(round(float(value))), None
    except Exception:
        return None, f"PRAZO_INVALIDO: {clean_text(value)}"


def build_codigo(
    obra_codigo: str,
    area_codigo: str,
    tipo_item: str,
    no_original: str,
    descricao: str,
    linha_excel: int,
) -> str:
    """
    Código interno sugerido para OpenProject.

    Observação: usamos linha_excel no código para garantir unicidade nesta primeira carga.
    Também incluímos um hash curto para rastrear descrição/tipo/no_original.
    """
    tipo_code = "ETP" if tipo_item == "Etapa" else "SRV"
    no_slug = slugify(no_original, max_len=12)
    desc_hash = short_hash(area_codigo, tipo_item, no_original, descricao, linha_excel, size=6)
    return f"{obra_codigo}-CAIO-{area_codigo}-{tipo_code}-{linha_excel:03d}-{no_slug}-{desc_hash}"


def detect_revisao(wb) -> str:
    """Tenta extrair texto de atualização da célula B6 da primeira aba."""
    try:
        first_sheet = wb[wb.sheetnames[0]]
        revisao = clean_text(first_sheet.cell(6, 2).value)
        return revisao or "REVISAO_NAO_IDENTIFICADA"
    except Exception:
        return "REVISAO_NAO_IDENTIFICADA"


def process_sheet(ws, sheet_name: str, config: Dict[str, str], obra_codigo: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    rows: List[Dict[str, Any]] = []
    issues: List[Dict[str, Any]] = []

    area_codigo = config["area_codigo"]
    area_cronograma = config["area_cronograma"]
    current_parent_codigo = ""
    current_parent_descricao = ""
    ordem_na_aba = 0

    for linha_excel in range(FIRST_DATA_ROW, ws.max_row + 1):
        no_original = clean_text(ws.cell(linha_excel, COL_NO).value)
        descricao = clean_text(ws.cell(linha_excel, COL_DESCRICAO).value)

        # Ignora linhas completamente vazias da tabela principal.
        if not no_original and not descricao:
            continue

        ordem_na_aba += 1

        tipo_item, tipo_issue = normalize_tipo(ws.cell(linha_excel, COL_TIPO).value, descricao)
        status_original = clean_text(ws.cell(linha_excel, COL_STATUS).value)
        status_openproject, status_issue = normalize_status(status_original)
        responsavel, resp_issue = normalize_responsavel(ws.cell(linha_excel, COL_RESP).value)
        data_inicio, inicio_issue = excel_date_to_iso(ws.cell(linha_excel, COL_INICIO).value)
        data_fim, fim_issue = excel_date_to_iso(ws.cell(linha_excel, COL_FIM).value)
        prazo_dias_uteis, prazo_issue = safe_int(ws.cell(linha_excel, COL_PRAZO).value)

        codigo_atividade = build_codigo(
            obra_codigo=obra_codigo,
            area_codigo=area_codigo,
            tipo_item=tipo_item,
            no_original=no_original,
            descricao=descricao,
            linha_excel=linha_excel,
        )

        parent_codigo = ""
        parent_descricao = ""
        if tipo_item == "Etapa":
            current_parent_codigo = codigo_atividade
            current_parent_descricao = descricao
        elif tipo_item == "Serviço":
            parent_codigo = current_parent_codigo
            parent_descricao = current_parent_descricao

        linha_issues = []

        if not no_original:
            linha_issues.append("NO_ORIGINAL_VAZIO")
        if not descricao:
            linha_issues.append("DESCRICAO_VAZIA")
        if tipo_issue:
            linha_issues.append(tipo_issue)
        if status_issue:
            linha_issues.append(status_issue)
        if inicio_issue:
            linha_issues.append(inicio_issue)
        if fim_issue:
            linha_issues.append(fim_issue)
        if prazo_issue:
            linha_issues.append(prazo_issue)
        if resp_issue:
            linha_issues.append(resp_issue)

        # Checagem de data invertida.
        if data_inicio and data_fim:
            dt_ini = pd.to_datetime(data_inicio)
            dt_fim = pd.to_datetime(data_fim)
            if dt_ini > dt_fim:
                linha_issues.append("DATA_INICIO_MAIOR_QUE_DATA_FIM")

        # Serviços sem etapa-pai anterior.
        if tipo_item == "Serviço" and not parent_codigo:
            linha_issues.append("SERVICO_SEM_ETAPA_PAI_ANTERIOR")

        # Etapas normalmente não têm responsável; não é erro.
        if tipo_item == "Serviço" and not responsavel:
            linha_issues.append("SERVICO_SEM_RESPONSAVEL")

        severity = "OK"
        if any(i.startswith("DATA_INVALIDA") or i in {
            "DESCRICAO_VAZIA",
            "DATA_INICIO_MAIOR_QUE_DATA_FIM",
            "SERVICO_SEM_ETAPA_PAI_ANTERIOR",
        } for i in linha_issues):
            severity = "ERRO"
        elif linha_issues:
            severity = "AVISO"

        row = {
            "codigo_atividade": codigo_atividade,
            "obra_codigo": obra_codigo,
            "projeto_nome": "CAIO - Área de convivência",
            "aba_origem": sheet_name,
            "area_codigo": area_codigo,
            "area_cronograma": area_cronograma,
            "linha_excel": linha_excel,
            "ordem_na_aba": ordem_na_aba,
            "no_original": no_original,
            "tipo_item": tipo_item,
            "atividade": descricao,
            "status_original": status_original,
            "status_openproject": status_openproject,
            "responsavel_original": clean_text(ws.cell(linha_excel, COL_RESP).value),
            "responsavel_normalizado": responsavel,
            "data_inicio": data_inicio,
            "data_fim": data_fim,
            "prazo_dias_uteis": prazo_dias_uteis,
            "parent_codigo_sugerido": parent_codigo,
            "parent_descricao_sugerida": parent_descricao,
            "nivel_importacao": "PAI" if tipo_item == "Etapa" else "FILHO",
            "acao_sugerida": "CRIAR_OU_ATUALIZAR",
            "validacao_status": severity,
            "validacao_observacoes": "; ".join(linha_issues),
        }
        rows.append(row)

        if linha_issues:
            issues.append({
                "aba_origem": sheet_name,
                "linha_excel": linha_excel,
                "codigo_atividade": codigo_atividade,
                "tipo": severity,
                "atividade": descricao,
                "observacoes": linha_issues,
            })

    return rows, issues


def normalize_workbook(input_file: Path, obra_codigo: str, out_dir: Path, report_dir: Path) -> Dict[str, Any]:
    out_dir.mkdir(parents=True, exist_ok=True)
    report_dir.mkdir(parents=True, exist_ok=True)

    # data_only=True lê os valores calculados salvos na planilha, não as fórmulas.
    wb = load_workbook(input_file, data_only=True)
    revisao = detect_revisao(wb)

    all_rows: List[Dict[str, Any]] = []
    all_issues: List[Dict[str, Any]] = []
    sheet_summary: List[Dict[str, Any]] = []

    for sheet_name, config in SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            all_issues.append({
                "aba_origem": sheet_name,
                "linha_excel": None,
                "codigo_atividade": "",
                "tipo": "ERRO",
                "atividade": "",
                "observacoes": ["ABA_NAO_ENCONTRADA"],
            })
            continue

        rows, issues = process_sheet(wb[sheet_name], sheet_name, config, obra_codigo)
        all_rows.extend(rows)
        all_issues.extend(issues)

        sheet_summary.append({
            "aba_origem": sheet_name,
            "area_codigo": config["area_codigo"],
            "area_cronograma": config["area_cronograma"],
            "total_linhas": len(rows),
            "total_etapas": sum(1 for r in rows if r["tipo_item"] == "Etapa"),
            "total_servicos": sum(1 for r in rows if r["tipo_item"] == "Serviço"),
            "total_erros": sum(1 for r in rows if r["validacao_status"] == "ERRO"),
            "total_avisos": sum(1 for r in rows if r["validacao_status"] == "AVISO"),
            "inicio_min": min([r["data_inicio"] for r in rows if r["data_inicio"]], default=""),
            "fim_max": max([r["data_fim"] for r in rows if r["data_fim"]], default=""),
        })

    df = pd.DataFrame(all_rows)
    issues_df = pd.DataFrame(all_issues)
    summary_df = pd.DataFrame(sheet_summary)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = out_dir / f"cronograma_caio_rev09_normalizado_{timestamp}.csv"
    json_path = report_dir / f"relatorio_cronograma_caio_rev09_validacao_{timestamp}.json"
    xlsx_path = report_dir / f"relatorio_cronograma_caio_rev09_dryrun_{timestamp}.xlsx"

    if not df.empty:
        df.insert(3, "revisao_cronograma", revisao)
        df.insert(4, "arquivo_origem", input_file.name)
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    else:
        csv_path.write_text("", encoding="utf-8")

    total_erros = int((df["validacao_status"] == "ERRO").sum()) if not df.empty else 0
    total_avisos = int((df["validacao_status"] == "AVISO").sum()) if not df.empty else 0

    report = {
        "arquivo_origem": str(input_file),
        "arquivo_origem_nome": input_file.name,
        "obra_codigo": obra_codigo,
        "revisao_cronograma": revisao,
        "processado_em": datetime.now().isoformat(timespec="seconds"),
        "abas_processadas": list(SHEET_CONFIG.keys()),
        "total_atividades": int(len(df)),
        "total_etapas": int((df["tipo_item"] == "Etapa").sum()) if not df.empty else 0,
        "total_servicos": int((df["tipo_item"] == "Serviço").sum()) if not df.empty else 0,
        "total_erros": total_erros,
        "total_avisos": total_avisos,
        "status_geral": "BLOQUEADO_PARA_APPLY" if total_erros > 0 else "APTO_PARA_DRYRUN",
        "csv_normalizado": str(csv_path),
        "relatorio_excel": str(xlsx_path),
        "relatorio_json": str(json_path),
    }

    json_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Resumo", index=False)
        df.to_excel(writer, sheet_name="Cronograma_Normalizado", index=False)
        if not issues_df.empty:
            issues_df.to_excel(writer, sheet_name="Inconsistencias", index=False)
        else:
            pd.DataFrame([{"status": "Sem inconsistências identificadas."}]).to_excel(
                writer, sheet_name="Inconsistencias", index=False
            )

    return report


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Normaliza cronograma Excel multiabas para OpenProject."
    )
    parser.add_argument("--arquivo", required=True, help="Caminho do arquivo Excel multiabas")
    parser.add_argument("--obra-codigo", default="OBRA-001", help="Código interno da obra")
    parser.add_argument("--saida", default="data/staging/openproject", help="Pasta de saída do CSV")
    parser.add_argument("--relatorios", default="outputs/reports/openproject", help="Pasta de relatórios")

    args = parser.parse_args()

    input_file = Path(args.arquivo)
    if not input_file.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {input_file}")

    report = normalize_workbook(
        input_file=input_file,
        obra_codigo=args.obra_codigo,
        out_dir=Path(args.saida),
        report_dir=Path(args.relatorios),
    )

    print(json.dumps(report, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
